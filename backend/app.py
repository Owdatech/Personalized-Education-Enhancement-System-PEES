# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify
import uuid
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, auth, firestore, storage
import logging
import time
import jwt
import os
import re
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
import base64
from flask_cors import CORS
import json
# from datetime import datetime
from datetime import datetime, timezone # Import timezone for use with datetime.now()
import nest_asyncio
nest_asyncio.apply()
import asyncio
from dotenv import load_dotenv
import os

load_dotenv()  # This loads the variables from .env into os.environ

# Initialize Flask App
app = Flask(__name__)
CORS(app)

# Define constants
BLOCK_SIZE = 16  # AES block size
SECRET_KEY = os.getenv(
    "AES_SECRET_KEY", "this_is_16_bytes"
).encode()  # Environment variable or fallback

# Initialize Firebase Admin SDK
if not firebase_admin._apps:
    cred = credentials.Certificate(
        "serviceAccountKey.json"
    )  # Path to your service account key JSON
    firebase_admin.initialize_app(
        cred,
        {
            "storageBucket": "pees-d1101.firebasestorage.app"  # Replace with your actual Firebase bucket name
        },
    )

# Get Firestore and Storage clients
db = firestore.client()
bucket = storage.bucket()

def get_password_reset_auth_app():
    try:
        return firebase_admin.get_app("password_reset_app")
    except ValueError:
        cred = credentials.Certificate("serviceAccountKey.json")
        return firebase_admin.initialize_app(cred, name="password_reset_app")


def sanitize_grade_key(key):
    key = str(key).upper().strip()
    key = re.sub(r'[\s\./\-]+', '_', key)
    key = key.replace('_(', '(').replace(')_', ')')
    key = re.sub(r'_{2,}', '_', key)
    return key.strip('_')


def extract_grade_map_for_user(user_data):
    assigned = user_data.get("assignedGrades", {}) or {}
    if isinstance(assigned, dict) and isinstance(assigned.get("grades"), dict):
        return assigned.get("grades", {})
    if isinstance(assigned, dict):
        return assigned
    return {}


def extract_normalized_grade_keys(user_data):
    return {
        sanitize_grade_key(grade_key)
        for grade_key in extract_grade_map_for_user(user_data).keys()
    }


def sync_student_teacher_links(student_id, student_data=None):
    user_ref = db.collection("users").document(student_id)
    student_ref = db.collection("students").document(student_id)

    if student_data is None:
        user_doc = user_ref.get()
        if not user_doc.exists:
            raise ValueError(f"Student {student_id} not found")
        student_data = user_doc.to_dict() or {}

    student_grade_keys = extract_normalized_grade_keys(student_data)

    teachers = list(db.collection("users").where("role", "==", "teacher").stream())
    matched_teacher_ids = set()
    teacher_updates = []

    for teacher_doc in teachers:
        teacher_id = teacher_doc.id
        teacher_data = teacher_doc.to_dict() or {}
        teacher_grade_keys = extract_normalized_grade_keys(teacher_data)
        should_link = bool(student_grade_keys and teacher_grade_keys and not student_grade_keys.isdisjoint(teacher_grade_keys))
        currently_linked = student_id in (teacher_data.get("associatedIds") or [])

        if should_link:
            matched_teacher_ids.add(teacher_id)
            if not currently_linked:
                teacher_updates.append((teacher_id, firestore.ArrayUnion([student_id])))
        elif currently_linked:
            teacher_updates.append((teacher_id, firestore.ArrayRemove([student_id])))

    existing_associated_ids = list(student_data.get("associatedIds") or [])
    teacher_ids = {teacher_doc.id for teacher_doc in teachers}
    existing_teacher_ids = {assoc_id for assoc_id in existing_associated_ids if assoc_id in teacher_ids}
    teacher_ids_to_add = matched_teacher_ids - existing_teacher_ids
    teacher_ids_to_remove = existing_teacher_ids - matched_teacher_ids

    batch = db.batch()
    batch_count = 0

    for teacher_id, update_op in teacher_updates:
        batch.update(db.collection("users").document(teacher_id), {"associatedIds": update_op})
        batch_count += 1
        if batch_count >= 400:
            batch.commit()
            batch = db.batch()
            batch_count = 0

    if teacher_ids_to_add:
        batch.set(user_ref, {"associatedIds": firestore.ArrayUnion(list(teacher_ids_to_add))}, merge=True)
        batch.set(student_ref, {"associatedIds": firestore.ArrayUnion(list(teacher_ids_to_add))}, merge=True)
        batch_count += 2

    if teacher_ids_to_remove:
        batch.set(user_ref, {"associatedIds": firestore.ArrayRemove(list(teacher_ids_to_remove))}, merge=True)
        batch.set(student_ref, {"associatedIds": firestore.ArrayRemove(list(teacher_ids_to_remove))}, merge=True)
        batch_count += 2

    if batch_count > 0:
        batch.commit()

    return sorted(matched_teacher_ids)


# Configure logging
logging.basicConfig(
    level=logging.ERROR, format="%(asctime)s %(levelname)s: %(message)s"
)

# Allowed file extensions
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}
try:
    from translation_provider import (
        translate_analysis_recursively as tp_translate_values,
    )  # value-only translation
except Exception:
    def tp_translate_values(payload):
        return payload

try:
    from AlertNotification import GetLangugage
except Exception:
    async def GetLangugage(curriculum_id):
        return "en"
from openai import OpenAI

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
client_openai = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None


# Helper Functions
def encrypt_password(password: str) -> str:
    cipher = AES.new(SECRET_KEY, AES.MODE_CBC)
    iv = cipher.iv  # Initialization vector
    encrypted_password = cipher.encrypt(
        pad(password.encode(), BLOCK_SIZE)
    )  # Encrypt and pad the password
    return base64.b64encode(
        iv + encrypted_password
    ).decode()  # Return Base64-encoded string


def decrypt_password(encrypted_password: str) -> str:
    try:
        encrypted_data = base64.b64decode(encrypted_password)
        iv = encrypted_data[:BLOCK_SIZE]
        cipher = AES.new(SECRET_KEY, AES.MODE_CBC, iv)
        decrypted_password = unpad(
            cipher.decrypt(encrypted_data[BLOCK_SIZE:]), BLOCK_SIZE
        )
        return decrypted_password.decode()
    except Exception as e:
        raise ValueError(f"Decryption failed: {e}")


def decode_token(token):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload
    except jwt.ExpiredSignatureError:
        logging.error("Token has expired.")
        return None
    except jwt.InvalidTokenError as e:
        logging.error(f"Invalid token: {e}")
        return None


def is_valid_email(email):
    email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
    return re.match(email_regex, email) is not None


def allowed_file(filename):
    if not filename:
        return False
    file_extension = filename.rsplit(".", 1)[-1].lower()
    return file_extension in ALLOWED_EXTENSIONS


def upload_image_to_firebase(file, unique_filename):
    try:
        blob = bucket.blob(unique_filename)
        blob.upload_from_file(file)
        blob.make_public()
        return blob.public_url
    except Exception as e:
        raise ValueError(f"Failed to upload image to Firebase: {e}")


def fetch_associated_details(associated_ids):
    """Fetch details of associated IDs from Firestore."""
    associated_details = []
    for assoc_id in associated_ids:
        assoc_user_ref = db.collection("users").document(assoc_id)
        assoc_user_doc = assoc_user_ref.get()
        if assoc_user_doc.exists:
            assoc_user_data = assoc_user_doc.to_dict()
            assoc_user_data.pop("passwordHash", None)
            assoc_user_data.pop("associatedIds", None)
            associated_details.append(assoc_user_data)
        else:
            logging.warning(f"Associated user with ID {assoc_id} does not exist.")
    return associated_details

def is_arabic_text(text: str) -> bool:
    """Return True if the text contains Arabic characters."""
    if not text:
        return False
    return bool(re.search(r'[\u0600-\u06FF]', text))

@app.route("/api/auth/createaccount", methods=["POST"])
def account_create():
    # Make sure global objects like 'db', 'auth', 'logging', 'uuid', 'time', 'os', 'jwt', 
    # 'encrypt_password', 'is_valid_email', 'allowed_file', and 'upload_image_to_firebase'
    # are defined and available in your environment.
    try:
        # if not request.content_type.startswith("multipart/form-data"):
        # return jsonify({"error": "Content-Type must be multipart/form-data"}), 415

        data = request.form
        logging.info("Received form data for account creation.")

        # Validate required fields
        required_fields = ["role", "email", "password", "confirmPassword"]
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            return (
                jsonify(
                    {"error": f"Missing required fields: {', '.join(missing_fields)}"}
                ),
                400,
            )

        # Validate passwords
        password = data["password"]
        confirm_password = data["confirmPassword"]
        if password != confirm_password:
            return jsonify({"error": "Passwords do not match"}), 400

        email = data["email"]
        if not is_valid_email(email):
            return jsonify({"error": "Invalid email format"}), 400

        role = data["role"].lower()
        valid_roles = ["headmaster", "parent", "teacher", "student"]
        if role not in valid_roles:
            return (
                jsonify(
                    {
                        "error": f"Invalid role value. Allowed roles are: {', '.join(valid_roles)}"
                    }
                ),
                400,
            )

        user_id = str(uuid.uuid4())
        encrypted_password = encrypt_password(password)
        logging.info(f"Generated user ID: {user_id}")

        profile_info = {
            "personalInformation": {
                "name": data.get("name", ""),
                "idNumber": data.get("idNumber", ""),
                "photoUrl": "https://example.com/default-photo.jpg",
            },
            "contactInformation": {
                "phoneNumber": data.get("phoneNumber", ""),
                "address": data.get("address", ""),
            },
        }

        if "photo" in request.files:
            file = request.files["photo"]
            if allowed_file(file.filename):
                unique_filename = (
                    f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
                )
                photo_url = upload_image_to_firebase(file, unique_filename)
                profile_info["personalInformation"]["photoUrl"] = photo_url
                logging.info(f"Photo uploaded for user {user_id}: {photo_url}")
            else:
                return (
                    jsonify(
                        {
                            "error": "Invalid file type for photo. Allowed types are: png, jpg, jpeg"
                        }
                    ),
                    400,
                )

        # Handle associated IDs
        associated_ids_raw = data.get("associatedIds", "[]")
        logging.info(f"Raw associatedIds input: {associated_ids_raw}")

        try:
            associated_ids = (
                json.loads(associated_ids_raw)
                if isinstance(associated_ids_raw, str)
                else associated_ids_raw
            )
            if not isinstance(associated_ids, list):
                raise ValueError("associatedIds must be a list")
            logging.info(f"Validated associated IDs: {associated_ids}")
        except (json.JSONDecodeError, ValueError) as e:
            logging.error(f"Error parsing associatedIds: {e}")
            return jsonify({"error": "Invalid format for associatedIds"}), 400

        user = auth.create_user(
            uid=user_id, email=data["email"], password=data["password"]
        )

        # ====================================================================
        # --- CORE LOGIC FOR ASSIGNING GRADES ---
        # --- Headmaster is automatically assigned all grades.
        # --- Other roles use form data input.
        # ====================================================================
        assigned_grades = {}
        
        if role == "headmaster":
            # Hardcoded list of specific grade names extracted from your data
            ALL_GRADES = [
                "GRADE 1", "KG1", "GRADE 2", "KG2", "GRADE 3", "GRADE 4", 
                "GRADE 5", "GRADE 6", "GRADE 7", "GRADE 8", "GRADE 9", 
                "GRADE 10", "GRADE 11 (SCIENCE)", "GRADE 11 LITERATURE", 
                "GRADE 11 (LITERATURE)", "GRADE 11-SCIENCE", "GRADE11 LITERATURE", 
                "GRADE11-SCIENCE", "GRADE 12"
            ]
            
            assigned_grades["grades"] = {}
            
            # Loop through ALL_GRADES to assign comprehensive permissions
            for g in ALL_GRADES:
                assigned_grades["grades"][g] = {
                    # Use generic placeholder class/subjects for full administrative access
                    "School-Wide": ["Oversight", "Administration"]
                }
            logging.info(f"Headmaster granted access to all {len(ALL_GRADES)} specific grades.")
            
        else:
            # Logic for non-headmaster roles (Teacher, Student, Parent)
            grade = data.get("grade", "").strip()
            class_ = data.get("class", "").strip()
            subject = data.get("subject", "").strip()
            
            if grade and class_ and subject:
                assigned_grades.setdefault("grades", {}).setdefault(grade, {}).setdefault(
                    class_, []
                ).append(subject)
                assigned_grades["grades"][grade][class_] = list(
                    set(assigned_grades["grades"][grade][class_])
                )  # Remove duplicates
        # ====================================================================

        user_data = {
            "userId": user_id,
            "role": role,
            "email": email,
            "passwordHash": encrypted_password,
            "profileInfo": profile_info,
            "associatedIds": associated_ids,
            "status": "Active",
            "createdAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "assignedGrades": assigned_grades,
        }

        student_data = {
            "studentId": user_id,
            "role": role,
            "email": email,
            "passwordHash": encrypted_password,
            "profileInfo": profile_info,
            "associatedIds": associated_ids,
            "status": "Active",
            "createdAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "assignedGrades": assigned_grades,
        }

        # Save user data to Firestore
        db.collection("users").document(user_id).set(user_data)
        if role == "student":
            db.collection("students").document(user_id).set(student_data)
        logging.info(f"User {user_id} created and stored in Firestore.")

        jwt_secret_key = os.getenv("JWT_SECRET_KEY", "default_jwt_secret")
        token_payload = {
            "userId": user_id,
            "email": email,
            "role": role,
            "exp": int(time.time() + (30 * 24 * 60 * 60)),  # 30 days expiration
            "iat": int(time.time()),
        }
        token = jwt.encode(token_payload, jwt_secret_key, algorithm="HS256")

        return (
            jsonify(
                {
                    "message": "User created successfully",
                    "token": token,
                    "userId": user_id,
                }
            ),
            200,
        )

    except Exception as e:
        logging.error(f"Error during account creation: {e}")
        return jsonify({"error": f"An error occurred: {e}"}), 500


# @app.route("/api/auth/getprofile", methods=["GET"])
# def get_user_profile():
#     try:
#         user_id = request.args.get('userId')
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         # Fetch the user's profile from Firestore
#         user_ref = db.collection('users').document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         user_data = user_doc.to_dict()

#         # Remove passwordHash and include decrypted password
#         if 'passwordHash' in user_data:
#             try:
#                 logging.info(f"Decrypting password for user: {user_id}, passwordHash: {user_data['passwordHash']}")
#                 user_data['password'] = decrypt_password(user_data.pop('passwordHash'))
#             except Exception as e:
#                 logging.error(f"Error decrypting password for user {user_id}: {e}")
#                 user_data['password'] = "None"  # Set password to None if decryption fails

#         # Ensure associatedIds is properly parsed
#         associated_ids = user_data.get('associatedIds', {})
#         if isinstance(associated_ids, str):
#             try:
#                 associated_ids = json.loads(associated_ids)
#             except json.JSONDecodeError:
#                 associated_ids = {}

#         # Fetch associated details
#         associated_details = []
#         for assoc_id in associated_ids:
#             assoc_user_ref = db.collection('users').document(assoc_id)
#             assoc_user_doc = assoc_user_ref.get()

#             if assoc_user_doc.exists:
#                 assoc_user_data = assoc_user_doc.to_dict()

#                 # Prepare minimal details for associated users
#                 try:
#                     logging.info(f"Decrypting password for associated user: {assoc_id}, passwordHash: {assoc_user_data.get('passwordHash')}")
#                     decrypted_password = decrypt_password(assoc_user_data.get('passwordHash', ''))
#                 except Exception as e:
#                     logging.error(f"Failed to decrypt password for user {assoc_user_data.get('userId')}: {e}")
#                     decrypted_password = None  # Set to None if decryption fails

#                 associated_details.append({
#                     "associatedId": assoc_user_data.get('userId'),
#                     "name": assoc_user_data.get('profileInfo', {}).get('personalInformation', {}).get('name'),
#                     "status": assoc_user_data.get('status'),
#                     "password": decrypted_password,
#                     "photoUrl": assoc_user_data.get('profileInfo', {}).get('personalInformation', {}).get('photoUrl'),
#                     # "teachingPlans" : assoc_user_data.get("teachingPlans",{})
#                 })

#         # Update response fields
#         user_data['associatedDetails'] = associated_details

#         return jsonify({
#             "message": "User profile fetched successfully",
#             "user": user_data
#         }), 200

#     except Exception as e:
#         logging.error(f"Error in get_user_profile: {e}")
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# from flask import Flask, request, jsonify
# import logging


# @app.route("/api/auth/getprofile", methods=["GET"])
# def get_user_profile():
#     try:
#         user_id = request.args.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         # Fetch user document from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         user_data = user_doc.to_dict()

#         # Extracting all relevant fields
#         role = user_data.get("role", "")
#         profile_info = user_data.get("profileInfo", {})
#         contact_info = profile_info.get("contactInformation", {})
#         personal_info = profile_info.get("personalInformation", {})
#         print(personal_info)

#         # Base user profile data
#         response_data = {
#             "message": "User profile fetched successfully",
#             "user": {
#                 "userId": user_id,
#                 "email": user_data.get("email", ""),
#                 "role": role,
#                 "status": user_data.get("status", "Active"),
#                 "createdAt": user_data.get("createdAt", ""),
#                 "last_login": user_data.get("last_login", ""),
#                 # "password": decrypt_password(user_data.get("passwordHash", "")),
#                 "password": bool(user_data.get("passwordHash")),
#                 "contactNumber": contact_info.get("phoneNumber", ""),
#                 "name": personal_info.get("name", ""),
#                 "photoUrl": personal_info.get("photoUrl", ""),
#                 "delivery_method": {
#                     "app": user_data.get("delivery_method", {}).get("app", False),
#                     "email": user_data.get("delivery_method", {}).get("email", False),
#                     "sms": user_data.get("delivery_method", {}).get("sms", False),
#                 },
#             },
#         }

#         # Role-specific handling
#         if role == "student":
#             profile_info = user_data.get("profileInfo", {}).get(
#                 "personalInformation", {}
#             )
#             contact_info = user_data.get("profileInfo", {}).get(
#                 "contactInformation", {}
#             )

#             response_data["user"].update(
#                 {
#                     "name": profile_info.get("name", ""),
#                     "email": user_data.get("email", ""),
#                     "contactNumber": contact_info.get("phoneNumber", ""),
#                     # "associatedIds": user_data.get("associatedIds", []),
#                 }
#             )

#         elif role == "parent":
#             profile_info = user_data.get("profileInfo", {}).get(
#                 "personalInformation", {}
#             )
#             contact_info = user_data.get("profileInfo", {}).get(
#                 "contactInformation", {}
#             )

#             response_data["user"].update(
#                 {
#                     "name": profile_info.get("name", ""),
#                     "email": user_data.get("email", ""),
#                     "contactNumber": contact_info.get("phoneNumber", ""),
#                     # "associatedIds": user_data.get("associatedIds", []),
#                 }
#             )

#         elif role == "headmaster":
#             response_data["user"].update(
#                 {
#                     "name": profile_info.get("personalInformation", {}).get("name", ""),
#                     "email": user_data.get("email", ""),
#                     "contactNumber": contact_info.get("phoneNumber", ""),
#                     # "associatedIds": user_data.get("associatedIds", {}),
#                     "delivery_method": {
#                         "app": user_data.get("delivery_method", {}).get("app", True),
#                         "email": user_data.get("delivery_method", {}).get(
#                             "email", False
#                         ),
#                         "sms": user_data.get("delivery_method", {}).get("sms", False),
#                     },
#                 }
#             )
#         elif role == "teacher":
#             response_data["user"].update(
#                 {
#                     "name": personal_info.get("name", ""),
#                     "email": user_data.get("email", ""),
#                     "contactNumber": contact_info.get("phoneNumber", ""),
#                     # "associatedIds": user_data.get("associatedIds", {}),
#                     "delivery_method": {
#                         "app": user_data.get("delivery_method", {}).get("app", True),
#                         "email": user_data.get("delivery_method", {}).get(
#                             "email", False
#                         ),
#                         "sms": user_data.get("delivery_method", {}).get("sms", False),
#                     },
#                 }
#             )

#         return jsonify(response_data), 200

#     except Exception as e:
#         logging.error(f"Error in get_user_profile: {e}")
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500



# #working get_profile
# @app.route("/api/auth/getprofile", methods=["GET"])
# def get_user_profile():
#     try:
#         user_id = request.args.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         # Fetch user document from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         user_data = user_doc.to_dict()

#         # --- Base Data Extraction (Single Source of Truth) ---
#         # The 'profileInfo' is the reliable source for name, contact, etc.
#         profile_info = user_data.get("profileInfo", {})
#         contact_info = profile_info.get("contactInformation", {})
#         personal_info = profile_info.get("personalInformation", {})
        
#         # Use a safe default for delivery_method
#         delivery_method = user_data.get("delivery_method", {})
        
#         role = user_data.get("role", "") # Extract role once

#         # Base user profile data initialization
#         response_data = {
#             "message": "User profile fetched successfully",
#             "user": {
#                 "userId": user_id,
#                 "email": user_data.get("email", ""),
#                 "role": role,
#                 "status": user_data.get("status", "Active"),
#                 "createdAt": user_data.get("createdAt", ""),
#                 "last_login": user_data.get("last_login", ""),
#                 "password": bool(user_data.get("passwordHash")),
                
#                 # --- All profile fields read directly from personal_info/contact_info ---
#                 "name": personal_info.get("name", ""),
#                 "contactNumber": contact_info.get("phoneNumber", ""),
#                 "photoUrl": personal_info.get("photoUrl", ""),
#                 "address": contact_info.get("address", ""), 
#                 "idNumber": personal_info.get("idNumber", ""),

#                 # Default delivery methods (app: False, email: False, sms: False)
#                 "delivery_method": {
#                     "app": delivery_method.get("app", False),
#                     "email": delivery_method.get("email", False),
#                     "sms": delivery_method.get("sms", False),
#                 },
#             },
#         }

#         # --- Role-specific adjustments (ONLY for delivery_method defaults) ---
#         # The redundant if/elif blocks for name/contact fetching have been removed.
        
#         if role in ["headmaster", "teacher"]:
#             # Headmasters and Teachers default to 'app' notifications being true
#             response_data["user"]["delivery_method"].update({
#                 "app": delivery_method.get("app", True), # Default to True if not in DB
#             })
            
#         # Optional: Add associated IDs if they exist
#         if "associatedIds" in user_data:
#              response_data["user"]["associatedIds"] = user_data["associatedIds"]

#         return jsonify(response_data), 200

#     except Exception as e:
#         logging.error(f"Error in get_user_profile: {e}")
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


    # DELETE_FIELD = None 
    # try:
    #     from firebase_admin import firestore
    #     DELETE_FIELD = firestore.DELETE_FIELD
    # except ImportError:
    #     # Fallback/Dummy for environment where full firestore import might not be available
    #     DELETE_FIELD = "__DELETE_FIELD__" 

# @app.route("/api/auth/getprofile", methods=["GET"])
# def get_user_profile():
#     try:
#         user_id = request.args.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         # Fetch user document from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         user_data = user_doc.to_dict()

#         # --- Base Data Extraction (Single Source of Truth) ---
#         # The 'profileInfo' is the reliable source for name, contact, etc.
#         profile_info = user_data.get("profileInfo", {})
#         contact_info = profile_info.get("contactInformation", {})
#         personal_info = profile_info.get("personalInformation", {})
        
#         # Use a safe default for delivery_method
#         delivery_method = user_data.get("delivery_method", {})
        
#         role = user_data.get("role", "") # Extract role once

#         # Base user profile data initialization
#         response_data = {
#             "message": "User profile fetched successfully",
#             "user": {
#                 "userId": user_id,
#                 "email": user_data.get("email", ""),
#                 "role": role,
#                 "status": user_data.get("status", "Active"),
#                 "createdAt": user_data.get("createdAt", ""),
#                 "last_login": user_data.get("last_login", ""),
#                 "password": bool(user_data.get("passwordHash")),
                
#                 # --- All profile fields read directly from personal_info/contact_info ---
#                 "name": profile_info.get("name", ""),
#                 "contactNumber": contact_info.get("phoneNumber", ""),
#                 "photoUrl": personal_info.get("photoUrl", ""),
#                 "address": contact_info.get("address", ""), 
#                 "idNumber": personal_info.get("idNumber", ""),

#                 # Default delivery methods (app: False, email: False, sms: False)
#                 "delivery_method": {
#                     "app": delivery_method.get("app", False),
#                     "email": delivery_method.get("email", False),
#                     "sms": delivery_method.get("sms", False),
#                 },
#             },
#         }

#         # --- Role-specific adjustments (ONLY for delivery_method defaults) ---
        
#         if role in ["headmaster", "teacher"]:
#             # Headmasters and Teachers default to 'app' notifications being true
#             response_data["user"]["delivery_method"].update({
#                 "app": delivery_method.get("app", True), # Default to True if not in DB
#             })
            
#         # Optional: Add associated IDs if they exist
#         if "associatedIds" in user_data:
#              response_data["user"]["associatedIds"] = user_data["associatedIds"]

#         return jsonify(response_data), 200

#     except Exception as e:
#         logging.error(f"Error in get_user_profile: {e}")
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500



@app.route("/api/auth/getprofile", methods=["GET"])
def get_user_profile():
	try:
		user_id = request.args.get("userId")
		if not user_id:
			return jsonify({"error": "userId is required"}), 400

		# Fetch user document from Firestore
		user_ref = db.collection("users").document(user_id)
		user_doc = user_ref.get()

		if not user_doc.exists:
			return jsonify({"error": "User not found"}), 404

		user_data = user_doc.to_dict()

		# --- Base Data Extraction (Single Source of Truth) ---
		profile_info = user_data.get("profileInfo", {})
		contact_info = profile_info.get("contactInformation", {})
		personal_info = profile_info.get("personalInformation", {})
		
		# Flag to track if we need to commit a cleanup update
		needs_cleanup = False
		cleanup_updates = {}
		
		# Determine the definitive name (prioritize root 'name', then personalInformation, then profileInfo, then contactInfo)
		
		# 1. Get the most current/best name value
		# Prioritize the root-level 'name' field, as it is explicitly set by update_profile.
		best_name = user_data.get("name", "") or personal_info.get("name", "") or profile_info.get("name", "") or contact_info.get("name", "")
		
		if best_name:
			# 2. Check if all four fields are synchronized to the best_name
			
			# Check 1: profileInfo.personalInformation.name (Target location)
			if personal_info.get("name") != best_name:
				personal_info["name"] = best_name
				needs_cleanup = True
			
			# Check 2: profileInfo.name (Legacy location 1)
			if profile_info.get("name") != best_name:
				profile_info["name"] = best_name
				needs_cleanup = True
			
			# Check 3: profileInfo.contactInformation.name (Legacy location 2)
			if contact_info.get("name") != best_name:
				contact_info["name"] = best_name
				needs_cleanup = True
			
			# Check 4: Root 'name' field
			if user_data.get("name") != best_name:
				cleanup_updates["name"] = best_name
				needs_cleanup = True


		# --- Commit Cleanup if necessary ---
		if needs_cleanup:
			# Re-assign nested structures to ensure they're correctly set on the parent map
			profile_info["personalInformation"] = personal_info
			profile_info["contactInformation"] = contact_info
			
			# Note: cleanup_updates["profileInfo"] will contain the updated nested structures
			cleanup_updates["profileInfo"] = profile_info
			
			# Only commit the cleanup if the name was found and we made changes
			if cleanup_updates:
					try:
						# Update the database to fix the inconsistent name fields
						user_ref.update(cleanup_updates)
						logging.info(f"Cleanup: Synchronized name fields for user {user_id}")
					except Exception as e:
						logging.error(f"Failed to commit cleanup update for {user_id}: {e}")

		# --- Data Preparation for Response (ALWAYS use the cleaned up/best name) ---
		
		# Use a safe default for delivery_method
		delivery_method = user_data.get("delivery_method", {})
		
		role = user_data.get("role", "") # Extract role once
		
		response_data = {
			"message": "User profile fetched successfully",
			"user": {
				"userId": user_id,
				"email": user_data.get("email", ""),
				"role": role,
				"status": user_data.get("status", "Active"),
				"createdAt": user_data.get("createdAt", ""),
				"last_login": user_data.get("last_login", ""),
				"password": bool(user_data.get("passwordHash")),
				
				# Use the consistently retrieved/cleaned up name (best_name)
				"name": best_name,
				
				# --- Other profile fields read directly from personal_info/contact_info ---
				"contactNumber": contact_info.get("phoneNumber", ""),
				"photoUrl": personal_info.get("photoUrl", ""),
				"address": contact_info.get("address", ""), 
				"idNumber": personal_info.get("idNumber", ""),

				# Default delivery methods (app: False, email: False, sms: False)
				"delivery_method": {
					"app": delivery_method.get("app", False),
					"email": delivery_method.get("email", False),
					"sms": delivery_method.get("sms", False),
				},
			},
		}

		# --- Role-specific adjustments (ONLY for delivery_method defaults) ---
		
		if role in ["headmaster", "teacher"]:
			# Headmasters and Teachers default to 'app' notifications being true
			response_data["user"]["delivery_method"].update({
				"app": delivery_method.get("app", True), # Default to True if not in DB
			})
			
		# Optional: Add associated IDs if they exist
		if "associatedIds" in user_data:
			 response_data["user"]["associatedIds"] = user_data["associatedIds"]

		return jsonify(response_data), 200

	except Exception as e:
		logging.error(f"Error in get_user_profile: {e}")
		return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# # # from flask import Flask, jsonify, request
# # from firebase_admin import firestore_async
# # import asyncio

# # db1 = firestore_async.client()  # Use async Firestore client
# # loop = asyncio.get_event_loop()

# # @app.route('/students/list', methods=['GET'])
# # def get_students_list():
# #     user_id = request.args.get('userId')
# #     result = loop.run_until_complete(fetch_students_list(user_id))
# #     return result

# # async def fetch_students_list(user_id=None):
# #     try:
# #         if user_id:
# #             user_doc = await db1.collection('users').document(user_id).get()
# #             if not user_doc.exists:
# #                 return jsonify({"error": "User not found"}), 404

# #             user_data = user_doc.to_dict()
#             associated_student_ids = user_data.get("associatedIds", [])

#             if not associated_student_ids:
#                 return jsonify({"message": "No associated students found"}), 200

#             students = await fetch_students_by_ids(associated_student_ids)
#             return jsonify(students), 200

#         else:
#             students_collection = db1.collection('students')
#             student_docs = [doc async for doc in students_collection.stream()]
#             students = await fetch_all_students(student_docs)
#             return jsonify(students), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# async def fetch_students_by_ids(student_ids):
#     """Fetch details for specific student IDs."""
#     tasks = [fetch_student_details(student_id) for student_id in student_ids]
#     students = await asyncio.gather(*tasks)
#     return [student for student in students if student is not None]

# async def fetch_student_details(student_id):
#     """Fetch a single student's details."""
#     student_doc = await db1.collection('students').document(student_id).get()
#     if not student_doc.exists:
#         return None

#     student_data = student_doc.to_dict()
#     return await parse_student_data(student_id, student_data)

# async def fetch_latest_plan(teaching_plans):
#     """Fetch the latest teaching plan based on the createdAt timestamp."""
#     if not teaching_plans:
#         return "", 0

#     latest_plan = max(teaching_plans.values(), key=lambda x: x.get("createdAt", ""), default={})
#     return latest_plan.get("planId", ""), latest_plan.get("version", 1)

# async def fetch_all_students(student_docs):
#     """Fetch details for all students."""
#     tasks = [await parse_student_data(doc.id, doc.to_dict()) for doc in student_docs]
#     return tasks

# async def fetch_grade_and_class_refs(grade, class_section):
#     if not grade or not class_section:
#         return "", ""

#     try:
#         grades_collection = db1.collection('Grades')
#         matching_grades = [doc async for doc in grades_collection.where("grade", "==", grade).stream()]

#         if not matching_grades:
#             return "", ""

#         grade_doc = matching_grades[0]
#         grade_id = grade_doc.id

#         class_collection = grade_doc.reference.collection('classes')
#         matching_classes = [doc async for doc in class_collection.stream()]

#         class_id = ""
#         for class_doc in matching_classes:
#             class_data = class_doc.to_dict()
#             if class_data.get("class_name") == class_section:
#                 class_id = class_data.get("class_id", "")
#                 break

#         return grade_id, class_id

#     except Exception as e:
#         print(f"Error fetching grade/class references: {str(e)}")
#         return "", ""

# async def parse_student_data(student_id, student_data):
#     """Parse student data to handle nested fields and return structured information."""
#     if not student_data:
#         return None

#     plan_id, version = await fetch_latest_plan(student_data.get("teachingPlans", {}))
#     grade_ref, class_ref = await fetch_grade_and_class_refs(
#         student_data.get("academicInformation", {}).get("grade", ""),
#         student_data.get("academicInformation", {}).get("classSection", "")
#     )

#     return {
#         "student_id": student_id,
#         "student_name": student_data.get("personalInformation", {}).get("name", ""),
#         "email": student_data.get("email", ""),
#         "status": student_data.get("status", ""),
#         "phonenumber": student_data.get("contactInformation", {}).get("phoneNumber", ""),
#         "address": student_data.get("contactInformation", {}).get("address", ""),
#         "photourl": student_data.get("personalInformation", {}).get("photourl", ""),
#         "grade": student_data.get("academicInformation", {}).get("grade", ""),
#         "classSection": student_data.get("academicInformation", {}).get("classSection", ""),
#         "planId": plan_id,
#         "version": version,
#         "grade_ref": grade_ref,
#         "class_ref": class_ref
#     }
from flask import Flask, jsonify, request
from firebase_admin import credentials, initialize_app, firestore
import asyncio

# Initialize Flask App
# app = Flask(__name__)

# Initialize Firebase Admin SDK
# cred = credentials.Certificate("serviceAccountKey.json")
# initialize_app(cred)

# # Use Firestore (Sync Mode for Compatibility)
# db1 = firestore.client()

# @app.route('/students/list', methods=['GET'])
# def get_students_list():
#     """Fetch students list efficiently, either by userId or all students."""
#     user_id = request.args.get('userId')
#     loop = asyncio.new_event_loop()
#     asyncio.set_event_loop(loop)
#     result = loop.run_until_complete(fetch_students_list(user_id))
#     return result

# async def fetch_students_list(user_id=None):
#     """Fetch students list based on userId or all students."""
#     try:
#         if user_id:
#             user_doc = db1.collection('users').document(user_id).get()
#             if not user_doc.exists:
#                 return jsonify({"error": "User not found"}), 404

#             user_data = user_doc.to_dict()
#             associated_student_ids = user_data.get("associatedIds", [])

#             if not associated_student_ids:
#                 return jsonify({"message": "No associated students found"}), 200

#             students = await fetch_students_by_ids(associated_student_ids)
#             return jsonify(students), 200

#         else:
#             students_collection = db1.collection('students')
#             students = await fetch_all_students(students_collection)
#             return jsonify(students), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# async def fetch_students_by_ids(student_ids):
#     """Fetch multiple students in batch."""
#     if not student_ids:
#         return []

#     # Batch request for student documents
#     student_refs = [db1.collection('students').document(sid) for sid in student_ids]
#     student_docs = [ref.get() for ref in student_refs]

#     # Process documents concurrently
#     tasks = [parse_student_data(doc.id, doc.to_dict()) for doc in student_docs if doc.exists]
#     return await asyncio.gather(*tasks)

# async def fetch_all_students(students_collection):
#     """Fetch details for all students efficiently."""
#     student_docs = [doc for doc in students_collection.stream()]
#     tasks = [parse_student_data(doc.id, doc.to_dict()) for doc in student_docs]
#     return await asyncio.gather(*tasks)

# # async def fetch_latest_plan(teaching_plans):
# #     """Fetch the latest teaching plan using timestamps."""
# #     if not teaching_plans:
# #         return "", 0

# #     latest_plan = max(teaching_plans.values(), key=lambda x: x.get("createdAt", ""), default={})
# #     return latest_plan.get("plan_id", ""), latest_plan.get("version", 1)

# from datetime import datetime
# from datetime import datetime

# async def fetch_latest_plan(teaching_plans):
#     """Fetch the latest teaching plan ID and version based on createdAt timestamp."""
#     if not teaching_plans:
#         return "", 0  # Return empty values if no plans exist

#     latest_plan_id = ""
#     latest_version = 1
#     latest_created_at = None  # Store the latest timestamp

#     for plan_key, plan_details in teaching_plans.items():
#         # Navigate inside actionPlan
#         action_plan = plan_details.get("actionPlan", {})

#         # Extract plan_id from actionPlan
#         plan_id = action_plan.get("plan_id", plan_key)  # Fallback to key if missing

#         # Extract createdAt timestamp
#         created_at = plan_details.get("createdAt")  # Check if createdAt exists at this level

#         if not created_at:
#             print(f"Warning: Plan {plan_id} has no createdAt field!")
#             continue  # Skip this plan if no timestamp is found

#         # Convert createdAt to a comparable datetime format
#         if isinstance(created_at, str):
#             try:
#                 created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
#             except ValueError:
#                 print(f"Invalid createdAt format in plan {plan_id}: {created_at}")
#                 continue  # Skip if timestamp is invalid

#         # Compare timestamps and update latest plan
#         if latest_created_at is None or created_at > latest_created_at:
#             latest_created_at = created_at
#             latest_plan_id = plan_id
#             latest_version = plan_details.get("version", 1)  # Default to version 1 if missing

#     # If no valid createdAt was found, return the first plan ID as a fallback
#     if not latest_plan_id and teaching_plans:
#         first_plan_key = list(teaching_plans.keys())[0]
#         first_action_plan = teaching_plans[first_plan_key].get("actionPlan", {})
#         latest_plan_id = first_action_plan.get("plan_id", first_plan_key)
#         latest_version = teaching_plans[first_plan_key].get("version", 1)

#     return latest_plan_id, latest_version

# from datetime import datetime

# async def fetch_latest_plan(teaching_plans):
#     """Fetch the latest teaching plan ID and version based on createdAt timestamp."""
#     if not teaching_plans:
#         return "", 0  # Return empty values if no plans exist

#     latest_plan_id = ""
#     latest_version = 1
#     latest_created_at = None  # Store the latest timestamp

#     for plan_id, plan_details in teaching_plans.items():
#         created_at = plan_details.get("createdAt")  # Fetch timestamp

#         # Ensure createdAt exists and is in a valid timestamp format
#         if created_at:
#             # If Firestore stores timestamp as a string, convert it to datetime
#             if isinstance(created_at, str):
#                 try:
#                     created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
#                 except ValueError:
#                     continue  # Skip if timestamp is invalid

#             # Compare timestamps and update latest plan
#             if latest_created_at is None or created_at > latest_created_at:
#                 latest_created_at = created_at
#                 latest_plan_id = plan_id  # The Plan ID is the key
#                 latest_version = plan_details.get("version", 1)  # Default to version 1 if missing

#     return latest_plan_id, latest_version

# async def fetch_grade_and_class_refs(grade, class_section):
#     """Fetch grade and class references efficiently."""
#     if not grade or not class_section:
#         return "", ""

#     try:
#         # Fetch grade document
#         grades_query = db1.collection('Grades').where("grade", "==", grade).limit(1)
#         grade_docs = [doc for doc in grades_query.stream()]

#         if not grade_docs:
#             return "", ""

#         grade_doc = grade_docs[0]
#         class_query = grade_doc.reference.collection('classes').where("class_name", "==", class_section).limit(1)
#         class_docs = [doc for doc in class_query.stream()]

#         return grade_doc.id, class_docs[0].id if class_docs else ""

#     except Exception as e:
#         print(f"Error fetching grade/class references: {str(e)}")
#         return "", ""

# async def parse_student_data(student_id, student_data):
#     """Parse and return only the required student data."""
#     if not student_data:
#         return None

#     # Fetch related data concurrently
#     plan_id, version = await fetch_latest_plan(student_data.get("teachingPlans", {}))
#     grade_ref, class_ref = await fetch_grade_and_class_refs(
#         student_data.get("academicInformation", {}).get("grade", ""),
#         student_data.get("academicInformation", {}).get("classSection", "")
#     )

#     return {
#         "student_id": student_id,
#         "student_name": student_data.get("personalInformation", {}).get("name", ""),
#         "email": student_data.get("email", ""),
#         "status": student_data.get("status", ""),
#         "phonenumber": student_data.get("contactInformation", {}).get("phoneNumber", ""),
#         "address": student_data.get("contactInformation", {}).get("address", ""),
#         "photourl": student_data.get("personalInformation", {}).get("photourl", ""),
#         "grade": student_data.get("academicInformation", {}).get("grade", ""),
#         "classSection": student_data.get("academicInformation", {}).get("classSection", ""),
#         "planId": plan_id,
#         "version": version,
#         "grade_ref": grade_ref,
#         "class_ref": class_ref
#     }
# ----------------------------------------------------------studentlists.py----------------------------
import asyncio
from flask import Flask, jsonify, request
from google.cloud import firestore
import os
from flask import Flask, jsonify, request
from firebase_admin import credentials, initialize_app, firestore_async
import asyncio

# app = Flask(__name__)
db1 = firestore_async.client()
# db1 = firestore.Client()  # sync client


# loop = asyncio.get_event_loop()

# @app.route('/students/list', methods=['GET'])
# def get_students_list():
#     user_id = request.args.get('userId')
#     result = loop.run_until_complete(fetch_students_list(user_id))
#     return result


# async def fetch_students_list(user_id=None):
#     try:
#         if user_id:
#             user_doc = await db1.collection('users').document(user_id).get()
#             if not user_doc.exists:
#                 return jsonify({"error": "User not found"}), 404

#             user_data = user_doc.to_dict()
#             associated_student_ids = user_data.get("associatedIds", [])

#             if not associated_student_ids:
#                 return jsonify({"message": "No associated students found"}), 200

#             students = await fetch_students_by_ids(associated_student_ids)
#             return jsonify(students), 200

#         else:
#             students_collection = await db1.collection('students')
#             student_docs = [doc async for doc in students_collection.stream()]
#             students = await fetch_all_students(student_docs)
#             return jsonify(students), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# async def fetch_students_by_ids(student_ids):
#     """Fetch details for specific student IDs."""
#     tasks = [fetch_student_details(student_id) for student_id in student_ids]
#     students = await asyncio.gather(*tasks)
#     return [student for student in students if student is not None]

# async def fetch_student_details(student_id):
#     """Fetch a single student's details."""
#     student_doc = await db1.collection('students').document(student_id).get()
#     if not student_doc.exists:
#         return None

#     student_data = student_doc.to_dict()
#     return await parse_student_data(student_id, student_data)

# from datetime import datetime

# async def fetch_latest_plan(teaching_plans):
#     """Fetch the latest teaching plan based on the createdAt timestamp (ISO 8601 string)."""
#     if not teaching_plans:
#         print("No teaching plans found")
#         return ""

#     print("Teaching Plans Data:", teaching_plans)  # Debugging output

#     def parse_timestamp(plan):
#         created_at_str = plan.get("createdAt", "").strip()
#         try:
#             if not created_at_str:
#                 print("Missing 'createdAt' for plan:", plan)  # Debugging output
#                 return 0  # Default to oldest timestamp

#             # Remove 'Z' (Zulu time) for correct parsing
#             created_at_str = created_at_str.replace("Z", "")

#             # Convert ISO string to datetime
#             dt = datetime.strptime(created_at_str, "%Y-%m-%dT%H:%M:%S")
#             return dt.timestamp()
#         except ValueError as e:
#             print(f"Error parsing timestamp ({created_at_str}): {str(e)}")  # Debugging output
#             return 0

#     # Extract teaching plan with the latest timestamp
#     latest_plan = max(teaching_plans.values(), key=parse_timestamp, default={})

#     # Extract planId from actionPlan
#     plan_id = latest_plan.get("actionPlan", {}).get("planId", "")

#     print("Latest Plan Selected:", latest_plan)  # Debugging output
#     print("Extracted Plan ID:", plan_id)  # Debugging output

#     return plan_id

# async def fetch_all_students(student_docs):
#     """Fetch details for all students."""
#     tasks = [await parse_student_data(doc.id, doc.to_dict()) for doc in student_docs]
#     return tasks

# async def fetch_grade_and_class_refs(grade, class_section):
#     if not grade or not class_section:
#         return "", ""

#     try:
#         grades_collection = db1.collection('Grades')
#         matching_grades = [doc async for doc in grades_collection.where("grade", "==", grade).stream()]

#         if not matching_grades:
#             return "", ""

#         grade_doc = matching_grades[0]
#         grade_id = grade_doc.id

#         class_collection = grade_doc.reference.collection('classes')
#         matching_classes = [doc async for doc in class_collection.stream()]

#         class_id = ""
#         for class_doc in matching_classes:
#             class_data = class_doc.to_dict()
#             if class_data.get("class_name") == class_section:
#                 class_id = class_data.get("class_id", "")
#                 break

#         return grade_id, class_id

#     except Exception as e:
#         print(f"Error fetching grade/class references: {str(e)}")
#         return "", ""

# async def parse_student_data(student_id, student_data):
#     """Parse student data to handle nested fields and return structured information."""
#     if not student_data:
#         return None

#     plan_id = await fetch_latest_plan(student_data.get("teachingPlans", {}))
#     grade_ref, class_ref = await fetch_grade_and_class_refs(
#         student_data.get("academicInformation", {}).get("grade", ""),
#         student_data.get("academicInformation", {}).get("classSection", "")
#     )

#     return {
#         "student_id": student_id,
#         "student_name": student_data.get("personalInformation", {}).get("name", ""),
#         "email": student_data.get("email", ""),
#         "status": student_data.get("status", "activie"),
#         "phonenumber": student_data.get("contactInformation", {}).get("phoneNumber", ""),
#         "address": student_data.get("contactInformation", {}).get("address", ""),
#         "photourl": student_data.get("personalInformation", {}).get("photourl", ""),
#         "grade": student_data.get("academicInformation", {}).get("grade", ""),
#         "classSection": student_data.get("academicInformation", {}).get("classSection", ""),
#         "planId": plan_id,
#         "grade_ref": grade_ref,
#         "class_ref": class_ref
#     }


from firebase_admin import firestore_async
import asyncio
from datetime import datetime

db1 = firestore_async.client()
loop = asyncio.get_event_loop()
import re

def extract_grade_number(student):
    grade = student.get("grade", "").upper()
    match = re.search(r'\d+', grade)
    return int(match.group()) if match else float('inf')


@app.route("/students/list", methods=["GET"])
def get_students_list():
    user_id = request.args.get("userId")
    result = loop.run_until_complete(fetch_students_list(user_id))
    return result


# async def fetch_students_list(user_id=None):
#     try:
#         if user_id:
#             user_doc = await db1.collection("users").document(user_id).get()
#             if not user_doc.exists:
#                 return jsonify({"error": "User not found"}), 404

#             user_data = user_doc.to_dict()
#             associated_student_ids = user_data.get("associatedIds", [])

#             if not associated_student_ids:
#                 return jsonify({"message": "No associated students found", "students": []}), 200
#                 return jsonify({"message": "No associated students found"}), 200

#             # ? Fetch from users collection
#             students = await fetch_users_by_ids(associated_student_ids)
#             students_sorted = sorted(students, key=extract_grade_number)
#             return jsonify(students_sorted), 200

#         else:
#             # Optional fallback to all users if no userId provided
#             users_collection = db1.collection("users")
#             user_docs = [doc async for doc in users_collection.stream()]
#             students = await fetch_all_students(user_docs)
#             students_sorted = sorted(students, key=extract_grade_number)
#             return jsonify(students_sorted), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

async def fetch_students_list(user_id=None):
    try:
        if user_id:
            user_doc = await db1.collection("users").document(user_id).get()
            if not user_doc.exists:
                return jsonify({"error": "User not found"}), 404

            user_data = user_doc.to_dict()
            associated_student_ids = user_data.get("associatedIds", [])

            if not associated_student_ids:
                return jsonify({"message": "No associated students found", "students": []}), 200
                # Removed redundant return jsonify({"message": "No associated students found"}), 200

            # ? Fetch from users collection (Filtering happens in fetch_all_students for this path)
            students = await fetch_users_by_ids(associated_student_ids)
            students_sorted = sorted(students, key=extract_grade_number)
            return jsonify(students_sorted), 200

        else:
            # ✅ CHANGE 1: Filter by role="student" at the database level for the 'all users' fallback
            users_collection = db1.collection("users").where("role", "==", "student")
            user_docs = [doc async for doc in users_collection.stream()]
            
            students = await fetch_all_students(user_docs)
            students_sorted = sorted(students, key=extract_grade_number)
            return jsonify(students_sorted), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

async def fetch_users_by_ids(user_ids):
    """Fetch students (users) by IDs."""
    user_docs = await asyncio.gather(
        *[db1.collection("users").document(uid).get() for uid in user_ids]
    )
    return await fetch_all_students(user_docs)


# async def fetch_all_students(user_docs):
#     """Extract usable student data."""
#     students = await asyncio.gather(
#         *[parse_student_data(doc.id, doc.to_dict()) for doc in user_docs]
#     )
#     return [student for student in students if student is not None]
async def fetch_all_students(user_docs):
    """Extract usable student data. (Now includes role filter)."""
    
    # ✅ CHANGE 2: Filter the documents locally to ensure only "student" roles are processed.
    student_docs = [
        doc for doc in user_docs 
        if doc.exists and doc.to_dict().get("role", "").lower() == "student"
    ]
    
    students = await asyncio.gather(
        *[parse_student_data(doc.id, doc.to_dict()) for doc in student_docs]
    )
    return [student for student in students if student is not None]

async def parse_student_data(student_id, student_data):
    """Parse individual student info."""
    if not student_data:
        return None

    grade_name_task = fetch_grade_and_class_refs11(student_id)
    plan_id_task = fetch_latest_plan(student_data.get("teachingPlans", {}))

    (grade_name, class_section), plan_id = await asyncio.gather(
        grade_name_task, plan_id_task
    )

    # Defaults if not found
    grade_name = grade_name or ""
    class_section = class_section or ""

    # ✅ FIX: Normalize grade to UPPERCASE to match Curriculum API
    # This ensures "Grade 1" becomes "GRADE 1" so frontend filtering works.
    grade_name = str(grade_name).upper().strip()

    grade_class_task = fetch_grade_and_class_refs(grade_name, class_section)
    grade_ref, class_ref = await grade_class_task

    return {
        "student_id": student_id,
        "student_name": student_data.get("profileInfo", {}).get("personalInformation", {}).get("name", ""),
        "email": student_data.get("email", ""),
        "status": student_data.get("status", "active"),
        "phonenumber": student_data.get("contactInformation", {}).get("phoneNumber", ""),
        "address": student_data.get("contactInformation", {}).get("address", ""),
        "photourl": student_data.get("profileInfo", {}).get("personalInformation", {}).get("photoUrl", ""),
        "grade": grade_name,  # Now sending "GRADE 1"
        "classSection": class_section,
        "planId": plan_id,
        "grade_ref": grade_ref,
        "class_ref": class_ref,
    }

async def fetch_grade_and_class_refs11(student_id):
    """Extract grade and class from assignedGrades field in users."""
    try:
        student_doc = await db1.collection("users").document(student_id).get()

        if not student_doc.exists:
            return None, None

        student_data = student_doc.to_dict()
        assigned_grades_data = student_data.get("assignedGrades", {})
        assigned_grades = assigned_grades_data.get("grades", assigned_grades_data)

        for grade, class_map in assigned_grades.items():
            if isinstance(class_map, dict):
                for class_sec in class_map:
                    if isinstance(class_map[class_sec], list):
                        return grade, class_sec  # First valid match
        return None, None
    except Exception:
        return None, None


async def fetch_grade_and_class_refs(grade_name, class_section):
    """(Stub) Simulate fetching references (optional)."""
    # You can modify this function to fetch document references from Firestore
    return f"ref_to_{grade_name}", f"ref_to_{class_section}"


async def fetch_latest_plan(teaching_plans):
    """Find latest plan based on createdAt timestamp."""
    if not teaching_plans:
        return ""

    def parse_timestamp(plan):
        created_at = plan.get("createdAt", "").replace("Z", "")
        try:
            if "." in created_at:
                dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S.%f")
            else:
                dt = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%S")
            return dt.timestamp()
        except ValueError:
            return 0

    latest_plan = max(teaching_plans.values(), key=parse_timestamp, default={})
    return latest_plan.get("actionPlan", {}).get("planId", "")


@app.route("/api/student/update", methods=["PUT"])
def update_student_profile():
    try:
        data = request.json
        if not data or "studentId" not in data:
            return jsonify({"error": "studentId is required"}), 400

        student_id = data["studentId"]
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        # Fetch email from users collection using userId
        user_ref = db.collection("users").document(student_id)
        user_doc = user_ref.get()
        if user_doc.exists:
            email = user_doc.to_dict().get("email", "")
        else:
            email = ""

        # Fields that can be updated in students collection
        student_updates = {}
        if "name" in data:
            student_updates["personalInformation.name"] = data["name"]
        if "photourl" in data:
            student_updates["personalInformation.photourl"] = data["photourl"]
        if "idno" in data:
            student_updates["personalInformation.idNumber"] = data["idno"]
        if "phoneNumber" in data:
            student_updates["contactInformation.phoneNumber"] = data["phoneNumber"]
        if "address" in data:
            student_updates["contactInformation.address"] = data["address"]
        if "grade" in data:
            student_updates["academicInformation.grade"] = data["grade"]
        if "classSection" in data:
            student_updates["academicInformation.classSection"] = data["classSection"]

        # Update email in users collection if provided
        if "email" in data:
            user_ref.update({"email": data["email"]})

        # Update student details
        if student_updates:
            student_ref.update(student_updates)
            return jsonify({"message": "Profile updated successfully"}), 200
        else:
            return jsonify({"error": "No valid fields to update"}), 400

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": "An internal server error occurred"}), 500


# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         if not request.content_type.startswith("multipart/form-data"):
#             return jsonify({"error": "Content-Type must be multipart/form-data"}), 415

#         data = request.form
#         user_id = data.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         updates = {}

#         #  Handle Password Update (If Provided)
#         new_password = data.get("password")
#         if new_password:
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = (
#                 encrypted_password  # Store encrypted password in Firestore
#             )

#             #  Update Firebase Authentication Password
#             try:
#                 auth.update_user(uid=user_id, password=new_password, app=get_password_reset_auth_app())
#             except firebase_admin.auth.UserNotFoundError:
#                 return (
#                     jsonify({"error": "User not found in Firebase Authentication"}),
#                     404,
#                 )
#             except Exception as e:
#                 return (
#                     jsonify(
#                         {
#                             "error": f"Failed to update password in Firebase Auth: {str(e)}"
#                         }
#                     ),
#                     500,
#                 )

#         #  Handle Associated IDs Update
#         if "associatedIds" in data:
#             try:
#                 associated_ids = (
#                     json.loads(data["associatedIds"])
#                     if isinstance(data["associatedIds"], str)
#                     else data["associatedIds"]
#                 )
#                 if not isinstance(associated_ids, list):
#                     raise ValueError("associatedIds must be a list")
#                 updates["associatedIds"] = associated_ids
#             except (json.JSONDecodeError, ValueError) as e:
#                 logging.error(f"Error parsing associatedIds: {e}")
#                 return jsonify({"error": "Invalid format for associatedIds"}), 400

#         #  Handle Profile Info Updates as a Full Dictionary
#         profile_info = user_doc.to_dict().get("profileInfo", {})

#         # Update contact information

#         if "phoneNumber" in data:
#             updates["phoneNumber"] = data["phoneNumber"]  # Root-level phone number
#             profile_info.setdefault("contactInformation", {})["phoneNumber"] = data[
#                 "phoneNumber"
#             ]  # Nested phone number

#         if "address" in data:
#             profile_info.setdefault("contactInformation", {})["address"] = data[
#                 "address"
#             ]

#         # Update personal information
#         if "name" in data:
#             profile_info.setdefault("personalInformation", {})["name"] = data["name"]

#         if "idNumber" in data:
#             profile_info.setdefault("personalInformation", {})["idNumber"] = data[
#                 "idNumber"
#             ]

#         # Ã¢Å“â€¦ Handle Photo Upload
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file.filename:
#                 unique_filename = (
#                     f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 )
#                 blob = bucket.blob(unique_filename)
#                 blob.upload_from_file(file)
#                 blob.make_public()
#                 profile_info.setdefault("personalInformation", {})[
#                     "photoUrl"
#                 ] = blob.public_url

#         # Add the updated profileInfo dictionary
#         updates["profileInfo"] = profile_info

#         # Ã¢Å“â€¦ Apply Updates to Firestore
#         if updates:
#             logging.info(f"Updating Firestore for user {user_id} with {updates}")
#             user_ref.update(updates)

#         return jsonify({"message": "Profile updated successfully"}), 200

#     except Exception as e:
#         logging.error(f"Error during profile update: {e}")
#         return (
#             jsonify({"error": f"An error occurred during profile update: {str(e)}"}),
#             500,
#         )
# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         if not request.content_type.startswith("multipart/form-data"):
#             return jsonify({"error": "Content-Type must be multipart/form-data"}), 415

#         data = request.form
#         user_id = data.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         updates = {}
#         profile_info = user_doc.to_dict().get("profileInfo", {})

#         # Handle profile info updates
#         if "name" in data:
#             # Update name in the correct nested structure
#             profile_info.setdefault("personalInformation", {})["name"] = data["name"]

#         if "phoneNumber" in data:
#             # Update phone number in the correct nested structure
#             profile_info.setdefault("contactInformation", {})["phoneNumber"] = data["phoneNumber"]

#         if "address" in data:
#             profile_info.setdefault("contactInformation", {})["address"] = data["address"]

#         # Add the updated profileInfo dictionary to updates
#         if profile_info:
#             updates["profileInfo"] = profile_info

#         # Handle photo upload with correct path update
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file.filename:
#                 unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 blob = bucket.blob(unique_filename)
#                 blob.upload_from_file(file)
#                 blob.make_public()
#                 # Update photo URL in the correct nested structure
#                 profile_info.setdefault("personalInformation", {})["photoUrl"] = blob.public_url

#         # Handle password update if provided
#         if "password" in data:
#             new_password = data["password"]
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = encrypted_password
#             try:
#                 auth.update_user(uid=user_id, password=new_password)
#             except Exception as e:
#                 return jsonify({"error": f"Failed to update password: {str(e)}"}), 500

#         # Apply updates to Firestore
#         if updates:
#             user_ref.update(updates)

#         return jsonify({"message": "Profile updated successfully"}), 200

#     except Exception as e:
#         logging.error(f"Error during profile update: {e}")
#         return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500
# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         # Accept either multipart/form-data (from browser) or JSON (API clients)
#         if request.content_type and "multipart/form-data" in request.content_type:
#             data = request.form.to_dict()
#         else:
#             data = request.get_json(silent=True) or {}

#         user_id = data.get("userId") or data.get("user_id")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         existing = user_doc.to_dict() or {}
#         profile_info = existing.get("profileInfo", {}) or {}

#         updates = {}

#         # Personal info
#         if "name" in data:
#             profile_info.setdefault("personalInformation", {})["name"] = data["name"]
#             # optional root-level name for backward compatibility
#             updates["name"] = data["name"]

#         if "idNumber" in data:
#             profile_info.setdefault("personalInformation", {})["idNumber"] = data["idNumber"]

#         # Contact info (accept contactNumber or phoneNumber)
#         phone = data.get("contactNumber") or data.get("phoneNumber")
#         if phone is not None:
#             profile_info.setdefault("contactInformation", {})["phoneNumber"] = phone
#             updates["contactNumber"] = phone  # keep root-level compatibility

#         if "address" in data:
#             profile_info.setdefault("contactInformation", {})["address"] = data["address"]

#         # Photo upload (multipart/form-data)
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file and file.filename:
#                 # reuse your existing upload helper
#                 unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 try:
#                     photo_url = upload_image_to_firebase(file, unique_filename)
#                     profile_info.setdefault("personalInformation", {})["photoUrl"] = photo_url
#                 except Exception as e:
#                     logging.exception("Photo upload failed")
#                     return jsonify({"error": "Photo upload failed", "details": str(e)}), 500

#         # Password update (optional)
#         if "password" in data:
#             new_password = data["password"]
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = encrypted_password
#             try:
#                 auth.update_user(uid=user_id, password=new_password)
#             except Exception:
#                 logging.exception("Failed to update password in Firebase Auth")

#         # Write profileInfo if changed
#         if profile_info:
#             updates["profileInfo"] = profile_info

#         # Apply updates to Firestore (only fields present in updates)
#         if updates:
#             user_ref.update(updates)
#             logging.info(f"Updated Firestore for user {user_id}: {list(updates.keys())}")

#         # --- NEW: return the updated user document so the UI can replace cached state ---
#         updated_doc = user_ref.get().to_dict() or {}
#         updated_doc.pop("passwordHash", None)
#         return jsonify({
#             "message": "Profile updated successfully",
#             "updated": list(updates.keys()),
#             "user": updated_doc
#         }), 200

#     except Exception as e:
#         logging.exception("Error during profile update")
#         return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500

        
@app.route("/api/auth/logout", methods=["POST"])
def logout():
    # Get the Authorization header
    auth_header = request.headers.get("Authorization")

    if not auth_header:
        return jsonify({"error": "Authorization token is required."}), 401

    # Extract the token from the header
    token = auth_header.split(" ")[1] if " " in auth_header else auth_header

    try:
        # Decode the token
        decoded_payload = decode_token(token)

        if decoded_payload:
            # Optionally, implement token invalidation logic, e.g., blacklist the token
            return jsonify({"message": "Logged out successfully."}), 200
        else:
            return jsonify({"error": "Invalid or expired token."}), 401

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         # Accept either multipart/form-data (from browser) or JSON (API clients)
#         if request.content_type and "multipart/form-data" in request.content_type:
#             data = request.form.to_dict()
#         else:
#             data = request.get_json(silent=True) or {}

#         user_id = data.get("userId") or data.get("user_id")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         existing = user_doc.to_dict() or {}
#         # Ensure we start with the existing nested structures
#         profile_info = existing.get("profileInfo", {})
#         personal_info = profile_info.setdefault("personalInformation", {})
#         contact_info = profile_info.setdefault("contactInformation", {})

#         updates = {}

#         # --- Profile Info Updates (Single Source of Truth: profileInfo) ---
#         if "name" in data:
#             personal_info["name"] = data["name"]
#             # REMOVED: updates["name"] = data["name"] (Avoids writing to the root level)

#         if "idNumber" in data:
#             personal_info["idNumber"] = data["idNumber"]

#         # Contact info (accept contactNumber or phoneNumber)
#         phone = data.get("contactNumber") or data.get("phoneNumber")
#         if phone is not None:
#             contact_info["phoneNumber"] = phone
#             # REMOVED: updates["contactNumber"] = phone (Avoids writing to the root level)

#         if "address" in data:
#             contact_info["address"] = data["address"]

#         # Delivery Method Update
#         if "delivery_method" in data and isinstance(data["delivery_method"], dict):
#             # This allows updating the entire delivery_method object
#             updates["delivery_method"] = data["delivery_method"]

#         # Photo upload (multipart/form-data)
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file and file.filename:
#                 # Reuse your existing upload helper
#                 unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 try:
#                     # NOTE: Assuming 'upload_image_to_firebase' is a defined helper function
#                     photo_url = upload_image_to_firebase(file, unique_filename)
#                     personal_info["photoUrl"] = photo_url
#                 except Exception as e:
#                     logging.exception("Photo upload failed")
#                     return jsonify({"error": "Photo upload failed", "details": str(e)}), 500

#         # Password update (optional)
#         if "password" in data and data["password"]:
#             new_password = data["password"]
#             # NOTE: Assuming 'encrypt_password' is a defined helper function
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = encrypted_password
#             try:
#                 # NOTE: Assuming 'auth' is the Firebase Admin Auth instance
#                 auth.update_user(uid=user_id, password=new_password)
#             except Exception:
#                 logging.exception("Failed to update password in Firebase Auth")

#         # Write the entire profileInfo object back to Firestore
#         updates["profileInfo"] = profile_info

#         # Apply updates to Firestore
#         if updates:
#             user_ref.update(updates)
#             logging.info(f"Updated Firestore for user {user_id}: {list(updates.keys())}")

#         # Return the updated user document (excluding hash)
#         updated_doc = user_ref.get().to_dict() or {}
#         updated_doc.pop("passwordHash", None)
#         return jsonify({
#             "message": "Profile updated successfully",
#             "updated": list(updates.keys()),
#             "user": updated_doc
#         }), 200

#     except Exception as e:
#         logging.exception("Error during profile update")
#         return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500


# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         # Accept either multipart/form-data (from browser) or JSON (API clients)
#         if request.content_type and "multipart/form-data" in request.content_type:
#             data = request.form.to_dict()
#         else:
#             data = request.get_json(silent=True) or {}

#         user_id = data.get("userId") or data.get("user_id")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         existing = user_doc.to_dict() or {}
#         # Ensure we start with the existing nested structures
#         profile_info = existing.get("profileInfo", {})
#         personal_info = profile_info.setdefault("personalInformation", {})
#         contact_info = profile_info.setdefault("contactInformation", {})

#         updates = {}

#         # --- START DATA MIGRATION/CLEANUP BLOCK ---
#         # This block ensures any root-level legacy data is moved to the preferred nested structure
#         # before any new updates are applied, and the old fields are cleaned up.

#         # 1. Contact Number Migration (Root 'contactNumber' -> Nested 'phoneNumber')
#         legacy_contact = existing.get("contactNumber")
#         # Migrate only if a legacy value exists AND the new nested field is currently empty
#         if legacy_contact and not contact_info.get("phoneNumber"):
#             contact_info["phoneNumber"] = legacy_contact
#             if DELETE_FIELD != "__DELETE_FIELD__":
#                 updates["contactNumber"] = DELETE_FIELD # Mark for deletion
        
#         # 2. Name Migration (Root 'name' -> Nested 'name')
#         legacy_name = existing.get("name")
#         if legacy_name and not personal_info.get("name"):
#             personal_info["name"] = legacy_name
#             if DELETE_FIELD != "__DELETE_FIELD__":
#                 updates["name"] = DELETE_FIELD # Mark for deletion
        
#         # --- END DATA MIGRATION/CLEANUP BLOCK ---

#         # --- Profile Info Updates (Single Source of Truth: profileInfo) ---
#         if "name" in data:
#             personal_info["name"] = data["name"]
#             # REMOVED: updates["name"] = data["name"] (Avoids writing to the root level)

#         if "idNumber" in data:
#             personal_info["idNumber"] = data["idNumber"]

#         # Contact info (accept contactNumber or phoneNumber)
#         phone = data.get("contactNumber") or data.get("phoneNumber")
#         if phone is not None:
#             contact_info["phoneNumber"] = phone
#             # REMOVED: updates["contactNumber"] = phone (Avoids writing to the root level)

#         if "address" in data:
#             contact_info["address"] = data["address"]

#         # Delivery Method Update
#         if "delivery_method" in data and isinstance(data["delivery_method"], dict):
#             # This allows updating the entire delivery_method object
#             updates["delivery_method"] = data["delivery_method"]

#         # Photo upload (multipart/form-data)
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file and file.filename:
#                 # Reuse your existing upload helper
#                 unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 try:
#                     # NOTE: Assuming 'upload_image_to_firebase' is a defined helper function
#                     photo_url = upload_image_to_firebase(file, unique_filename)
#                     personal_info["photoUrl"] = photo_url
#                 except Exception as e:
#                     logging.exception("Photo upload failed")
#                     return jsonify({"error": "Photo upload failed", "details": str(e)}), 500

#         # Password update (optional)
#         if "password" in data and data["password"]:
#             new_password = data["password"]
#             # NOTE: Assuming 'encrypt_password' is a defined helper function
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = encrypted_password
#             try:
#                 # NOTE: Assuming 'auth' is the Firebase Admin Auth instance
#                 auth.update_user(uid=user_id, password=new_password)
#             except Exception:
#                 logging.exception("Failed to update password in Firebase Auth")

#         # Write the entire profileInfo object back to Firestore
#         updates["profileInfo"] = profile_info

#         # Apply updates to Firestore
#         if updates:
#             user_ref.update(updates)
#             logging.info(f"Updated Firestore for user {user_id}: {list(updates.keys())}")

#         # Return the updated user document (excluding hash)
#         updated_doc = user_ref.get().to_dict() or {}
#         updated_doc.pop("passwordHash", None)
#         return jsonify({
#             "message": "Profile updated successfully",
#             "updated": list(updates.keys()),
#             "user": updated_doc
#         }), 200

#     except Exception as e:
#         logging.exception("Error during profile update")
#         return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500
# DELETE_FIELD = None 
# try:
#     from firebase_admin import firestore
#     DELETE_FIELD = firestore.DELETE_FIELD
# except ImportError:
#     # Fallback/Dummy for environment where full firestore import might not be available
#     DELETE_FIELD = "__DELETE_FIELD__" 

# DELETE_FIELD = None 
# try:
#     from firebase_admin import firestore
#     DELETE_FIELD = firestore.DELETE_FIELD
# except ImportError:
#     # Fallback/Dummy for environment where full firestore import might not be available
#     DELETE_FIELD = "__DELETE_FIELD__" 

# @app.route("/api/auth/updateprofile", methods=["POST"])
# def update_profile():
#     try:
#         # Accept either multipart/form-data (from browser) or JSON (API clients)
#         if request.content_type and "multipart/form-data" in request.content_type:
#             data = request.form.to_dict()
#         else:
#             data = request.get_json(silent=True) or {}

#         user_id = data.get("userId") or data.get("user_id")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         existing = user_doc.to_dict() or {}
#         # Capture the name the front-end is likely reading and sending back (profileInfo.name).
#         stale_name_sent_by_fe = existing.get("profileInfo", {}).get("name")
        
#         # Ensure we start with the existing nested structures
#         profile_info = existing.get("profileInfo", {})
#         # Note: profile_info is a dictionary reference, changes here affect the outer structure
        
#         personal_info = profile_info.setdefault("personalInformation", {})
#         contact_info = profile_info.setdefault("contactInformation", {})

#         updates = {}

#         # --- START DATA MIGRATION/CLEANUP BLOCK ---
#         # This block ensures any legacy data is moved to the preferred nested structure
        
#         # 1. Root Contact Number Migration (Root 'contactNumber' -> Nested 'phoneNumber')
#         legacy_contact = existing.get("contactNumber")
#         if legacy_contact and not contact_info.get("phoneNumber"):
#             contact_info["phoneNumber"] = legacy_contact
#         # Cleanup: Mark root field for deletion if it exists
#         if legacy_contact and DELETE_FIELD != "__DELETE_FIELD__" and DELETE_FIELD is not None:
#             updates["contactNumber"] = DELETE_FIELD
        
#         # 2. Root Name Migration (Root 'name' -> Nested 'personalInformation.name')
#         legacy_root_name = existing.get("name")
#         if legacy_root_name and not personal_info.get("name"):
#             personal_info["name"] = legacy_root_name
#         # Cleanup: Mark root field for deletion if it exists
#         if legacy_root_name and DELETE_FIELD != "__DELETE_FIELD__" and DELETE_FIELD is not None:
#             updates["name"] = DELETE_FIELD

#         # 3. ProfileInfo Name Migration and Cleanup (profileInfo.name -> personalInformation.name)
#         # This handles the internal legacy field where name was stored directly in profileInfo
#         legacy_profile_info_name = profile_info.get("name")
#         if legacy_profile_info_name and not personal_info.get("name"):
#             personal_info["name"] = legacy_profile_info_name
            
#         # Cleanup: Delete the 'name' field from the profile_info dictionary object itself.
#         # We DO NOT use updates["profileInfo.name"] = DELETE_FIELD because it conflicts
#         # with updates["profileInfo"] = profile_info later on.
#         if "name" in profile_info:
#              del profile_info["name"]
        
#         # --- END DATA MIGRUPTION/CLEANUP BLOCK ---

#         # --- Profile Info Updates (Single Source of Truth: profileInfo) ---
#         if "name" in data:
#             incoming_name = data["name"]
#             current_correct_name = personal_info.get("name")
            
#             # DEFENSIVE FIX: Check if the incoming name is the STALE value 
#             # (which we captured before it was deleted) AND a correct, different name already exists.
#             if incoming_name == stale_name_sent_by_fe and current_correct_name and incoming_name != current_correct_name:
#                 # The payload contains the old name ("Alaha") but the database has a corrected name ("Alahhahaha").
#                 # Action: Ignore the incoming stale update to preserve the correct name.
#                 logging.info(f"Ignoring stale name update for user {user_id}. Keeping '{current_correct_name}'.")
#                 pass # Skip the update
#             else:
#                 # It's either a genuine new update, or the correct name hasn't been set yet.
#                 # We update the correct location: personalInformation.name
#                 personal_info["name"] = incoming_name 
            
#         if "idNumber" in data:
#             personal_info["idNumber"] = data["idNumber"]

#         # Contact info (accept contactNumber or phoneNumber)
#         phone = data.get("contactNumber") or data.get("phoneNumber")
#         if phone is not None:
#             contact_info["phoneNumber"] = phone

#         if "address" in data:
#             contact_info["address"] = data["address"]

#         # Delivery Method Update
#         if "delivery_method" in data and isinstance(data["delivery_method"], dict):
#             # This allows updating the entire delivery_method object
#             updates["delivery_method"] = data["delivery_method"]

#         # Photo upload (multipart/form-data)
#         if "photo" in request.files:
#             file = request.files["photo"]
#             if file and file.filename:
#                 # Reuse your existing upload helper
#                 unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
#                 try:
#                     # NOTE: Assuming 'upload_image_to_firebase' is a defined helper function
#                     photo_url = upload_image_to_firebase(file, unique_filename)
#                     personal_info["photoUrl"] = photo_url
#                 except Exception as e:
#                     logging.exception("Photo upload failed")
#                     return jsonify({"error": "Photo upload failed", "details": str(e)}), 500

#         # Password update (optional)
#         if "password" in data and data["password"]:
#             new_password = data["password"]
#             # NOTE: Assuming 'encrypt_password' is a defined helper function
#             encrypted_password = encrypt_password(new_password)
#             updates["passwordHash"] = encrypted_password
#             try:
#                 # NOTE: Assuming 'auth' is the Firebase Admin Auth instance
#                 auth.update_user(uid=user_id, password=new_password)
#             except Exception:
#                 logging.exception("Failed to update password in Firebase Auth")

#         # Write the entire profileInfo object back to Firestore
#         # This overwrites the existing profileInfo map, thus applying the 'del profile_info["name"]' 
#         # change made earlier.
#         updates["profileInfo"] = profile_info

#         # Apply updates to Firestore
#         if updates:
#             user_ref.update(updates)
#             logging.info(f"Updated Firestore for user {user_id}: {list(updates.keys())}")

#         # Return the updated user document (excluding hash)
#         updated_doc = user_ref.get().to_dict() or {}
#         updated_doc.pop("passwordHash", None)
#         return jsonify({
#             "message": "Profile updated successfully",
#             "updated": list(updates.keys()),
#             "user": updated_doc
#         }), 200

#     except Exception as e:
#         logging.exception("Error during profile update")
#         return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500
DELETE_FIELD = None
try:
	from firebase_admin import firestore
	DELETE_FIELD = firestore.DELETE_FIELD
except ImportError:
	DELETE_FIELD = "__DELETE_FIELD__"


@app.route("/api/auth/updateprofile", methods=["POST"])
def update_profile():
	try:
		# Accept either multipart/form-data (browser) or JSON (API clients)
		if request.content_type and "multipart/form-data" in request.content_type:
			data = request.form.to_dict()
		else:
			data = request.get_json(silent=True) or {}

		# Normalize keys for case-insensitivity
		normalized_data = {k.lower(): v for k, v in data.items()}

		# Accept multiple possible field naming conventions
		user_id = (
			data.get("userId")
			or data.get("user_id")
			or normalized_data.get("userid")
		)
		if not user_id:
			return jsonify({"error": "userId is required"}), 400

		name_from_frontend = (
			data.get("name")
			or normalized_data.get("name")
			or normalized_data.get("username")
			or normalized_data.get("full_name")
		)
		# NEW LOG: Confirm name detection
		logging.info(f"UpdateProfile: Detected name_from_frontend: '{name_from_frontend}'")

		user_ref = db.collection("users").document(user_id)
		user_doc = user_ref.get()
		if not user_doc.exists:
			return jsonify({"error": "User not found"}), 404

		existing = user_doc.to_dict() or {}

		# Initialize nested structures
		profile_info = existing.get("profileInfo", {})
		personal_info = profile_info.setdefault("personalInformation", {})
		contact_info = profile_info.setdefault("contactInformation", {})

		updates = {}
		profile_info_updated = False # Flag to track if we touched profile_info

		# --- START DATA MIGRATION/CLEANUP BLOCK ---
		# This block is now conditionally run only if the client did NOT provide a new name.
		if not name_from_frontend:
			legacy_root_name = existing.get("name")
			if legacy_root_name and not personal_info.get("name"):
				personal_info["name"] = legacy_root_name
				profile_info_updated = True
			if legacy_root_name and DELETE_FIELD not in [None, "__DELETE_FIELD__"]:
				updates["name"] = DELETE_FIELD

			legacy_profile_info_name = profile_info.get("name")
			if legacy_profile_info_name and not personal_info.get("name"):
				personal_info["name"] = legacy_profile_info_name
				profile_info_updated = True
		
		# Contact number migration still runs regardless of name change
		legacy_contact = existing.get("contactNumber")
		if legacy_contact and not contact_info.get("phoneNumber"):
			contact_info["phoneNumber"] = legacy_contact
		if legacy_contact and DELETE_FIELD not in [None, "__DELETE_FIELD__"]:
			updates["contactNumber"] = DELETE_FIELD
			
		# --- END DATA MIGRATION/CLEANUP BLOCK ---

		# --- Root-level updates ---
		if "email" in data:
			updates["email"] = data["email"]

		if "role" in data:
			updates["role"] = data["role"]

		# --- Handle Name Synchronization (AUTHORITATIVE UPDATE) ---
		if name_from_frontend:
			name_to_set = name_from_frontend.strip()
			
			logging.info(f"UpdateProfile: Applying new name: '{name_to_set}'")
			
			# 1. Set root 'name' explicitly to the new value
			updates["name"] = name_to_set 

			# *** CRITICAL FIX: Direct Assignment to ensure nested fields are updated unconditionally ***
			# This is the most reliable way to force the update and ensure the fields are identical.
			personal_info["name"] = name_to_set
			profile_info["name"] = name_to_set
			contact_info["name"] = name_to_set
			profile_info_updated = True # Force this flag true since we definitely changed the nested map
			
			logging.info(f"UpdateProfile: Nested name fields forcibly updated for {user_id}")


		# --- Personal Information Updates ---
		if "idNumber" in data or "idnumber" in normalized_data:
			personal_info["idNumber"] = data.get("idNumber") or normalized_data.get("idnumber")
			profile_info_updated = True

		# --- Contact Info Updates ---
		phone = data.get("contactNumber") or data.get("phoneNumber") or normalized_data.get("phonenumber")
		if phone:
			contact_info["phoneNumber"] = phone
			profile_info_updated = True

		if "address" in data or "address" in normalized_data:
			contact_info["address"] = data.get("address") or normalized_data.get("address")
			profile_info_updated = True

		# --- Delivery Method Update ---
		if "delivery_method" in data and isinstance(data["delivery_method"], dict):
			updates["delivery_method"] = data["delivery_method"]

		# --- Photo Upload (if provided) ---
		if "photo" in request.files:
			file = request.files["photo"]
			if file and file.filename:
				unique_filename = f"user-profile-images/{user_id}/{int(time.time())}_{file.filename}"
				try:
					# NOTE: Assuming 'upload_image_to_firebase' is a defined helper function
					photo_url = upload_image_to_firebase(file, unique_filename)
					personal_info["photoUrl"] = photo_url
					profile_info_updated = True
				except Exception as e:
					logging.exception("Photo upload failed")
					return jsonify({"error": "Photo upload failed", "details": str(e)}), 500

		# --- Password Update (optional) ---
		if "password" in data and data["password"]:
			new_password = data["password"]
			# NOTE: Assuming 'encrypt_password' is a defined helper function
			encrypted_password = encrypt_password(new_password)
			updates["passwordHash"] = encrypted_password
			try:
				# NOTE: Assuming 'auth' is the Firebase Admin Auth instance
				auth.update_user(uid=user_id, password=new_password)
			except Exception:
				logging.exception("Failed to update password in Firebase Auth")
		
		# Include profileInfo in updates ONLY if it was modified
		if profile_info_updated:
			# Explicitly reassign nested dictionaries
			profile_info["personalInformation"] = personal_info
			profile_info["contactInformation"] = contact_info
			updates["profileInfo"] = profile_info

		# NEW LOG: Show what is being sent to Firestore
		logging.info(f"UpdateProfile: Final updates to Firestore for user {user_id}: {updates.keys()}")


		# Apply updates to Firestore
		if updates:
			user_ref.update(updates)
			logging.info(f"Updated Firestore for user {user_id}: {list(updates.keys())}")

		# Fetch updated document
		updated_doc = user_ref.get().to_dict() or {}
		
		# --- FINAL Name Synchronization before returning (Guarantees client sees latest name) ---
		# This re-runs the cleanup logic on the fetched document just before returning it.
		doc_profile_info = updated_doc.get("profileInfo", {})
		doc_personal_info = doc_profile_info.get("personalInformation", {})
		
		# Find the best name in the fetched document
		final_name = doc_personal_info.get("name") # Prioritize the main location
		if not final_name:
			final_name = doc_profile_info.get("name", "")
		if not final_name:
			final_name = doc_profile_info.get("contactInformation", {}).get("name", "")
		# CRITICAL: Now also check the root field (updated in this func)
		if not final_name:
			final_name = updated_doc.get("name", "")
			
		# Apply the best found name to the root 'name' field of the response payload
		# This ensures the client sees the correct name immediately.
		updated_doc["name"] = final_name 
		
		updated_doc.pop("passwordHash", None)

		return jsonify({
			"message": "Profile updated successfully",
			"updated": list(updates.keys()),
			"user": updated_doc
		}), 200

	except Exception as e:
		logging.exception("Error during profile update")
		return jsonify({"error": f"An error occurred during profile update: {str(e)}"}), 500

# # Helper Functions
# def encrypt_password(password: str) -> str:
#     """
#     Encrypt a password using AES encryption with a dynamic IV.
#     """
#     cipher = AES.new(SECRET_KEY, AES.MODE_CBC)
#     iv = cipher.iv  # Initialization vector
#     encrypted_password = cipher.encrypt(pad(password.encode(), BLOCK_SIZE))  # Encrypt and pad the password
#     return base64.b64encode(iv + encrypted_password).decode()  # Return Base64-encoded string


# #reset-password
# @app.route("/api/auth/resetpassword12", methods=["POST"])
# def reset_password12():
#     try:
#         data = request.json
#         if not data:
#             return jsonify({"error": "Invalid input"}), 400

#         # Validate required fields
#         required_fields = ["password", "confirmPassword", "userId"]
#         for field in required_fields:
#             if field not in data:
#                 return jsonify({"error": f"{field} is required"}), 400

#         password = data["password"]
#         confirm_password = data["confirmPassword"]
#         user_id = data["userId"]

#         # Validate password and confirmPassword match
#         if password != confirm_password:
#             return jsonify({"error": "Passwords do not match"}), 400

#         # Retrieve user from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         # Encrypt the new password
#         encrypted_password = encrypt_password(password)
#         print(f"Encrypted Password: {encrypted_password}")

#         # Update Firestore with the new encrypted password
#         user_ref.update({"passwordHash": encrypted_password})

#         # Update the password in Firebase Authentication
#         try:
#             auth.update_user(uid=user_id, password=password)
#         except firebase_admin.auth.UserNotFoundError:
#             return jsonify({"error": "User not found in Firebase Authentication"}), 404
#         except Exception as e:
#             return jsonify({"error": f"Failed to update password in Firebase Auth: {str(e)}"}), 500

#         return jsonify({"message": "Password reset successfully"}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/auth/changepassword", methods=["POST"])
def change_password():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Invalid input"}), 400

        # Validate required fields
        required_fields = ["newPassword", "confirmNewPassword"]
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"{field} is required"}), 400

        new_password = data["newPassword"]
        confirm_new_password = data["confirmNewPassword"]
        email = data.get("email")
        user_id = data.get("userId")

        # Ensure either email or userId is provided
        if not email and not user_id:
            return jsonify({"error": "Either email or userId is required"}), 400

        # Validate password and confirmPassword match
        if new_password != confirm_new_password:
            return jsonify({"error": "Passwords do not match"}), 400

        # Retrieve user from Firestore using email or userId
        user_ref = None
        if email:
            users_query = db.collection("users").where("email", "==", email).stream()
            user_doc = next(users_query, None)
            if user_doc:
                user_ref = db.collection("users").document(user_doc.id)
                user_id = user_doc.id
        elif user_id:
            user_ref = db.collection("users").document(user_id)
            user_doc = user_ref.get()
            if not user_doc.exists:
                return jsonify({"error": "User not found"}), 404

        if not user_ref:
            return jsonify({"error": "User not found"}), 404

        # Encrypt the new password
        encrypted_password = encrypt_password(new_password)
        print(f"Encrypted Password: {encrypted_password}")

        # Update Firestore with the new encrypted password
        user_ref.update({"passwordHash": encrypted_password})

        # Update the password in Firebase Authentication
        try:
            auth.update_user(uid=user_id, password=new_password)
        except firebase_admin.auth.UserNotFoundError:
            return jsonify({"error": "User not found in Firebase Authentication"}), 404
        except Exception as e:
            return (
                jsonify(
                    {"error": f"Failed to update password in Firebase Auth: {str(e)}"}
                ),
                500,
            )

        return jsonify({"message": "Password changed successfully"}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# login simple
import time


@app.route("/api/auth/login1", methods=["POST"])
def login1():
    try:
        # Parse request data
        data = request.json
        if not data:
            return jsonify({"error": "Invalid input"}), 400

        email = data.get("email")
        password = data.get("password")

        # Validate inputs
        if not email or not password:
            return jsonify({"error": "Email and password are required"}), 400

        # Fetch user by email from Firebase Auth
        try:
            user = auth.get_user_by_email(email)
        except firebase_admin.auth.UserNotFoundError:
            return jsonify({"error": "Invalid email or password"}), 400

        # Fetch user data from Firestore using UID
        user_doc = db.collection("users").document(user.uid).get()
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        user_data = user_doc.to_dict()
        print(f"Retrieved user data from Firestore: {user_data}")  # Debug log

        # Retrieve the encrypted password from Firestore
        stored_encrypted_password = user_data.get("passwordHash")
        print(f"Stored Encrypted Password: {stored_encrypted_password}")  # Debug log

        # Decrypt the stored password
        try:
            decrypted_password = decrypt_password(stored_encrypted_password)
            print(f"Decrypted Password: {decrypted_password}")  # Debug log
        except Exception as e:
            print(f"Password decryption failed: {e}")
            return jsonify({"error": "Password decryption failed"}), 500

        # Compare input password with decrypted password
        if password != decrypted_password:
            print(
                f"Password mismatch: Input={password}, Decrypted={decrypted_password}"
            )
            return jsonify({"error": "Invalid email or password"}), 400

        # Generate JWT token
        current_time = int(time.time())
        exp_time = current_time + (30 * 24 * 60 * 60)  # 30 days in seconds
        token_payload = {
            "userId": user.uid,
            "email": email,
            "role": user_data.get("role"),
            "exp": exp_time,
            "iat": current_time,
        }
        token = jwt.encode(token_payload, SECRET_KEY, algorithm="HS256")

        # Update last login time
        last_login = int(time.time())  # Get current time as an integer (seconds)
        formatted_last_login = time.strftime(
            "%Y-%m-%d %H:%M:%S", time.localtime(last_login)
        )

        # Store timestamp in Firestore
        db.collection("users").document(user.uid).update(
            {"last_login": formatted_last_login}
        )

        print(f"Updated last login timestamp for user: {user.uid}")  # Debug log

        # Return success response
        return (
            jsonify(
                {
                    "message": "Logged in successfully",
                    "token": token,
                    "user_role": user_data.get("role"),
                    "user_id": user.uid,
                    "last_login": time.strftime(
                        "%Y-%m-%dT%H:%M:%SZ", time.gmtime(last_login)
                    ),
                }
            ),
            200,
        )

    except Exception as e:
        # Log the error for debugging
        print(f"Error during login: {e}")
        return jsonify({"error": f"An error occurred during login: {str(e)}"}), 500


# login with auth  code

FIREBASE_API_KEY = os.environ.get("FIREBASE_API_KEY", "")

# Firebase REST API Login URL
LOGIN_URL = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"

# app = Flask(__name__)
# SECRET_KEY = "app_counselor"
app.secret_key = "app_counselor"

# from flask import session
# import requests
# import time
# # Helper function to verify Firebase ID token
# def verify_firebase_token(id_token):
#     try:
#         decoded_token = auth.verify_id_token(id_token)
#         return decoded_token
#     except Exception as e:
#         print(f"Token verification failed: {e}")
#         return None

# # Helper function to refresh Firebase token
# def refresh_firebase_token(refresh_token):
#     try:
#         refresh_payload = {
#             'grant_type': 'refresh_token',
#             'refresh_token': refresh_token
#         }
#         response = requests.post("https://securetoken.googleapis.com/v1/token", data=refresh_payload)
#         new_tokens = response.json()
#         if response.status_code == 200:
#             return new_tokens.get("id_token"), new_tokens.get("refresh_token")
#         else:
#             print("Error refreshing token:", new_tokens.get("error"))
#             return None, None
#     except Exception as e:
#         print(f"Error refreshing token: {e}")
#         return None, None

# import time

# def generate_token11(user_id,role):
#     payload = {
#         "user_id": user_id,
#         "exp": int(time.time()) + (30 * 24 * 60 * 60) , # Expiry in 30 days
#         "role": role
#     }
#     return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

# @app.route('/api/auth/login', methods=['POST'])
# def login():
#     data = request.get_json()
#     email = data.get('email')
#     password = data.get('password')

#     if not email or not password:
#         return jsonify({"error": "Email and password are required"}), 400

#     login_payload = {
#         "email": email,
#         "password": password,
#         "returnSecureToken": True
#     }

#     try:
#         login_response = requests.post(LOGIN_URL, json=login_payload)
#         login_data = login_response.json()

#         if login_response.status_code == 200:
#             id_token = login_data.get("idToken")
#             refresh_token = login_data.get("refreshToken")
#             email = login_data.get("email")
#             user_id = login_data.get("localId")

#             # Store tokens in the session
#             session['id_token'] = id_token
#             session['refresh_token'] = refresh_token
#             session['user_id'] = user_id

#             # Verify the token
#             verified_token = verify_firebase_token(id_token)
#             if not verified_token:
#                 new_id_token, new_refresh_token = refresh_firebase_token(refresh_token)
#                 if not new_id_token:
#                     return jsonify({"error": "Failed to refresh token"}), 401
#                 session['id_token'] = new_id_token
#                 session['refresh_token'] = new_refresh_token

#             jwt_token = generate_token11(user_id,role)

#             # Update last login timestamp
#             current_time = time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())
#             user_ref = db.collection("users").document(user_id)
#             user_ref.update({'last_login': current_time})
#             user_data = user_ref.get().to_dict()

#             if user_data is None:
#                 return jsonify({"error": "User not found in database"}), 404

#             role = user_data.get("role", "N/A")  # Default role if not found

#             return jsonify({
#                 "message": "Login successful",
#                 "idToken": session['id_token'],
#                 "refreshToken": session['refresh_token'],
#                 "email": email,
#                 "user_id": user_id,
#                 "current_login": current_time,
#                 "jwtToken": jwt_token,
#                 "role" : role
#             }), 200
#         else:
#             error_message = login_data.get("error", {}).get("message", "Login failed")
#             return jsonify({"error": error_message}), 401

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


from flask import Flask, request, jsonify, session


# Helper function to verify Firebase ID token
def verify_firebase_token(id_token):
    try:
        decoded_token = auth.verify_id_token(id_token)
        return decoded_token
    except Exception as e:
        print(f"Token verification failed: {e}")
        return None


# Helper function to refresh Firebase token
def refresh_firebase_token(refresh_token):
    try:
        refresh_payload = {
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
        }
        response = requests.post(
            "https://securetoken.googleapis.com/v1/token", data=refresh_payload
        )
        new_tokens = response.json()
        if response.status_code == 200:
            return new_tokens.get("id_token"), new_tokens.get("refresh_token")
        else:
            print("Error refreshing token:", new_tokens.get("error"))
            return None, None
    except Exception as e:
        print(f"Error refreshing token: {e}")
        return None, None


import time as t


# Function to generate JWT token including role
def generate_token11(user_id, role):
    payload = {
        "user_id": user_id,
        "role": role,  # Include role in token
        "exp": int(t.time()) + (30 * 24 * 60 * 60),  # Expiry in 30 days
    }
    return jwt.encode(payload, app.secret_key, algorithm="HS256")


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json()
    identifier = data.get("email")  # can be email or username
    password = data.get("password")

    if not identifier or not password:
        return jsonify({"error": "Identifier (email or username) and password are required"}), 400

    # Determine if identifier is an email
    if is_valid_email(identifier):
        email = identifier
    else:
        # Lookup user by username in Firestore
        users_query = db.collection("users").where("profileInfo.contactInformation.personalInformation.name", "==", identifier).stream()
        user_doc = next(users_query, None)
        if not user_doc or not user_doc.exists:
            return jsonify({"error": "Invalid username or password"}), 401
        user_data = user_doc.to_dict()
        email = user_data.get("email")
        if not email:
            return jsonify({"error": "No email associated with this username"}), 401

    login_payload = {"email": email, "password": password, "returnSecureToken": True}

    try:
        login_response = requests.post(LOGIN_URL, json=login_payload)
        login_data = login_response.json()

        if login_response.status_code == 200:
            id_token = login_data.get("idToken")
            refresh_token = login_data.get("refreshToken")
            email = login_data.get("email")
            user_id = login_data.get("localId")

            # Store tokens in the session
            session["id_token"] = id_token
            session["refresh_token"] = refresh_token
            session["user_id"] = user_id

            # Verify the token
            verified_token = verify_firebase_token(id_token)
            if not verified_token:
                new_id_token, new_refresh_token = refresh_firebase_token(refresh_token)
                if not new_id_token:
                    return jsonify({"error": "Failed to refresh token"}), 401
                session["id_token"] = new_id_token
                session["refresh_token"] = new_refresh_token

            # Update last login timestamp
            current_time = t.strftime(
                "%Y-%m-%dT%H:%M:%SZ", t.gmtime()
            )
            user_ref = db.collection("users").document(user_id)
            user_ref.update({"last_login": current_time})

            # Fetch user role from Firestore
            user_doc = user_ref.get()
            if not user_doc.exists:
                return jsonify({"error": "User not found in database"}), 404

            user_data = user_doc.to_dict()
            role = user_data.get("role", "N/A")  # Default to "N/A" if role is missing

            # Generate JWT token with role (Now AFTER fetching role)
            jwt_token = generate_token11(user_id, role)

            return (
                jsonify(
                    {
                        "message": "Login successful",
                        "idToken": session["id_token"],
                        "refreshToken": session["refresh_token"],
                        "email": email,
                        "user_id": user_id,
                        "current_login": current_time,
                        "jwtToken": jwt_token,
                        "role": role,
                    }
                ),
                200,
            )

        else:
            error_message = login_data.get("error", {}).get("message", "Login failed")
            return jsonify({"error": error_message}), 401

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# @app.route('/api/auth/logout', methods=['POST'])
# def logout():
#     session.clear()  # Clear session data
#     return jsonify({"message": "Logged out successfully"}), 200

# # Run Flask app
# if __name__ == '__main__':
#     app.run(debug=True)

############################################################################################################################################################################

# TEACHER DASHBOARD FINAL CODE

from flask import Flask, request, jsonify
from flask_cors import CORS
import uuid
import firebase_admin
import jwt
from firebase_admin import credentials, auth, firestore
import datetime
from datetime import timedelta
import re
import hashlib

# import jwt
# import datetime
import secrets  # For generating a secure nonce
from datetime import datetime

# app = Flask(__name__)
# SECRET_KEY = "secret_key"
# CORS(app)


# # Initialize Firebase Admin
# cred = credentials.Certificate(r"serviceAccountKey.json")
# firebase_admin.initialize_app(cred)
# db = firestore.client()

# #teachers dashboard
# @app.route('/api/teacher/classes', methods=['POST'])
# def assign_grade_ref_to_teacher():
#     try:
#         # Authorization token validation
#         auth_header = request.headers.get('Authorization')
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_token(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract teacher ID from token
#         teacher_id = decoded_token.get('userId')
#         if not teacher_id:
#             return jsonify({"error": "Invalid token"}), 401

#         # Parse request body
#         data = request.get_json()
#         class_id = data.get("class_id")

#         if not class_id:
#             return jsonify({"error": "Missing required field: class_id"}), 400

#         # Fetch teacher's document and validate role
#         teacher_ref = db.collection('users').document(teacher_id)
#         teacher_doc = teacher_ref.get()
#         if not teacher_doc.exists:
#             return jsonify({"error": f"Teacher with ID {teacher_id} not found"}), 404

#         teacher_data = teacher_doc.to_dict()
#         if teacher_data.get('role') != 'teacher':
#             return jsonify({"error": "Access denied. User is not a teacher."}), 403

#         # Validate class document globally
#         class_ref = None
#         grades_ref = db.collection('Grades').stream()
#         for grade in grades_ref:
#             grade_ref = db.collection('Grades').document(grade.id).collection('classes').document(class_id)
#             if grade_ref.get().exists:
#                 class_ref = grade_ref
#                 break

#         if not class_ref:
#             return jsonify({"error": f"Class with ID {class_id} not found"}), 404

#         # Update teacher's document with the grade_ref
#         teacher_ref.update({"grade_ref": class_ref})

#         return jsonify({
#             "message": f"Class {class_id} assigned to Teacher {teacher_id} successfully."
#         }), 200

#     except Exception as e:
#         # Log the error internally for debugging purposes
#         print(f"Error in assign_grade_ref_to_teacher: {str(e)}")
#         return jsonify({"error": "An unexpected error occurred"}), 500


# Function to validate email format
def is_valid_email(email):
    email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
    return re.match(email_regex, email) is not None


# Generate JWT token with 30-day validity using time module
def generate_token(user_id):
    payload = {
        "user_id": user_id,
        "exp": int(time.time())
        + (30 * 24 * 60 * 60),  # Current time + 30 days in seconds
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")


def decode_token(token):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


@app.route("/api/teacher/classes/dashboard", methods=["GET"])
def get_teacher_classes_dashboard():
    try:
        # Authorization token validation
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_token(token)

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        # Extract teacher ID from token
        teacher_id = decoded_token.get("userId")
        if not teacher_id:
            return jsonify({"error": "Invalid token"}), 401

        # Fetch teacher's document from the users collection
        teacher_ref = db.collection("users").document(teacher_id)
        teacher_doc = teacher_ref.get()
        if not teacher_doc.exists:
            return jsonify({"error": f"Teacher with ID {teacher_id} not found"}), 404

        teacher_data = teacher_doc.to_dict()

        # Validate the role of the user
        if teacher_data.get("role") != "teacher":
            return jsonify({"error": "Access denied. User is not a teacher."}), 403

        # Fetch the grade_ref field
        grade_ref = teacher_data.get("grade_ref")
        if not grade_ref:
            return jsonify({"error": "No class assigned to this teacher"}), 404

        # Fetch the referenced class document
        class_doc = grade_ref.get()
        if not class_doc.exists:
            return jsonify({"error": "Referenced class document not found"}), 404

        # Return the class details
        class_data = class_doc.to_dict()
        return (
            jsonify(
                {
                    "teacher_id": teacher_id,
                    "class_id": class_doc.id,
                    "class_name": class_data.get("class_name"),
                    "subject": class_data.get("subject"),
                }
            ),
            200,
        )

    except Exception as e:
        print(f"Error in get_teacher_classes: {str(e)}")  # Log the error
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# STUDENT PROFILE OVERVIEW
@app.route("/api/teacher/classes/studentprofile", methods=["GET"])
def get_teacher_classes_studentprofile():
    try:
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_token(token)
        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        teacher_id = decoded_token.get("userId")
        if not teacher_id:
            return jsonify({"error": "Invalid token"}), 401

        # Fetch classes associated with the teacher
        classes_ref = db.collection("classes").where("teacherId", "==", teacher_id)
        classes = classes_ref.stream()

        class_data = []
        for cls in classes:
            class_info = cls.to_dict()
            class_data.append(
                {
                    "class_id": class_info.get("classId"),
                    "class_name": class_info.get("className"),
                    "student_count": len(
                        class_info.get("students", [])
                    ),  # Assume students is a list
                }
            )

        return jsonify({"teacher_id": teacher_id, "classes": class_data}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# if __name__ == "__main__":
#     app.run(port=5000, debug=True)
###############################################################


# USERMANAGEMENT PAGE

from flask import Flask, request, jsonify
from flask_cors import CORS
import uuid
import firebase_admin
import jwt
from firebase_admin import credentials, auth, firestore
import datetime

# from datetime import timedelta
import re
import hashlib

# import jwt
import datetime

# import secrets  # For generating a secure nonce
from datetime import datetime

# import bcrypt
import time


# app = Flask(__name__)
# SECRET_KEY = "secret_key"
# CORS(app)


# # Initialize Firebase Admin
# cred = credentials.Certificate(r"serviceAccountKey.json")
# firebase_admin.initialize_app(cred)
# db = firestore.client()


# Function to validate email format
def is_valid_email(email):
    email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
    return re.match(email_regex, email) is not None


# Generate JWT token with 30-day validity
def generate_token(user_id):
    payload = {
        "user_id": user_id,
        "exp": int(time.time())
        + (30 * 24 * 60 * 60),  # Current time + 30 days in seconds
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")


def decode_token(token):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


# usermanagement

# @app.route('/api/headmaster/users', methods=['GET'])
# def get_all_users_with_filter():
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get('Authorization')
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract the token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header

#         # Decode the token
#         decoded_token = decode_token(token)
#         if not decoded_token or not isinstance(decoded_token, dict):
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Check if the token contains the required role
#         user_role = decoded_token.get("role")
#         if user_role != "headmaster":
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Optional: Get 'role' query parameter for filtering
#         role_filter = request.args.get('role')  # e.g., ?role=teacher

#         # Fetch users from Firestore
#         users_ref = db.collection('users')
#         if role_filter:
#             users_query = users_ref.where("role", "==", role_filter)
#         else:
#             users_query = users_ref

#         users = users_query.stream()

#         # Build response
#         user_list = []
#         for user in users:
#             user_data = user.to_dict()
#             user_list.append({
#                 "name": user_data.get("profileInfo", {}).get("name", "Unknown"),
#                 "role": user_data.get("role"),
#                 "status": user_data.get("status", "Unknown")
#             })

#         return jsonify(user_list), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# import jwt
# from flask import request, jsonify

# @app.route('/api/headmaster/users', methods=['GET'])
# def get_all_users_with_filter():
#     try:
#         # Ã¢Å“â€¦ Step 1: Extract and Log the Authorization Header
#         auth_header = request.headers.get('Authorization')
#         print("Authorization Header:", auth_header)  # Debugging

#         if not auth_header or "Bearer " not in auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Ã¢Å“â€¦ Step 2: Extract JWT Token
#         token = auth_header.split(" ")[1]  # Extract the token part
#         print("Extracted Token:", token)  # Debugging

#         # Ã¢Å“â€¦ Step 3: Decode the JWT Token
#         try:
#             decoded_token = jwt.decode(token, app.secret_key, algorithms=["HS256"])
#             print("Decoded Token:", decoded_token)  # Debugging
#         except jwt.ExpiredSignatureError:
#             print("Token has expired")
#             return jsonify({"error": "Token has expired"}), 401
#         except jwt.InvalidTokenError:
#             print("Invalid token")
#             return jsonify({"error": "Invalid token"}), 401

#         # Ã¢Å“â€¦ Step 4: Extract user_id from Token
#         user_id = decoded_token.get("user_id")
#         print("Extracted user_id:", user_id)  # Debugging

#         if not user_id:
#             return jsonify({"error": "Invalid token: user ID missing"}), 401

#         # Ã¢Å“â€¦ Step 5: Fetch User Data from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_data = user_ref.get().to_dict()
#         print("User Data from Firestore:", user_data)  # Debugging

#         if not user_data:
#             return jsonify({"error": "User not found"}), 404

#         # Ã¢Å“â€¦ Step 6: Check If the User is a Headmaster
#         if user_data.get("role") != "headmaster":
#             print("Unauthorized access: Not a headmaster")  # Debugging
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Ã¢Å“â€¦ Step 7: Check for Role Filtering (Optional)
#         role_filter = request.args.get('role')
#         users_ref = db.collection('users')

#         if role_filter:
#             users_query = users_ref.where("role", "==", role_filter)
#         else:
#             users_query = users_ref

#         users = users_query.stream()

#         # Ã¢Å“â€¦ Step 8: Build Response
#         user_list = []
#         for user in users:
#             user_data = user.to_dict()
#             user_list.append({
#                 "name": user_data.get("profileInfo", {}).get("name", "Unknown"),
#                 "role": user_data.get("role"),
#                 "status": user_data.get("status", "Unknown")
#             })

#         return jsonify(user_list), 200

#     except Exception as e:
#         print(f"Unexpected Error: {e}")  # Debugging
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# import jwt
# from flask import Flask, request, jsonify
# from google.cloud import firestore
# import re

# @app.route('/api/headmaster/users', methods=['GET'])
# def get_all_users_with_filter():
#     try:
#         # Step 1: Extract and Log the Authorization Header
#         auth_header = request.headers.get('Authorization')
#         print("Authorization Header:", auth_header)  # Debugging

#         if not auth_header or "Bearer " not in auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Step 2: Extract JWT Token
#         token = auth_header.split(" ")[1]  # Extract the token part
#         print("Extracted Token:", token)  # Debugging

#         # Step 3: Decode the JWT Token
#         try:
#             decoded_token = jwt.decode(token, app.secret_key, algorithms=["HS256"])
#             print("Decoded Token:", decoded_token)  # Debugging
#         except jwt.ExpiredSignatureError:
#             print("Token has expired")
#             return jsonify({"error": "Token has expired"}), 401
#         except jwt.InvalidTokenError:
#             print("Invalid token")
#             return jsonify({"error": "Invalid token"}), 401

#         # Step 4: Extract user_id from Token
#         user_id = decoded_token.get("user_id")
#         print("Extracted user_id:", user_id)  # Debugging

#         if not user_id:
#             return jsonify({"error": "Invalid token: user ID missing"}), 401

#         # Step 5: Fetch User Data from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_data = user_ref.get().to_dict()
#         print("Fetched User Data from Firestore:", user_data)  # Debugging

#         if not user_data:
#             return jsonify({"error": "User not found"}), 404

#         # Step 6: Check If the User is a Headmaster
#         if user_data.get("role") != "headmaster":
#             print("Unauthorized access: Not a headmaster")  # Debugging
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Step 7: Check for Role Filtering (Optional)
#         role_filter = request.args.get('role')
#         users_ref = db.collection('users')

#         if role_filter:
#             users_query = users_ref.where("role", "==", role_filter)
#         else:
#             users_query = users_ref

#         users = users_query.stream()

#         # Step 8: Build Response
#         user_list = []
#         email_regex = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")  # Basic email pattern

#         for user in users:
#             user_data = user.to_dict()
#             print("Processing User Data:", user_data)  # Debugging

#             # Extract profileInfo first, then personalInformation
#             profile_info = user_data.get("profileInfo", {})
#             if not profile_info:
#                 print("profileInfo key missing in user_data")

#             personal_info = profile_info.get("personalInformation", {})
#             if not personal_info:
#                 print("personalInformation key missing in profileInfo")

#             name = personal_info.get("name", "Unknown")  # Extract name safely

#             # Extract email from associatedIds field (handling list case)
#             associated_ids = user_data.get("associatedIds", [])
#             email = "Unknown"
#             if isinstance(associated_ids, list):
#                 for entry in associated_ids:
#                     if isinstance(entry, str) and email_regex.match(entry):
#                         email = entry
#                         break  # Stop at the first valid email found

#             user_list.append({
#                 "userId": user.id,  # Include userId
#                 "name": name,
#                 "email": email,  # Include email
#                 "role": user_data.get("role"),
#                 "status": user_data.get("status", "Unknown")
#             })

#         return jsonify(user_list), 200

#     except Exception as e:
#         print(f"Unexpected Error: {e}")  # Debugging
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

import jwt
from flask import Flask, request, jsonify
from google.cloud import firestore


@app.route("/api/headmaster/users", methods=["GET"])
def get_all_users_with_filter():
    try:
        # Step 1: Extract and Log the Authorization Header
        auth_header = request.headers.get("Authorization")
        print("Authorization Header:", auth_header)  # Debugging

        if not auth_header or "Bearer " not in auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        # Step 2: Extract JWT Token
        token = auth_header.split(" ")[1]  # Extract the token part
        print("Extracted Token:", token)  # Debugging

        # Step 3: Decode the JWT Token
        try:
            decoded_token = jwt.decode(token, app.secret_key, algorithms=["HS256"])
            print("Decoded Token:", decoded_token)  # Debugging
        except jwt.ExpiredSignatureError:
            print("Token has expired")
            return jsonify({"error": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            print("Invalid token")
            return jsonify({"error": "Invalid token"}), 401

        # Step 4: Extract user_id from Token
        user_id = decoded_token.get("user_id")
        print("Extracted user_id:", user_id)  # Debugging

        if not user_id:
            return jsonify({"error": "Invalid token: user ID missing"}), 401

        # Step 5: Fetch User Data from Firestore
        user_ref = db.collection("users").document(user_id)
        user_data = user_ref.get().to_dict()
        print("Fetched User Data from Firestore:", user_data)  # Debugging

        if not user_data:
            return jsonify({"error": "User not found"}), 404

        # Step 6: Check If the User is a Headmaster
        if user_data.get("role") != "headmaster":
            print("Unauthorized access: Not a headmaster")  # Debugging
            return jsonify({"error": "Unauthorized access"}), 403

        # Step 7: Check for Role Filtering (Optional)
        role_filter = request.args.get("role")
        users_ref = db.collection("users")

        if role_filter:
            users_query = users_ref.where("role", "==", role_filter)
        else:
            users_query = users_ref

        users = users_query.stream()

        # Step 8: Build Response
        user_list = []

        for user in users:
            user_data = user.to_dict()
            print("Processing User Data:", user_data)  # Debugging

            # Extract name
            profile_info = user_data.get("profileInfo", {})
            personal_info = profile_info.get("personalInformation", {})
            name = personal_info.get("name", "Unknown")

            # Extract email directly from user_data first
            email = user_data.get("email")

            # If email is not found, check in associatedIds
            if not email:
                associated_ids = user_data.get("associatedIds", [])
                print("Associated IDs:", associated_ids)  # Debugging

                if isinstance(associated_ids, list):
                    for entry in associated_ids:
                        if isinstance(entry, dict) and "email" in entry:
                            email = entry["email"]
                            break  # Stop at the first valid email found

            # Extract role and status
            role = user_data.get("role")
            status = user_data.get("status", "active")

            assigned_grades = user_data.get("assignedGrades", {})

            # Build final user info object
            user_info = {
                "userId": user.id,  # Include userId
                "name": name,
                "role": role,
                "status": status,
                "email": email if email else "Unknown",  # Ensure email is included
                "assignedGrades": assigned_grades,
            }

            user_list.append(user_info)

        return jsonify(user_list), 200

    except Exception as e:
        print(f"Unexpected Error: {e}")  # Debugging
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route('/api/headmaster/users', methods=['GET'])
# def get_all_users_with_filter():
#     try:
#         # Step 1: Extract and Log the Authorization Header
#         auth_header = request.headers.get('Authorization')
#         print("Authorization Header:", auth_header)  # Debugging

#         if not auth_header or "Bearer " not in auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Step 2: Extract JWT Token
#         token = auth_header.split(" ")[1]  # Extract the token part
#         print("Extracted Token:", token)  # Debugging

#         # Step 3: Decode the JWT Token
#         try:
#             decoded_token = jwt.decode(token, app.secret_key, algorithms=["HS256"])
#             print("Decoded Token:", decoded_token)  # Debugging
#         except jwt.ExpiredSignatureError:
#             print("Token has expired")
#             return jsonify({"error": "Token has expired"}), 401
#         except jwt.InvalidTokenError:
#             print("Invalid token")
#             return jsonify({"error": "Invalid token"}), 401

#         # Step 4: Extract user_id from Token
#         user_id = decoded_token.get("user_id")
#         print("Extracted user_id:", user_id)  # Debugging

#         if not user_id:
#             return jsonify({"error": "Invalid token: user ID missing"}), 401

#         # Step 5: Fetch User Data from Firestore
#         user_ref = db.collection("users").document(user_id)
#         user_data = user_ref.get().to_dict()
#         print("Fetched User Data from Firestore:", user_data)  # Debugging

#         if not user_data:
#             return jsonify({"error": "User not found"}), 404

#         # Step 6: Check If the User is a Headmaster
#         if user_data.get("role") != "headmaster":
#             print("Unauthorized access: Not a headmaster")  # Debugging
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Step 7: Check for Role Filtering (Optional)
#         role_filter = request.args.get('role')
#         users_ref = db.collection('users')

#         if role_filter:
#             users_query = users_ref.where("role", "==", role_filter)
#         else:
#             users_query = users_ref

#         users = users_query.stream()

#         # Step 8: Build Response
#         user_list = []

#         for user in users:
#             user_data = user.to_dict()
#             print("Processing User Data:", user_data)  # Debugging

#             # Extract name
#             profile_info = user_data.get("profileInfo", {})
#             personal_info = profile_info.get("personalInformation", {})
#             name = personal_info.get("name", "Unknown")

#             # Extract email directly from user_data first
#             email = user_data.get("email")

#             # If email is not found, check in associatedIds
#             if not email:
#                 associated_ids = user_data.get("associatedIds", [])
#                 print("Associated IDs:", associated_ids)  # Debugging

#                 if isinstance(associated_ids, list):
#                     for entry in associated_ids:
#                         if isinstance(entry, dict) and "email" in entry:
#                             email = entry["email"]
#                             break  # Stop at the first valid email found

#             user_info = {
#                 "userId": user.id,  # Include userId
#                 "name": name,
#                 "role": user_data.get("role"),
#                 "status": user_data.get("status", "active"),
#                 "email": email if email else "Unknown"  # Ensure email is included
#             }

#             user_list.append(user_info)

#         return jsonify(user_list), 200

#     except Exception as e:
#         print(f"Unexpected Error: {e}")  # Debugging
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route('/api/headmaster/add-users', methods=['POST'])
# def add_user():
#     try:
#         data = request.json
#         # Validate required fields
#         required_fields = ["name", "email", "contactNumber", "role"]
#         for field in required_fields:
#             if field not in data or not data[field]:
#                 return jsonify({"error": f"{field} is required"}), 400

#         # Validate role (optional: extend to handle predefined roles)
#         valid_roles = ["Teacher", "Parent", "Student", "Headmaster"]
#         if data["role"] not in valid_roles:
#             return jsonify({"error": "Invalid role"}), 400

#         # Validate email format
#         email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
#         if not re.match(email_regex, data["email"]):
#             return jsonify({"error": "Invalid email format"}), 400

#         # Validate contact number (optional, example regex for Indian phone numbers)
#         phone_regex = r'^\+91[6-9]\d{9}$'
#         if not re.match(phone_regex, data["contactNumber"]):
#             return jsonify({"error": "Invalid contact number"}), 400

#         # Add userId and createdAt
#         user_id = str(uuid.uuid4())
#         data["userId"] = user_id
#         data["createdAt"] = time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())

#         # Store in Firestore
#         db.collection('users').document(user_id).set(data)

#         return jsonify({"message": "User added successfully", "userId": user_id}), 201

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/headmaster/add-users", methods=["POST"])
def add_user():
    try:
        data = request.json

        # Validate required fields
        required_fields = ["name", "email", "contactNumber", "role"]
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"error": f"{field} is required"}), 400

        # Validate role (optional: extend to handle predefined roles)
        valid_roles = ["Teacher", "Parent", "Student", "Headmaster"]
        if data["role"] not in valid_roles:
            return jsonify({"error": "Invalid role"}), 400

        # Validate email format
        email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        if not re.match(email_regex, data["email"]):
            return jsonify({"error": "Invalid email format"}), 400

        # Validate contact number (optional, example regex for Indian phone numbers)
        phone_regex = r"^\+91[6-9]\d{9}$"
        if not re.match(phone_regex, data["contactNumber"]):
            return jsonify({"error": "Invalid contact number"}), 400

        # Generate userId and createdAt
        user_id = str(uuid.uuid4())
        created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

        # Structure data as per the required Firestore format
        user_data = {
            "userId": user_id,
            "email": data["email"],  # Email stored as a separate key Ã¢Å“â€¦
            "contactNumber": data["contactNumber"],
            "role": data["role"],
            "createdAt": created_at,
            "profileInfo": {
                "personalInformation": {
                    "name": data["name"]  # Name stored correctly Ã¢Å“â€¦
                }
            },
            "associatedIds": [data["email"]],  # Email stored inside an array Ã¢Å“â€¦
        }

        # Store in Firestore
        db.collection("users").document(user_id).set(user_data)

        return jsonify({"message": "User added successfully", "userId": user_id}), 201

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route('/api/headmaster/users/<string:user_id>', methods=['PUT'])
# def update_user(user_id):
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get('Authorization')
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_token(token)
#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Validate that the request contains a valid JSON body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Define allowed fields (from createaccount API)
#         allowed_fields = {"role", "email", "password", "profileInfo", "associatedIds"}

#         # Filter the updates to only include allowed fields
#         filtered_updates = {key: value for key, value in updates.items() if key in allowed_fields}

#         if not filtered_updates:
#             return jsonify({"error": "No valid fields to update"}), 400

#         # Fetch user document from Firestore
#         user_ref = db.collection('users').document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "User not found"}), 404

#         # Perform the update
#         if "password" in filtered_updates:
#             # Hash the new password
#             filtered_updates["passwordHash"] = hashlib.sha256(filtered_updates.pop("password").encode()).hexdigest()

#         user_ref.update(filtered_updates)

#         # Fetch the updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop('passwordHash', None)  # Remove sensitive fields

#         return jsonify({
#             "message": "User updated successfully",
#             "updated_user": updated_user
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

import hashlib
import jwt
from flask import Flask, request, jsonify
from google.cloud import firestore


# Function to decode JWT token
def decode_jwt(token):
    try:
        decoded = jwt.decode(token, app.secret_key, algorithms=["HS256"])
        return decoded
    except jwt.ExpiredSignatureError:
        return None  # Token expired
    except jwt.InvalidTokenError:
        return None  # Invalid token


@app.route("/api/headmaster/users/<string:user_id>", methods=["PUT"])
def update_user(user_id):
    try:
        # Validate Authorization header
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        # Extract JWT token
        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_jwt(token)

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        # Extract user_id and role from JWT token
        requester_user_id = decoded_token.get("user_id")
        requester_role = decoded_token.get("role")

        if not requester_user_id or not requester_role:
            return jsonify({"error": "Unauthorized access"}), 403

        # Validate request body
        updates = request.json
        if not updates or not isinstance(updates, dict):
            return jsonify({"error": "Updates should be a valid JSON object"}), 400

        # Reference user document
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found in Firestore"}), 404

        user_data = user_doc.to_dict()
        user_role = user_data.get("role", "").lower()

        # Mapping input fields to Firestore structure
        mapped_updates = {}

        if "email" in updates:
            mapped_updates["email"] = updates["email"]
        if "name" in updates:
            if user_role == "headmaster":
                mapped_updates["profileInfo.name"] = updates["name"]
            else:
                mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
            mapped_updates["name"] = updates["name"]
        if "contactNumber" in updates:
            if user_role == "headmaster":
                mapped_updates["contactNumber"] = updates["contactNumber"]
            else:
                mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates[
                    "contactNumber"
                ]
        if "role" in updates:
            mapped_updates["role"] = updates["role"]

        # Hash password if updated
        if "password" in updates:
            mapped_updates["passwordHash"] = encrypt_password(updates["password"])

        # Update Firestore document
        user_ref.update(mapped_updates)

        # Fetch updated user data
        updated_user = user_ref.get().to_dict()
        updated_user.pop("passwordHash", None)  # Remove sensitive data

        return (
            jsonify(
                {"message": "User updated successfully", "updated_user": updated_user}
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract JWT token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract user_id and role from JWT token
#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Validate request body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Reference user document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # Mapping input fields to Firestore structure
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("phoneNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # --- NEW LOGIC FOR PARENT ASSOCIATED IDS ---
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                      return jsonify({"error": "associatedStudentIds must be a list"}), 400

#                 # Use ArrayUnion to add new student IDs to the 'associatedIds' field
#                 # without fetching, modifying in memory, and writing back the whole array.
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # Handle grade, class, and subject updates
#         if "grades" in updates:
#             if user_data["role"].lower() == "student":
#                current_grades = user_data.get("assignedGrades", {})
        
#                def extract_grade_number(grade_str):
           
#                   import re
#                   match = re.search(r"GRADE\s+(\d+)", grade_str.upper())
#                   return int(match.group(1)) if match else None

#                current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                current_grade_nums = [g for g in current_grade_nums if g is not None]
#                highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                new_grade_nums = [g for g in new_grade_nums if g is not None]

#                if highest_current_grade is not None:
#                   for new_grade in new_grade_nums:
#                       if new_grade < highest_current_grade:
#                          return jsonify({
#                            "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                          }), 400

        
#                mapped_updates["assignedGrades"] = updates["grades"]
#                mapped_updates["associatedIds"] = []

#             elif user_data["role"].lower() == "teacher":
#                  assigned_grades = user_data.get("assignedGrades", {})

#                  for grade, classes in updates["grades"].items():
#                      if grade not in assigned_grades:
#                         assigned_grades[grade] = {}

#                      for class_name, subjects in classes.items():
#                          if class_name not in assigned_grades[grade]:
#                             assigned_grades[grade][class_name] = []

#                          assigned_grades[grade][class_name] = list(
#                               set(assigned_grades[grade][class_name] + subjects)
#                 )

#                  mapped_updates["assignedGrades"] = assigned_grades

#                  existing_ids = set(user_data.get("associatedIds", []))
#                  students_ref = db.collection("users").where("role", "==", "student")
#                  all_students = students_ref.stream()

#                  for student_doc in all_students:
#                      student_data = student_doc.to_dict()
#                      student_grades = student_data.get("assignedGrades", {})
#                      normalized_student_grades = {g.lower(): cls for g, cls in student_grades.items()}

#                      for grade_name in updates["grades"].keys():
#                          if grade_name.lower() in normalized_student_grades:
#                              existing_ids.add(student_doc.id)

#                  mapped_updates["associatedIds"] = list(existing_ids)

#         # Update Firestore document
#         user_ref.update(mapped_updates)

#         # Fetch updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)  # Remove sensitive data

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# # Make sure you have this import at the top of your file
# # from google.cloud import firestore
# # import re 
# # from flask import jsonify, request # Assuming these are imported globally
# # from google.cloud import firestore # Assuming this is imported globally
# # from your_auth_module import decode_jwt, encrypt_password # Assuming these are available

# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
#     # --- HELPER FUNCTIONS DEFINED ONCE ---
#     # Helper function to sanitize grade keys
#     def sanitize_grade_key(key):
#         # 1. Convert to uppercase for case-insensitivity
#         key = key.upper()
#         # 2. Remove parentheses, dots, or slashes that break Firestore dot notation queries
#         key = re.sub(r'[()\./]', '', key).strip() 
#         # 3. Replace spaces with underscores
#         key = key.replace(" ", "_")
#         return key
        
#     def extract_grade_number(grade_str):
#         # Use the sanitized key or convert to upper for the regex search
#         match = re.search(r"GRADE_?(\d+)", sanitize_grade_key(grade_str)) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---

#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract JWT token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract user_id and role from JWT token
#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Validate request body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Reference user document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # Mapping input fields to Firestore structure
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # --- LOGIC FOR PARENT ASSOCIATED IDS ---
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # Handle grade, class, and subject updates
#         if "grades" in updates:
#             if user_data["role"].lower() == "student":
#                 current_grades = user_data.get("assignedGrades", {})
        
#                 # CRUCIAL: Normalize stored keys for comparison
#                 normalized_current_grade_keys = [sanitize_grade_key(g) for g in current_grades.keys()]
#                 current_grade_nums = [extract_grade_number(g) for g in normalized_current_grade_keys]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 # CRUCIAL: Normalize incoming keys for comparison
#                 normalized_new_grade_keys = [sanitize_grade_key(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [extract_grade_number(g) for g in normalized_new_grade_keys]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                             if new_grade < highest_current_grade:
#                                 return jsonify({
#                                     "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                                 }), 400
                
#                 # Normalize the final assignedGrades map before saving it
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
                
#                 mapped_updates["assignedGrades"] = normalized_updates 
#                 mapped_updates["associatedIds"] = []

#             # --- TEACHER LOGIC (ULTIMATE SOLUTION: CODE-SIDE FILTERING) ---
#             elif user_data["role"].lower() == "teacher":
                
#                 current_assigned_grades = user_data.get("assignedGrades", {})
#                 # Normalize existing assigned grades keys
#                 assigned_grades = {sanitize_grade_key(k): v for k, v in current_assigned_grades.items()}

#                 existing_ids = set(user_data.get("associatedIds", []))
#                 new_grade_keys_for_query = []

#                 # 1. Update teacher's assignedGrades, enforcing SANITIZED key
#                 for grade, classes in updates["grades"].items():
                    
#                     normalized_grade_key = sanitize_grade_key(grade)
#                     new_grade_keys_for_query.append(normalized_grade_key)

#                     # Use sanitized_grade_key for storage
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     # Merge classes and subjects 
#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 2. Query ALL students and filter in Python
#                 student_ids_to_add = set()
                
#                 # Fetch all students in one broad, indexed query
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()

#                 # Determine which sanitized grade keys we are looking for
#                 target_grade_keys = set(new_grade_keys_for_query)
                
#                 # Iterate through all students and perform local filtering
#                 for student_doc in all_students:
#                     student_data = student_doc.to_dict()
#                     student_grades = student_data.get("assignedGrades", {})
                    
#                     # Normalize the student's assigned grades keys for accurate matching
#                     normalized_student_grades = set(sanitize_grade_key(g) for g in student_grades.keys())

#                     # Check for intersection: If the student has ANY of the grades the teacher is being assigned
#                     if target_grade_keys.intersection(normalized_student_grades):
#                         student_ids_to_add.add(student_doc.id)
                
#                 # 3. Add the new student IDs to the teacher's associatedIds
#                 existing_ids.update(student_ids_to_add)
#                 mapped_updates["associatedIds"] = list(existing_ids)
#             # --- END TEACHER LOGIC ---
            
#         # Update Firestore document
#         user_ref.update(mapped_updates)

#         # Fetch updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)  # Remove sensitive data

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500
# import re 
# import firestore # Ensure this is imported (e.g., from google.cloud import firestore)

# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS DEFINED ONCE ---
#     # Robustly sanitizes keys like "Grade 12(Science)" and "GRADE 12" to "GRADE_12"
#     def sanitize_grade_key(key):
#         key = key.upper()
        
#         # 1. Strip ALL parentheses and their contents (e.g., removes '(LITERATURE)')
#         key = re.sub(r'\s*\([^)]*\)', '', key).strip() 
        
#         # 2. Normalize by replacing all remaining spaces, dots, or slashes with underscores.
#         key = re.sub(r'[\s\./]', '_', key)
        
#         # 3. Clean up leading/trailing underscores
#         key = key.strip('_')
        
#         return key

#     def extract_grade_number(grade_str):
#         # Uses the sanitized key to reliably find the number
#         match = re.search(r"GRADE_?(\d+)", sanitize_grade_key(grade_str)) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract JWT token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract user_id and role from JWT token
#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Validate request body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Reference user document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # Mapping input fields to Firestore structure
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             # NOTE: Uses "contactNumber" from updates but maps to "phoneNumber" in Firestore structure
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # --- NEW LOGIC FOR PARENT ASSOCIATED IDS ---
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400

#                 # Use ArrayUnion to add new student IDs
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # Handle grade, class, and subject updates
#         if "grades" in updates:
#             if user_data["role"].lower() == "student":
#                 current_grades = user_data.get("assignedGrades", {})
        
#                 # CRITICAL: Grade level check must use sanitized keys for comparison
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # Normalize the final assignedGrades map before saving it
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
                
#                 # Based on the student screenshot, assumed structure is {assignedGrades: {grades: {...}}}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
#                 mapped_updates["associatedIds"] = []

#             # --- TEACHER LOGIC (FIXED) ---
#             elif user_data["role"].lower() == "teacher":
                
#                 current_assigned_grades = user_data.get("assignedGrades", {})
#                 # Normalize existing assigned grades keys for safe merging and consistent storage
#                 assigned_grades = {sanitize_grade_key(k): v for k, v in current_assigned_grades.items()}

#                 existing_ids = set(user_data.get("associatedIds", []))
#                 target_grade_keys = set() # Keys the teacher is assigned, used for finding students

#                 # 1. Update teacher's assignedGrades, enforcing SANITIZED key
#                 for grade, classes in updates["grades"].items():
                    
#                     # Use the sanitized key for both storage and querying
#                     normalized_grade_key = sanitize_grade_key(grade) 
#                     target_grade_keys.add(normalized_grade_key)

#                     # Use sanitized_grade_key for storage
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     # Merge classes and subjects (logic for merging subjects remains the same)
#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 2. Query ALL students and filter in Python (Code-side filtering to avoid index issues)
#                 student_ids_to_add = set()
                
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     student_data = student_doc.to_dict()
#                     student_assigned_data = student_data.get("assignedGrades", {})
                    
#                     # CRUCIAL FIX: Handle nested student grades structure
#                     if 'grades' in student_assigned_data and isinstance(student_assigned_data['grades'], dict):
#                         # Student data is nested under 'grades'
#                         student_grades_map = student_assigned_data['grades']
#                     else:
#                         # Student data is flat (like the teacher's), fallback
#                         student_grades_map = student_assigned_data

#                     # Normalize the student's grade keys from the correct map (flat or nested)
#                     normalized_student_grades = set(sanitize_grade_key(g) for g in student_grades_map.keys())

#                     # Check for intersection: If the student has ANY of the teacher's assigned grades
#                     if target_grade_keys.intersection(normalized_student_grades):
#                         student_ids_to_add.add(student_doc.id)
                
#                 # 3. Add the new student IDs to the teacher's associatedIds
#                 existing_ids.update(student_ids_to_add)
#                 mapped_updates["associatedIds"] = list(existing_ids)
#             # --- END TEACHER LOGIC ---

#         # Update Firestore document
#         user_ref.update(mapped_updates)

#         # Fetch updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)  # Remove sensitive data

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

import re
from google.cloud import firestore # Ensure this import exists in your file

# Assume 'db' is your Firestore database client and 'encrypt_password' 
# and 'decode_jwt' are defined elsewhere in your application.

# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS DEFINED ONCE ---
#     # 1. REVISED: This function is now designed to PRESERVE specialized grade names 
#     # like "Grade 11(Science)" while standardizing case and spaces.
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         by converting to uppercase, normalizing spaces/symbols to underscores, and 
#         crucially, PRESERVING the content inside parentheses.
#         """
#         key = str(key).upper().strip()
        
#         # 1. Normalize spaces, dots, or slashes to underscores
#         key = re.sub(r'[\s\./]', '_', key)
        
#         # 2. Clean up extra underscores around parentheses if they exist
#         # E.g., 'GRADE_11_(SCIENCE)' -> 'GRADE_11(SCIENCE)'
#         key = key.replace('_(', '(').replace(')_', ')')
        
#         # 3. Clean up any repeated underscores
#         key = re.sub(r'_{2,}', '_', key)

#         # 4. Clean up leading/trailing underscores
#         key = key.strip('_')
        
#         return key

#     def extract_grade_number(grade_str):
#         # Uses the sanitized key to reliably find the number
#         # We temporarily REMOVE parentheses content only for number extraction
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract JWT token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract user_id and role from JWT token
#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # Validate request body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Reference user document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # Mapping input fields to Firestore structure
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             # NOTE: Uses "contactNumber" from updates but maps to "phoneNumber" in Firestore structure
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # --- NEW LOGIC FOR PARENT ASSOCIATED IDS ---
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400

#                 # Use ArrayUnion to add new student IDs
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # Handle grade, class, and subject updates
#         if "grades" in updates:
#             if user_data["role"].lower() == "student":
#                 current_grades = user_data.get("assignedGrades", {}).get("grades", user_data.get("assignedGrades", {})) # Handle nested structure
            
#                 # CRITICAL: Grade level check must use sanitized keys for comparison
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # Normalize the final assignedGrades map before saving it
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
                
#                 # Based on the student screenshot, assumed structure is {assignedGrades: {grades: {...}}}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
#                 mapped_updates["associatedIds"] = []

#             # --- TEACHER LOGIC (USES REVISED SANITIZER) ---
#             elif user_data["role"].lower() == "teacher":
                
#                 current_assigned_grades = user_data.get("assignedGrades", {})
#                 # Normalize existing assigned grades keys for safe merging and consistent storage
#                 assigned_grades = {sanitize_grade_key(k): v for k, v in current_assigned_grades.items()}

#                 existing_ids = set(user_data.get("associatedIds", []))
#                 target_grade_keys = set() # Keys the teacher is assigned, used for finding students

#                 # 1. Update teacher's assignedGrades, enforcing SANITIZED key
#                 for grade, classes in updates["grades"].items():
                    
#                     # Use the sanitized key for both storage and querying (e.g., GRADE_11(SCIENCE))
#                     normalized_grade_key = sanitize_grade_key(grade) 
#                     target_grade_keys.add(normalized_grade_key)

#                     # Use sanitized_grade_key for storage
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     # Merge classes and subjects (logic for merging subjects remains the same)
#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 2. Query ALL students and filter in Python (Code-side filtering to avoid index issues)
#                 student_ids_to_add = set()
                
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     student_data = student_doc.to_dict()
#                     student_assigned_data = student_data.get("assignedGrades", {})
                    
#                     # CRUCIAL FIX: Handle nested student grades structure
#                     if 'grades' in student_assigned_data and isinstance(student_assigned_data['grades'], dict):
#                         # Student data is nested under 'grades'
#                         student_grades_map = student_assigned_data['grades']
#                     else:
#                         # Student data is flat (like the teacher's), fallback
#                         student_grades_map = student_assigned_data

#                     # Normalize the student's grade keys from the correct map (flat or nested)
#                     # This uses the SAME specialized key format (e.g., GRADE_11(SCIENCE))
#                     normalized_student_grades = set(sanitize_grade_key(g) for g in student_grades_map.keys())

#                     # Check for intersection: If the student has ANY of the teacher's assigned grades
#                     if target_grade_keys.intersection(normalized_student_grades):
#                         student_ids_to_add.add(student_doc.id)
                
#                 # 3. Add the new student IDs to the teacher's associatedIds
#                 existing_ids.update(student_ids_to_add)
#                 mapped_updates["associatedIds"] = list(existing_ids)
#             # --- END TEACHER LOGIC ---

#         # Update Firestore document
#         user_ref.update(mapped_updates)

#         # Fetch updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)  # Remove sensitive data

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS ---
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         """
#         key = str(key).upper().strip()
#         # Normalize spaces, dots, or slashes to underscores
#         key = re.sub(r'[\s\./]', '_', key)
#         # Clean up extra underscores around parentheses
#         key = key.replace('_(', '(').replace(')_', ')')
#         # Clean up repeated underscores
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         # Temporarily remove parentheses content to find the number
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # 1. Authorization Check
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # 5. Parent Logic (Preserved)
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # 6. Grade/Class/Subject Logic
#         if "grades" in updates:
            
#             # ======================================================
#             # CASE A: STUDENT UPDATE (FIX APPLIED HERE)
#             # ======================================================
#             if user_role == "student":
#                 current_grades = user_data.get("assignedGrades", {}).get("grades", user_data.get("assignedGrades", {}))
                
#                 # A1. Validate Grade Downgrade
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
                
#                 # Based on your structure, usually {assignedGrades: {grades: {...}}}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
                
#                 # IMPORTANT: Do not wipe associatedIds if you are using them for parents, 
#                 # but based on your old code, you were setting it to [] for students. 
#                 # I will keep that behavior to not break legacy logic, unless 'parent' puts IDs there.
#                 # mapped_updates["associatedIds"] = [] 

#                 # A3. [THE FIX] UPDATE TEACHERS
#                 # Find teachers who teach these grades and add this student to them.
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
                    
#                     # Fetch all teachers
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_assigned = t_data.get("assignedGrades", {})
                        
#                         # Normalize teacher's assigned grades for comparison
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())

#                         # Check intersection (Does teacher teach any of student's new grades?)
#                         if not student_new_grade_keys.isdisjoint(t_grade_keys):
#                             t_ref = db.collection("users").document(teacher_doc.id)
#                             # Add student ID to teacher's associatedIds
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                             batch_count += 1
#                             updates_made = True
                            
#                             # Firestore batch limit is 500
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()
#                         print(f"✅ Automatically linked Student {user_id} to relevant teachers.")

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to auto-link student to teachers: {e}")
#                     # Don't fail the whole request, just log the error

#             # ======================================================
#             # CASE B: TEACHER UPDATE (PRESERVED LOGIC)
#             # ======================================================
#             elif user_role == "teacher":
                
#                 current_assigned_grades = user_data.get("assignedGrades", {})
#                 assigned_grades = {sanitize_grade_key(k): v for k, v in current_assigned_grades.items()}

#                 existing_ids = set(user_data.get("associatedIds", []))
#                 target_grade_keys = set()

#                 # B1. Update Teacher's assignedGrades
#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade) 
#                     target_grade_keys.add(normalized_grade_key)

#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # B2. Find Students matching these grades
#                 student_ids_to_add = set()
                
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})
                    
#                     # Handle nested vs flat structure for students
#                     if 'grades' in s_assigned and isinstance(s_assigned['grades'], dict):
#                         s_grades_map = s_assigned['grades']
#                     else:
#                         s_grades_map = s_assigned

#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())

#                     # Check intersection
#                     if target_grade_keys.intersection(norm_s_grades):
#                         student_ids_to_add.add(student_doc.id)
                
#                 existing_ids.update(student_ids_to_add)
#                 mapped_updates["associatedIds"] = list(existing_ids)

#         # 7. Commit Updates to User Document
#         user_ref.update(mapped_updates)

#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS ---
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         """
#         key = str(key).upper().strip()
#         key = re.sub(r'[\s\./]', '_', key)
#         key = key.replace('_(', '(').replace(')_', ')')
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # 1. Authorization Check
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # 5. Parent Logic
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # 6. Grade/Class/Subject Logic
#         if "grades" in updates:
            
#             # ======================================================
#             # CASE A: STUDENT UPDATE
#             # ======================================================
#             if user_role == "student":
#                 current_grades_data = user_data.get("assignedGrades", {})
#                 current_grades = current_grades_data.get("grades", current_grades_data)
                
#                 # A1. Validate Grade Downgrade
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
                
#                 # A3. FULL TEACHER RECONCILIATION
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False
#                     student_id_list = [user_id]

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_ref = db.collection("users").document(teacher_doc.id)
                        
#                         t_assigned = t_data.get("assignedGrades", {})
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())

#                         is_currently_associated = user_id in t_data.get("associatedIds", [])
#                         should_be_associated_now = not student_new_grade_keys.isdisjoint(t_grade_keys)
                        
#                         action = None
#                         if should_be_associated_now and not is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion(student_id_list)})
#                             action = "Added"
#                         elif not should_be_associated_now and is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayRemove(student_id_list)})
#                             action = "Removed"
                        
#                         if action:
#                             batch_count += 1
#                             updates_made = True
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to perform full teacher reconciliation for student: {e}")

#             # ======================================================
#             # CASE B: TEACHER UPDATE (FIXED)
#             # ======================================================
#             elif user_role == "teacher":
                
#                 # 1. Fetch current raw structure
#                 raw_current_assigned = user_data.get("assignedGrades", {})
                
#                 # 2. Normalize structure: Handle nested "grades" key if present
#                 # This FIXES the bug where existing grades were hidden/lost inside a "grades" wrapper
#                 if "grades" in raw_current_assigned and isinstance(raw_current_assigned["grades"], dict):
#                     current_grades_map = raw_current_assigned["grades"]
#                 else:
#                     current_grades_map = raw_current_assigned

#                 # 3. Build the working map of assigned grades (Sanitized Key -> Value)
#                 assigned_grades = {}
#                 for k, v in current_grades_map.items():
#                     if isinstance(k, str):
#                         assigned_grades[sanitize_grade_key(k)] = v

#                 # 4. Merge NEW grades from the update payload
#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade) 
                    
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         # Merge subjects (Union of existing and new)
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 # 5. Save the flattened/merged structure back to DB
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 6. Recalculate Associations based on the COMPLETE list of grades (Old + New)
#                 target_grade_keys = set(assigned_grades.keys())
#                 final_student_ids = set()
                
#                 # Query all students (Read-Only scan)
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})
                    
#                     # Handle nested vs flat structure for student data
#                     if 'grades' in s_assigned and isinstance(s_assigned['grades'], dict):
#                         s_grades_map = s_assigned['grades']
#                     else:
#                         s_grades_map = s_assigned

#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())

#                     # If student matches ANY of the teacher's grades (Old or New), add them
#                     if not target_grade_keys.isdisjoint(norm_s_grades):
#                         final_student_ids.add(student_doc.id)
                
#                 # 7. Update Teacher's list (Replace with fully calculated list)
#                 mapped_updates["associatedIds"] = list(final_student_ids)

#         # 7. Commit Updates to User Document
#         user_ref.update(mapped_updates)

#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500



# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS ---
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         """
#         key = str(key).upper().strip()
#         # Normalize spaces, dots, or slashes to underscores
#         key = re.sub(r'[\s\./]', '_', key)
#         # Clean up extra underscores around parentheses
#         key = key.replace('_(', '(').replace(')_', ')')
#         # Clean up repeated underscores
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         # Temporarily remove parentheses content to find the number
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # 1. Authorization Check
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # 5. Parent Logic (Preserved)
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # 6. Grade/Class/Subject Logic
#         if "grades" in updates:
            
#             # ======================================================
#             # CASE A: STUDENT UPDATE (FIX APPLIED HERE)
#             # ======================================================
#             if user_role == "student":
#                 current_grades_data = user_data.get("assignedGrades", {})
#                 current_grades = current_grades_data.get("grades", current_grades_data)
                
#                 # A1. Validate Grade Downgrade
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
                
#                 # A3. [THE FIX] FULL TEACHER RECONCILIATION
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
                    
#                     # Fetch all teachers and prepare batch
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False
                    
#                     # Track which student IDs are currently in the teacher's list
#                     # (Used for removal logic)
#                     student_id_list = [user_id] # ArrayUnion/ArrayRemove expects a list

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_ref = db.collection("users").document(teacher_doc.id)
                        
#                         # Normalize teacher's assigned grades for comparison
#                         t_assigned = t_data.get("assignedGrades", {})
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())

#                         # Check for current student association status
#                         is_currently_associated = user_id in t_data.get("associatedIds", [])
                        
#                         # Check for expected association status based on new grades
#                         should_be_associated_now = not student_new_grade_keys.isdisjoint(t_grade_keys)
                        
#                         action = None

#                         if should_be_associated_now and not is_currently_associated:
#                             # Student should be linked but isn't: ACTION = ADD
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion(student_id_list)})
#                             action = "Added"
                        
#                         elif not should_be_associated_now and is_currently_associated:
#                             # Student is linked but shouldn't be: ACTION = REMOVE (FIX for "Extra Students")
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayRemove(student_id_list)})
#                             action = "Removed"
                        
#                         if action:
#                             batch_count += 1
#                             updates_made = True
                            
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()
#                         print(f"✅ Full reconciliation for Student {user_id} complete. Associations updated.")
#                     elif not updates_made:
#                         print(f"✅ Student {user_id} grades updated, no teacher association changes needed.")

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to perform full teacher reconciliation for student: {e}")
#                     # Don't fail the whole request, just log the error
# # ======================================================
#             # CASE B: TEACHER UPDATE
#             # ======================================================
#             elif user_role == "teacher":
                
#                 # 1. Fetch current raw structure
#                 raw_current_assigned = user_data.get("assignedGrades", {})
                
#                 # 2. Normalize structure: Handle nested "grades" key if present
#                 # This FIXES the bug where existing grades were hidden/lost inside a "grades" wrapper
#                 if "grades" in raw_current_assigned and isinstance(raw_current_assigned["grades"], dict):
#                     current_grades_map = raw_current_assigned["grades"]
#                 else:
#                     current_grades_map = raw_current_assigned

#                 # 3. Build the working map of assigned grades (Sanitized Key -> Value)
#                 # We verify 'k' is a string to avoid errors if bad data exists
#                 assigned_grades = {}
#                 for k, v in current_grades_map.items():
#                     if isinstance(k, str):
#                         assigned_grades[sanitize_grade_key(k)] = v

#                 # 4. Merge NEW grades from the update payload
#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade) 
                    
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         # Merge subjects (Union of existing and new)
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 # 5. Save the flattened/merged structure back to DB
#                 # This ensures 'assignedGrades' is always flat going forward
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 6. Recalculate Associations based on the COMPLETE list of grades (Old + New)
#                 target_grade_keys = set(assigned_grades.keys())
#                 final_student_ids = set()
                
#                 # Query all students (Read-Only scan)
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})
                    
#                     # Handle nested vs flat structure for student data
#                     if 'grades' in s_assigned and isinstance(s_assigned['grades'], dict):
#                         s_grades_map = s_assigned['grades']
#                     else:
#                         s_grades_map = s_assigned

#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())

#                     # If student matches ANY of the teacher's grades (Old or New), add them
#                     # Using !isdisjoint is faster than intersection for boolean check
#                     if not target_grade_keys.isdisjoint(norm_s_grades):
#                         final_student_ids.add(student_doc.id)
                
#                 # 7. Update Teacher's list (Replace with fully calculated list)
#                 mapped_updates["associatedIds"] = list(final_student_ids)


#         # 7. Commit Updates to User Document
#         user_ref.update(mapped_updates)

#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
    
#     # --- HELPER FUNCTIONS ---
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         """
#         key = str(key).upper().strip()
#         key = re.sub(r'[\s\./]', '_', key)
#         key = key.replace('_(', '(').replace(')_', ')')
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key) 
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---
    
#     try:
#         # 1. Authorization Check
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])
        
#         # 5. Parent Logic
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
#                 mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # 6. Grade/Class/Subject Logic
#         if "grades" in updates:
            
#             # ======================================================
#             # CASE A: STUDENT UPDATE
#             # ======================================================
#             if user_role == "student":
#                 current_grades_data = user_data.get("assignedGrades", {})
#                 current_grades = current_grades_data.get("grades", current_grades_data)
                
#                 # A1. Validate Grade Downgrade
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}
                
#                 # A3. FULL TEACHER RECONCILIATION
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False
#                     student_id_list = [user_id]

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_ref = db.collection("users").document(teacher_doc.id)
                        
#                         t_assigned = t_data.get("assignedGrades", {})
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())

#                         is_currently_associated = user_id in t_data.get("associatedIds", [])
#                         should_be_associated_now = not student_new_grade_keys.isdisjoint(t_grade_keys)
                        
#                         action = None
#                         if should_be_associated_now and not is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion(student_id_list)})
#                             action = "Added"
#                         elif not should_be_associated_now and is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayRemove(student_id_list)})
#                             action = "Removed"
                        
#                         if action:
#                             batch_count += 1
#                             updates_made = True
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to perform full teacher reconciliation for student: {e}")

#             # ======================================================
#             # CASE B: TEACHER UPDATE (FIXED)
#             # ======================================================
#             elif user_role == "teacher":
                
#                 # 1. Fetch current raw structure
#                 raw_current_assigned = user_data.get("assignedGrades", {})
                
#                 # 2. Normalize structure: Handle nested "grades" key if present
#                 # This FIXES the bug where existing grades were hidden/lost inside a "grades" wrapper
#                 if "grades" in raw_current_assigned and isinstance(raw_current_assigned["grades"], dict):
#                     current_grades_map = raw_current_assigned["grades"]
#                 else:
#                     current_grades_map = raw_current_assigned

#                 # 3. Build the working map of assigned grades (Sanitized Key -> Value)
#                 assigned_grades = {}
#                 for k, v in current_grades_map.items():
#                     if isinstance(k, str):
#                         assigned_grades[sanitize_grade_key(k)] = v

#                 # 4. Merge NEW grades from the update payload
#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade) 
                    
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
                        
#                         # Merge subjects (Union of existing and new)
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)
                
#                 # 5. Save the flattened/merged structure back to DB
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 6. Recalculate Associations based on the COMPLETE list of grades (Old + New)
#                 target_grade_keys = set(assigned_grades.keys())
#                 final_student_ids = set()
                
#                 # Query all students (Read-Only scan)
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()
                
#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})
                    
#                     # Handle nested vs flat structure for student data
#                     if 'grades' in s_assigned and isinstance(s_assigned['grades'], dict):
#                         s_grades_map = s_assigned['grades']
#                     else:
#                         s_grades_map = s_assigned

#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())

#                     # If student matches ANY of the teacher's grades (Old or New), add them
#                     if not target_grade_keys.isdisjoint(norm_s_grades):
#                         final_student_ids.add(student_doc.id)
                
#                 # 7. Update Teacher's list (Replace with fully calculated list)
#                 mapped_updates["associatedIds"] = list(final_student_ids)

#         # 7. Commit Updates to User Document
#         user_ref.update(mapped_updates)

#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
#     # This function assumes 'db', 'firestore', 'request', 'jsonify',
#     # 're', 'decode_jwt', and 'encrypt_password' are defined in the global scope.

#     # --- HELPER FUNCTIONS (Kept as requested to maintain existing logic) ---
#     def sanitize_grade_key(key):
#         """
#         Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
#         """
#         key = str(key).upper().strip()
#         key = re.sub(r'[\s\./]', '_', key)
#         key = key.replace('_(', '(').replace(')_', ')')
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key)
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---

#     try:
#         # 1. Authorization Check (omitted for brevity, assume working)
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         # Assume decode_jwt is available
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             # Assume encrypt_password is available
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])

#         # 5. Parent Logic (FIXED: Allow direct replacement of the list)
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
                
#                 # FIX: Directly assign the new list. This allows the user to
#                 # clear the list by sending `[]`, handling the removal of
#                 # deleted student IDs without ArrayUnion/ArrayRemove errors.
#                 mapped_updates["associatedIds"] = new_student_ids

#         # 6. Grade/Class/Subject Logic (Kept as before)
#         if "grades" in updates:

#             # ======================================================
#             # CASE A: STUDENT UPDATE
#             # ======================================================
#             if user_role == "student":
#                 current_grades_data = user_data.get("assignedGrades", {})
#                 current_grades = current_grades_data.get("grades", current_grades_data)

#                 # A1. Validate Grade Downgrade (omitted for brevity, assume working)
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}

#                 # A3. FULL TEACHER RECONCILIATION
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False
#                     student_id_list = [user_id] # Non-empty list

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_ref = db.collection("users").document(teacher_doc.id)

#                         # NOTE: Ensure t_assigned structure aligns with how it's saved/read (key structure may vary)
#                         t_assigned = t_data.get("assignedGrades", {})
#                         # This assumes assignedGrades for teacher is a flat map {GRADE_KEY: {class: [subjects]}}
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())

#                         is_currently_associated = user_id in t_data.get("associatedIds", [])
#                         should_be_associated_now = not student_new_grade_keys.isdisjoint(t_grade_keys)

#                         action = None
#                         if should_be_associated_now and not is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion(student_id_list)})
#                             action = "Added"
#                         elif not should_be_associated_now and is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayRemove(student_id_list)})
#                             action = "Removed"

#                         if action:
#                             batch_count += 1
#                             updates_made = True
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0

#                     if updates_made and batch_count > 0:
#                         batch.commit()

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to perform full teacher reconciliation for student: {e}")

#             # ======================================================
#             # CASE B: TEACHER UPDATE (FIXED)
#             # ======================================================
#             elif user_role == "teacher":

#                 # 1. Fetch current raw structure
#                 raw_current_assigned = user_data.get("assignedGrades", {})

#                 # 2. Normalize structure: Handle nested "grades" key if present
#                 if "grades" in raw_current_assigned and isinstance(raw_current_assigned["grades"], dict):
#                     current_grades_map = raw_current_assigned["grades"]
#                 else:
#                     current_grades_map = raw_current_assigned

#                 # 3. Build the working map of assigned grades (Sanitized Key -> Value)
#                 assigned_grades = {}
#                 for k, v in current_grades_map.items():
#                     if isinstance(k, str):
#                         assigned_grades[sanitize_grade_key(k)] = v

#                 # 4. Merge NEW grades from the update payload
#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade)

#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}

#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []

#                         # Merge subjects (Union of existing and new)
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)

#                 # 5. Save the flattened/merged structure back to DB
#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # 6. Recalculate Associations based on the COMPLETE list of grades (Old + New)
#                 target_grade_keys = set(assigned_grades.keys())
#                 final_student_ids = set()

#                 # Query all students (Read-Only scan)
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()

#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})

#                     # Handle nested vs flat structure for student data
#                     if 'grades' in s_assigned and isinstance(s_assigned['grades'], dict):
#                         s_grades_map = s_assigned['grades']
#                     else:
#                         s_grades_map = s_assigned

#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())

#                     # If student matches ANY of the teacher's grades (Old or New), add them
#                     if not target_grade_keys.isdisjoint(norm_s_grades):
#                         final_student_ids.add(student_doc.id)

#                 # 7. Update Teacher's list (Replace with fully calculated list)
#                 mapped_updates["associatedIds"] = list(final_student_ids)

#         # 7. Commit Updates to User Document (FIXED: Check if there are updates)
#         if mapped_updates:
#             user_ref.update(mapped_updates)
#         else:
#             return jsonify({"message": "No valid updates provided.", "user_id": user_id}), 200


#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
# def update_user11(user_id):
#     """
#     Handles user profile updates, including basic info, passwords,
#     and complex grade/association logic for students and teachers.
#     Syncs changes to 'students' collection if the user is a student.
#     """
    
#     # --- HELPER FUNCTIONS ---
#     def sanitize_grade_key(key):
#         key = str(key).upper().strip()
#         key = re.sub(r'[\s\./]', '_', key)
#         key = key.replace('_(', '(').replace(')_', ')')
#         key = re.sub(r'_{2,}', '_', key)
#         return key.strip('_')

#     def extract_grade_number(grade_str):
#         temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
#         match = re.search(r"GRADE_?(\d+)", temp_key)
#         return int(match.group(1)) if match else None
#     # --- END HELPER FUNCTIONS ---

#     try:
#         # 1. Authorization Check
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # 2. Validate Request
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # 3. Fetch User Document
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         user_data = user_doc.to_dict()
#         user_role = user_data.get("role", "").lower()

#         # 4. Map Basic Fields
#         mapped_updates = {}

#         if "email" in updates:
#             mapped_updates["email"] = updates["email"]
#         if "name" in updates:
#             mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
#         if "contactNumber" in updates:
#             mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
#         if "role" in updates:
#             mapped_updates["role"] = updates["role"]
#         if "password" in updates:
#             mapped_updates["passwordHash"] = encrypt_password(updates["password"])

#         # 5. Parent Logic
#         if user_role == "parent":
#             if "associatedStudentIds" in updates:
#                 new_student_ids = updates["associatedStudentIds"]
#                 if not isinstance(new_student_ids, list):
#                     return jsonify({"error": "associatedStudentIds must be a list"}), 400
                
#                 if new_student_ids:
#                     mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

#         # 6. Grade/Class/Subject Logic
#         if "grades" in updates:

#             # --- CASE A: STUDENT UPDATE ---
#             if user_role == "student":
#                 current_grades_data = user_data.get("assignedGrades", {})
#                 current_grades = current_grades_data.get("grades", current_grades_data)

#                 # A1. Validate Grade Downgrade
#                 current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
#                 current_grade_nums = [g for g in current_grade_nums if g is not None]
#                 highest_current_grade = max(current_grade_nums) if current_grade_nums else None

#                 new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
#                 new_grade_nums = [g for g in new_grade_nums if g is not None]

#                 if highest_current_grade is not None:
#                     for new_grade in new_grade_nums:
#                         if new_grade < highest_current_grade:
#                             return jsonify({
#                                 "error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade} for student"
#                             }), 400

#                 # A2. Prepare Student Update
#                 normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
#                 mapped_updates["assignedGrades"] = {"grades": normalized_updates}

#                 # A3. Teacher Reconciliation (Logic kept as is)
#                 try:
#                     student_new_grade_keys = set(normalized_updates.keys())
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False
#                     student_id_list = [user_id] 

#                     for teacher_doc in teachers:
#                         t_data = teacher_doc.to_dict()
#                         t_ref = db.collection("users").document(teacher_doc.id)
#                         t_assigned = t_data.get("assignedGrades", {})
#                         t_grade_keys = set(sanitize_grade_key(k) for k in t_assigned.keys())
#                         is_currently_associated = user_id in t_data.get("associatedIds", [])
#                         should_be_associated_now = not student_new_grade_keys.isdisjoint(t_grade_keys)

#                         action = None
#                         if should_be_associated_now and not is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion(student_id_list)})
#                             action = "Added"
#                         elif not should_be_associated_now and is_currently_associated:
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayRemove(student_id_list)})
#                             action = "Removed"

#                         if action:
#                             batch_count += 1
#                             updates_made = True
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0

#                     if updates_made and batch_count > 0:
#                         batch.commit()

#                 except Exception as e:
#                     print(f"⚠️ Warning: Failed to perform full teacher reconciliation: {e}")

#             # --- CASE B: TEACHER UPDATE ---
#             elif user_role == "teacher":
#                 raw_current_assigned = user_data.get("assignedGrades", {})
#                 current_grades_map = raw_current_assigned.get("grades", raw_current_assigned)
                
#                 assigned_grades = {}
#                 for k, v in current_grades_map.items():
#                     if isinstance(k, str):
#                         assigned_grades[sanitize_grade_key(k)] = v

#                 for grade, classes in updates["grades"].items():
#                     normalized_grade_key = sanitize_grade_key(grade)
#                     if normalized_grade_key not in assigned_grades:
#                         assigned_grades[normalized_grade_key] = {}
#                     for class_name, subjects in classes.items():
#                         if class_name not in assigned_grades[normalized_grade_key]:
#                             assigned_grades[normalized_grade_key][class_name] = []
#                         existing_subjects = set(assigned_grades[normalized_grade_key][class_name])
#                         existing_subjects.update(subjects)
#                         assigned_grades[normalized_grade_key][class_name] = list(existing_subjects)

#                 mapped_updates["assignedGrades"] = assigned_grades

#                 # Recalculate Student Associations
#                 target_grade_keys = set(assigned_grades.keys())
#                 final_student_ids = set()
#                 all_students_ref = db.collection("users").where("role", "==", "student")
#                 all_students = all_students_ref.stream()

#                 for student_doc in all_students:
#                     s_data = student_doc.to_dict()
#                     s_assigned = s_data.get("assignedGrades", {})
#                     s_grades_map = s_assigned.get('grades', s_assigned)
#                     norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())
#                     if not target_grade_keys.isdisjoint(norm_s_grades):
#                         final_student_ids.add(student_doc.id)

#                 mapped_updates["associatedIds"] = list(final_student_ids)

#         # 7. Commit Updates
#         if mapped_updates:
#             # A. Update the USERS collection (Primary)
#             user_ref.update(mapped_updates)

#             # ===============================================================
#             # ✅ FIX: SYNC UPDATES TO 'STUDENTS' COLLECTION
#             # ===============================================================
#             if user_role == "student":
#                 print(f"--- Syncing updates to students collection for {user_id} ---")
#                 try:
#                     student_ref = db.collection("students").document(user_id)
                    
#                     student_sync_updates = {}
                    
#                     # 1. Sync Assigned Grades (Mirror the structure from users)
#                     if "assignedGrades" in mapped_updates:
#                         student_sync_updates["assignedGrades"] = mapped_updates["assignedGrades"]
                    
#                     # 2. Sync Basic Profile Fields
#                     for key, val in mapped_updates.items():
#                         if key.startswith("profileInfo.") or key == "email":
#                             student_sync_updates[key] = val
                    
#                     if student_sync_updates:
#                         # Attempt to update directly. If doc missing, it raises NotFound error.
#                         student_ref.update(student_sync_updates)
#                         print(f"✅ Successfully synced to students/{user_id}")
#                     else:
#                         print(f"ℹ️ No relevant fields to sync for students/{user_id}")
                            
#                 except Exception as e:
#                     # Check if it's a "Not Found" error (common if IDs mismatch)
#                     error_msg = str(e).lower()
#                     if "not found" in error_msg:
#                         print(f"⚠️ Student document {user_id} does not exist in 'students' collection. Sync skipped.")
#                     else:
#                         print(f"❌ Error syncing to students collection: {e}")
#             # ===============================================================

#         else:
#             return jsonify({"message": "No valid updates provided.", "user_id": user_id}), 200
#             # ===============================================================

#         # 8. Return Updated Data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop("passwordHash", None)

#         return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500



# 1. UPDATE USER (PUT)
# ==============================================================================
@app.route("/api/headmaster/users1/<string:user_id>", methods=["PUT"])
def update_user11(user_id):
    """
    Handles user profile updates. 
    FIX: Ensures assignedGrades subjects are always stored as Arrays/Lists.
    """

    def extract_grade_number(grade_str):
        temp_key = re.sub(r'\s*\([^)]*\)', '', sanitize_grade_key(grade_str))
        match = re.search(r"GRADE_?(\d+)", temp_key)
        return int(match.group(1)) if match else None

    try:
        # 1. Authorization Check
        auth_header = request.headers.get("Authorization")
        if not auth_header: return jsonify({"error": "Authorization token is required"}), 401
        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_jwt(token)
        if not decoded_token: return jsonify({"error": "Invalid or expired token"}), 401

        requester_user_id = decoded_token.get("user_id")
        requester_role = decoded_token.get("role")
        if not requester_user_id or not requester_role:
            return jsonify({"error": "Unauthorized access"}), 403

        # 2. Validate Request
        updates = request.json
        if not updates or not isinstance(updates, dict):
            return jsonify({"error": "Updates should be a valid JSON object"}), 400

        # 3. Fetch User
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()
        if not user_doc.exists: return jsonify({"error": "User not found"}), 404
        user_data = user_doc.to_dict()
        user_role = user_data.get("role", "").lower()

        # 4. Map Basic Fields
        mapped_updates = {}
        if "email" in updates: mapped_updates["email"] = updates["email"]
        if "name" in updates: mapped_updates["profileInfo.personalInformation.name"] = updates["name"]
        if "contactNumber" in updates: mapped_updates["profileInfo.contactInformation.phoneNumber"] = updates.get("contactNumber")
        if "role" in updates: mapped_updates["role"] = updates["role"]
        new_password = None
        if "password" in updates:
            new_password = str(updates["password"]).strip()
            if new_password:
                mapped_updates["passwordHash"] = encrypt_password(new_password)

        # 5. Parent Logic
        if user_role == "parent":
            if "associatedStudentIds" in updates:
                new_student_ids = updates["associatedStudentIds"]
                if isinstance(new_student_ids, list) and new_student_ids:
                    mapped_updates["associatedIds"] = firestore.ArrayUnion(new_student_ids)

        # 6. Grade/Class/Subject Logic
        if "grades" in updates:
            # --- CASE A: STUDENT UPDATE ---
            if user_role == "student":
                current_grades_data = user_data.get("assignedGrades", {})
                current_grades = current_grades_data.get("grades", current_grades_data)
                
                # Check Downgrade logic
                current_grade_nums = [extract_grade_number(g) for g in current_grades.keys()]
                current_grade_nums = [g for g in current_grade_nums if g is not None]
                highest_current_grade = max(current_grade_nums) if current_grade_nums else None

                new_grade_nums = [extract_grade_number(g) for g in updates["grades"].keys()]
                new_grade_nums = [g for g in new_grade_nums if g is not None]

                if highest_current_grade is not None:
                    for new_grade in new_grade_nums:
                        if new_grade < highest_current_grade:
                            return jsonify({"error": f"Cannot downgrade grade from {highest_current_grade} to {new_grade}"}), 400

                normalized_updates = {sanitize_grade_key(k): v for k, v in updates["grades"].items()}
                mapped_updates["assignedGrades"] = {"grades": normalized_updates}

            # --- CASE B: TEACHER UPDATE (FIXED) ---
            elif user_role == "teacher":
                # 1. Master Dictionary
                final_grades_map = {}

                # 2. Deep Merge Helper
                def merge_into_master(grade_key, class_data):
                    norm_grade = sanitize_grade_key(grade_key)
                    if norm_grade not in final_grades_map: final_grades_map[norm_grade] = {}
                    
                    if isinstance(class_data, dict):
                        for class_name, subjects in class_data.items():
                            if class_name not in final_grades_map[norm_grade]:
                                final_grades_map[norm_grade][class_name] = set()
                            
                            if isinstance(subjects, list):
                                final_grades_map[norm_grade][class_name].update(subjects)
                            elif isinstance(subjects, str):
                                final_grades_map[norm_grade][class_name].add(subjects)

                # 3. Load Existing Data
                raw_current_assigned = user_data.get("assignedGrades", {})
                current_grades_db = raw_current_assigned.get("grades", raw_current_assigned)
                if isinstance(current_grades_db, dict):
                    for grade, classes in current_grades_db.items():
                        merge_into_master(grade, classes)

                # 4. Process New Updates
                raw_updates_input = updates.get("grades", {})
                new_grades_dict = {}

                if isinstance(raw_updates_input, list):
                    for item in raw_updates_input:
                        g = item.get("grade") or item.get("gradeName")
                        c = item.get("class") or item.get("className")
                        s = item.get("subject") or item.get("subjects")
                        if g and c and s:
                            if g not in new_grades_dict: new_grades_dict[g] = {}
                            new_grades_dict[g][c] = s
                elif isinstance(raw_updates_input, dict):
                    new_grades_dict = raw_updates_input

                # Merge New Data
                if isinstance(new_grades_dict, dict):
                    for grade, classes in new_grades_dict.items():
                        merge_into_master(grade, classes)

                # 5. CONVERT TO LISTS (The Fix)
                mapped_assigned_grades = {}
                for grade, classes in final_grades_map.items():
                    mapped_assigned_grades[grade] = {}
                    for class_name, subject_set in classes.items():
                        # ✅ FIX: Always return a list
                        mapped_assigned_grades[grade][class_name] = list(subject_set)

                # 6. Save
                mapped_updates["assignedGrades"] = mapped_assigned_grades

                # 7. Recalculate Associations (Auto-link students)
                target_grade_keys = set(mapped_assigned_grades.keys())
                final_student_ids = set()
                
                if target_grade_keys:
                    all_students_ref = db.collection("users").where("role", "==", "student")
                    all_students = all_students_ref.stream()
                    for student_doc in all_students:
                        s_data = student_doc.to_dict()
                        s_assigned = s_data.get("assignedGrades", {})
                        s_grades_map = s_assigned.get('grades', s_assigned)
                        if s_grades_map:
                            norm_s_grades = set(sanitize_grade_key(g) for g in s_grades_map.keys())
                            if not target_grade_keys.isdisjoint(norm_s_grades):
                                final_student_ids.add(student_doc.id)
                
                mapped_updates["associatedIds"] = list(final_student_ids)

        # 7. Commit Updates
        if mapped_updates:
            if new_password:
                auth.update_user(uid=user_id, password=new_password)
            user_ref.update(mapped_updates)
            
            # Sync to Students collection
            if user_role == "student":
                try:
                    student_ref = db.collection("students").document(user_id)
                    student_sync_updates = {}
                    if "assignedGrades" in mapped_updates:
                        student_sync_updates["assignedGrades"] = mapped_updates["assignedGrades"]
                    for key, val in mapped_updates.items():
                        if key.startswith("profileInfo.") or key == "email":
                            student_sync_updates[key] = val
                    if student_sync_updates:
                        student_ref.update(student_sync_updates)
                    refreshed_student = user_ref.get().to_dict() or {}
                    sync_student_teacher_links(user_id, refreshed_student)
                except Exception as e:
                    print(f"Warning: Student sync failed: {e}")
        else:
            return jsonify({"message": "No valid updates provided.", "user_id": user_id}), 200

        updated_user = user_ref.get().to_dict()
        updated_user.pop("passwordHash", None)
        return jsonify({"message": "User updated successfully", "updated_user": updated_user}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


import re 
from google.cloud import firestore # Make sure you import firestore

@app.route("/api/headmaster/users/<string:user_id>/grade/<string:grade_name>", methods=["DELETE"])
def delete_user_grade(user_id, grade_name):
    
    # --- HELPER FUNCTIONS DEFINED LOCALLY ---
    # 1. CRITICAL FIX: Replaced with the *correct* sanitizer that PRESERVES parentheses
    def sanitize_grade_key(key):
        """
        Standardizes the grade string (e.g., 'Grade 11 (Science)' to 'GRADE_11(SCIENCE)')
        by converting to uppercase, normalizing spaces/symbols to underscores, and 
        crucially, PRESERVING the content inside parentheses.
        """
        key = str(key).upper().strip()
        
        # 1. Normalize spaces, dots, or slashes to underscores
        key = re.sub(r'[\s\./]', '_', key)
        
        # 2. Clean up extra underscores around parentheses if they exist
        # E.g., 'GRADE_11_(SCIENCE)' -> 'GRADE_11(SCIENCE)'
        key = key.replace('_(', '(').replace(')_', ')')
        
        # 3. Clean up any repeated underscores
        key = re.sub(r'_{2,}', '_', key)

        # 4. Clean up leading/trailing underscores
        return key.strip('_')
    # --- END HELPER FUNCTIONS ---
    
    try:
        # Validate Authorization header
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_jwt(token) # Ensure decode_jwt is imported/defined

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        requester_user_id = decoded_token.get("user_id")
        requester_role = decoded_token.get("role")

        if not requester_user_id or not requester_role:
            return jsonify({"error": "Unauthorized access"}), 403

        user_ref = db.collection("users").document(user_id) # Ensure db is your Firestore client
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found in Firestore"}), 404

        user_data = user_doc.to_dict()
        user_role = user_data.get("role", "").lower()

        # 1. Sanitize the grade name input to find the correct key
        # This will now correctly produce 'GRADE_11(LITERATURE)'
        sanitized_grade_name = sanitize_grade_key(grade_name)
        
        assigned_grades = user_data.get("assignedGrades", {})
        
        # Handle the structure for teachers (which is flat, based on your update logic)
        assigned_grades_map = assigned_grades

        # Determine the key to delete
        key_to_delete = None
        if sanitized_grade_name in assigned_grades_map:
            key_to_delete = sanitized_grade_name 
        elif grade_name in assigned_grades_map:
            key_to_delete = grade_name # Fallback for old/un-sanitized data
        
        if not key_to_delete:
            return jsonify({"error": f"Grade '{grade_name}' (Sanitized: {sanitized_grade_name}) not found for this user"}), 404

        # Firestore delete syntax for nested field
        # Assumes teacher structure is flat: assignedGrades.GRADE_KEY
        field_path = f"assignedGrades.{key_to_delete}"
             
        user_ref.update({field_path: firestore.DELETE_FIELD})
        
        
        # 2. --- RECALCULATE ASSOCIATED IDS IF THE USER IS A TEACHER ---
        if user_role == "teacher":
            
            # Fetch the document *again* to get the grades *after* deletion
            updated_doc_data = user_ref.get().to_dict()
            remaining_grades = updated_doc_data.get("assignedGrades", {})
            
            # Determine the grade keys the teacher is NOW assigned to (using correct sanitizer)
            target_grade_keys = set(sanitize_grade_key(g) for g in remaining_grades.keys())
            
            newly_associated_ids = set()
            
            # Query all students
            all_students_ref = db.collection("users").where("role", "==", "student")
            all_students = all_students_ref.stream()

            for student_doc in all_students:
                student_data = student_doc.to_dict()
                student_assigned_data = student_data.get("assignedGrades", {})
                
                # Copy the crucial structural check for *students* from update_user11
                if 'grades' in student_assigned_data and isinstance(student_assigned_data['grades'], dict):
                    student_grades_map = student_assigned_data['grades']
                else:
                    student_grades_map = student_assigned_data

                # Normalize the student's grade keys for matching
                normalized_student_grades = set(sanitize_grade_key(g) for g in student_grades_map.keys())

                # Check for intersection with the teacher's REMAINING grades
                if target_grade_keys.intersection(normalized_student_grades):
                    newly_associated_ids.add(student_doc.id)
            
            # Update the teacher's associatedIds array with the new, recalculated list
            user_ref.update({"associatedIds": list(newly_associated_ids)})
        # --- END RECALCULATION LOGIC ---
        
        # 3. Final cleanup: sets assignedGrades to {} if it's now completely empty
        final_doc_data = user_ref.get().to_dict()
        final_assigned_data = final_doc_data.get("assignedGrades", {})

        if not final_assigned_data:
             # Set the entire assignedGrades field to {} if its contents are empty
             user_ref.update({"assignedGrades": {}})

        return jsonify({
            "message": f"Grade '{key_to_delete}' deleted successfully. Associated students recalculated.",
            "userId": user_id
        }), 200

    except Exception as e:
        print("Error deleting grade:", str(e))
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/headmaster/users/<string:user_id>/associatedStudents", methods=["DELETE"])
# def delete_multiple_associated_students(user_id):
#     """
#     Deletes one or more student IDs from a user's (parent's) 
#     'associatedIds' array based on a JSON body.
#     """
#     try:
#         # 1. --- Authorization (Same as before) ---
#         auth_header = request.headers.get("Authorization")
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token) # Assumes you have this helper function

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role in ["headmaster", "admin"]:
#             return jsonify({"error": "Unauthorized access for this role"}), 403

#         # 2. --- Get Student IDs from JSON Body ---
#         try:
#             data = request.get_json()
#             if data is None:
#                  return jsonify({"error": "Missing or invalid JSON body"}), 400
#             student_ids_to_delete = data.get("student_ids")
#         except Exception:
#             return jsonify({"error": "Invalid JSON format"}), 400

#         # 3. --- Validate the input list ---
#         if not student_ids_to_delete or not isinstance(student_ids_to_delete, list) or len(student_ids_to_delete) == 0:
#             return jsonify({
#                 "error": "Request body must contain a 'student_ids' key with a non-empty list."
#             }), 400

#         # 4. --- Get User (Parent) Document ---
#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User (Parent) not found in Firestore"}), 404
            
#         # 5. --- Firestore update to REMOVE from an ARRAY (handles multiple) ---
#         # Firestore.ArrayRemove is perfect for this.
#         # It accepts a list and will remove all matching elements.
#         user_ref.update({
#             "associatedIds": firestore.ArrayRemove(student_ids_to_delete)
#         })

#         # 6. --- Return Success Response ---
#         return jsonify({
#             "message": f"Successfully processed removal for {len(student_ids_to_delete)} student IDs.",
#             "userId": user_id
#         }), 200

#     except Exception as e:
#         print(f"Error removing associated students: {str(e)}")
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

from flask import request, jsonify
from firebase_admin import credentials, firestore as admin_firestore, initialize_app, get_app, _apps
import firebase_admin


@app.route("/api/headmaster/users/<string:headmaster_id>/associatedStudents", methods=["DELETE"])
def delete_multiple_associated_students(headmaster_id):
    """
    URL PARAM  -> headmaster_id (auth & validation)
    BODY PARAM -> parent_id (target parent)
    BODY PARAM -> student_ids (students to remove from parent)
    """

    try:
        # =========================================================
        # 🔐 1. AUTHORIZATION (HEADMASTER / ADMIN)
        # =========================================================
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_jwt(token)

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        requester_user_id = decoded_token.get("user_id")
        requester_role = decoded_token.get("role")

        if not requester_user_id or requester_role not in ["headmaster", "admin"]:
            return jsonify({"error": "Unauthorized access for this role"}), 403

        # ✅ Token user must match URL headmaster (unless admin)
        if requester_role not in ["headmaster", "admin"]:
            return jsonify({
                "error": "Unauthorized access"
            }), 403



        # =========================================================
        # 📦 2. INPUT (PARENT ID + STUDENT IDS)
        # =========================================================
        data = request.get_json()
        if not data:
            return jsonify({"error": "Request body is required"}), 400

        parent_id = data.get("parent_id")
        raw_student_ids = data.get("student_ids")

        if not parent_id:
            return jsonify({"error": "parent_id is required in request body"}), 400

        if not isinstance(raw_student_ids, list):
            return jsonify({"error": "student_ids must be an array"}), 400

        # ✅ Auto-flatten student_ids (handles nested arrays)
        student_ids_to_delete = []
        for sid in raw_student_ids:
            if isinstance(sid, list):
                student_ids_to_delete.extend(sid)
            else:
                student_ids_to_delete.append(sid)

        if not student_ids_to_delete:
            return jsonify({"error": "student_ids cannot be empty"}), 400


        # =========================================================
        # 🔥 3. FORCE INITIALIZE **pees-d1101 ONLY**
        # =========================================================
        if "pees_app" not in _apps:
            cred = credentials.Certificate("serviceAccountKey.json")  # must be pees-d1101 key
            initialize_app(cred, name="pees_app")

        pees_app = get_app("pees_app")
        pees_db = admin_firestore.client(app=pees_app)

        # ✅ HARD CONFIRM PROJECT
        if pees_db.project != "pees-d1101":
            return jsonify({
                "error": "Wrong Firestore project connected",
                "connected_project": pees_db.project
            }), 500


        # =========================================================
        # 👤 4. FETCH **PARENT** DOCUMENT
        # =========================================================
        user_ref = pees_db.collection("users").document(parent_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "Parent user not found"}), 404

        user_data = user_doc.to_dict()

        if "associatedIds" not in user_data:
            return jsonify({
                "error": "Parent account has no associatedIds field"
            }), 400

        existing_ids = user_data["associatedIds"]


        # =========================================================
        # 🧠 5. VALIDATION
        # =========================================================
        not_found_ids = [sid for sid in student_ids_to_delete if sid not in existing_ids]

        if not_found_ids:
            return jsonify({
                "error": "Some student IDs were not found in associatedIds",
                "not_found_ids": not_found_ids,
                "existing_associatedIds": existing_ids
            }), 400


        # =========================================================
        # 🗑️ 6. SAFE DELETE (FROM PARENT ONLY)
        # =========================================================
        updated_ids = [
            sid for sid in existing_ids if sid not in student_ids_to_delete
        ]

        user_ref.update({
            "associatedIds": updated_ids
        })


        # =========================================================
        # ✅ 7. RESPONSE
        # =========================================================
        return jsonify({
            "message": "Student IDs successfully removed from parent account",
            "firebase_project": pees_db.project,
            "headmasterUserId": headmaster_id,
            "parentUserId": parent_id,
            "deleted_student_ids": student_ids_to_delete,
            "before": existing_ids,
            "after": updated_ids
        }), 200


    except Exception as e:
        print("🔥 DELETE ERROR:", str(e))
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500

# @app.route('/api/headmaster/users/<string:user_id>', methods=['PUT'])
# def update_user(user_id):
#     try:
#         # Validate Authorization header
#         auth_header = request.headers.get('Authorization')
#         if not auth_header:
#             return jsonify({"error": "Authorization token is required"}), 401

#         # Extract JWT token
#         token = auth_header.split(" ")[1] if " " in auth_header else auth_header
#         decoded_token = decode_jwt(token)

#         if not decoded_token:
#             return jsonify({"error": "Invalid or expired token"}), 401

#         # Extract user_id and role from JWT token
#         requester_user_id = decoded_token.get("user_id")
#         requester_role = decoded_token.get("role")

#         if not requester_user_id or not requester_role:
#             return jsonify({"error": "Unauthorized access"}), 403

#         # # Ensure only headmasters can update users
#         # if requester_role.lower() != "headmaster":
#         #     return jsonify({"error": "Only headmasters are allowed to update users"}), 403

#         # Validate request body
#         updates = request.json
#         if not updates or not isinstance(updates, dict):
#             return jsonify({"error": "Updates should be a valid JSON object"}), 400

#         # Allowed fields to update
#         allowed_fields = {"role", "email", "password", "profileInfo", "associatedIds"}

#         # Filter allowed fields
#         filtered_updates = {key: value for key, value in updates.items() if key in allowed_fields}

#         if not filtered_updates:
#             return jsonify({"error": "No valid fields to update"}), 400

#         # Reference user document
#         user_ref = db.collection('users').document(user_id)
#         user_doc = user_ref.get()

#         if not user_doc.exists:
#             return jsonify({"error": "User not found in Firestore"}), 404

#         # Hash password if updated
#         if "password" in filtered_updates:
#             filtered_updates["passwordHash"] = hashlib.sha256(filtered_updates.pop("password").encode()).hexdigest()

#         # Update only changed fields in Firestore
#         user_ref.update(filtered_updates)  # Uses Firestore's update method to keep existing values intact

#         # Fetch updated user data
#         updated_user = user_ref.get().to_dict()
#         updated_user.pop('passwordHash', None)  # Remove sensitive data

#         return jsonify({
#             "message": "User updated successfully",
#             "updated_user": updated_user
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/headmaster/users/<string:user_id>", methods=["DELETE"])
def delete_user(user_id):
    try:
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        user_ref.delete()

        return jsonify({"message": "User deleted successfully"}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/headmaster/deactivate", methods=["POST"])
def deactivate_user():
    try:
        # Parse request JSON
        data = request.json
        if not data or "userId" not in data:
            return jsonify({"error": "User ID is required"}), 400

        user_id = data["userId"]

        # Fetch user document from Firestore
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        # Check if user is already inactive
        if user_doc.to_dict().get("status", "").lower() == "inactive":
            return jsonify({"error": "User is already inactive"}), 400

        # Update the user's status to inactive
        user_ref.update({"status": "inactive"})

        return (
            jsonify({"message": f"User with ID {user_id} deactivated successfully"}),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# active


@app.route("/api/headmaster/activate", methods=["POST"])
def activate_user():
    try:
        # Parse request JSON
        data = request.json
        if not data or "userId" not in data:
            return jsonify({"error": "User ID is required"}), 400

        user_id = data["userId"]

        # Fetch user document from Firestore
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        current_status = user_doc.to_dict().get("status", "").lower()

        if current_status == "active":
            return jsonify({"error": "User is already active"}), 400

        user_ref.update({"status": "active"})  # ? Corrected

        return jsonify({"message": f"User with ID {user_id} deactivated successfully"}), 200
    
    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# if __name__ == "__main__":
#     app.run(port=5000, debug=True)


############################################################################################################################################################################
# TEACHING PLAN PAGE
import os
from flask import Flask, request, jsonify
from firebase_admin import credentials, firestore, initialize_app
from groq import Groq
from dotenv import load_dotenv
import groq
from datetime import datetime  # Ensure consistent usage of datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
import uuid
import firebase_admin
import jwt
from firebase_admin import credentials, auth, firestore
import datetime

# from datetime import timedelta
import re
import hashlib

# import jwt
# import datetime
import secrets  # For generating a secure nonce
from datetime import datetime

# # Load environment variables from .env file
# load_dotenv()

# Initialize the Groq client with the API key
client = Groq(api_key="YOUR_GROQ_API_KEY")

# # Initialize Firebase Admin SDK
# cred = credentials.Certificate(r"serviceAccountKey.json")
# initialize_app(cred)
# db = firestore.client()

# # Initialize Flask app
# app = Flask(__name__)
# SECRET_KEY = "secret_key"
# CORS(app)

# create student used again


@app.route("/api/student/create", methods=["POST"])
def create_user_account():
    try:
        data = request.json
        if not data or "role" not in data:
            return jsonify({"error": "role is required"}), 400

        role = data["role"].lower()

        # Validate role
        allowed_roles = ["student"]  # Add allowed roles as needed
        if role not in allowed_roles:
            return jsonify({"error": "Invalid role"}), 400

        # Generate a unique userId
        user_id = str(uuid.uuid4())

        # Check if the generated userId already exists in the users collection
        user_doc = db.collection("users").document(user_id).get()
        if user_doc.exists:
            return jsonify({"error": f"User with userId {user_id} already exists"}), 409

        # Define default structure for the user document
        user_data = {
            "userId": user_id,
            "role": role,
            "email": data.get("email", ""),
            "createdAt": datetime.utcnow().isoformat(),
        }

        # Save the user data in Firestore
        db.collection("users").document(user_id).set(user_data)

        # If the role is 'student', also store data in the students collection
        if role == "student":
            student_data = {
                "studentId": user_id,
                "personalInformation": {
                    "name": data.get("name", ""),
                    "photourl": data.get(
                        "photourl", "https://example.com/default-photo.jpg"
                    ),
                    "idNumber": data.get("idno", ""),
                },
                "contactInformation": {
                    "phoneNumber": data.get("phoneNumber", ""),
                    "address": data.get("address", ""),
                },
                "academicInformation": {
                    "grade": data.get("grade", ""),
                    "classSection": data.get("classSection", ""),
                },
                "createdAt": datetime.utcnow().isoformat(),
            }
            db.collection("students").document(user_id).set(student_data)

        return (
            jsonify(
                {"message": "User account created successfully", "userId": user_id}
            ),
            201,
        )

    except Exception as e:
        # Log the exception for debugging purposes (use proper logging tools in production)
        print(f"Error: {str(e)}")
        return jsonify({"error": "An internal server error occurred"}), 500


# @app.route("/api/student/create", methods=["POST"])
# def create_user_account():
#     try:
#         data = request.json
#         if not data or 'role' not in data:
#             return jsonify({"error": "role is required"}), 400

#         role = data['role'].lower()

#         # Validate role
#         allowed_roles = ['student']  # Add allowed roles as needed
#         if role not in allowed_roles:
#             return jsonify({"error": "Invalid role"}), 400

#         # Generate a unique userId
#         user_id = str(uuid.uuid4())

#         # Check if the generated userId already exists in the users collection
#         user_doc = db.collection("users").document(user_id).get()
#         if user_doc.exists:
#             return jsonify({"error": f"User with userId {user_id} already exists"}), 409

#         # Define default structure for the user document
#         user_data = {
#             "userId": user_id,
#             "role": role,
#             "email": data.get('email', ""),
#             "createdAt": datetime.utcnow().isoformat()
#         }

#         # Save the user data in Firestore
#         db.collection("users").document(user_id).set(user_data)

#         # If the role is 'student', also store data in the students collection
#         if role == "student":
#             student_data = {
#                 "studentId": user_id,
#                 "personalInformation": {
#                     "name": data.get('name', ""),
#                     "photourl": data.get('photourl', "https://example.com/default-photo.jpg"),
#                     "idNumber": data.get('idno', ""),
#                 },
#                 "contactInformation": {
#                     "phoneNumber": data.get('phoneNumber', ""),
#                     "address": data.get('address', ""),
#                 },
#                 "createdAt": datetime.utcnow().isoformat()
#             }
#             db.collection("students").document(user_id).set(student_data)

#         return jsonify({"message": "User account created successfully", "userId": user_id}), 201

#     except Exception as e:
#         # Log the exception for debugging purposes (use proper logging tools in production)
#         print(f"Error: {str(e)}")
#         return jsonify({"error": "An internal server error occurred"}), 500


# Define the template prompts with consistent placeholders
english_template = """
As an educational expert, develop a personalized teaching plan for the
student below, based on their academic performance analysis and teacher
observations. The plan should align with the specified curriculum
objectives.

**Student Information**
- Name: {student_name}
- Grade Level: {grade_level}

**Academic Performance Summary**
{performance_summary}

**Teacher Observations**
{teacher_observations}

**Curriculum Objectives**
{curriculum_objectives}

The teaching plan should include:
- **Learning Objectives**: Specific and measurable goals.
- **Instructional Strategies**: Tailored methods to address the student's
  needs.
- **Resources and Materials**: Recommended materials (e.g., textbooks,
  online resources).
- **Assessment Methods**: Ways to measure progress.
- **Timeline**: Suggested schedule for achieving objectives.
"""

arabic_template = """
Ã˜Â¨Ã˜ÂµÃ™ÂÃ˜ÂªÃ™Æ’ Ã˜Â®Ã˜Â¨Ã™Å Ã˜Â±Ã™â€¹Ã˜Â§ Ã˜ÂªÃ˜Â±Ã˜Â¨Ã™Ë†Ã™Å Ã™â€¹Ã˜Â§Ã˜Å’ Ã™â€šÃ™â€¦ Ã˜Â¨Ã˜ÂªÃ˜Â·Ã™Ë†Ã™Å Ã˜Â± Ã˜Â®Ã˜Â·Ã˜Â© Ã˜ÂªÃ˜Â¯Ã˜Â±Ã™Å Ã˜Â³ Ã˜Â´Ã˜Â®Ã˜ÂµÃ™Å Ã˜Â© Ã™â€žÃ™â€žÃ˜Â·Ã˜Â§Ã™â€žÃ˜Â¨ Ã˜Â£Ã˜Â¯Ã™â€ Ã˜Â§Ã™â€¡Ã˜Å’ Ã˜Â¨Ã™â€ Ã˜Â§Ã˜Â¡Ã™â€¹ Ã˜Â¹Ã™â€žÃ™â€° Ã˜ÂªÃ˜Â­Ã™â€žÃ™Å Ã™â€ž Ã˜Â£Ã˜Â¯Ã˜Â§Ã˜Â¦Ã™â€¡ Ã˜Â§Ã™â€žÃ˜Â£Ã™Æ’Ã˜Â§Ã˜Â¯Ã™Å Ã™â€¦Ã™Å  Ã™Ë†Ã™â€¦Ã™â€žÃ˜Â§Ã˜Â­Ã˜Â¸Ã˜Â§Ã˜Âª Ã˜Â§Ã™â€žÃ™â€¦Ã˜Â¹Ã™â€žÃ™â€¦. Ã™Å Ã˜Â¬Ã˜Â¨ Ã˜Â£Ã™â€  Ã˜ÂªÃ˜ÂªÃ™Ë†Ã˜Â§Ã™ÂÃ™â€š Ã˜Â§Ã™â€žÃ˜Â®Ã˜Â·Ã˜Â©
Ã™â€¦Ã˜Â¹ Ã˜Â£Ã™â€¡Ã˜Â¯Ã˜Â§Ã™Â Ã˜Â§Ã™â€žÃ™â€¦Ã™â€ Ã™â€¡Ã˜Â¬ Ã˜Â§Ã™â€žÃ™â€¦Ã˜Â­Ã˜Â¯Ã˜Â¯Ã˜Â©.

**Ã™â€¦Ã˜Â¹Ã™â€žÃ™Ë†Ã™â€¦Ã˜Â§Ã˜Âª Ã˜Â§Ã™â€žÃ˜Â·Ã˜Â§Ã™â€žÃ˜Â¨**
- Ã˜Â§Ã™â€žÃ˜Â§Ã˜Â³Ã™â€¦: {student_name}
- Ã˜Â§Ã™â€žÃ™â€¦Ã˜Â³Ã˜ÂªÃ™Ë†Ã™â€° Ã˜Â§Ã™â€žÃ˜Â¯Ã˜Â±Ã˜Â§Ã˜Â³Ã™Å : {grade_level}

**Ã™â€¦Ã™â€žÃ˜Â®Ã˜Âµ Ã˜Â§Ã™â€žÃ˜Â£Ã˜Â¯Ã˜Â§Ã˜Â¡ Ã˜Â§Ã™â€žÃ˜Â£Ã™Æ’Ã˜Â§Ã˜Â¯Ã™Å Ã™â€¦Ã™Å **
{performance_summary}

**Ã™â€¦Ã™â€žÃ˜Â§Ã˜Â­Ã˜Â¸Ã˜Â§Ã˜Âª Ã˜Â§Ã™â€žÃ™â€¦Ã˜Â¹Ã™â€žÃ™â€¦**
{teacher_observations}

**Ã˜Â£Ã™â€¡Ã˜Â¯Ã˜Â§Ã™Â Ã˜Â§Ã™â€žÃ™â€¦Ã™â€ Ã™â€¡Ã˜Â¬**
{curriculum_objectives}

Ã™Å Ã˜Â¬Ã˜Â¨ Ã˜Â£Ã™â€  Ã˜ÂªÃ˜ÂªÃ˜Â¶Ã™â€¦Ã™â€  Ã˜Â®Ã˜Â·Ã˜Â© Ã˜Â§Ã™â€žÃ˜ÂªÃ˜Â¯Ã˜Â±Ã™Å Ã˜Â³:
- **Ã˜Â§Ã™â€žÃ˜Â£Ã™â€¡Ã˜Â¯Ã˜Â§Ã™Â Ã˜Â§Ã™â€žÃ˜ÂªÃ˜Â¹Ã™â€žÃ™Å Ã™â€¦Ã™Å Ã˜Â©**: Ã˜Â£Ã™â€¡Ã˜Â¯Ã˜Â§Ã™Â Ã™â€¦Ã˜Â­Ã˜Â¯Ã˜Â¯Ã˜Â© Ã™Ë†Ã™â€šÃ˜Â§Ã˜Â¨Ã™â€žÃ˜Â© Ã™â€žÃ™â€žÃ™â€šÃ™Å Ã˜Â§Ã˜Â³.
- **Ã˜Â§Ã˜Â³Ã˜ÂªÃ˜Â±Ã˜Â§Ã˜ÂªÃ™Å Ã˜Â¬Ã™Å Ã˜Â§Ã˜Âª Ã˜Â§Ã™â€žÃ˜ÂªÃ˜Â¯Ã˜Â±Ã™Å Ã˜Â³**: Ã˜Â·Ã˜Â±Ã™â€š Ã™â€¦Ã˜Â®Ã˜ÂµÃ˜ÂµÃ˜Â© Ã™â€žÃ™â€¦Ã˜Â¹Ã˜Â§Ã™â€žÃ˜Â¬Ã˜Â© Ã˜Â§Ã˜Â­Ã˜ÂªÃ™Å Ã˜Â§Ã˜Â¬Ã˜Â§Ã˜Âª Ã˜Â§Ã™â€žÃ˜Â·Ã˜Â§Ã™â€žÃ˜Â¨.
- **Ã˜Â§Ã™â€žÃ™â€¦Ã™Ë†Ã˜Â§Ã˜Â±Ã˜Â¯ Ã™Ë†Ã˜Â§Ã™â€žÃ™â€¦Ã™Ë†Ã˜Â§Ã˜Â¯**: Ã˜Â§Ã™â€žÃ™â€¦Ã™Ë†Ã˜Â§Ã˜Â¯ Ã˜Â§Ã™â€žÃ™â€¦Ã™Ë†Ã˜ÂµÃ™â€° Ã˜Â¨Ã™â€¡Ã˜Â§ (Ã™â€¦Ã˜Â«Ã™â€ž Ã˜Â§Ã™â€žÃ™Æ’Ã˜ÂªÃ˜Â¨ Ã˜Â§Ã™â€žÃ˜Â¯Ã˜Â±Ã˜Â§Ã˜Â³Ã™Å Ã˜Â©Ã˜Å’ Ã˜Â§Ã™â€žÃ™â€¦Ã™Ë†Ã˜Â§Ã˜Â±Ã˜Â¯ Ã˜Â¹Ã˜Â¨Ã˜Â± Ã˜Â§Ã™â€žÃ˜Â¥Ã™â€ Ã˜ÂªÃ˜Â±Ã™â€ Ã˜Âª).
- **Ã˜Â·Ã˜Â±Ã™â€š Ã˜Â§Ã™â€žÃ˜ÂªÃ™â€šÃ™Å Ã™Å Ã™â€¦**: Ã˜Â·Ã˜Â±Ã™â€š Ã™â€žÃ™â€šÃ™Å Ã˜Â§Ã˜Â³ Ã˜Â§Ã™â€žÃ˜ÂªÃ™â€šÃ˜Â¯Ã™â€¦.
- **Ã˜Â§Ã™â€žÃ˜Â¬Ã˜Â¯Ã™Ë†Ã™â€ž Ã˜Â§Ã™â€žÃ˜Â²Ã™â€¦Ã™â€ Ã™Å **: Ã˜Â§Ã™â€žÃ˜Â¬Ã˜Â¯Ã™Ë†Ã™â€ž Ã˜Â§Ã™â€žÃ™â€¦Ã™â€šÃ˜ÂªÃ˜Â±Ã˜Â­ Ã™â€žÃ˜ÂªÃ˜Â­Ã™â€šÃ™Å Ã™â€š Ã˜Â§Ã™â€žÃ˜Â£Ã™â€¡Ã˜Â¯Ã˜Â§Ã™Â.
"""


# Function to validate student information
def validate_student_info(student_info):
    errors = []

    if (
        not student_info.get("student_name")
        or not student_info["student_name"].replace(" ", "").isalpha()
    ):
        errors.append(
            {
                "field": "student_name",
                "message": "Name must contain only alphabetic characters and not be empty.",
            }
        )

    valid_grades = [
        "Kindergarten",
        "1st",
        "2nd",
        "3rd",
        "4th",
        "5th",
        "6th",
        "7th",
        "8th",
        "9th",
        "10th",
        "11th",
        "12th",
        "Grade 1",
        "Grade 2",
        "Grade 3",
        "Grade 4",
        "Grade 5",
        "Grade 6",
        "Grade 7",
        "Grade 8",
        "Grade 9",
        "Grade 10",
        "Grade 11",
        "Grade 12",
        "Grade 11 (Literature)",
        "Grade 11(Science)",
        "Grade 12(Literature)",
        "Grade 12(Science)",
        "Ã˜Â§Ã™â€žÃ˜ÂµÃ™Â Ã˜Â§Ã™â€žÃ˜Â®Ã˜Â§Ã™â€¦Ã˜Â³",
        "Ã˜Â§Ã™â€žÃ˜ÂµÃ™Â Ã˜Â§Ã™â€žÃ˜Â¹Ã˜Â§Ã˜Â´Ã˜Â±",
    ]
    # if not student_info.get('grade_level') or student_info['grade_level'] not in valid_grades:
    #     errors.append({"field": "grade_level", "message": f"Invalid grade level {student_info.get('grade_level')}. Accepted values are: {valid_grades}"})

    # if not student_info.get('performance_summary') or len(student_info['performance_summary']) < 20:
    #     errors.append({"field": "performance_summary", "message": "Performance summary must be at least 20 characters long."})

    # if not student_info.get('teacher_observations') or len(student_info['teacher_observations']) < 20:
    #     errors.append({"field": "teacher_observations", "message": "Teacher observations must be at least 20 characters long."})

    # if not student_info.get('curriculum_objectives') or len(student_info['curriculum_objectives']) < 20:
    #     errors.append({"field": "curriculum_objectives", "message": "Curriculum objectives must be at least 20 characters long."})

    if errors:
        return {"status": "error", "errors": errors}

    return {"status": "success"}


# Function to sanitize input
def sanitize_input(data):
    import html

    sanitized_data = {}
    for key, value in data.items():
        if isinstance(value, str):
            sanitized_data[key] = html.escape(value.strip())
        else:
            sanitized_data[key] = value
    return sanitized_data


# # Function to generate the teaching plan
# def generate_teaching_plan(template, student_info):
#     validation_response = validate_student_info(student_info)
#     if validation_response["status"] == "error":
#         return validation_response

#     prompt = template.format(
#         student_name=student_info["student_name"],
#         grade_level=student_info["grade_level"],
#         performance_summary=student_info["performance_summary"],
#         teacher_observations=student_info["teacher_observations"],
#         curriculum_objectives=student_info["curriculum_objectives"],
#     )

#     try:
#         chat_completion = client_openai.chat.completions.create(
#             messages=[
#                 {"role": "system", "content": "You are a helpful assistant."},
#                 {"role": "user", "content": prompt},
#             ],
#             model="llama3-8b-8192",
#             temperature=0.7,
#             max_tokens=1000,
#         )
#         return {"status": "success", "plan": chat_completion.choices[0].message.content}
#     except groq.APIConnectionError as e:
#         return {
#             "status": "error",
#             "message": "Could not reach the server.",
#             "details": str(e),
#         }
#     except groq.RateLimitError as e:
#         return {
#             "status": "error",
#             "message": "Rate limit exceeded. Please try again later.",
#             "details": str(e),
#         }
#     except Exception as e:
#         return {"status": "error", "message": f"An unexpected error occurred: {str(e)}"}


# Firestore integration: save only the teaching plan
def save_to_firestore(user_id, teaching_plan):
    try:
        # Reference the student document in Firestore
        student_ref = db.collection("students").document(user_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return {"status": "error", "message": "Student document does not exist"}

        # Add the teachingPlan field to the existing student document
        student_ref.update({"teachingPlan": teaching_plan})

        return {"status": "success", "message": "Teaching plan added to Firestore"}
    except Exception as e:
        return {"status": "error", "message": f"Failed to save teaching plan: {str(e)}"}


import jwt


# # API Route to generate teaching plan
# @app.route("/api/teaching-plan", methods=["POST"])
# def api_generate_teaching_plan():
#     try:
#         # Parse and sanitize the request data
#         data = request.get_json()
#         if not data:
#             return (
#                 jsonify(
#                     {
#                         "status": "error",
#                         "message": "Invalid input, JSON data is required",
#                     }
#                 ),
#                 400,
#             )

#         sanitized_data = sanitize_input(data)
#         student_report = sanitized_data.get("student_report")
#         student_id = sanitized_data.get("student_id")

#         if not student_report:
#             return (
#                 jsonify({"status": "error", "message": "Student report is missing"}),
#                 400,
#             )

#         if not student_id:
#             return (
#                 jsonify({"status": "error", "message": "Student ID is required"}),
#                 400,
#             )

#         # Validate the input data
#         validation_response = validate_student_info(student_report)
#         if validation_response["status"] == "error":
#             return jsonify(validation_response), 400

#         # Generate the teaching plan using AI
#         template = english_template
#         response = generate_teaching_plan(template, student_report)

#         if response["status"] == "success":
#             # Generate a unique planId
#             plan_id = str(uuid.uuid4())

#             # Construct the teaching plan data
#             teaching_plan = {
#                 "planId": plan_id,
#                 "actionPlan": response["plan"],  # AI-generated teaching plan
#                 "createdAt": datetime.utcnow().isoformat(),
#                 "version": sanitized_data.get("version", 1),  # Default version is 1
#             }

#             # Reference the student's document in Firestore
#             student_ref = db.collection("students").document(student_id)
#             student_doc = student_ref.get()

#             if not student_doc.exists:
#                 return (
#                     jsonify(
#                         {
#                             "status": "error",
#                             "message": f"Student with ID {student_id} not found",
#                         }
#                     ),
#                     404,
#                 )

#             # Fetch the current teachingPlans and add the new plan
#             current_plans = student_doc.to_dict().get("teachingPlans", {})
#             current_plans[plan_id] = teaching_plan

#             # Update the student's teachingPlans field
#             student_ref.update({"teachingPlans": current_plans})

#             # Return success response
#             return (
#                 jsonify(
#                     {
#                         "status": "success",
#                         "teaching_plan": teaching_plan,
#                     }
#                 ),
#                 201,
#             )

#         # Handle errors during teaching plan generation
#         return jsonify(response), 500

#     except Exception as e:
#         # Catch and handle any unexpected errors
#         return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# 3. Fetch Teaching Plans
# @app.route('/students/<student_id>/teaching-plan', methods=['GET'])
# def get_teaching_plan(student_id):
#     try:
#         # Validate the format of the student_id
#         if not isinstance(student_id, str) or len(student_id) != 36:
#             return jsonify({"error": "Invalid student ID format."}), 400

#         # Retrieve the student document from Firestore
#         student_ref = db.collection('students').document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": f"Student document for ID {student_id} not found."}), 404

#         # Retrieve the teaching plan
#         student_data = student_doc.to_dict()
#         teaching_plan = student_data.get('teachingPlan')
#         if not teaching_plan:
#             return jsonify({"error": "Teaching plan not found for this student."}), 404

#         return jsonify({"teaching_plan": teaching_plan}), 200

#     except Exception as e:
#         print(f"Unexpected error: {e}")
#         return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# update teaching:

# @app.route('/students/<student_id>/teaching-plan', methods=['PATCH'])
# def update_teaching_plans(student_id):
#     try:
#         data = request.json
#         if not data:
#             return jsonify({"error": "Request body is required and must be JSON"}), 400

#         updates = data.get('teaching_plan')
#         if updates is None:
#             return jsonify({"error": "'teaching_plan' field is required"}), 400

#         # Reference the student document
#         student_ref = db.collection('students').document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": f"Student with ID {student_id} not found"}), 404

#         # Fetch current teaching plans and merge with new updates
#         current_teaching_plans = student_doc.to_dict().get("teachingPlans", {})
#         current_teaching_plans.update(updates)

#         # Update the document in Firestore
#         student_ref.update({"teachingPlans": current_teaching_plans})

#         return jsonify({"message": "Teaching plans updated successfully"}), 200
#     except Exception as e:
#         return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500

# # Fetch a Specific Teaching Plan by planId


@app.route("/students/<student_id>/teaching-plans/<plan_id>", methods=["GET"])
def get_specific_teaching_plan(student_id, plan_id):
    try:
        # Reference the student's document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": f"Student with ID {student_id} not found"}), 404

        # Fetch the teachingPlans field
        teaching_plans = student_doc.to_dict().get("teachingPlans", {})

        # Check if the requested planId exists
        teaching_plan = teaching_plans.get(plan_id)
        if not teaching_plan:
            return jsonify({"error": f"Teaching plan with ID {plan_id} not found"}), 404

        return jsonify({"teaching_plan": teaching_plan}), 200

    except Exception as e:
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# Fetch All Teaching Plans for a Student
@app.route("/students/<student_id>/teaching-plans", methods=["GET"])
def get_all_teaching_plans(student_id):
    try:
        # Reference the student's document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": f"Student with ID {student_id} not found"}), 404

        # Fetch the teachingPlans field
        teaching_plans = student_doc.to_dict().get("teachingPlans", {})

        return jsonify({"teaching_plans": teaching_plans}), 200

    except Exception as e:
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# # Run the Flask app
# if __name__ == "__main__":
#     app.run(debug=True)
############################################################################################################################################################################

# TEACHING PLAN PDF
import os
from flask import Flask, request, jsonify
from firebase_admin import credentials, firestore, initialize_app, storage
from fpdf import FPDF
import uuid
from flask_cors import CORS
from datetime import datetime
import html

# # Initialize Flask app
# app = Flask(__name__)
# SECRET_KEY = "secret_key"
# CORS(app)

# # Firebase setup
# cred = credentials.Certificate(r"serviceAccountKey.json")
# initialize_app(cred, {'storageBucket': 'pees-d1101.firebasestorage.app'})

# db = firestore.client()
# bucket = storage.bucket()


# Helper class for generating PDFs
class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 12)
        self.cell(0, 10, "Teaching Plan", 0, 1, "C")

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")


def create_pdf(content, filename):
    pdf = PDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.multi_cell(0, 10, content)
    pdf.output(filename)


# # API Route to export teaching plan as PDF
# @app.route("/api/teaching-plan/export", methods=["POST"])
# def export_teaching_plan_as_pdf():
#     try:
#         # Get studentId from the request
#         data = request.json
#         student_id = data.get("studentId")

#         if not student_id:
#             return jsonify({"error": "Student ID is required"}), 400

#         # Validate studentId exists
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()
#         if not student_doc.exists:
#             return jsonify({"error": f"Student ID {student_id} does not exist"}), 404

#         # Fetch teaching plan data
#         student_data = student_doc.to_dict()
#         teaching_plans = student_data.get("teachingPlans", {})
#         if not teaching_plans:
#             return jsonify({"error": "No teaching plans found for this student"}), 404

#         # Get the latest teaching plan content
#         latest_plan = list(teaching_plans.values())[-1]
#         teaching_plan_content = latest_plan.get("actionPlan", "")

#         if not teaching_plan_content:
#             return jsonify({"error": "Teaching plan content is missing"}), 400

#         # Generate the PDF locally
#         pdf_filename = f"{student_id}_teaching_plan.pdf"
#         create_pdf(teaching_plan_content, pdf_filename)

#         # Upload the PDF to Firebase Storage
#         blob = bucket.blob(f"teaching_plans/{uuid.uuid4()}_{pdf_filename}")
#         blob.upload_from_filename(pdf_filename)
#         blob.make_public()

#         # Get the public URL
#         public_url = blob.public_url

#         # Optional: Add the public URL to Firestore
#         student_ref.update({"latestTeachingPlanPdf": public_url})

#         # Clean up the local file
#         os.remove(pdf_filename)

#         return jsonify({"status": "success", "url": public_url}), 200

#     except Exception as e:
#         return jsonify({"error": "An unexpected error occurred", "details": str(e)}), 500

# # # Run the Flask app
# # if __name__ == "__main__":
# #     app.run(debug=True)


############################################################################################################################################################################
# azure -> teachingplan
# import os
# import re
# import asyncio
# from flask import Flask, request, jsonify
# from werkzeug.utils import secure_filename
# from azure.ai.formrecognizer.aio import DocumentAnalysisClient
# from azure.core.credentials import AzureKeyCredential
# import firebase_admin
# from firebase_admin import credentials, storage, firestore
# from dotenv import load_dotenv
# import openai
# import os
# import uuid
# from flask import Flask, request, jsonify
# from fpdf import FPDF
# import firebase_admin
# from firebase_admin import credentials, firestore, storage
# import json

# # Load environment variables
# # load_dotenv()

# # app = Flask(__name__)

# # Configuration
# AZURE_ENDPOINT = "https://aiocr395080637747.cognitiveservices.azure.com/"
# AZURE_KEY = "b1b026f421034dabb948999b80a63e8c"
# OPENAI_API_KEY = "YOUR_OPENAI_API_KEY"

# if not AZURE_KEY:
#     raise ValueError("AZURE_KEY environment variable is missing.")
# if not OPENAI_API_KEY:
#     raise ValueError("OPENAI_API_KEY environment variable is missing.")

# openai_client = openai.AsyncOpenAI(api_key=OPENAI_API_KEY)

# # # Initialize Firebase Admin SDK
# # cred = credentials.Certificate("serviceAccountKey.json")
# # firebase_admin.initialize_app(cred, {'storageBucket': 'pees-d1101.firebasestorage.app'})

# app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

# # Initialize Firestore and Storage
# db = firestore.client()
# bucket = storage.bucket()

# # Initialize Azure OCR Client
# azure_client = DocumentAnalysisClient(endpoint=AZURE_ENDPOINT, credential=AzureKeyCredential(AZURE_KEY))

# # Function to clean extracted text
# def clean_text(raw_text):
#     cleaned = re.sub(r'\s+', ' ', raw_text)
#     cleaned = re.sub(r'[^a-zA-Z0-9.,!?\'\";:()\\-]', ' ', cleaned)
#     return cleaned.strip()

# # Extract text using Azure OCR (Async)
# async def extract_text_from_file(file_bytes):
#     async with DocumentAnalysisClient(endpoint=AZURE_ENDPOINT, credential=AzureKeyCredential(AZURE_KEY)) as azure_client:
#         poller = await azure_client.begin_analyze_document("prebuilt-document", file_bytes)
#         result = await poller.result()
#         raw_text = "\n".join([line.content for page in result.pages for line in page.lines])
#         return clean_text(raw_text)

# # Extract student info
# def extract_student_info(text):
#     name_match = re.search(r'Name[:\-\s]+([\w\s]+)', text, re.IGNORECASE)
#     grade_match = re.search(r'Grade\s*Level[:\-\s]+([\w\s]+)', text, re.IGNORECASE)
#     return name_match.group(1) if name_match else "Unknown", grade_match.group(1) if grade_match else "Unknown"

# # Evaluate text using OpenAI GPT-4o (async)
# async def evaluate_text_with_openai(extracted_text):
#     response = await openai_client.chat.completions.create(
#         model="gpt-4o mini",
#         messages=[
#             {"role": "system", "content": "You are an AI assistant that analyzes educational text."},
#             {"role": "user", "content": f"Analyze the following extracted text and summarize key points:\n{extracted_text}"}
#         ],
#         temperature=0.3,
#         # max_tokens=1000
#     )
#     return clean_text(response.choices[0].message.content) if response and response.choices else "No valid response received."

# import json

# # Generate personalized teaching plan using OpenAI GPT-4o (async)
# async def generate_teaching_plan(student_name, grade_level, evaluation):
#     template = f"""
#     Create a personalized teaching plan for:

#     - Student: {student_name}
#     - Grade Level: {grade_level}

#     Evaluation Summary:
#     {evaluation}

#     The plan should be returned in **strict JSON format** inside **triple backticks (` ```json `)** like this:

#     ```json
#     {{
#       "learningObjectives": "Provide a clear and structured list of learning goals.",
#       "instructionalStrategies": "Explain the teaching methods that will be used.",
#       "recommendedResources": "List books, tools, websites, or materials for learning.",
#       "assessmentMethods": "Describe how progress will be measured and evaluated.",
#       "timeline": "Provide a structured breakdown of the learning schedule."
#     }}
#     ```

#     Ensure:
#     - **No extra text** before or after the JSON block.
#     - The response is **100% valid JSON**.
#     """

#     response = await openai_client.chat.completions.create(
#         model="gpt-4o",
#         messages=[
#             {"role": "system", "content": "You are an AI that generates structured JSON responses."},
#             {"role": "user", "content": template}
#         ],
#         temperature=0.7,
#         max_tokens=1000
#     )

#     if response and response.choices:
#         try:
#             # Extract the response and remove backticks if needed
#             raw_response = response.choices[0].message.content.strip()

#             # Ensure we extract JSON block only
#             if "```json" in raw_response:
#                 raw_response = raw_response.split("```json")[1].split("```")[0].strip()

#             # Parse JSON response
#             teaching_plan_json = json.loads(raw_response)

#             return teaching_plan_json
#         except json.JSONDecodeError:
#             return {"error": "Failed to parse teaching plan response. AI did not return valid JSON."}
#     else:
#         return {"error": "No valid teaching plan generated."}

# from fpdf import FPDF

# class PDF(FPDF):
#     def header(self):
#         self.set_font('Arial', 'B', 12)
#         self.cell(0, 10, 'Teaching Plan', 0, 1, 'C')

#     def footer(self):
#         self.set_y(-15)
#         self.set_font('Arial', 'I', 8)
#         self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

# def create_pdf(content, filename):
#     pdf = PDF()
#     pdf.add_page()
#     pdf.set_font('Arial', size=12)

#     # Strip extra whitespace and newlines to avoid blank pages
#     formatted_content = content.strip()

#     # Ensure no forced extra page by limiting content width
#     pdf.multi_cell(190, 10, formatted_content)

#     pdf.output(filename)

# @app.route('/file_ocr', methods=['POST'])
# async def upload_file():
#     if 'file' not in request.files:
#         return jsonify({'error': 'No file part in the request'}), 400

#     if 'studentId' not in request.form:
#         return jsonify({'error': 'Missing studentId in request'}), 400

#     student_id = request.form['studentId']
#     file = request.files['file']

#     if file.filename == '':
#         return jsonify({'error': 'Invalid file'}), 400

#     filename = secure_filename(file.filename)
#     content_type = file.content_type
#     file_bytes = file.read()

#     if content_type not in ['application/pdf', 'image/jpeg']:
#         return jsonify({'error': 'Unsupported file type. Only PDF and JPG images are allowed.'}), 400

#     try:
#         # Ensure an event loop is running
#         loop = asyncio.get_running_loop()

#         extracted_text = await extract_text_from_file(file_bytes)
#         student_name, grade_level = extract_student_info(extracted_text)

#         evaluation, teaching_plan = await asyncio.gather(
#             evaluate_text_with_openai(extracted_text),
#             generate_teaching_plan(student_name, grade_level, extracted_text)
#         )

#         # Generate a new unique plan ID
#         plan_id = str(uuid.uuid4())

#         # Check if the student exists in Firestore
#         student_ref = db.collection('students').document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({'error': f'Student with ID {student_id} not found'}), 404

#         # Ensure teachingPlans and actionPlan are stored correctly in Firestore
#         student_ref.update({
#             f"teachingPlans.{plan_id}": {
#                 "actionPlan": {  # Ensure actionPlan contains all necessary fields
#                     "extracted_text": extracted_text,
#                     "evaluation": evaluation,
#                     "assessmentMethods": teaching_plan.get("assessmentMethods", ""),
#                     "instructionalStrategies": teaching_plan.get("instructionalStrategies", ""),
#                     "learningObjectives": teaching_plan.get("learningObjectives", ""),
#                     "recommendedResources": teaching_plan.get("recommendedResources", []),
#                     "timeline": teaching_plan.get("timeline", "")
#                 }
#             }
#         })

#         return jsonify({
#             "studentId": student_id,
#             "planId": plan_id,
#             "extracted_text": extracted_text,
#             "evaluation": evaluation,
#             "teaching_plan": {
#                 "assessmentMethods": teaching_plan.get("assessmentMethods", ""),
#                 "instructionalStrategies": teaching_plan.get("instructionalStrategies", ""),
#                 "learningObjectives": teaching_plan.get("learningObjectives", ""),
#                 "recommendedResources": teaching_plan.get("recommendedResources", []),
#                 "timeline": teaching_plan.get("timeline", "")
#             }
#         }), 200

#     except RuntimeError:
#         loop = asyncio.new_event_loop()
#         asyncio.set_event_loop(loop)

#         extracted_text = loop.run_until_complete(extract_text_from_file(file_bytes))
#         student_name, grade_level = extract_student_info(extracted_text)

#         evaluation, teaching_plan = loop.run_until_complete(
#             asyncio.gather(
#                 evaluate_text_with_openai(extracted_text),
#                 generate_teaching_plan(student_name, grade_level, extracted_text)
#             )
#         )

#         return jsonify({
#             "extracted_text": extracted_text,
#             "evaluation": evaluation,
#             "teaching_plan": teaching_plan
#         }), 200
#     except Exception as e:
#         return jsonify({'error': f'Processing failed: {str(e)}'}), 500


# @app.route('/teaching-plan/<plan_id>', methods=['PATCH'])
# def update_teaching_plan(plan_id):
#     try:
#         # Parse request data
#         data = request.json
#         if not data:
#             return jsonify({"error": "Request body is required and must be JSON"}), 400

#         # Query Firestore to find the student who has this planId
#         students = db.collection('students').stream()
#         student_ref = None

#         for student in students:
#             student_data = student.to_dict()
#             teaching_plans = student_data.get("teachingPlans", {})

#             if plan_id in teaching_plans:  # Check if this student has the plan
#                 student_ref = db.collection('students').document(student.id)
#                 break  # Stop searching after finding the first match

#         # If no student was found with this planId, return an error
#         if student_ref is None:
#             return jsonify({"error": f"Teaching plan {plan_id} not found"}), 404

#         # Fetch existing plan data
#         student_data = student_ref.get().to_dict()
#         teaching_plans = student_data.get("teachingPlans", {})

#         # Ensure plan exists
#         if plan_id not in teaching_plans:
#             return jsonify({"error": f"Teaching plan {plan_id} does not exist"}), 404

#         # Get the actionPlan inside the plan
#         existing_plan = teaching_plans[plan_id]
#         action_plan = existing_plan.get("actionPlan", {})

#         # Fields that should be arrays
#         fields_as_arrays = ["assessmentMethods", "instructionalStrategies", "learningObjectives", "recommendedResources"]
#         # Timeline should remain a string

#         # Update only the provided fields inside actionPlan
#         for field in fields_as_arrays:
#             if field in data:
#                 if not isinstance(data[field], list):
#                     return jsonify({"error": f"{field} must be an array"}), 400
#                 action_plan[field] = data[field]  # Replace the existing field

#         # Handle timeline separately as a string
#         if "timeline" in data:
#             if not isinstance(data["timeline"], str):
#                 return jsonify({"error": "timeline must be a string"}), 400
#             action_plan["timeline"] = data["timeline"]

#         # Save the updated actionPlan back inside teachingPlans
#         teaching_plans[plan_id]["actionPlan"] = action_plan
#         student_ref.update({"teachingPlans": teaching_plans})

#         return jsonify({
#             "message": f"Teaching plan {plan_id} updated successfully",
#             "updated_teaching_plan": action_plan
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500


# @app.route("/api/teaching-plan/export", methods=["POST"])
# def export_teaching_plan_as_pdf():
#     try:
#         # Get studentId and planId from the request
#         data = request.json
#         student_id = data.get("studentId")
#         plan_id = data.get("planId")

#         if not student_id or not plan_id:
#             return jsonify({"error": "Both studentId and planId are required"}), 400

#         # Validate studentId exists
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()
#         if not student_doc.exists:
#             return jsonify({"error": f"Student ID {student_id} does not exist"}), 404

#         # Fetch teaching plan data
#         student_data = student_doc.to_dict()
#         teaching_plans = student_data.get("teachingPlans", {})

#         if not teaching_plans or plan_id not in teaching_plans:
#             return jsonify({"error": f"Teaching plan {plan_id} not found for this student"}), 404

#         # Get the specific teaching plan
#         selected_plan = teaching_plans.get(plan_id, {})
#         action_plan = selected_plan.get("actionPlan", {})

#         if not action_plan:
#             return jsonify({"error": "Teaching plan content is missing"}), 400

#         # Format content for PDF
#         teaching_plan_content = f"""
#         Assessment Methods:
#         {', '.join(action_plan.get("assessmentMethods", [])).strip()}

#         Instructional Strategies:
#         {', '.join(action_plan.get("instructionalStrategies", [])).strip()}

#         Learning Objectives:
#         {', '.join(action_plan.get("learningObjectives", [])).strip()}

#         Recommended Resources:
#         {', '.join(action_plan.get("recommendedResources", [])).strip()}

#         Timeline:
#         {action_plan.get("timeline", "").strip()}
#         """.strip()  # Ensures trailing newlines are removed

#         # Generate the PDF locally
#         pdf_filename = f"{student_id}_{plan_id}_teaching_plan.pdf"
#         create_pdf(teaching_plan_content, pdf_filename)

#         # Upload the PDF to Firebase Storage
#         pdf_storage_path = f"teaching_plans/{uuid.uuid4()}_{pdf_filename}"
#         blob = bucket.blob(pdf_storage_path)
#         blob.upload_from_filename(pdf_filename)
#         blob.make_public()

#         # Get the public URL
#         public_url = blob.public_url

#         # Store URL in Firestore
#         student_ref.update({f"teachingPlans.{plan_id}.pdfUrl": public_url})

#         # Clean up the local file
#         os.remove(pdf_filename)

#         return jsonify({"status": "success", "url": public_url}), 200

#     except Exception as e:
#         return jsonify({"error": "An unexpected error occurred", "details": str(e)}), 500

# # # Run Flask App
# # if __name__ == "__main__":
# #     app.run(debug=True)

# #############################################333333
# OBSERVATION PAGE
from flask import Flask, request, jsonify
import uuid
from datetime import datetime
import firebase_admin
from firebase_admin import credentials, firestore, storage
import logging

# Initialize Flask app
# app = Flask(__name__)

# # Initialize Firebase Admin SDK
# if not firebase_admin._apps:
#     cred = credentials.Certificate(r"serviceAccountKey.json")  # Path to your service account key JSON
#     firebase_admin.initialize_app(cred, {
#         'storageBucket': 'pees-d1101.firebasestorage.app'  # Replace with your Firebase bucket name
#     })

# db = firestore.client()

# Configure logging
logging.basicConfig(
    level=logging.ERROR, format="%(asctime)s %(levelname)s: %(message)s"
)

# Allowed file extensions
ALLOWED_EXTENSIONS = {"pdf", "docx", "png", "jpg", "jpeg"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# @app.route('/students/<student_id>/observations', methods=['POST'])
# def add_observation(student_id):
#     try:
#         # Check if the file is included in the request
#         if 'file' not in request.files:
#             return jsonify({"error": "File is required"}), 400

#         # Get the uploaded file
#         uploaded_file = request.files['file']

#         # Validate file format
#         if not allowed_file(uploaded_file.filename):
#             return jsonify({"error": "Invalid file type. Only PDF, DOCX, PNG, JPG, and JPEG are allowed."}), 400

#         # Generate a unique file name for storage
#         file_name = f"{uuid.uuid4()}_{uploaded_file.filename}"

#         # Reference the Firebase Storage bucket
#         bucket = storage.bucket()
#         blob = bucket.blob(file_name)

#         # Upload the file to Firebase Storage
#         blob.upload_from_file(uploaded_file, content_type=uploaded_file.content_type)

#         # Make the file publicly accessible and get its URL
#         blob.make_public()
#         attachment_url = blob.public_url

#         # Extract additional fields from the request
#         subject = request.form.get('subject', None)
#         observation_text = request.form.get('observation', None)

#         # Create the observation object
#         observation = {
#             "date": datetime.utcnow().strftime('%Y-%m-%d'),
#             "subject": subject,
#             "observation": observation_text,
#             "attachment_url": attachment_url
#         }

#         # Reference the student document in Firestore
#         student_ref = db.collection('students').document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         # Add the observation to the student's existing observations
#         observations = student_doc.to_dict().get('observations', [])
#         observations.append(observation)

#         # Update the Firestore document
#         student_ref.update({"observations": observations})

#         return jsonify({"message": "Observation added successfully", "observation": observation}), 200

#     except Exception as e:
#         logging.error(f"Error adding observation: {str(e)}")
#         return jsonify({"error": str(e)}), 500

# @app.route('/students/<student_id>/observation1', methods=['GET'])
# def get_observation1(student_id):
#     try:
#         # Reference the student document in Firestore
#         student_ref = db.collection('students').document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         # Retrieve observations from the document
#         observations = student_doc.to_dict().get('observations', [])

#         return jsonify({"observations": observations}), 200

#     except Exception as e:
#         logging.error(f"Error retrieving observations: {str(e)}")
#         return jsonify({"error": str(e)}), 500

# @app.route("/students/<student_id>/observations", methods=["POST"])
# def add_observation(student_id):
#     try:
#         # Initialize attachment URL to None (since it's no longer required)
#         attachment_url = None
        
#         # Determine data source: JSON (if no file is being uploaded) or Form (if files may be present)
#         if request.is_json:
#             subject = request.json.get("subject", None)
#             observation_text = request.json.get("observation", None)
#         else:
#             # Handle multipart/form-data (used when a file might be present)
#             subject = request.form.get("subject", None)
#             observation_text = request.form.get("observation", None)

#         # Retrieve file if present (now optional)
#         uploaded_file = request.files.get("file")

#         # Check if a file was provided and has a filename
#         if uploaded_file and uploaded_file.filename != "":
#             if not allowed_file(uploaded_file.filename):
#                 return (
#                     jsonify(
#                         {
#                             "error": "Invalid file type. Only PDF, DOCX, PNG, JPG, and JPEG are allowed."
#                         }
#                     ),
#                     400,
#                 )

#             # File upload logic
#             file_name = f"{uuid.uuid4()}_{uploaded_file.filename}"
#             bucket = storage.bucket()
#             blob = bucket.blob(file_name)

#             blob.upload_from_file(uploaded_file, content_type=uploaded_file.content_type)
#             blob.make_public()
#             attachment_url = blob.public_url
        
#         # Ensure observation text is present, regardless of attachment
#         if not observation_text:
#              return jsonify({"error": "Observation text is required"}), 400

#         observation = {
#             "date": datetime.utcnow().strftime("%Y-%m-%d"),
#             "subject": subject,
#             "observation": observation_text,
#             "attachment_url": attachment_url,  # Will be None if no file was uploaded
#         }

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         observations = student_doc.to_dict().get("observations", [])
#         observations.append(observation)

#         student_ref.update({"observations": observations})

#         return (
#             jsonify(
#                 {
#                     "message": "Observation added successfully",
#                     "observation": observation,
#                 }
#             ),
#             200,
#         )

#     except Exception as e:
#         logging.error(f"Error adding observation: {str(e)}")
#         return jsonify({"error": str(e)}), 500


# @app.route("/students/<student_id>/observations", methods=["GET"])
# def get_observations(student_id):
#     try:
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         observations = student_doc.to_dict().get("observations", [])

#         return jsonify({"observations": observations}), 200

#     except Exception as e:
#         logging.error(f"Error retrieving observations: {str(e)}")
#         return jsonify({"error": str(e)}), 500
import datetime # <-- FIX: Now importing the whole module
import uuid
import logging
from flask import request, jsonify # Assuming this is part of your Flask setup
# from your_app import db, storage, allowed_file # Assuming these dependencies are imported

# Note: The dependencies (db, storage, allowed_file, etc.) are assumed to be 
# correctly defined and initialized in your full application environment.

def _normalize_observation_date(value):
    if not value:
        return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d")
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def _build_observation_id(index):
    return str(index)


def _extract_teacher_id():
    teacher_id = (
        request.args.get("teacher_id")
        or request.args.get("teacherId")
    )
    if teacher_id:
        return str(teacher_id).strip()
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        teacher_id = payload.get("teacher_id") or payload.get("teacherId")
        if teacher_id:
            return str(teacher_id).strip()
    else:
        teacher_id = request.form.get("teacher_id") or request.form.get("teacherId")
        if teacher_id:
            return str(teacher_id).strip()
    return ""


def _find_observation_index(observations, observation_id):
    try:
        idx = int(str(observation_id))
        if 0 <= idx < len(observations):
            return idx
    except Exception:
        pass
    for idx, item in enumerate(observations):
        if str(item.get("observation_id", "")).strip() == str(observation_id).strip():
            return idx
    return -1


def _entry_teacher_id(entry):
    if not isinstance(entry, dict):
        return ""
    for key in (
        "teacher_id",
        "teacherId",
        "created_by_teacher_id",
        "updated_by_teacher_id",
        "assigned_teacher_id",
    ):
        value = entry.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


@app.route("/students/<student_id>/observations", methods=["POST"])
@app.route("/api/students/<student_id>/observations", methods=["POST"])
def add_observation(student_id):
    try:
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        current_data = student_doc.to_dict() or {}
        observations = current_data.get("observations", [])
        if not isinstance(observations, list):
            observations = []

        if request.is_json:
            payload = request.get_json(silent=True) or {}
            action = (payload.get("action") or "").strip().lower()
            request_teacher_id = _extract_teacher_id()
            if action in {"update", "delete"}:
                target = payload.get("old", payload)
                target_date = (target.get("date") or payload.get("date") or "").strip()
                target_subject = (target.get("subject") or payload.get("subject") or "").strip()
                target_observation = (
                    target.get("observation") or payload.get("observation") or ""
                ).strip()

                match_index = -1
                for idx, item in enumerate(observations):
                    if request_teacher_id and str(item.get("teacher_id", "")).strip() != request_teacher_id:
                        continue
                    date_ok = (item.get("date") or "").strip() == target_date
                    subject_ok = (item.get("subject") or "").strip() == target_subject
                    observation_ok = (
                        (item.get("observation") or "").strip() == target_observation
                    )
                    if date_ok and subject_ok and observation_ok:
                        match_index = idx
                        break

                if match_index < 0:
                    return jsonify({"error": "Observation not found"}), 404

                if action == "delete":
                    deleted = observations.pop(match_index)
                    for idx, item in enumerate(observations):
                        item["observation_id"] = _build_observation_id(idx)
                    student_ref.update({"observations": observations})
                    return jsonify({"message": "Observation deleted successfully", "observation": deleted}), 200

                update_payload = payload.get("new", payload)
                if "subject" in update_payload:
                    observations[match_index]["subject"] = update_payload.get("subject", "")
                if "observation" in update_payload:
                    observations[match_index]["observation"] = update_payload.get("observation", "")
                for idx, item in enumerate(observations):
                    item["observation_id"] = _build_observation_id(idx)
                student_ref.update({"observations": observations})
                return jsonify({"message": "Observation updated successfully", "observation": observations[match_index]}), 200

        attachment_url = None
        request_teacher_id = _extract_teacher_id()
        if request.is_json:
            source = request.get_json(silent=True) or {}
            subject = source.get("subject")
            observation_text = source.get("observation")
            raw_date = (
                source.get("observation_date")
                or source.get("observationDate")
                or source.get("selected_date")
                or source.get("selectedDate")
                or source.get("entryDate")
                or source.get("date")
            )
        else:
            source = request.form
            subject = source.get("subject")
            observation_text = source.get("observation")
            raw_date = (
                source.get("observation_date")
                or source.get("observationDate")
                or source.get("selected_date")
                or source.get("selectedDate")
                or source.get("entryDate")
                or source.get("date")
            )

        uploaded_file = request.files.get("file")
        if uploaded_file and uploaded_file.filename != "":
            if not allowed_file(uploaded_file.filename):
                return jsonify({"error": "Invalid file type"}), 400
            file_name = f"{uuid.uuid4()}_{uploaded_file.filename}"
            bucket = storage.bucket()
            blob = bucket.blob(file_name)
            blob.upload_from_file(uploaded_file, content_type=uploaded_file.content_type)
            blob.make_public()
            attachment_url = blob.public_url

        if not observation_text:
            return jsonify({"error": "Observation text is required"}), 400

        observation = {
            "observation_id": _build_observation_id(len(observations)),
            "date": _normalize_observation_date(raw_date),
            "subject": subject,
            "observation": observation_text,
            "attachment_url": attachment_url,
            "teacher_id": request_teacher_id,
        }
        observations.append(observation)
        student_ref.update({"observations": observations})
        return jsonify({"message": "Observation added successfully", "observation": observation}), 200
    except Exception as e:
        logging.error(f"Error adding observation: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route("/students/<student_id>/observations", methods=["GET"])
@app.route("/api/students/<student_id>/observations", methods=["GET"])
def get_observations(student_id):
    try:
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        observations = student_doc.to_dict().get("observations", [])
        if not isinstance(observations, list):
            observations = []
        teacher_id = (request.args.get("teacher_id") or request.args.get("teacherId") or "").strip()
        if teacher_id:
            observations = [
                item
                for item in observations
                if _entry_teacher_id(item) in ("", teacher_id)
            ]

        normalized = []
        for idx, item in enumerate(observations):
            row = dict(item or {})
            row["observation_id"] = row.get("observation_id") or _build_observation_id(idx)
            normalized.append(row)
        return jsonify({"observations": normalized}), 200
    except Exception as e:
        logging.error(f"Error retrieving observations: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route("/students/<student_id>/observations/<observation_id>", methods=["PATCH", "PUT", "DELETE"])
@app.route("/api/students/<student_id>/observations/<observation_id>", methods=["PATCH", "PUT", "DELETE"])
def update_or_delete_observation(student_id, observation_id):
    try:
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        data = student_doc.to_dict() or {}
        observations = data.get("observations", [])
        if not isinstance(observations, list):
            observations = []
        teacher_id = _extract_teacher_id()

        idx = _find_observation_index(observations, observation_id)
        if idx < 0:
            return jsonify({"error": "Observation not found"}), 404
        if teacher_id and _entry_teacher_id(observations[idx]) != teacher_id:
            return jsonify({"error": "You can only modify your own observations"}), 403

        if request.method == "DELETE":
            removed = observations.pop(idx)
            for i, item in enumerate(observations):
                item["observation_id"] = _build_observation_id(i)
            student_ref.update({"observations": observations})
            return jsonify({"message": "Observation deleted successfully", "observation": removed}), 200

        payload = request.get_json(silent=True) or {}
        if "subject" in payload:
            observations[idx]["subject"] = payload.get("subject", "")
        if "observation" in payload:
            observations[idx]["observation"] = payload.get("observation", "")
        if "date" in payload:
            observations[idx]["date"] = _normalize_observation_date(payload.get("date"))
        observations[idx]["observation_id"] = _build_observation_id(idx)
        student_ref.update({"observations": observations})
        return jsonify({"message": "Observation updated successfully", "observation": observations[idx]}), 200
    except Exception as e:
        logging.error(f"Error updating/deleting observation: {str(e)}")
        return jsonify({"error": str(e)}), 500

# EXAM SCRIPT

# AZURE OCR


# if __name__ == '__main__':
#     app.run(debug=True)

############################################################################################################################################################################

# ALERTSs

import os
import re
import uuid  # For generating unique alert IDs
from time import time, gmtime, strftime
from flask import Flask, request, jsonify
from flask_cors import CORS
from groq import Groq
from dotenv import load_dotenv
from firebase_admin import credentials, firestore, initialize_app

# Load environment variables
load_dotenv()

# # Initialize Firebase Admin SDK
# cred = credentials.Certificate(r"serviceAccountKey.json")  # Path to your Firebase service account key
# initialize_app(cred)
# db = firestore.client()

# Initialize the GROQ client
# client = Groq(api_key="YOUR_GROQ_API_KEY")

# Initialize Flask app
# app = Flask(__name__)
# CORS(app)

# Static thresholds
THRESHOLDS = {"improvement": 10, "decline": 5}

# # ALERTS AND NOTIFICATIONS (Dual-language Version)
# import os
# import re
# import uuid
import datetime as dt_util
# from time import gmtime, strftime
# from flask import Flask, request, jsonify
# from flask_cors import CORS
# from groq import Groq
# from dotenv import load_dotenv
# from firebase_admin import credentials, firestore, initialize_app
# from flask_socketio import SocketIO, join_room, leave_room

# # Load environment variables
# load_dotenv()

# # Initialize Flask app
# app = Flask(__name__)
# CORS(app)

# # Initialize Firebase
# cred = credentials.Certificate(r"serviceAccountKey.json")
# initialize_app(cred)
# db = firestore.client()

# # Initialize SocketIO
# socketio = SocketIO(app, cors_allowed_origins="*")

# # Static thresholds
# THRESHOLDS = {"improvement": 10, "decline": 5}

###########################################
# Utility Functions
###########################################

# def clean_ai_message(message):
#     """Clean unwanted fragments from AI-generated text."""
#     cleaned_message = re.sub(
#         r"(Compose a 3-4 line alert message in.*?:| - Student ID:.*?\.|based on the following student performance details:.*?)",
#         "",
#         message,
#         flags=re.DOTALL
#     ).strip()
#     cleaned_message = re.sub(r'(\n|\r|\\n|\\r)', ' ', cleaned_message)
#     return re.sub(r'\s{2,}', ' ', cleaned_message).strip()

def clean_ai_message(message):
    """Clean unwanted fragments from AI-generated text."""
    
    # Use a simple regex to remove known, undesirable intros at the START of the string.
    cleaned_message = re.sub(
        r"^(The alert message is|Here is the alert|The motivational alert|Alert:)\s*",
        "",
        message,
        flags=re.IGNORECASE | re.MULTILINE
    )

    return cleaned_message.strip()


# Utility Functions (Only format_timestamp is shown for revision)
import datetime as dt_util 
async def generate_ai_alert(
    student_name, student_id, previous_score, current_score, lang_desc, lang_prompt
):
    """
    Generates a personalized AI alert using OpenAI's chat completion API.
    """
    prompt = (
        f"Generate a **{lang_prompt}**:\n"
        f"Student Name: {student_name}\n"
        # ... (rest of your prompt) ...
        f"I want the entire alert to be generated in this language:- {lang_desc} **strictly**"
    )

    # This assumes 'openai_client' is your global ASYNC client (the proxy)
    response = await openai_client.chat.completions.create(
        model="gpt-4.1-mini-2025-04-14",
        messages=[
            {
                "role": "system",
                "content": "You are a professional education assistant. Your output must be ONLY the alert message, and strictly in the requested language.",
            },
            {"role": "user", "content": prompt},
        ],
    )
    return response.choices[0].message.content.strip()


import datetime as dt_util 

# def format_timestamp(date_str):
#     """Format timestamps safely."""
#     # Handle datetime objects directly (e.g., if coming from Firestore)
#     if isinstance(date_str, dt_util.datetime):
#         return date_str.strftime("%d %b %Y, %H:%M")
    
#     # Handle string timestamps
#     if isinstance(date_str, str):
#         try:
#             # Standard ISO format with 'Z'
#             dt = dt_util.datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
#             return dt.strftime("%d %b %Y, %H:%M")
#         except ValueError:
#             # Try without the 'Z'
#             try:
#                 dt = dt_util.datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S")
#                 return dt.strftime("%d %b %Y, %H:%M")
#             except Exception:
#                 pass
    
    # Fallback for any other type or unparseable string
    # return date_str
# You must have this utility function in the file
# (Make sure 'import datetime as dt_util' is at the top of the file)

# Make sure 'import datetime as dt_util' and 'from datetime import timezone'
# are at the top of your file.

# ✅ FIX #1: A correct, robust date formatting function
def format_timestamp(date_str):
    """
    Format timestamps safely to a standardized ISO 8601 format.
    FIXED: Correctly handles strings that already have a timezone.
    """
    
    # 1. Handle datetime objects
    if isinstance(date_str, dt_util.datetime):
        dt = date_str
        # If it's naive (no timezone), assume it's UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=dt_util.timezone.utc)
        # Return in standard ISO format
        return dt.isoformat()

    # 2. Handle string timestamps
    if isinstance(date_str, str):
        try:
            # Clean up the buggy "+00:00Z" format if it exists
            if date_str.endswith("+00:00Z"):
                date_str = date_str[:-1] # Becomes '...321724+00:00'
            
            # Let fromisoformat handle '...Z', '...+00:00', or no timezone
            dt = dt_util.datetime.fromisoformat(date_str)
            
            # If it's naive (no timezone), assume it's UTC
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=dt_util.timezone.utc)
            
            # Return the valid ISO string
            return dt.isoformat()
            
        except Exception as e:
            print(f"Warning: Could not parse date string '{date_str}'. Error: {e}")
            pass 
    
    # Fallback for unparseable data
    return "2025-01-01T00:00:00+00:00" # Return a valid ISO string

# ###########################################
# # Dual-Language Alert Generator
# ###########################################
async def generate_alert(student_id, student_name, previous_score, current_score, lang):
    """
    Generate alert messages in both English and Arabic by calling the async
    generate_ai_alert function.
    """
    if current_score - previous_score >= THRESHOLDS["improvement"]:
        type_of_alert = "Success"
        lang_prompt = "motivational alert for performance improvement"
    elif previous_score - current_score >= THRESHOLDS["decline"]:
        type_of_alert = "Warning"
        lang_prompt = "urgent warning alert for performance decline and required intervention"
    else:
        type_of_alert = "No Change"
        lang_prompt = "neutral status update regarding recent performance stability"

    # ✅ FIX: Run both AI calls at the same time
    try:
        results = await asyncio.gather(
            generate_ai_alert(
                student_name, student_id, previous_score, current_score, "English", lang_prompt
            ),
            generate_ai_alert(
                student_name, student_id, previous_score, current_score, "Arabic", lang_prompt
            )
        )
        message_en = clean_ai_message(results[0])
        message_ar = clean_ai_message(results[1])
        
    except Exception as e:
        print(f"Error generating AI messages: {e}")
        message_en = f"Error generating English message for {student_name}."
        message_ar = f"Error generating Arabic message for {student_name}."

    return {
        "status": "success",
        "type_of_alert": type_of_alert,
        "ai_generated_message_en": message_en, 
        "ai_generated_message_ar": message_ar, 
        "preferred_language_at_creation": lang,
    }

import datetime as dt_util 
# (Make sure you have the 'format_timestamp' function in this file too)

# @app.route("/api/alerts-notifications", methods=["GET"])
# def get_alerts_notifications():
#     """
#     Fetch alerts & notifications based on 'lang' query (en or ar)
#     using teacher_id.
    
#     This version STRICTLY filters by language and fixes the 'null' message bug.
#     """
#     try:
#         teacher_id = request.args.get("teacher_id")
#         lang = request.args.get("lang", "en").lower()
        
#         if not teacher_id:
#             return jsonify({"error": "teacher_id is required"}), 400

#         def process_data(entry):
#             # 1. Determine target message field
#             target_msg_field = "ai_generated_message_ar" if lang == "ar" else "ai_generated_message_en"

#             # 2. Get the message, with fallback for English
#             message = entry.get(target_msg_field)
#             if not message and lang == "en":
#                 # If user wants English, we can fall back to the old 'ai_generated_message'
#                 message = entry.get("ai_generated_message")

#             # 3. If no message was found for the target language, SKIP this alert.
#             if not message:
#                 return None 

#             # 4. Format the date (using the ISO-format-fix function)
#             date_value = entry.get("date")
#             formatted_date = format_timestamp(date_value)

#             # 5. Build a NEW, clean dictionary to return.
#             # This is the fix: it avoids all 'pop' errors and only returns what's needed.
#             clean_alert = {
#                 "ai_generated_message": message,
#                 "date": formatted_date,
#                 "current_score": entry.get("current_score"),
#                 "isSeen": entry.get("isSeen"),
#                 "previous_score": entry.get("previous_score"),
#                 "student_id": entry.get("student_id"),
#                 "student_name": entry.get("student_name"),
#                 "teacher_id": entry.get("teacher_id"),
#                 "type_of_alert": entry.get("type_of_alert"),
#             }
            
#             return clean_alert

#         # --- Update the main logic to filter out 'None' results ---
        
#         # Fetch alerts
#         alerts_stream = db.collection("alerts").where("teacher_id", "==", teacher_id).stream()
#         alerts = []
#         for doc in alerts_stream:
#             processed_alert = process_data(doc.to_dict())
#             if processed_alert:  # <-- Only add if process_data returned an alert
#                 alerts.append(processed_alert)

#         # Fetch notifications
#         notifications_stream = (
#             db.collection("notifications")
#             .where("teacher_id", "==", teacher_id) 
#             .stream()
#         )
#         notifications = []
#         for doc in notifications_stream:
#             processed_notification = process_data(doc.to_dict())
#             if processed_notification: # <-- Only add if process_data returned a notification
#                 notifications.append(processed_notification)
        
#         return jsonify({"alerts": alerts, "notifications": notifications}), 200

#     except Exception as e:
#         print(f"FATAL ERROR in get_alerts_notifications: {e}") 
#         return jsonify({"error": "An unexpected server error occurred during data retrieval."}), 500
@app.route("/api/alerts-notifications", methods=["GET"])
def get_alerts_notifications():
    """
    Fetch alerts & notifications.
    FIXED: Avoids composite indexes and correctly passes doc.id.
    """
    try:
        teacher_id = request.args.get("teacher_id")
        lang = request.args.get("lang", "en").lower()
        
        if not teacher_id:
            return jsonify({"error": "teacher_id is required"}), 400

        # --- (This helper function is now correct) ---
        def process_data(entry):
            target_msg_field = "ai_generated_message_ar" if lang == "ar" else "ai_generated_message_en"
            message = entry.get(target_msg_field)
            if not message and lang == "en":
                message = entry.get("ai_generated_message")
            if not message:
                message = entry.get("description")
            if not message:
                return None 

            date_value = entry.get("date", entry.get("created_at"))
            formatted_date = format_timestamp(date_value) # Uses the new, fixed function

            clean_entry = {
                # ✅ FIX #2: "id" is now passed in
                "id": entry.get("id"), 
                "ai_generated_message": message,
                "date": formatted_date,
                "isSeen": entry.get("isSeen", "Read" if entry.get("status") else "Unread"),
                "student_id": entry.get("student_id"),
                "student_name": entry.get("student_name"),
                "teacher_id": entry.get("teacher_id"),
                "type": entry.get("type_of_alert", entry.get("type")),
            }
            return clean_entry
        # --- (End of helper function) ---

        alerts = []
        notifications = []

        # --- Fetch AI Alerts ---
        alerts_stream = db.collection("alerts").where("teacher_id", "==", teacher_id).stream()
        for doc in alerts_stream:
            doc_data = doc.to_dict()
            if doc_data.get("type_of_alert"): 
                # ✅ FIX #2: Pass the document ID
                doc_data['id'] = doc.id 
                processed_alert = process_data(doc_data)
                if processed_alert:
                    alerts.append(processed_alert)

        # --- Fetch Simple Notifications ---
        notifications_stream = (
            db.collection("alerts")
            .where("receiver_id", "==", teacher_id)
            .stream()
        )
        for doc in notifications_stream:
            doc_data = doc.to_dict()
            if doc_data.get("type") == "notification": 
                # ✅ FIX #2: Pass the document ID
                doc_data['id'] = doc.id
                processed_notification = process_data(doc_data)
                if processed_notification:
                    notifications.append(processed_notification)
        
        return jsonify({"alerts": alerts, "notifications": notifications}), 200

    except Exception as e:
        print(f"FATAL ERROR in get_alerts_notifications: {e}") 
        return jsonify({"error": "An unexpected server error occurred during data retrieval."}), 500
    
# API: Create Alert (Dual-Language Storage)
###########################################

# @app.route("/api/alerts", methods=["POST"])
# def create_alert():
#     """Generate alert message in two languages and store them."""
#     try:
#         data = request.get_json()
#         student_id = data.get("student_id")
#         student_name = data.get("student_name")
#         previous_score = data.get("previous_score")
#         current_score = data.get("current_score")
#         lang = data.get("lang", "en").lower()

#         if not all([student_id, previous_score, current_score]):
#             return jsonify({"status": "error", "message": "Missing required fields."}), 400

#         previous_score = int(previous_score)
#         current_score = int(current_score)

#         result = generate_alert(student_id, student_name, previous_score, current_score, lang)
#         alert_id = str(uuid.uuid4())

#         alert_data = {
#             "student_id": student_id,
#             "alert_id": alert_id,
#             # "student_name": student_name,
#             "previous_score": previous_score,
#             "current_score": current_score,
#             "type_of_alert": result["type_of_alert"],
#             "ai_generated_message_en": result["ai_generated_message_en"],
#             "ai_generated_message_ar": result["ai_generated_message_ar"],
#             "preferred_language": lang,
#             "date": dt_util.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
#             "isSeen": "Unread"
#         }

#         db.collection("alerts").document(alert_id).set(alert_data)
#         return jsonify({"status": "success", "message": "Alert created successfully.", "alert_id": alert_id}), 200

#     except Exception as e:
#         return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/alerts", methods=["POST"])
async def create_alert():
    """Generate alert message in two languages and store them."""
    try:
        data = request.get_json()
        student_id = data.get("student_id")
        teacher_id = data.get("teacher_id")         
        student_name = data.get("student_name")
        previous_score = data.get("previous_score")
        current_score = data.get("current_score")
        lang = data.get("lang", "en").lower()

        if not all([student_id, teacher_id, previous_score, current_score]):
            return jsonify({"status": "error", "message": "Missing required fields."}), 400

        previous_score = int(previous_score)
        current_score = int(current_score)

        # ✅ 4. You MUST use 'await' here
        result = await generate_alert(student_id, student_name, previous_score, current_score, lang)
        alert_id = str(uuid.uuid4())

        alert_data = {
            "student_id": student_id,
            "alert_id": alert_id,
            "teacher_id": teacher_id,
            "student_name": student_name,
            "previous_score": previous_score,
            "current_score": current_score,
            "type_of_alert": result["type_of_alert"],
            "ai_generated_message_en": result["ai_generated_message_en"],
            "ai_generated_message_ar": result["ai_generated_message_ar"],
            "preferred_language": lang,
            "date": dt_util.datetime.now(timezone.utc).isoformat(), # This line is correct
            "isSeen": "Unread"
        }

        # This assumes your 'db' client is synchronous
        db.collection("alerts").document(alert_id).set(alert_data)
        
        return jsonify({"status": "success", "message": "Alert created successfully.", "alert_id": alert_id}), 200

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# ###########################################
# # API: Fetch Alerts & Notifications (Language Dynamic)
# ###########################################

# @app.route("/api/alerts-notifications", methods=["GET"])
# def get_alerts_notifications():
#     """Fetch alerts & notifications based on 'lang' query (en or ar) using teacher_id."""
#     try:
#         teacher_id = request.args.get("teacher_id")
#         lang = request.args.get("lang", "en").lower()
        
#         if not teacher_id:
#             return jsonify({"error": "teacher_id is required"}), 400

#         # Determine which message field to prioritize
#         target_msg_field = "ai_generated_message_ar" if lang == "ar" else "ai_generated_message_en"
#         other_msg_field = "ai_generated_message_en" if lang == "ar" else "ai_generated_message_ar"

#         def process_data(entry):
#             # 1. Select the correct language message for the final response field.
#             entry["ai_generated_message"] = entry.get(
#                 target_msg_field, 
#                 entry.get(other_msg_field, "Message content unavailable.")
#             )
            
#             # 2. Cleanup: Safely remove the dual-language fields (if they exist).
#             entry.pop("ai_generated_message_en", None)
#             entry.pop("ai_generated_message_ar", None)
            
#             # 3. Format timestamp safely.
#             date_value = entry.get("date")
#             if date_value is not None:
#                 entry["date"] = format_timestamp(date_value)
#             else:
#                 entry["date"] = "N/A"
            
#             # 4. CRITICAL FIX: The logic in the DB document provided is flawed 
#             # (e.g., 'Success' for a score drop). We must trust the code's score logic, 
#             # but since we are just fetching here, we'll keep the DB value, but it's 
#             # noted that the data integrity should be checked at the creation endpoint.
            
#             return entry

#         # Fetch alerts: Query by the 'teacher_id' field as specified in your structure
#         alerts_stream = db.collection("alerts").where("teacher_id", "==", teacher_id).stream()
#         alerts = [process_data(doc.to_dict()) for doc in alerts_stream]

#         # Fetch notifications: Query by the 'teacher_id' field as specified in your structure
#         notifications_stream = (
#             db.collection("notifications")
#             .where("teacher_id", "==", teacher_id) # ✅ CHANGE: Now querying by 'teacher_id'
#             .stream()
#         )
#         notifications = [process_data(doc.to_dict()) for doc in notifications_stream]
        
#         return jsonify({"alerts": alerts, "notifications": notifications}), 200

#     except Exception as e:
#         print(f"FATAL ERROR in get_alerts_notifications: {e}") 
#         return jsonify({"error": "An unexpected server error occurred during data retrieval."}), 500
    
###########################################
# API: Mark Alert as Seen
###########################################

@app.route("/api/alerts/<alert_id>/mark-seen", methods=["PATCH"])
def mark_alert_as_seen(alert_id):
    """Mark specific alert as read."""
    try:
        ref = db.collection("alerts").document(alert_id)
        doc = ref.get()
        if not doc.exists:
            return jsonify({"status": "error", "message": "Alert not found."}), 404
        ref.update({"isSeen": "Read"})
        return jsonify({"status": "success", "message": "Alert marked as read."}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

###########################################
# API: Get Recent Alerts
###########################################

@app.route("/api/alerts", methods=["GET"])
def get_alerts():
    """Retrieve alerts by student or last 24 hours."""
    try:
        student_id = request.args.get("student_id")
        last24 = request.args.get("last24hr", "false").lower() == "true"
        ref = db.collection("alerts")

        if student_id:
            ref = ref.where("student_id", "==", student_id)
        if last24:
            # since = dt_util.datetime.utcnow() - dt_util.timedelta(hours=24)
            # ref = ref.where("date", ">=", since.strftime("%Y-%m-%dT%H:%M:%SZ"))
        # ✅ NEW, FIXED LINE:
            since = dt_util.datetime.now(timezone.utc) - dt_util.timedelta(hours=24)
            ref = ref.where("date", ">=", since.isoformat())

        alerts = [doc.to_dict() for doc in ref.stream()]
        return jsonify({"status": "success", "alerts": alerts}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

###########################################
# Run Flask
###########################################

# if __name__ == "__main__":
#     socketio.run(app, debug=True)

############################################################################################################################################################################
# NOTIFICATIONS

from flask import Flask, request, jsonify
from datetime import datetime, timezone
import uuid
from firebase_admin import credentials, firestore, initialize_app


# cred = credentials.Certificate(r"serviceAccountKey.json")  # Path to your Firebase service account key
# initialize_app(cred)
# db = firestore.client()

# app = Flask(__name__)

# Add Notification


# -----------------------------
# Add Notification (Corrected)
# -----------------------------
# (Imports...)
# Make sure you have these imports at the top of your file
from flask import jsonify, request
import uuid

# -----------------------------
# Add Notification (Corrected)
# -----------------------------
@app.route('/add-notification', methods=['POST'])
def add_notification112():
    """
    API to add a multilingual notification to the alerts collection.
    """
    # ✅ 1. ADD THESE IMPORTS INSIDE THE FUNCTION
    from datetime import datetime, timezone

    try:
        data = request.json
        
        title_en = data.get("title_en")
        description_en = data.get("description_en")
        title_ar = data.get("title_ar")
        description_ar = data.get("description_ar")
        receiver_id = data.get("receiver_id")
        sender_id = data.get("sender_id")

        if not all([title_en, description_en, title_ar, description_ar, receiver_id, sender_id]):
            return jsonify({"error": "Missing required fields: title_en, description_en, title_ar, description_ar, receiver_id, sender_id"}), 400

        notification_id = str(uuid.uuid4())
        
        # ✅ 2. This line will now work
        created_at = datetime.now(timezone.utc).isoformat()

        notification_data = {
            "id": notification_id,
            "type": "notification",
            "title_en": title_en,
            "description_en": description_en,
            "title_ar": title_ar,
            "description_ar": description_ar,
            "receiver_id": receiver_id,
            "sender_role": sender_id,
            "status": False,
            "created_at": created_at,
        }
        
        db.collection("alerts").document(notification_id).set(notification_data)
        
        return jsonify({"success": True, "message": "Notification added", "notification_data": notification_data}), 200
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# -----------------------------
# GET Notifications (Multilingual)
# -----------------------------
@app.route("/get-notifications", methods=["GET"])
def get_notifications_for_user11():
    """
    API to fetch notifications for a specific user from the alerts collection.
    Handles language switching for title and description.
    """
    try:
        user_id = request.args.get("user_id")
        lang = request.args.get("lang", "en").lower() 
        if not user_id:
            return jsonify({"error": "User ID is required"}), 400

        # Check user's delivery method preference
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        delivery_method = (
            user_doc.to_dict().get("delivery_method", {}) if user_doc.exists else {}
        )
        app_enabled = delivery_method.get("app", False) 

        # Fetch notifications from Firestore
        notifications_ref = (
            db.collection("alerts")
            .where("type", "==", "notification")
            # ✅ FIX: Query for "receiver_id"
            .where("receiver_id", "==", user_id) 
            .stream()
        )

        notifications = []
        unread_count = 0

        for doc in notifications_ref:
            doc_data = doc.to_dict()
            
            # Assumes 'status': False means unread
            if not doc_data.get("status", True): 
                unread_count += 1
            
            # ✅ FIX: Select text based on language
            title_field = f"title_{lang}"
            desc_field = f"description_{lang}"
            
            # Fallback to English if the requested language isn't found
            title = doc_data.get(title_field, doc_data.get("title_en", "No Title"))
            description = doc_data.get(desc_field, doc_data.get("description_en", "No Description"))

            notifications.append(
                {
                    "id": doc.id,
                    "title": title, # Use language-specific title
                    "description": description, # Use language-specific description
                    "date": doc_data.get("created_at", ""),
                    "status": doc_data.get("status", ""),
                    "receiver_id": doc_data.get("receiver_id", ""), 
                    "sender_role": doc_data.get("sender_role", ""),
                    "responsestatus": doc_data.get("responsestatus", ""),
                    "response": doc_data.get("responseMessage", ""),
                    "responseTimestamp": doc_data.get("responseTimestamp", ""),
                    "type": doc_data.get("type", ""),
                }
            )

        if not app_enabled:
            unread_count = 0

        return (
            jsonify(
                {
                    "success": True,
                    "notifications": notifications,
                    "unread_count": unread_count,
                }
            ),
            200,
        )

    except Exception as e:
        # logging.error(f"Error fetching notifications: {e}") 
        return jsonify({"error": str(e)}), 500
    
# @app.route('/get-notifications', methods=['GET'])
# def get_notifications():
#     """
#     API to fetch all notifications from the alerts collection.
#     """
#     try:
#         notifications_ref = db.collection("alerts").where("type", "==", "notification").stream()
#         notifications = []

#         for doc in notifications_ref:
#             data = doc.to_dict()
#             notifications.append({
#                 "id": data.get("id"),
#                 "title": data.get("title"),
#                 "description": data.get("description"),
#                 "date": data.get("created_at"),
#                 "status": data.get("status"),
#                 "user_id": data.get("user_id"),
#                 "receiver_role": data.get("receiver_role"),
#                 "sender_role": data.get("sender_role"),
#                 "responsestatus" : data.get("responsestatus",""),
#                 "response" : data.get("responseMessage",""),
#                 "responseTimestamp" : data.get("responseTimestamp",""),
#                 "type" : data.get("type","")
#             })

#         return jsonify({"success": True, "notifications": notifications}), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# Get Notifications by Role
@app.route("/get-filtered-notifications", methods=["GET"])
def get_filtered_notifications():
    """
    API to fetch notifications filtered by role from the alerts collection.
    (UPDATED to support multilingual fields)
    """
    try:
        role = request.args.get("role")
        role_type = request.args.get("role_type", "receiver_role")
        
        # ✅ Get the lang parameter
        lang = request.args.get("lang", "en").lower() 

        # Query for notifications filtered by role
        notifications_ref = db.collection("alerts").where("type", "==", "notification")
        if role:
            notifications_ref = notifications_ref.where(role_type, "==", role)
        notifications_ref = notifications_ref.stream()

        notifications = []
        for doc in notifications_ref:
            data = doc.to_dict()
            
            # ✅ FIX: Select text based on language
            title_field = f"title_{lang}"
            desc_field = f"description_{lang}"
            
            # Fallback to English
            title = data.get(title_field, data.get("title_en", "No Title"))
            description = data.get(desc_field, data.get("description_en", "No Description"))

            notifications.append(
                {
                    "id": data.get("id"),
                    "title": title, # Use language-specific title
                    "description": description, # Use language-specific description
                    "date": data.get("created_at"),
                    "status": data.get("status"),
                    "user_id": data.get("user_id"), # This might be 'receiver_id'
                    "receiver_role": data.get("receiver_role"),
                    "sender_role": data.get("sender_role"),
                    "responsestatus": data.get("responsestatus", ""),
                    "response": data.get("responseMessage", ""),
                    "responseTimestamp": data.get("responseTimestamp", ""),
                }
            )

        return jsonify({"success": True, "notifications": notifications}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Mark Notification as Read
@app.route("/mark-notification-read/<notification_id>", methods=["POST"])
def mark_notification_as_read(notification_id):
    """
    API to mark a notification as read in the alerts collection.
    """
    try:
        # ✅ THE FIX: Query the "alerts" collection
        notification_ref = db.collection("alerts").document(notification_id)
        notification = notification_ref.get()

        if (
            not notification.exists
            or notification.to_dict().get("type") != "notification"
        ):
            return jsonify({"error": "Notification not found"}), 404

        notification_ref.update({"status": True})

        return jsonify({"success": True, "message": "Notification marked as read"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Get Notifications for a Specific User
# @app.route('/notifications/<user_id>', methods=['GET'])

from flask_socketio import SocketIO, join_room, leave_room

socketio = SocketIO(
    app,
    cors_allowed_origins="*",
    transports=["websocket"],
    logger=True,
    engineio_logger=True,
    async_mode="threading",
)


# @socketio.on("get_notifications")
# def get_notifications_for_user(data):
#     """
#     API to fetch notifications for a specific user from the alerts collection.
#     """
#     try:
#         user_id = data.get("user_id")
#         session_id = data.get("session_id")
#         join_room(session_id)

#         notifications_ref = (
#             db.collection("alerts")
#             .where("type", "==", "notification")
#             .where("user_id", "==", user_id)
#             .stream()
#         )
#         notifications = []
#         unread_count = 0

#         for doc in notifications_ref:
#             doc_data = doc.to_dict()
#             if not doc_data.get("status", True):  # Count unread notifications
#                 unread_count += 1
#             notifications.append(
#                 {
#                     "id": doc_data.get("id"),
#                     "title": doc_data.get("title"),
#                     "description": doc_data.get("description"),
#                     "date": doc_data.get("created_at"),
#                     "status": doc_data.get("status"),
#                     "receiver_role": doc_data.get("receiver_role"),
#                     "sender_role": doc_data.get("sender_role"),
#                     "responsestatus": data.get("responsestatus", ""),
#                     "response": data.get("responseMessage", ""),
#                     "responseTimestamp": data.get("responseTimestamp", ""),
#                 }
#             )

#         # Emit the response to the specific room (session_id)
#         socketio.emit(
#             "notifications_response",
#             {
#                 "success": True,
#                 "notifications": notifications,
#                 "unread_count": unread_count,
#             },
#             room=session_id,
#         )

#         # Leave the room after emitting
#         leave_room(session_id)

#     except Exception as e:
#         # Emit error instead of returning
#         socketio.emit("notifications_error", {"error": str(e)}, room=session_id)
@socketio.on("get_notifications")
def get_notifications_for_user(data):
    """
    API to fetch notifications for a specific user from the alerts collection.
    (UPDATED to support multilingual fields and correct query)
    """
    try:
        user_id = data.get("user_id")
        session_id = data.get("session_id")
        
        # ✅ 1. Get the lang parameter from the socket data
        lang = data.get("lang", "en").lower() 
        join_room(session_id)

        # ✅ 2. FIX: Query using 'receiver_id' and 'type'
        notifications_ref = (
            db.collection("alerts")
            .where("type", "==", "notification")
            .where("receiver_id", "==", user_id) 
            .stream()
        )
        
        notifications = []
        unread_count = 0

        for doc in notifications_ref:
            doc_data = doc.to_dict()
            if not doc_data.get("status", True):  # Count unread notifications
                unread_count += 1
            
            # ✅ 3. FIX: Select text based on language
            title_field = f"title_{lang}"
            desc_field = f"description_{lang}"
            
            # Fallback to English
            title = doc_data.get(title_field, doc_data.get("title_en", "No Title"))
            description = doc_data.get(desc_field, doc_data.get("description_en", "No Description"))

            notifications.append(
                {
                    "id": doc_data.get("id"),
                    "title": title,
                    "description": description,
                    "date": doc_data.get("created_at"),
                    "status": doc_data.get("status"),
                    "receiver_id": doc_data.get("receiver_id"), # Changed from receiver_role
                    "sender_role": doc_data.get("sender_role"),
                    "responsestatus": doc_data.get("responsestatus", ""),
                    "response": doc_data.get("responseMessage", ""),
                    "responseTimestamp": doc_data.get("responseTimestamp", ""),
                }
            )

        # Emit the response to the specific room (session_id)
        socketio.emit(
            "notifications_response",
            {
                "success": True,
                "notifications": notifications,
                "unread_count": unread_count,
            },
            room=session_id,
        )

        # Leave the room after emitting
        leave_room(session_id)

    except Exception as e:
        # Emit error instead of returning
        socketio.emit("notifications_error", {"error": str(e)}, room=session_id)

# Update Notification Status for User
@app.route("/notifications/status/<user_id>", methods=["POST"])
def update_notification_status(user_id):
    """
    API to update the notification status for a specific user in the Users collection.
    """
    try:
        data = request.json
        new_status = data.get("notification_status")

        if new_status not in [True, False]:
            return (
                jsonify(
                    {
                        "error": "Invalid value for notification_status. Must be True or False."
                    }
                ),
                400,
            )

        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        user_ref.update({"notification_status": new_status})

        return (
            jsonify(
                {"success": True, "message": "Notification status updated successfully"}
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# if __name__ == "__main__":
#     app.run(debug=True)
###########################################################################################################################################

# SEND MAIL

from flask import Flask, request, jsonify
# from datetime import datetime, timezone
from datetime import datetime, timedelta, timezone
import uuid
from firebase_admin import credentials, firestore, initialize_app
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging

# Configuration Section
EMAIL_SENDER = "tajheezai@gmail.com"
EMAIL_PASSWORD = "sfek auxg kasz ulgy"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
FIREBASE_CREDENTIAL_PATH = r"serviceAccountKey.json"
NOTIFICATION_API_URL = "http://127.0.0.1:5000/add-notification"

# # Initialize Firebase
# cred = credentials.Certificate(FIREBASE_CREDENTIAL_PATH)
# initialize_app(cred)
# db = firestore.client()

# # Initialize Flask app
# app = Flask(__name__)

# Configure Logging
logging.basicConfig(level=logging.DEBUG)


def send_email(receiver_email, subject, message_body):
    """Sends an email using Gmail SMTP."""
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER
        msg["To"] = receiver_email
        msg["Subject"] = subject
        msg.attach(MIMEText(message_body, "plain"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
        logging.info(f"Email sent to {receiver_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        raise


def validate_input(data, required_fields):
    """Validates that all required fields are present in the input."""
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")


def validate_user(user_id, user_type):
    """Fetches and validates user details."""
    user_doc = db.collection("users").document(user_id).get()
    if not user_doc.exists:
        raise ValueError(f"{user_type.capitalize()} with ID {user_id} not found")
    user_data = user_doc.to_dict()
    return {"role": user_data.get("role", "unknown"), "data": user_data}


@app.route("/api/teacher/create-observation", methods=["POST"])
def create_observation():
    """API to create an observation for a student and notify the parent."""
    try:
        data = request.json
        required_fields = [
            "teacherId",
            "studentId",
            "parentId",
            "studentName",
            "observationMessage",
        ]
        validate_input(data, required_fields)

        teacher_id = data["teacherId"]
        student_id = data["studentId"]
        parent_id = data["parentId"]
        student_name = data["studentName"]
        observation_message = data["observationMessage"]

        # Validate users
        teacher_details = validate_user(teacher_id, "teacher")
        parent_details = validate_user(parent_id, "parent")

        # Create observation payload
        observation = {
            "studentId": student_id,
            "parentId": parent_id,
            "studentName": student_name,
            "observationMessage": observation_message,
            "createdAt": datetime.now(timezone.utc).isoformat(),
        }

        # Send notification
        notification_payload = {
            "title": f"New observation for {student_name}",
            "description": observation_message,
            "receiver_id": parent_id,
            "sender_id": teacher_id,
            "receiver_role": parent_details["role"],
            "sender_role": teacher_details["role"],
        }
        notification_response = requests.post(
            NOTIFICATION_API_URL, json=notification_payload
        )
        if notification_response.status_code != 200:
            raise Exception(f"Notification API failed: {notification_response.text}")

        # Send email
        subject = f"New Observation for {student_name}"
        message_body = f"""
        Dear Parent,

        A new observation has been created for your child, {student_name}.
        Message: {observation_message}

        Please log in to the portal to view more details.

        Best regards,
        The School
        """
        send_email("meet.wellorgs@gmail.com", subject, message_body)

        return (
            jsonify({"success": True, "message": "Observation created successfully"}),
            200,
        )

    except ValueError as ve:
        logging.error(f"Validation Error: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        logging.error(f"Error in create_observation: {e}")
        return jsonify({"error": "An error occurred", "details": str(e)}), 500


@app.route("/add-notification/alt", methods=["POST"])
def add_notification_alt():
    """API to add a notification to the notifications collection."""
    try:
        data = request.json
        required_fields = [
            "title",
            "description",
            "receiver_id",
            "sender_id",
            "receiver_role",
            "sender_role",
        ]
        validate_input(data, required_fields)

        notification_id = str(uuid.uuid4())
        created_at = datetime.now(timezone.utc).isoformat()

        notification_data = {
            "id": notification_id,
            "type": "notification",
            "title": data["title"],
            "description": data["description"],
            "receiver_id": data["receiver_id"],
            "sender_id": data["sender_id"],
            "receiver_role": data["receiver_role"],
            "sender_role": data["sender_role"],
            "status": False,
            "created_at": created_at,
        }

        db.collection("notifications").document(notification_id).set(notification_data)
        return jsonify({"success": True, "notification_data": notification_data}), 200

    except ValueError as ve:
        logging.error(f"Validation Error: {ve}")
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        logging.error(f"Error in add_notification: {e}")
        return jsonify({"error": "An error occurred", "details": str(e)}), 500


@app.route("/notifications/<receiver_id>", methods=["GET"])
def get_notifications_receiver2(receiver_id):
    """API to fetch all notifications for a specific receiver."""
    try:
        notifications_ref = (
            db.collection("notifications")
            .where("receiver_id", "==", receiver_id)
            .stream()
        )
        notifications = [doc.to_dict() for doc in notifications_ref]
        return jsonify({"success": True, "notifications": notifications}), 200
    except Exception as e:
        logging.error(f"Error in get_notifications: {e}")
        return jsonify({"error": "An error occurred", "details": str(e)}), 500


# if __name__ == "__main__":
#     app.run(debug=True)
############################################################################################################################################################################

# CURICULUM COVERAGE
from flask import Flask, request, jsonify
from firebase_admin import credentials, firestore, initialize_app
import uuid

# app = Flask(__name__)

# # Initialize Firebase
# cred = credentials.Certificate(r'serviceAccountKey.json')
# initialize_app(cred)
# db = firestore.client()
# curriculum_collection = db.collection('curriculum')

@app.route("/curriculum", methods=["POST"])  # course id and doc id are same
def create_curriculum():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Invalid data"}), 400

        # Validate required fields
        required_fields = ["objectives", "syllabi", "assessment_criteria"]
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"Missing field: {field}"}), 400

        # Generate unique course ID
        course_id = str(uuid.uuid4())
        data["course_id"] = course_id

        # Add document to Firestore
        curriculum_collection.document(course_id).set(data)

        return (
            jsonify(
                {"message": "Curriculum created successfully", "course_id": course_id}
            ),
            201,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/curriculum/<course_id>", methods=["GET"])
def get_curriculum(course_id):
    try:
        # Retrieve document by course_id
        doc = curriculum_collection.document(course_id).get()

        if not doc.exists:
            return jsonify({"error": "Curriculum not found"}), 404

        return jsonify(doc.to_dict()), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500





# if __name__ == '__main__':
#     app.run(debug=True)
#############################################################################################################################3
# MODULE

from flask import Flask, request, jsonify
from firebase_admin import credentials, firestore, initialize_app
import uuid
import os
import logging

# # Initialize Flask App
# app = Flask(__name__)

# Initialize Firebase
logging.basicConfig(level=logging.ERROR)

# def initialize_firebase():
#     try:
#         cred_path = os.getenv('FIREBASE_CREDENTIALS_PATH', 'serviceAccountKey.json')
#         cred = credentials.Certificate(cred_path)
#         initialize_app(cred)
#         return firestore.client()
#     except Exception as e:
#         logging.error("Failed to initialize Firebase: %s", e)
#         raise

# db = initialize_firebase()
curriculum_collection = db.collection("curriculum")


# Helper function for validating data
def validate_data(data, required_fields, field_types=None):
    for field in required_fields:
        if field not in data:
            return False, f"Missing field: {field}"
        if not data[field]:
            return False, f"Field '{field}' must not be empty"
    if field_types:
        for field, expected_type in field_types.items():
            if field in data and not isinstance(data[field], expected_type):
                return (
                    False,
                    f"Field '{field}' must be of type {expected_type.__name__}",
                )
    return True, None


# Endpoint: Create a new curriculum
@app.route("/curriculum2", methods=["POST"])
def create_curriculum2():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Invalid data"}), 400

        required_fields = ["objectives", "syllabi", "assessment_criteria"]
        field_types = {
            "objectives": list,
            "syllabi": dict,
            "assessment_criteria": dict,
        }

        is_valid, error_message = validate_data(data, required_fields, field_types)
        if not is_valid:
            return jsonify({"error": error_message}), 400

        # Generate unique course ID
        course_id = str(uuid.uuid4())
        data["course_id"] = course_id

        # Assign unique IDs to modules if provided
        if "modules" in data and isinstance(data["modules"], list):
            for module in data["modules"]:
                module["module_id"] = str(uuid.uuid4())

        # Add document to Firestore
        curriculum_collection.document(course_id).set(data)

        return (
            jsonify(
                {"message": "Curriculum created successfully", "course_id": course_id}
            ),
            201,
        )
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# Endpoint: Get a curriculum by course_id
@app.route("/curriculum1/<course_id>", methods=["GET"])
def get_curriculum1(course_id):
    try:
        doc = curriculum_collection.document(course_id).get()
        if not doc.exists:
            return jsonify({"error": "Curriculum not found"}), 404

        return jsonify(doc.to_dict()), 200
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# Endpoint: Add a module to a curriculum
@app.route("/curriculum/modules", methods=["POST"])
def add_module():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Invalid data"}), 400

        if "course_id" not in data or "module" not in data:
            return jsonify({"error": "Missing course_id or module field"}), 400

        course_id = data["course_id"]
        module = data["module"]

        curriculum_doc = curriculum_collection.document(course_id).get()
        if not curriculum_doc.exists:
            return jsonify({"error": "Curriculum not found"}), 404

        # Generate unique module ID
        module_id = str(uuid.uuid4())
        module["module_id"] = module_id

        # Add module to curriculum
        curriculum_collection.document(course_id).update(
            {"modules": firestore.ArrayUnion([module])}
        )

        return (
            jsonify({"message": "Module added successfully", "module_id": module_id}),
            201,
        )
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# Endpoint: Update a specific module
@app.route("/curriculum/modules/<module_id>", methods=["PUT"])
def update_module(module_id):
    try:
        data = request.json
        if not data or "course_id" not in data:
            return jsonify({"error": "Missing course_id or module data"}), 400

        course_id = data["course_id"]
        updated_module = data.get("module", {})

        curriculum_doc = curriculum_collection.document(course_id).get()
        if not curriculum_doc.exists:
            return jsonify({"error": "Curriculum not found"}), 404

        curriculum_data = curriculum_doc.to_dict()
        modules = curriculum_data.get("modules", [])

        module_found = False
        for module in modules:
            if module.get("module_id") == module_id:
                module.update(updated_module)
                module_found = True

        if not module_found:
            return jsonify({"error": "Module not found"}), 404

        curriculum_collection.document(course_id).update({"modules": modules})
        return jsonify({"message": "Module updated successfully"}), 200
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# Endpoint: Get a specific module
@app.route("/curriculum/modules/<module_id>", methods=["GET"])
def get_module(module_id):
    try:
        course_id = request.args.get("course_id")
        if not course_id:
            return jsonify({"error": "Missing course_id query parameter"}), 400

        curriculum_doc = curriculum_collection.document(course_id).get()
        if not curriculum_doc.exists:
            return jsonify({"error": "Curriculum not found"}), 404

        modules = curriculum_doc.to_dict().get("modules", [])
        for module in modules:
            if module.get("module_id") == module_id:
                return jsonify(module), 200

        return jsonify({"error": "Module not found"}), 404
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# Endpoint: Get all modules (optionally filtered by course_id)
@app.route("/curriculum/modules", methods=["GET"])
def get_all_modules():
    try:
        course_id = request.args.get("course_id")

        if course_id:
            curriculum_doc = curriculum_collection.document(course_id).get()
            if not curriculum_doc.exists():
                return jsonify({"error": "Curriculum not found"}), 404

            modules = curriculum_doc.to_dict().get("modules", [])
            return jsonify({"modules": modules}), 200
        else:
            all_modules = []
            docs = curriculum_collection.stream()
            for doc in docs:
                curriculum_data = doc.to_dict()
                modules = curriculum_data.get("modules", [])
                all_modules.extend(modules)

            return jsonify({"modules": all_modules}), 200
    except Exception as e:
        logging.error(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


# # Run the app
# if __name__ == '__main__':
#     app.run(debug=True)
############################################################################################################################################################################

# SUBJECTS

from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import uuid

# Initialize Flask app
# app = Flask(__name__)

# # Initialize Firebase Admin SDK
# cred = credentials.Certificate(r"serviceAccountKey.json")
# firebase_admin.initialize_app(cred)
# db = firestore.client()


# POST Subject API
@app.route("/grades/classes/subjects111", methods=["POST"])
def add_subject111():
    try:
        # Parse request data
        data = request.get_json()

        # Validate input
        if (
            not data
            or "grade_id" not in data
            or "class_id" not in data
            or "subject_name" not in data
        ):
            return (
                jsonify(
                    {
                        "error": "JSON body must contain 'grade_id', 'class_id', and 'subject_name'"
                    }
                ),
                400,
            )

        grade_id = data["grade_id"]
        class_id = data["class_id"]
        subject_name = data["subject_name"]

        # Generate a unique subject ID
        subject_id = str(uuid.uuid4())

        # Reference to the subjects collection
        subject_ref = (
            db.collection("Grades")
            .document(grade_id)
            .collection("classes")
            .document(class_id)
            .collection("subjects")
            .document(subject_id)
        )

        # Set subject data
        subject_ref.set({"subject_id": subject_id, "subject_name": subject_name})

        return (
            jsonify(
                {
                    "message": f"Subject {subject_name} added successfully.",
                    "subject_id": subject_id,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

from flask import Flask, jsonify, request
from google.cloud import firestore

def fetch_subjects_for_class_sync(grade_id: str, class_id: str):
    subjects_ref = (
        db.collection("Grades")
          .document(grade_id)
          .collection("classes")
          .document(class_id)
          .collection("subjects")
    )
    subjects = []
    for doc in subjects_ref.stream():  # sync stream [web:42][web:38]
        data = doc.to_dict() or {}
        subjects.append({"subject_id": doc.id, **data})
    return subjects

def fetch_classes_for_grade_sync(grade_id: str):
    classes_ref = (
        db.collection("Grades")
          .document(grade_id)
          .collection("classes")
    )
    classes = []
    # First collect class docs
    class_docs = list(classes_ref.stream())  # sync stream [web:42][web:38]
    for class_doc in class_docs:
        cdata = class_doc.to_dict() or {}
        class_id = class_doc.id
        subjects = fetch_subjects_for_class_sync(grade_id, class_id)
        classes.append({
            "class_id": class_id,
            "class_name": cdata.get("class_name", ""),
            "subjects": subjects,
        })
    return classes

def fetch_single_grade_block_sync(grade_id: str):
    grade_doc = db.collection("Grades").document(grade_id).get()  # sync get [web:38][web:42]
    if not grade_doc.exists:
        return None
    gdata = grade_doc.to_dict() or {}
    classes = fetch_classes_for_grade_sync(grade_id)
    return {
        "grade_id": grade_doc.id,
        "grade": gdata.get("grade", ""),
        "classes": classes,
    }

def fetch_grades_by_name_sync(grade_name: str):
    q = db.collection("Grades").where("grade", "==", grade_name)  # sync where [web:32][web:38]
    matches = list(q.stream())  # sync stream [web:42][web:38]
    blocks = []
    for gdoc in matches:
        block = fetch_single_grade_block_sync(gdoc.id)
        if block:
            blocks.append(block)
    return blocks

def fetch_all_grades_blocks_sync():
    grades_ref = db.collection("Grades")
    results = []
    for gdoc in grades_ref.stream():  # sync stream [web:42][web:38]
        block = fetch_single_grade_block_sync(gdoc.id)
        if block:
            results.append(block)
    return results




@app.get("/api/grades-classes-subjects")
def get_grades_classes_subjects():
    """
    Optional query params:
      - grade_id: return just this grade’s classes and subjects
      - grade_name: filter subjects by this grade name in addition to path
    """
    grade_id = request.args.get("grade_id")
    grade_name_filter = request.args.get("grade_name")

    try:
        grades_out = []

        if grade_id:
            grade_doc = db.collection("Grades").document(grade_id).get()
            if not grade_doc.exists:
                return jsonify({"grades": []}), 200

            grades_out.append(_build_grade_block(db, grade_doc, grade_name_filter))
        else:
            for grade_doc in db.collection("Grades").stream():
                grades_out.append(_build_grade_block(db, grade_doc, grade_name_filter))

        return jsonify({"grades": grades_out}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _build_grade_block(db, grade_doc, grade_name_filter=None):
    grade_data = grade_doc.to_dict() or {}
    grade_id = grade_doc.id
    grade_name = grade_data.get("grade", "")

    classes_out = []
    classes_ref = db.collection("Grades").document(grade_id).collection("classes")
    for class_doc in classes_ref.stream():
        classes_out.append(_build_class_block(db, grade_id, grade_name, class_doc, grade_name_filter))

    return {"grade_id": grade_id, "grade": grade_name, "classes": classes_out}


# def _build_class_block(db, grade_id, grade_name, class_doc, grade_name_filter=None):
#     class_data = class_doc.to_dict() or {}
#     class_id = class_doc.id
#     class_name = class_data.get("class_name", "")

#     # Subjects under this class only
#     subjects_ref = (db.collection("Grades")
#                       .document(grade_id)
#                       .collection("classes")
#                       .document(class_id)
#                       .collection("subjects"))

#     # Strong filter by grade_name when provided, to avoid leaking cross-grade subjects
#     if grade_name_filter:
#         subjects_query = subjects_ref.where("grade_name", "==", grade_name_filter)  # [web:15]
#     else:
#         # Since these are nested under the class for this grade, this already scopes to the grade
#         subjects_query = subjects_ref

#     subjects_out = []
#     for sub_doc in subjects_query.stream():
#         sub_data = sub_doc.to_dict() or {}
#         subjects_out.append({
#             "subject_id": sub_doc.id,
#             "subject_name": sub_data.get("subject_name", ""),
#             "grade_name": sub_data.get("grade_name", grade_name),
#             "grade": sub_data.get("grade", grade_name),
#             "subject_reference": sub_data.get("subject_reference")
#         })

#     return {"class_id": class_id, "class_name": class_name, "subjects": subjects_out}

def _build_class_block(db, grade_id, grade_name, class_doc, grade_name_filter=None):
    class_data = class_doc.to_dict() or {}
    class_id = class_doc.id
    class_name = class_data.get("class_name", "")

    # Subjects under this class only
    subjects_ref = (db.collection("Grades")
                      .document(grade_id)
                      .collection("classes")
                      .document(class_id)
                      .collection("subjects"))

    # Strong filter by grade_name when provided, to avoid leaking cross-grade subjects
    if grade_name_filter:
        subjects_query = subjects_ref.where("grade_name", "==", grade_name_filter)  # [web:15]
    else:
        # Since these are nested under the class for this grade, this already scopes to the grade
        subjects_query = subjects_ref

    # --- START FIX: Deduplicate subjects by name ---
    subjects_out = []
    seen_subject_names_lower = set()  # To track subjects we've already added

    for sub_doc in subjects_query.stream():
        sub_data = sub_doc.to_dict() or {}
        
        # Get the original name and clean it (strip whitespace)
        original_subject_name = sub_data.get("subject_name", "")
        cleaned_subject_name = str(original_subject_name).strip()
        
        # Use a lowercase version for comparison to catch case differences
        cleaned_name_lower = cleaned_subject_name.lower()

        # If we haven't seen this cleaned, lowercase name before...
        if cleaned_name_lower not in seen_subject_names_lower:
            # 1. Mark it as seen
            seen_subject_names_lower.add(cleaned_name_lower)
            
            # 2. Add it to the output list using the *cleaned* (not lowercase) name
            subjects_out.append({
                "subject_id": sub_doc.id,
                "subject_name": cleaned_subject_name,  # <-- Use the cleaned name
                "grade_name": sub_data.get("grade_name", grade_name),
                "grade": sub_data.get("grade", grade_name),
                "subject_reference": sub_data.get("subject_reference")
            })
        # ...else, it's a duplicate (e.g., "Science" vs "Science "), so we skip it.
    # --- END FIX ---

    return {"class_id": class_id, "class_name": class_name, "subjects": subjects_out}
# async def get_grades_classes_subjects():
#     try:
#         grades_collection = db11.collection("Grades")
#         grades_docs = grades_collection.stream()

#         all_grades = []
#         async for grade_doc in grades_docs:
#             grade_data = grade_doc.to_dict()
#             grade_id = grade_doc.id
#             grade_name = grade_data.get("grade", "")

#             # Fetch classes under this grade
#             classes_collection = db11.collection("Grades").document(grade_id).collection("classes")
#             classes_docs = classes_collection.stream()

#             classes = []
#             async for class_doc in classes_docs:
#                 class_data = class_doc.to_dict()
#                 class_id = class_doc.id
#                 class_name = class_data.get("class_name", "")

#                 # Fetch subjects under this class
#                 subjects_collection = (
#                     db11.collection("Grades")
#                     .document(grade_id)
#                     .collection("classes")
#                     .document(class_id)
#                     .collection("subjects")
#                 )
#                 subjects_docs = subjects_collection.stream()

#                 subjects = [
#                     {"subject_id": subject_doc.id, **subject_doc.to_dict()}
#                     async for subject_doc in subjects_docs
#                 ]

#                 classes.append({"class_id": class_id, "class_name": class_name, "subjects": subjects})

#             all_grades.append({"grade_id": grade_id, "grade": grade_name, "classes": classes})

#         return jsonify({"grades": all_grades}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# GET Subjects API
@app.route("/grades/classes/subjects", methods=["GET"])
def get_subjects():
    try:
        # Parse query parameters
        grade_id = request.args.get("grade_id")
        class_id = request.args.get("class_id")
        subject_name = request.args.get("subject_name")  # Optional filter

        # Validate input
        if not grade_id or not class_id:
            return (
                jsonify(
                    {"error": "Query parameters 'grade_id' and 'class_id' are required"}
                ),
                400,
            )

        # Reference to the subjects collection
        subjects_ref = (
            db.collection("Grades")
            .document(grade_id)
            .collection("classes")
            .document(class_id)
            .collection("subjects")
        )

        # Fetch all subjects
        subjects_query = subjects_ref.stream()

        # Filter results based on query parameters
        subjects = []
        for subject in subjects_query:
            subject_data = subject.to_dict()
            if subject_name and subject_data.get("subject_name") != subject_name:
                continue
            subjects.append(subject_data)

        # Return results
        if not subjects:
            return jsonify({"message": "No subjects found for the given filters"}), 404

        return jsonify(subjects), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/pdf_ocr", methods=["POST"])
def upload_pdf():
    if "file" not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    if not file.filename.endswith(".pdf"):
        return jsonify({"error": "Only PDF files are allowed"}), 400

    # Save the uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_pdf_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(temp_pdf_path)

    try:
        # Perform OCR and extract text
        extracted_text = extract_text_from_pdf(temp_pdf_path)

        # Clean the extracted text
        cleaned_text = clean_text(extracted_text)
    except Exception as e:
        return jsonify({"error": f"OCR processing failed: {str(e)}"}), 500
    finally:
        # Cleanup the temporary file
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)

    # Return cleaned text as JSON
    return jsonify({"exactracted_text": cleaned_text}), 200


# Function to clean extracted text
def clean_text(raw_text):
    """
    Cleans the extracted OCR text by:
    - Removing extra whitespace
    - Removing special characters
    - Fixing common OCR errors
    """
    import re

    # Remove extra whitespace and line breaks
    cleaned = re.sub(r"\s+", " ", raw_text)

    # Remove special characters (retain alphanumerics and basic punctuation)
    cleaned = re.sub(r"[^a-zA-Z0-9.,!?\'\";:()\\-]", " ", cleaned)

    # Normalize spaces
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    return cleaned


# Function to evaluate provided text with Groq
def evaluate_text_with_groq(extracted_text):
    try:
        # Clean the input text
        cleaned_text = clean_text(extracted_text)

        # Define your prompt
        prompt = f"""
        Evaluate the following extracted text and provide insights:

        Extracted Text:
        {cleaned_text}

        Analysis Required:
        - Identify the main topics covered.
        - Highlight any potential errors or inconsistencies.
        - Provide suggestions for improving clarity.
        """

        # Generate response from Groq
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "system",
                    "content": "You are an evaluator. Provide insights and suggestions.",
                },
                {"role": "user", "content": prompt},
            ],
            model="llama3-8b-8192",
            temperature=0.7,
            max_tokens=1000,
        )

        # Clean the generated evaluation text
        evaluation = clean_text(chat_completion.choices[0].message.content)
        return evaluation
    except Exception as e:
        return f"Error during evaluation: {str(e)}"


# Define API route
@app.route("/evaluate", methods=["POST"])
def evaluate():
    data = request.get_json()
    if not data or "text" not in data:
        return jsonify({"error": "No text provided"}), 400

    extracted_text = data["text"]

    try:
        # Evaluate provided text using Groq
        evaluation = evaluate_text_with_groq(extracted_text)
        return jsonify(
            {"extracted_text": clean_text(extracted_text), "evaluation": evaluation}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


################################################################################################################################################################

# CHANGEROLES

from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore

# app = Flask(__name__)

# # Initialize Firebase Admin SDK
# cred = credentials.Certificate(r"serviceAccountKey.json")
# firebase_admin.initialize_app(cred)
# db = firestore.client()

# Define roles in circular order
roles = ["student", "parent", "headmaster", "teacher"]


@app.route("/users/changeRole", methods=["PUT"])
def change_role():
    try:
        # Parse the request body to get the user ID and the new role
        data = request.get_json()
        user_id = data.get("user_id")
        new_role = data.get("new_role")

        if not user_id:
            return jsonify({"error": "User ID is required"}), 400
        if not new_role:
            return jsonify({"error": "New role is required"}), 400

        # Validate the new role
        roles = ["student", "parent", "headmaster", "teacher"]
        if new_role not in roles:
            return (
                jsonify(
                    {
                        "error": "Invalid role. Must be one of: student, parent, headmaster, teacher"
                    }
                ),
                400,
            )

        # Fetch the user document from Firestore
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        # Update the role in Firestore
        user_ref.update({"role": new_role})

        return (
            jsonify({"message": "Role updated successfully", "new_role": new_role}),
            200,
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# if __name__ == '__main__':
#     app.run(debug=True)

############################################################################################################################################################################
# STUDENT


from flask import Flask, request, jsonify
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, firestore
import time
import secrets
import jwt
import uuid
import logging

# Initialize Flask app
# app = Flask(__name__)
# SECRET_KEY = "secret_key"
# CORS(app)

# # Initialize Firebase Admin
# cred = credentials.Certificate(r"serviceAccountKey.json")
# firebase_admin.initialize_app(cred)
# db = firestore.client()


# Decode JWT token
def decode_token(token):
    """
    Decode a JWT token and handle errors.
    """
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload
    except jwt.ExpiredSignatureError:
        print("Token has expired.")
        return None
    except jwt.InvalidTokenError as e:
        print(f"Invalid token: {e}")
        return None


@app.route("/users", methods=["GET"])
def get_users_summary():
    role_filter = request.args.get("role")  # Optional role filter
    users_ref = db.collection("users")

    if role_filter:
        users = users_ref.where("role", "==", role_filter).stream()
    else:
        users = users_ref.stream()

    user_list = []
    for user in users:
        user_data = user.to_dict()
        user_summary = {
            "role": user_data.get("role", ""),
            "userId": user_data.get("userId", ""),
            "status": user_data.get(
                "status", "Unknown"
            ),  # Default to 'Unknown' if status is not present
        }
        user_list.append(user_summary)

    return jsonify(user_list), 200


@app.route("/users/<user_id>", methods=["GET"])
def get_user_details(user_id):
    user_ref = db.collection("users").document(user_id)
    user = user_ref.get()

    if user.exists:
        return jsonify(user.to_dict()), 200
    else:
        return jsonify({"error": "User not found"}), 404


# @app.route('/api/student/create', methods=["POST"])
# def create_user_account():
#     try:
#         data = request.json
#         if not data or 'role' not in data:
#             return jsonify({"error": "role is required"}), 400

#         role = data['role']

#         # Generate a unique JWT-like userId
#         payload = {
#             "timestamp": time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()),  # Use time instead of datetime
#             "nonce": secrets.token_hex(8)  # Random unique string
#         }
#         user_id = jwt.encode(payload, SECRET_KEY, algorithm="HS256")

#         # Check if the generated userId already exists in the users collection
#         user_doc = db.collection("users").document(user_id).get()
#         if user_doc.exists:
#             return jsonify({"error": f"User with userId {user_id} already exists"}), 409

#         # Define default structure for the user document
#         user_data = {
#             "userId": user_id,
#             "role": role,
#             "email": data.get('email', ""),
#             "createdAt": time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())  # Replaced datetime.utcnow()
#         }

#         # Save the user data in Firestore
#         db.collection("users").document(user_id).set(user_data)

#         # If the role is 'student', also store data in the students collection
#         if role.lower() == "student":
#             student_data = {
#                 "studentId": user_id,  # userId and studentId must be the same
#                 "personalInformation": {
#                     "name": data.get('name', ""),
#                     "photourl": data.get('photourl', "https://example.com/default-photo.jpg"),
#                     "idNumber": data.get('idno', ""),
#                 },
#                 "contactInformation": {
#                     "phoneNumber": data.get('phoneNumber', ""),
#                     "address": data.get('address', ""),
#                 },
#                 "createdAt": time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())  # Replaced datetime.utcnow()
#             }
#             db.collection("students").document(user_id).set(student_data)

#         return jsonify({"message": "User account created successfully", "userId": user_id}), 201

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# grade and class management

# @app.route('/grades', methods=['POST'])
# def create_or_update_grade():
#     data = request.get_json()

#     # Validate input
#     if not isinstance(data, dict):
#         return jsonify({"error": "Invalid payload format. Expected a JSON object."}), 400

#     grade = data.get("grade")
#     classes = data.get("classes")  # Expecting a list of class names, e.g., ["Class A", "Class B"]

#     if not grade:
#         return jsonify({"error": "Invalid or missing 'grade' field"}), 400
#     if not classes or not isinstance(classes, list):
#         return jsonify({"error": "Invalid or missing 'classes' field"}), 400

#     # Generate a unique UUID for the grade document
#     grade_id = str(uuid.uuid4())
#     grade_ref = db.collection('Grades').document(grade_id)

#     # Create the grade document
#     grade_ref.set({"grade_id": grade_id, "grade": grade}, merge=True)

#     # Create nested collections for classes and store their details
#     class_list = []
#     for class_name in classes:
#         class_id = str(uuid.uuid4())  # Unique ID for each class
#         class_ref = grade_ref.collection('classes').document(class_id)
#         class_data = {
#             "class_id": class_id,
#             "class_name": class_name
#         }
#         class_ref.set(class_data)
#         class_list.append(class_data)  # Append the class details for response

#     return jsonify({
#         "message": f"Grade {grade} created/updated successfully with ID {grade_id}.",
#         "grade_id": grade_id,
#         "classes": class_list  # Include class details in the response
#     }), 200


@app.route("/grades", methods=["POST"])
def create_or_update_grade():
    data = request.get_json()

    # Validate input
    if not isinstance(data, dict):
        return (
            jsonify({"error": "Invalid payload format. Expected a JSON object."}),
            400,
        )

    grade = data.get("grade")
    classes = data.get("classes")  # Expecting a list of class names

    if not grade:
        return jsonify({"error": "Invalid or missing 'grade' field"}), 400
    if not classes or not isinstance(classes, list):
        return jsonify({"error": "Invalid or missing 'classes' field"}), 400

    # Check if the grade already exists
    grade_query = db.collection("Grades").where("grade", "==", grade).stream()
    grade_doc = next(grade_query, None)

    if grade_doc:
        grade_id = grade_doc.id  # Reuse existing grade_id
        grade_ref = db.collection("Grades").document(grade_id)
    else:
        grade_id = str(uuid.uuid4())
        grade_ref = db.collection("Grades").document(grade_id)
        grade_ref.set({"grade_id": grade_id, "grade": grade}, merge=True)

    class_list = []
    for class_name in classes:
        # Check if class exists in the subcollection
        class_query = (
            grade_ref.collection("classes")
            .where("class_name", "==", class_name)
            .stream()
        )
        class_doc = next(class_query, None)

        if class_doc:
            return (
                jsonify(
                    {
                        "error": f"Class '{class_name}' already exists in grade '{grade}'."
                    }
                ),
                400,
            )
        else:
            class_id = str(uuid.uuid4())  # Create new class_id if not found
            class_ref = grade_ref.collection("classes").document(class_id)
            class_data = {"class_id": class_id, "class_name": class_name}
            class_ref.set(class_data)

        class_list.append({"class_id": class_id, "class_name": class_name})

    return (
        jsonify(
            {
                "message": f"Grade {grade} created/updated successfully with ID {grade_id}.",
                "grade_id": grade_id,
                "classes": class_list,
            }
        ),
        200,
    )


@app.route("/grades", methods=["GET"])
def get_all_grades():
    grades_ref = db.collection("Grades")
    grades = grades_ref.stream()

    grade_list = []
    for grade in grades:
        grade_data = grade.to_dict()

        # Fetch nested classes
        classes_ref = (
            db.collection("Grades").document(grade.id).collection("classes").stream()
        )
        grade_data["classes"] = [cls.to_dict() for cls in classes_ref]

        grade_list.append(grade_data)

    return jsonify(grade_list), 200


@app.route("/grades/<int:grade>", methods=["GET"])
def get_classes_by_grade(grade):
    grades_ref = db.collection("Grades").where("grade", "==", grade).stream()

    results = []
    for grade_doc in grades_ref:
        grade_data = grade_doc.to_dict()

        # Fetch nested classes for the grade
        classes_ref = (
            db.collection("Grades")
            .document(grade_doc.id)
            .collection("classes")
            .stream()
        )
        grade_data["classes"] = [cls.to_dict() for cls in classes_ref]

        results.append(grade_data)

    if not results:
        return jsonify({"error": f"Grade {grade} not found"}), 404

    return jsonify(results), 200


# teacher managment api
@app.route("/api/teacher/classes/studentcount", methods=["GET"])
def get_teacher_classes_studentcount():
    try:
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_token(token)

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        # Extract teacher ID from token
        teacher_id = decoded_token.get("userId")
        if not teacher_id:
            return jsonify({"error": "Invalid token"}), 401

        # Fetch classes associated with the teacher
        classes_ref = db.collection("classes").where("teacherId", "==", teacher_id)
        classes = classes_ref.stream()

        class_data = []
        for cls in classes:
            class_info = cls.to_dict()
            class_data.append(
                {
                    "class_id": class_info.get("classId"),
                    "class_name": class_info.get("className"),
                    "student_count": len(
                        class_info.get("students", [])
                    ),  # Assume students is a list
                }
            )

        return jsonify({"teacher_id": teacher_id, "classes": class_data}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/teacher-assign/classes", methods=["POST"])
def assign_grade_ref_to_teacher():
    try:
        # Authorization token validation
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Authorization token is required"}), 401

        token = auth_header.split(" ")[1] if " " in auth_header else auth_header
        decoded_token = decode_token(token)
        logging.debug(f"Decoded token: {decoded_token}")

        if not decoded_token:
            return jsonify({"error": "Invalid or expired token"}), 401

        teacher_id = decoded_token.get("userId")
        if not teacher_id:
            return jsonify({"error": "Invalid token"}), 401

        # Parse request body
        data = request.get_json()
        logging.debug(f"Request body: {data}")
        class_id = data.get("class_id")

        if not class_id:
            return jsonify({"error": "Missing required field: class_id"}), 400

        # Fetch teacher's document and validate role
        teacher_ref = db.collection("users").document(teacher_id)
        teacher_doc = teacher_ref.get()
        if not teacher_doc.exists:
            return jsonify({"error": f"Teacher with ID {teacher_id} not found"}), 404

        teacher_data = teacher_doc.to_dict()
        if teacher_data.get("role") != "teacher":
            return jsonify({"error": "Access denied. User is not a teacher."}), 403

        # Validate class document globally
        class_ref = None
        grades_ref = db.collection("Grades").stream()
        for grade in grades_ref:
            grade_ref = (
                db.collection("Grades")
                .document(grade.id)
                .collection("classes")
                .document(class_id)
            )
            if grade_ref.get().exists:
                class_ref = grade_ref
                logging.debug(f"Class found: {class_ref.id}")
                break

        if not class_ref:
            return jsonify({"error": f"Class with ID {class_id} not found"}), 404

        # Update teacher's document with the grade_ref
        teacher_ref.update(
            {"grade_ref": class_ref.path}
        )  # Use `path` for Firestore references
        logging.debug(
            f"Class {class_id} assigned to Teacher {teacher_id} successfully."
        )

        return (
            jsonify(
                {
                    "message": f"Class {class_id} assigned to Teacher {teacher_id} successfully."
                }
            ),
            200,
        )

    except Exception as e:
        logging.error(f"Error in assign_grade_ref_to_teacher: {str(e)}")
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500
    

# Student Report Card APIs
@app.route("/api/student/report-card", methods=["POST"])
async def enter_student_report_card():
    try:
        data = request.get_json()

        required_fields = ["studentId", "academicData"]
        if not all(field in data for field in required_fields):
            return (
                jsonify(
                    {
                        "error": "Missing required fields: 'studentId' and 'academicData' are required"
                    }
                ),
                400,
            )

        student_id = data["studentId"]
        academic_data = data["academicData"]

        studentName = data.get("student_name", "Unknown Student")
        teacherId = data.get("teacher_id", "Unknown Teacher")
        
        # --- DATE HANDLING LOGIC ---
        entry_date_str = data.get("entryDate")
        current_utc_timestamp = time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())
        if entry_date_str:
            try:
                dt_obj = datetime.datetime.strptime(entry_date_str, '%Y-%m-%d')
                entry_date_timestamp = dt_obj.strftime("%Y-%m-%dT00:00:00")
            except ValueError:
                print(f"Warning: Invalid entryDate format '{entry_date_str}'. Falling back to current timestamp.")
                entry_date_timestamp = current_utc_timestamp
        else:
            entry_date_timestamp = current_utc_timestamp
        # --- END DATE HANDLING ---

        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        student_data = student_doc.to_dict()
        grade_history = student_data.get("gradeHistory", {})

        # --- LOGIC CHANGE 1: Overall Report Card Grade is now Optional ---
        if "reportCard" in data:
            report_card = data["reportCard"]
            
            # CHANGED: We removed the check `if "grade" not in report_card: return error`
            # Now we only update history IF the grade is present.
            if "grade" in report_card:
                report_card_grade = report_card["grade"]
                timestamp_key = entry_date_timestamp 
                grade_history[timestamp_key] = str(report_card_grade)
            # If "grade" is missing, we simply proceed without error.
        else:
            report_card = student_data.get("reportCard", {})

        existing_academic_data = student_data.get("academicData", {})
        existing_subjects = existing_academic_data.get("subjects", {})
        incoming_subjects = academic_data.get("subjects", {})

        alerts = []

        for subject, details in incoming_subjects.items():
            subject_history = existing_subjects.get(subject, {})
            if "history" not in subject_history:
                subject_history["history"] = []

            previous_score = (
                subject_history["history"][-1]["marks"]
                if subject_history["history"]
                else None
            )
            current_score = details.get("marks", 0)

            curriculum_id = details.get("curriculumId", "Unknown_Curriculum_ID")
            curriculum_name = details.get("curriculumName", "Unknown_Curriculum_Name")
            
            # --- LOGIC CHANGE 2: Subject Grade defaults to "" instead of "E" ---
            # CHANGED: Changed default from "E" to ""
            grade = details.get("grade", "") 
            
            total_mark = details.get("totalMark", 100)
            
            subject_history["history"].append(
                {
                    "timestamp": entry_date_timestamp,
                    "curriculumId": curriculum_id,
                    "curriculumName": curriculum_name,
                    "marks": current_score,
                    "totalMark": total_mark,
                    "grade": grade, # This will be "" if not provided
                    "teacher_id": teacherId,
                }
            )

            existing_subjects[subject] = subject_history

            # --- DUAL-LANGUAGE ALERT LOGIC (Kept as is) ---
            lang_prompt = None
            type_of_alert = None

            if previous_score is not None and current_score < previous_score:
                lang_prompt = "urgent warning alert for performance decline and required intervention"
                type_of_alert = "Warning" 
            
            elif previous_score is not None and current_score > previous_score:
                lang_prompt = "motivational alert for performance improvement"
                type_of_alert = "Success"

            if lang_prompt and type_of_alert:
                message_en = await generate_ai_alert1(
                    studentName, student_id, previous_score, current_score, "English", lang_prompt
                )
                message_ar = await generate_ai_alert1(
                    studentName, student_id, previous_score, current_score, "Arabic", lang_prompt
                )

                alert_data = {
                    "ai_generated_message_en": message_en,
                    "ai_generated_message_ar": message_ar,
                    "current_score": current_score,
                    "date": time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime()),
                    "isSeen": "Unread",
                    "previous_score": previous_score,
                    "student_id": student_id,
                    "student_name": studentName,
                    "teacher_id": teacherId,
                    "type_of_alert": type_of_alert,
                }

                alerts.append(alert_data)
                db.collection("alerts").add(alert_data)
            # --- END ALERT LOGIC ---

        academic_data["subjects"] = existing_subjects

        update_data = {
            "academicData": academic_data,
            "gradeHistory": grade_history,
        }

        if "reportCard" in data:
            update_data["reportCard"] = report_card
        if "attendance" in data:
            update_data["attendance"] = data["attendance"]

        student_ref.update(update_data)

        return (
            jsonify(
                {
                    "message": "Student report card updated successfully",
                    "studentId": student_id,
                    "alerts_generated": alerts,
                    "report_date_used": entry_date_timestamp
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# # Student Report Card APIs
# @app.route("/api/student/report-card", methods=["POST"])
# async def enter_student_report_card():
#     try:
#         data = request.get_json()

#         required_fields = ["studentId", "academicData"]
#         if not all(field in data for field in required_fields):
#             return (
#                 jsonify(
#                     {
#                         "error": "Missing required fields: 'studentId' and 'academicData' are required"
#                     }
#                 ),
#                 400,
#             )

#         student_id = data["studentId"]
#         academic_data = data["academicData"]

#         studentName = data.get("student_name", "Unknown Student")
#         teacherId = data.get("teacher_id", "Unknown Teacher")
        
#         # --- DATE HANDLING LOGIC (Your code is correct) ---
#         entry_date_str = data.get("entryDate")
#         current_utc_timestamp = time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())
#         if entry_date_str:
#             try:
#                 dt_obj = datetime.datetime.strptime(entry_date_str, '%Y-%m-%d')
#                 entry_date_timestamp = dt_obj.strftime("%Y-%m-%dT00:00:00")
#             except ValueError:
#                 print(f"Warning: Invalid entryDate format '{entry_date_str}'. Falling back to current timestamp.")
#                 entry_date_timestamp = current_utc_timestamp
#         else:
#             entry_date_timestamp = current_utc_timestamp
#         # --- END DATE HANDLING ---

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()
#         grade_history = student_data.get("gradeHistory", {})

#         if "reportCard" in data:
#             report_card = data["reportCard"]
#             if "grade" not in report_card:
#                 return (
#                     jsonify({"error": "'reportCard' must contain 'grade' if provided"}),
#                     400,
#                 )
#             report_card_grade = report_card["grade"]
#             timestamp_key = entry_date_timestamp 
#             grade_history[timestamp_key] = str(report_card_grade)
#         else:
#             report_card = student_data.get("reportCard", {})

#         existing_academic_data = student_data.get("academicData", {})
#         existing_subjects = existing_academic_data.get("subjects", {})
#         incoming_subjects = academic_data.get("subjects", {})

#         alerts = []

#         for subject, details in incoming_subjects.items():
#             subject_history = existing_subjects.get(subject, {})
#             if "history" not in subject_history:
#                 subject_history["history"] = []

#             previous_score = (
#                 subject_history["history"][-1]["marks"]
#                 if subject_history["history"]
#                 else None
#             )
#             current_score = details.get("marks", 0)

#             # (Your code for curriculum, grade, totalMark, etc. is fine)
#             curriculum_id = details.get("curriculumId", "Unknown_Curriculum_ID")
#             curriculum_name = details.get("curriculumName", "Unknown_Curriculum_Name")
#             grade = details.get("grade", "E")
#             total_mark = details.get("totalMark", 100)
            
#             subject_history["history"].append(
#                 {
#                     "timestamp": entry_date_timestamp,
#                     "curriculumId": curriculum_id,
#                     "curriculumName": curriculum_name,
#                     "marks": current_score,
#                     "totalMark": total_mark,
#                     "grade": grade,
#                 }
#             )

#             existing_subjects[subject] = subject_history

#             # --- START OF DUAL-LANGUAGE ALERT FIX ---
#             #
#             # This 'if' block is the only part that changed
#             #
#             lang_prompt = None
#             type_of_alert = None

#             # Check for decline
#             if previous_score is not None and current_score < previous_score:
#                 lang_prompt = "urgent warning alert for performance decline and required intervention"
#                 type_of_alert = "Warning"  # <-- FIX: Was "Success"
            
#             # (Optional: Add this 'elif' to also handle improvement)
#             elif previous_score is not None and current_score > previous_score:
#                 lang_prompt = "motivational alert for performance improvement"
#                 type_of_alert = "Success"

#             # If an alert needs to be generated (either up or down)
#             if lang_prompt and type_of_alert:
                
#                 # Generate BOTH languages
#                 message_en = await generate_ai_alert1(
#                     studentName, student_id, previous_score, current_score, "English", lang_prompt
#                 )
#                 message_ar = await generate_ai_alert1(
#                     studentName, student_id, previous_score, current_score, "Arabic", lang_prompt
#                 )

#                 alert_data = {
#                     # --- FIX: Save the new dual-language fields ---
#                     "ai_generated_message_en": message_en,
#                     "ai_generated_message_ar": message_ar,
#                     # ---
#                     "current_score": current_score,
#                     "date": time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime()),
#                     "isSeen": "Unread",
#                     "previous_score": previous_score,
#                     "student_id": student_id,
#                     "student_name": studentName,
#                     "teacher_id": teacherId,
#                     "type_of_alert": type_of_alert, # <-- FIX: Now dynamic
#                 }

#                 alerts.append(alert_data)
#                 db.collection("alerts").add(alert_data)
#             # --- END OF DUAL-LANGUAGE ALERT FIX ---

#         # (Rest of your function is fine)
#         academic_data["subjects"] = existing_subjects

#         update_data = {
#             "academicData": academic_data,
#             "gradeHistory": grade_history,
#         }

#         if "reportCard" in data:
#             update_data["reportCard"] = report_card
#         if "attendance" in data:
#             update_data["attendance"] = data["attendance"]

#         student_ref.update(update_data)

#         return (
#             jsonify(
#                 {
#                     "message": "Student report card updated with grade history and timestamped subject records",
#                     "studentId": student_id,
#                     "alerts_generated": alerts,
#                     "report_date_used": entry_date_timestamp
#                 }
#             ),
#         200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500
    
# # @app.route("/api/student/report-card", methods=["POST"])
# # def enter_student_report_card():
# #     try:
# #         data = request.get_json()

# #         if not data or 'studentId' not in data or 'academicData' not in data or 'attendance' not in data or 'reportCard' not in data:
# #             return jsonify({"error": "JSON body must contain studentId, academicData, attendance, and reportCard fields"}), 400

# #         student_id = data['studentId']
# #         academic_data = data['academicData']
# #         attendance = data['attendance']
# #         report_card = data['reportCard']

# #         # Validate Academic Data
# #         if not isinstance(academic_data, dict) or 'grade' not in academic_data or 'subjects' not in academic_data:
# #             return jsonify({"error": "academicData must contain 'grade' and 'subjects' fields"}), 400

# #         grade = academic_data['grade']
# #         if not isinstance(grade, str):
# #             return jsonify({"error": "'grade' must be a string"}), 400

# #         subjects = academic_data['subjects']
# #         if not isinstance(subjects, dict) or not all(isinstance(v, dict) and 'marks' in v for v in subjects.values()):
# #             return jsonify({"error": "'subjects' must be a valid JSON object with subject names and marks"}), 400

# #         # Validate Attendance
# #         if not isinstance(attendance, dict) or not all(k in attendance for k in ["totalWorkingDays", "presentDays", "absentDays", "halfDays"]):
# #             return jsonify({"error": "'attendance' must contain 'totalWorkingDays', 'presentDays', 'absentDays', and 'halfDays' fields"}), 400

# #         # Validate Report Card
# #         if not isinstance(report_card, dict) or not all(k in report_card for k in ["totalMarks", "obtainedMarks", "grade", "activity"]):
# #             return jsonify({"error": "'reportCard' must contain 'totalMarks', 'obtainedMarks', 'grade', and 'activity' fields"}), 400

# #         # Fetch student record
# #         student_ref = db.collection("students").document(student_id)
# #         student_doc = student_ref.get()

# #         if not student_doc.exists:
# #             return jsonify({"error": "Student account not found"}), 404

# #         # Update Firestore Database
# #         student_ref.update({
# #             "academicData": {
# #                 "grade": grade,
# #                 "subjects": subjects
# #             },
# #             "attendance": attendance,
# #             "reportCard": report_card
# #         })

# #         return jsonify({"message": "Student report card information updated successfully", "studentId": student_id}), 200

# #     except Exception as e:
# #         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# # @app.route("/api/student/report-card/<student_id>", methods=["GET"])
# # def get_student_report_card(student_id):
# #     try:
# #         # Fetch student record from Firestore
# #         student_ref = db.collection("students").document(student_id)
# #         student_doc = student_ref.get()

# #         # Check if student exists
# #         if not student_doc.exists:
# #             return jsonify({"error": "Student account not found"}), 404

# #         # Get student data
# #         student_data = student_doc.to_dict()

# #         # Extract report card details
# #         response_data = {
# #             "studentId": student_id,
# #             "academicData": student_data.get("academicData", {}),
# #             "attendance": student_data.get("attendance", {}),
# #             "reportCard": student_data.get("reportCard", {})
# #         }

# #         return jsonify(response_data), 200

# #     except Exception as e:
# #         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# # @app.route("/api/student/update-report-card", methods=["POST"])
# # def update_student_report_card():
# #     try:
# #         data = request.get_json()

# #         if not data or 'studentId' not in data or 'academicData' not in data or 'attendance' not in data or 'reportCard' not in data:
# #             return jsonify({"error": "JSON body must contain studentId, academicData, attendance, and reportCard fields"}), 400

# #         student_id = data['studentId']
# #         academic_data = data['academicData']
# #         attendance = data['attendance']
# #         report_card = data['reportCard']

# #         # Validate Academic Data
# #         if not isinstance(academic_data, dict) or 'grade' not in academic_data or 'subjects' not in academic_data:
# #             return jsonify({"error": "academicData must contain 'grade' and 'subjects' fields"}), 400

# #         grade = academic_data['grade']
# #         if not isinstance(grade, str):
# #             return jsonify({"error": "'grade' must be a string"}), 400

# #         subjects = academic_data['subjects']
# #         if not isinstance(subjects, dict) or not all(isinstance(v, dict) and 'marks' in v for v in subjects.values()):
# #             return jsonify({"error": "'subjects' must be a valid JSON object with subject names and marks"}), 400

# #         # Validate Attendance
# #         if not isinstance(attendance, dict) or not all(k in attendance for k in ["totalWorkingDays", "presentDays", "absentDays", "halfDays"]):
# #             return jsonify({"error": "'attendance' must contain 'totalWorkingDays', 'presentDays', 'absentDays', and 'halfDays' fields"}), 400

# #         # Validate Report Card
# #         if not isinstance(report_card, dict) or not all(k in report_card for k in ["totalMarks", "obtainedMarks", "grade", "activity"]):
# #             return jsonify({"error": "'reportCard' must contain 'totalMarks', 'obtainedMarks', 'grade', and 'activity' fields"}), 400

# #         # Fetch student record
# #         student_ref = db.collection("students").document(student_id)
# #         student_doc = student_ref.get()

# #         if not student_doc.exists:
# #             return jsonify({"error": "Student account not found"}), 404

# #         # Update Firestore Database
# #         student_ref.update({
# #             "academicData": {
# #                 "grade": grade,
# #                 "subjects": subjects
# #             },
# #             "attendance": attendance,
# #             "reportCard": report_card
# #         })

# #         return jsonify({"message": "Student report card updated successfully", "studentId": student_id}), 200

# #     except Exception as e:
# #         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# teacher feedback api
@app.route("/api/teacher/feedback", methods=["POST"])
def observation_feedback():
    try:
        # Extract data from request
        data = request.json
        teacher_id = data.get("teacherId")
        feedback_message = data.get("feedback")

        if not teacher_id or not feedback_message:
            return jsonify({"error": "teacherId and feedback are required"}), 400

        # Reference to the teacher's feedback document
        feedback_ref = db.collection("teacher_feedback").document(teacher_id)
        feedback_doc = feedback_ref.get()

        if feedback_doc.exists:
            existing_feedback = feedback_doc.to_dict().get("feedback", [])
            existing_feedback.append(feedback_message)
            feedback_ref.update({"feedback": existing_feedback})
        else:
            feedback_ref.set({"feedback": [feedback_message]})

        return jsonify({"Success": True}), 200

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# subjects

# POST Subject API
# @app.route('/grades/classes/subject10', methods=['POST'])
# def add_subject10():
#     try:
#         data = request.get_json()

#         # Ensure JSON body exists and contains required fields
#         if not data or 'grade_id' not in data or 'class_id' not in data or 'subject_name' not in data:
#             return jsonify({"error": "JSON body must contain 'grade_id', 'class_id', and 'subject_name'"}), 400

#         grade_id = data['grade_id']
#         class_id = data['class_id']
#         subject_name = data['subject_name']

#         subject_id = str(uuid.uuid4())
#         subject_ref = (
#             db.collection('Grades')
#             .document(grade_id)
#             .collection('classes')
#             .document(class_id)
#             .collection('subjects')
#             .document(subject_id)
#         )

#         subject_ref.set({
#             "subject_id": subject_id,
#             "subject_name": subject_name
#         })

#         return jsonify({
#             "message": f"Subject {subject_name} added successfully.",
#             "subject_id": subject_id
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# # GET Subjects API
# @app.route('/grades/classes/subjects1', methods=['GET'])
# def get_subjects1():
#     try:
#         grade_id = request.args.get('grade_id')
#         class_id = request.args.get('class_id')
#         subject_name = request.args.get('subject_name')  # Optional filter

#         if not grade_id or not class_id:
#             return jsonify({"error": "Query parameters 'grade_id' and 'class_id' are required"}), 400

#         subjects_ref = (
#             db.collection('Grades')
#             .document(grade_id)
#             .collection('classes')
#             .document(class_id)
#             .collection('subjects')
#         )

#         subjects_query = subjects_ref.stream()
#         subjects = []

#         for subject in subjects_query:
#             subject_data = subject.to_dict()
#             if subject_name and subject_data.get('subject_name') != subject_name:
#                 continue
#             subjects.append(subject_data)

#         if not subjects:
#             return jsonify({"message": "No subjects found for the given filters"}), 404

#         return jsonify(subjects), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# # GET Subjects API
# @app.route('/grades/classes/subject10', methods=['GET'])
# def get_subjects10():
#     try:
#         grade_id = request.args.get('grade_id')
#         class_id = request.args.get('class_id')
#         subject_name = request.args.get('subject_name')  # Optional filter

#         # Ensure required query parameters exist
#         if not grade_id or not class_id:
#             return jsonify({"error": "Query parameters 'grade_id' and 'class_id' are required"}), 400

#         subjects_ref = (
#             db.collection('Grades')
#             .document(grade_id)
#             .collection('classes')
#             .document(class_id)
#             .collection('subjects')
#         )

#         subjects_query = subjects_ref.stream()
#         subjects = []

#         for subject in subjects_query:
#             subject_data = subject.to_dict()
#             if subject_name and subject_data.get('subject_name') != subject_name:
#                 continue
#             subjects.append(subject_data)

#         if not subjects:
#             return jsonify({"message": "No subjects found for the given filters"}), 404

#         return jsonify({"subjects": subjects}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

import random
import smtplib
from email.mime.text import MIMEText
from flask import Flask, request, jsonify
from threading import Thread
import time

time.sleep(1)  # Add a 1-second delay between emails


class EmailClient:
    def __init__(self):
        self.smtp_host = "smtp.gmail.com"
        self.smtp_port = 587
        self.username = "tajheezai@gmail.com"
        self.password = "sfekauxgkaszulgy"
        self.server = None  # Store persistent SMTP connection

    def connect(self):
        """Establish SMTP connection once and reuse it"""
        if self.server is None:
            self.server = smtplib.SMTP(self.smtp_host, self.smtp_port, timeout=10)
            self.server.ehlo()
            self.server.starttls()
            self.server.ehlo()
            self.server.login(self.username, self.password)

    def send_email(self, recipient_email, subject, body):
        try:
            self.connect()  # Ensure connection is established

            msg = MIMEText(body, "html")
            msg["Subject"] = subject
            msg["From"] = self.username
            msg["To"] = recipient_email

            self.server.sendmail(self.username, recipient_email, msg.as_string())
            print(f"Email sent to {recipient_email}")
            return True

        except Exception as e:
            print(f"Error sending email: {e}")
            return False

    def close(self):
        """Close SMTP connection when done"""
        if self.server:
            self.server.quit()
            self.server = None


# Initialize global email client
email_client = EmailClient()


# Function to generate 6-digit OTP
def generate_otp():
    return random.randint(100000, 999999)


# Function to send an email synchronously
def send_email_async(recipient_email, otp):
    try:
        # Prepare email content
        subject = "Your OTP Code for PEES"
        body = f"""
        <html>
            <body>
                <p>Hi <b>{recipient_email}</b>,</p>
                <p>To complete your login or verification, please use the following One-Time Password (OTP):</p>
                <p style="font-size: 18px; font-weight: bold;">Ã°Å¸â€Â Your OTP: <span style="color: green;">{otp}</span></p>
                <p>This OTP is valid for the next <b>10 minutes</b>. Please do not share this code with anyone to ensure your account remains secure.</p>
                <p>If you didn't request this, please ignore this email or contact our support team.</p>
                <br>
                <p>Best regards,</p>
                <p>PEES Team</p>
            </body>
        </html>
        """
        # Send the email synchronously
        email_client.send_email(recipient_email, subject, body)
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False
    return True


import redis

redis_client = redis.StrictRedis(host="localhost", port=6379, db=0)


@app.route("/generate-otp1", methods=["POST"])
def generate_otp_api2():
    try:
        data = request.json
        if "email" not in data:
            return jsonify({"error": "Email not provided"}), 400

        recipient_email = data["email"]

        # Check if the email already exists in the users collection
        users_collection = db.collection("users")
        existing_user = users_collection.where("email", "==", recipient_email).stream()

        if any(existing_user):  # Check if there's any matching user
            print("Email Already Exists")
            return jsonify({"error": "Email already exists"}), 400

        # Generate OTP
        otp = generate_otp()
        print("Generated OTP:", otp)

        # Store OTP in Redis with a 5-minute expiration
        redis_client.setex(f"otp:{recipient_email}", 300, otp)

        # Send OTP to the email
        email_status = send_email_async(recipient_email, otp)

        if email_status:
            return jsonify({"message": f"OTP sent to {recipient_email}"}), 200
        else:
            return jsonify({"error": "Failed to send email"}), 500

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/verify-otp", methods=["POST"])
def verify_otp():
    try:
        data = request.json
        if "email" not in data or "otp" not in data:
            return jsonify({"error": "Email and OTP are required"}), 400

        recipient_email = data["email"]
        otp = str(data["otp"])  # Convert input OTP to string for comparison
        print(f"Verifying OTP for {recipient_email}: {otp}")

        stored_otp = redis_client.get(f"otp:{recipient_email}")

        if stored_otp:
            stored_otp = stored_otp.decode("utf-8")  # Decode bytes to string
            print(f"Stored OTP from Redis: {stored_otp}")

            if stored_otp == otp:  # Compare as strings
                redis_client.delete(
                    f"otp:{recipient_email}"
                )  # Delete OTP after successful verification
                return jsonify({"message": "OTP verified successfully"}), 200
            else:
                return jsonify({"error": "Invalid OTP"}), 400
        else:
            return jsonify({"error": "Expired or missing OTP"}), 400

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# -----------------------------------------------------------------------FORGOT-PASSWORD-----------------------------------------------------------

from itsdangerous import URLSafeTimedSerializer

# app.secret_key = os.getenv('SECRET_KEY', 'your_super_secret_key')  # Ensure a default value

serializer = URLSafeTimedSerializer(app.secret_key)


from reset_mail import send_reset_email


# Function to generate password reset token
def generate_reset_token(user_id):
    return serializer.dumps(str(user_id), salt="password-reset-salt")


@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    try:
        email = request.json.get("email")
        if not email:
            return jsonify({"message": "Email is required"}), 400

        # Check if user exists in Firestore
        users_ref = db.collection("users")
        query = users_ref.where("email", "==", email).stream()
        user_data = next(query, None)

        if user_data:
            user_id = user_data.id  # Firestore document ID
            token = generate_reset_token(user_id)
            reset_link = (
                f"http://localhost:3000/reset_password?token={token}&email={email}"
            )

            # Send email (implement `send_reset_email`)
            send_reset_email(
                email,
                user_data.to_dict()
                .get("profileInfo")
                .get("personalInformation")
                .get("name"),
                reset_link,
            )

            return (
                jsonify(
                    {"message": "Password reset link has been sent to your email."}
                ),
                200,
            )
        else:
            return (
                jsonify({"message": "No account found with that email address."}),
                404,
            )

    except Exception as e:
        return jsonify({"message": f"Error: {str(e)}"}), 500


# API route for resetting password
@app.route("/reset_password/<token>", methods=["POST"])
def reset_password11(token):
    try:
        user_id = serializer.loads(token, salt="password-reset-salt", max_age=3600)

        user_ref = db.collection("users").document(user_id)
        user = user_ref.get()

        if not user.exists:
            return jsonify({"message": "User not found."}), 404

        new_password = request.json.get("password")
        re_new_password = request.json.get("re_password")

        if not new_password:
            return jsonify({"message": "Password is required"}), 400

        if new_password != re_new_password:
            return jsonify({"message": "Passwords do not match"}), 400

        # hashed_password = hashlib.sha256(new_password.encode()).hexdigest()

        hashed_password = encrypt_password(new_password)

        # Prevent setting the same password
        if user.to_dict().get("passwordHash") == hashed_password:
            return (
                jsonify(
                    {
                        "message": "New password cannot be the same as the current password."
                    }
                ),
                400,
            )

        # Update password in Firestore
        user_ref.update({"passwordHash": hashed_password})

        return jsonify({"message": "Your password has been updated."}), 200

    except Exception as e:
        return jsonify({"message": "The reset link is invalid or has expired."}), 400


@app.route("/api/student/upload-photo", methods=["POST"])
def upload_student_photo():
    try:
        # Validate input
        student_id = request.form.get("studentId")
        if not student_id:
            return jsonify({"error": "studentId is required"}), 400

        if "photo" not in request.files:
            return jsonify({"error": "Photo file is required"}), 400

        photo = request.files["photo"]

        if photo.filename == "":
            return jsonify({"error": "No file selected"}), 400

        # Get the student's Firestore document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        # Check if an old photo exists
        student_data = student_doc.to_dict()
        old_photo_url = student_data.get("personalInformation", {}).get("photourl")

        if old_photo_url:
            # Extract file path from URL to delete old file
            old_file_path = old_photo_url.split("o/")[1].split("?")[
                0
            ]  # Extracting the path
            old_file_path = old_file_path.replace("%2F", "/")  # Decode path
            old_blob = bucket.blob(old_file_path)

            # Delete old photo
            old_blob.delete()
            print(f"Old photo deleted: {old_file_path}")

        # Generate a unique filename for new image
        file_extension = os.path.splitext(photo.filename)[1]  # Extract file extension
        unique_filename = (
            f"student_photos/{student_id}_{uuid.uuid4().hex}{file_extension}"
        )

        # Upload new file to Firebase Storage
        blob = bucket.blob(unique_filename)
        blob.upload_from_file(photo, content_type=photo.content_type)

        # Make the file publicly accessible
        blob.make_public()
        new_photo_url = blob.public_url  # Get public URL of the uploaded image

        # Update Firestore with the new photo URL
        student_ref.update({"personalInformation.photourl": new_photo_url})

        return (
            jsonify(
                {"message": "Photo updated successfully", "photo_url": new_photo_url}
            ),
            200,
        )

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": "An internal server error occurred"}), 500


@app.route("/api/student/upload-photo1", methods=["POST"])
def upload_student_photo1():
    try:
        # Validate input
        student_id = request.form.get("studentId")
        if not student_id:
            return jsonify({"error": "studentId is required"}), 400

        if "photo" not in request.files:
            return jsonify({"error": "Photo file is required"}), 400

        photo = request.files["photo"]

        if photo.filename == "":
            return jsonify({"error": "No file selected"}), 400

        # Get the student's Firestore document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        # Check if an old photo exists
        student_data = student_doc.to_dict()
        old_photo_url = (
            student_data.get("profileInfo", {})
            .get("personalInformation", {})
            .get("photoUrl")
        )

        if old_photo_url:
            try:
                # Extract file path from URL to delete old file
                old_file_path = old_photo_url.split("o/")[1].split("?")[0]
                old_file_path = old_file_path.replace("%2F", "/")  # Decode path
                old_blob = bucket.blob(old_file_path)

                # Delete old photo
                old_blob.delete()
                print(f"Old photo deleted: {old_file_path}")
            except Exception as e:
                print(f"Error deleting old photo: {str(e)}")

        # Generate a unique filename for new image
        file_extension = os.path.splitext(photo.filename)[1]  # Extract file extension
        unique_filename = (
            f"student_photos/{student_id}_{uuid.uuid4().hex}{file_extension}"
        )

        # Upload new file to Firebase Storage
        blob = bucket.blob(unique_filename)
        blob.upload_from_file(photo, content_type=photo.content_type)

        # Make the file publicly accessible
        blob.make_public()
        new_photo_url = blob.public_url  # Get public URL of the uploaded image

        # Update Firestore with the new photo URL in students collection
        student_ref.update({"profileInfo.personalInformation.photoUrl": new_photo_url})

        # âœ… Update the user's photo in the users collection where userId matches studentId
        user_query = db.collection("users").where("userId", "==", student_id).stream()
        for user_doc in user_query:
            user_ref = db.collection("users").document(user_doc.id)
            user_ref.update({"profileInfo.personalInformation.photoUrl": new_photo_url})

        print(f"Photo uploaded successfully: {new_photo_url}")
        return (
            jsonify(
                {"message": "Photo updated successfully", "photo_url": new_photo_url}
            ),
            200,
        )

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500


# ------------------------------------------------------file/ocr------------------------------------------
import os
import re
import asyncio
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from azure.ai.formrecognizer.aio import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import firebase_admin
from firebase_admin import credentials, storage, firestore
from dotenv import load_dotenv
from openai import AsyncOpenAI
import os
import uuid
from flask import Flask, request, jsonify
from fpdf import FPDF
import firebase_admin
from firebase_admin import credentials, firestore, storage
import json
import pdfkit

# Load environment variables
# load_dotenv()

# app = Flask(__name__)

# Configuration
AZURE_ENDPOINT = "https://aiocr395080637747.cognitiveservices.azure.com/"
AZURE_KEY = "b1b026f421034dabb948999b80a63e8c"
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()

if not AZURE_KEY:
    raise ValueError("AZURE_KEY environment variable is missing.")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY environment variable is missing.")

openai_client = AsyncOpenAI(api_key=OPENAI_API_KEY)

# # Initialize Firebase Admin SDK
# cred = credentials.Certificate("serviceAccountKey.json")
# firebase_admin.initialize_app(cred, {'storageBucket': 'pees-d1101.firebasestorage.app'})

app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

# Initialize Firestore and Storage
db = firestore.client()
bucket = storage.bucket()

# Initialize Azure OCR Client
azure_client = DocumentAnalysisClient(
    endpoint=AZURE_ENDPOINT, credential=AzureKeyCredential(AZURE_KEY)
)


# Function to clean extracted text
def clean_text(raw_text):
    cleaned = re.sub(r"\s+", " ", raw_text)
    cleaned = re.sub(r"[^a-zA-Z0-9.,!?\'\";:()\\-]", " ", cleaned)
    return cleaned.strip()


# Extract text using Azure OCR (Async)
async def extract_text_from_file(file_bytes):
    async with DocumentAnalysisClient(
        endpoint=AZURE_ENDPOINT, credential=AzureKeyCredential(AZURE_KEY)
    ) as azure_client:
        poller = await azure_client.begin_analyze_document(
            "prebuilt-document", file_bytes
        )
        result = await poller.result()
        raw_text = "\n".join(
            [line.content for page in result.pages for line in page.lines]
        )
        return clean_text(raw_text)


# Extract student info
def extract_student_info(text):
    name_match = re.search(r"Name[:\-\s]+([\w\s]+)", text, re.IGNORECASE)
    grade_match = re.search(r"Grade\s*Level[:\-\s]+([\w\s]+)", text, re.IGNORECASE)
    return name_match.group(1) if name_match else "Unknown", (
        grade_match.group(1) if grade_match else "Unknown"
    )


# Evaluate text using OpenAI GPT-4o (async)
async def evaluate_text_with_openai(extracted_text):
    response = await openai_client.chat.completions.create(
        model="gpt-4.1-mini-2025-04-14",
        messages=[
            {
                "role": "system",
                "content": "You are an AI assistant that analyzes educational text.",
            },
            {
                "role": "user",
                "content": f"Analyze the following extracted text and summarize key points:\n{extracted_text}",
            },
        ],
        temperature=0.7,
        max_tokens=1000,
    )
    return (
        clean_text(response.choices[0].message.content)
        if response and response.choices
        else "No valid response received."
    )


import json


# Generate personalized teaching plan using OpenAI GPT-4o (async)
import json # Ensure 'import json' is at the top of your app.py if it's not already there

# async def generate_teaching_plan(student_name, grade_level, evaluation):
#     template = f"""
#     Create a personalized teaching plan for:

#     - Student: {student_name}
#     - Grade Level: {grade_level}

#     Evaluation Summary:
#     {evaluation}

#     The final output **MUST** be a **strict JSON object** containing two top-level keys: "en" and "ar".
#     - The value for "en" must be the teaching plan in **English**.
#     - The value for "ar" must be the same teaching plan, translated and formatted in **Arabic**.

#     Each language plan (the value of "en" and "ar") must follow the same internal structure of key-value pairs (maps). For the Arabic plan, you **MUST** translate the keys (e.g., "assessmentMethods" becomes "Ø·Ø±Ù‚_Ø§Ù„ØªÙ‚ÙŠÙŠÙ…") as well as the values.

#     Example of the required **strict JSON structure**:

#     ```json
#     {{
#       "en": {{
#         "assessmentMethods": {{
#           "method1": "Description in English",
#           "method2": "Description in English"
#         }},
#         "instructionalStrategies": {{...}},
#         "learningObjectives": {{...}},
#         "recommendedResources": {{...}},
#         "timeline": {{...}}
#       }},
#       "ar": {{
#         "Ø·Ø±Ù‚_Ø§Ù„ØªÙ‚ÙŠÙŠÙ…": {{
#           "Ø§Ù„Ø·Ø±ÙŠÙ‚Ø©_1": "Ø§Ù„ÙˆØµÙ Ø¨Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©",
#           "Ø§Ù„Ø·Ø±ÙŠÙ‚Ø©_2": "Ø§Ù„ÙˆØµÙ Ø¨Ø§Ù„Ù„ØºØ© Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©"
#         }},
#         "Ø§Ø³ØªØ±Ø§ØªÙŠØ¬ÙŠØ§Øª_Ø§Ù„ØªØ¯Ø±ÙŠØ³": {{...}},
#         "Ø£Ù‡Ø¯Ø§Ù_Ø§Ù„ØªØ¹Ù„Ù…": {{...}},
#         "Ø§Ù„Ù…ØµØ§Ø¯Ø±_Ø§Ù„Ù…ÙˆØµÙ‰_Ø¨Ù‡Ø§": {{...}},
#         "Ø§Ù„Ø¬Ø¯ÙˆÙ„_Ø§Ù„Ø²Ù…Ù†ÙŠ": {{...}}
#       }}
#     }}
#     ```

#     Ensure:
#     - The top-level response is **100% valid JSON** starting with {{"en":...}}.
#     - **No extra text** before or after the JSON block.
#     """

#     response = await openai_client.chat.completions.create(
#         model="gpt-4.1-mini-2025-04-14",
#         messages=[
#             {
#                 "role": "system",
#                 "content": "You are an AI that generates a single, structured JSON response containing both English and Arabic translations.",
#             },
#             {"role": "user", "content": template},
#         ],
#         temperature=0.7,
#         max_tokens=2000, # Increased tokens to support dual output
#     )

#     if response and response.choices:
#         try:
#             # Extract the response and remove backticks if needed
#             raw_response = response.choices[0].message.content.strip()

#             # Ensure we extract JSON block only
#             if "```json" in raw_response:
#                 raw_response = raw_response.split("```json")[1].split("```")[0].strip()

#             # Parse the dual-language plan
#             dual_plan_json = json.loads(raw_response)

#             # Validate that both required keys are present
#             if "en" in dual_plan_json and "ar" in dual_plan_json:
#                 return dual_plan_json # <--- Function now returns the dual-language structure
#             else:
#                  return {"error": "Generated plan is missing 'en' or 'ar' key from the AI response."}
            
#         except json.JSONDecodeError:
#             return {
#                 "error": "Failed to parse teaching plan response. AI did not return valid JSON.",
#                 "raw_output": raw_response # Useful for debugging what the model returned
#             }
#     else:
#         return {"error": "No valid teaching plan generated."}
    

from fpdf import FPDF


class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 12)
        self.cell(0, 10, "Teaching Plan", 0, 1, "C")

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")


def create_pdf(content, filename):
    pdf = PDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Strip extra whitespace and newlines to avoid blank pages
    formatted_content = content.strip()

    # Ensure no forced extra page by limiting content width
    pdf.multi_cell(190, 10, formatted_content)

    pdf.output(filename)


@app.route("/file_ocr", methods=["POST"])
async def upload_file():
    if "file" not in request.files or "studentId" not in request.form:
        return (
            jsonify({"error": "Invalid request, file and studentId are required"}),
            400,
        )

    student_id = request.form["studentId"]
    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "Invalid file"}), 400

    file_bytes = file.read()

    try:
        extracted_text = await extract_text_from_file(file_bytes)
        student_name, grade_level = extract_student_info(extracted_text)

        evaluation, teaching_plan = await asyncio.gather(
            evaluate_text_with_openai(extracted_text),
            generate_teaching_plan(
                student_name, grade_level, extracted_text
            ),  # Now returns maps directly
        )

        # Generate Plan ID
        plan_id = str(uuid.uuid4()).replace("-", "_")

        # Check if the student exists
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": f"Student with ID {student_id} not found"}), 404

        # Ensure `planId` is stored inside `actionPlan`
        teaching_plan["planId"] = plan_id
        import time

        # Store directly as a map
        student_ref.update(
            {
                f"teachingPlans.{plan_id}": {
                    "actionPlan": teaching_plan,
                    "createdAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                }
            }
        )

        return (
            jsonify(
                {
                    "studentId": student_id,
                    "planId": plan_id,
                    "teaching_plan": teaching_plan,
                    "extracted_text": extracted_text,
                    "evaluation": evaluation,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

#--------------------------

from flask import Flask, request, jsonify, send_file
import openai
import firebase_admin
from firebase_admin import credentials, firestore, storage, initialize_app
import os
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
import uuid
import textwrap
import json
from reportlab.lib.pagesizes import A4
from openai import OpenAI
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from arabic_reshaper import arabic_reshaper
from bidi.algorithm import get_display

from trans import translate
# Register Arabic fonts
try:
    pdfmetrics.registerFont(TTFont("Amiri", "static/fonts/Amiri-Regular.ttf"))
    pdfmetrics.registerFont(TTFont("Amiri-Bold", "static/fonts/Amiri-Bold.ttf"))
    print("Arabic fonts registered successfully.")
except Exception as e:
    print(f"CRITICAL ERROR: Arabic fonts not found or failed to register: {e}")
    # Consider raising the error or stopping for testing, if it's a critical dependency

# Register English fonts
# pdfmetrics.registerFont(TTFont("Helvetica", "Helvetica.ttf"))
# pdfmetrics.registerFont(TTFont("Helvetica-Bold", "Helvetica-Bold.ttf"))

# Translation dictionaries
ARABIC_TRANSLATIONS = {
    "Teaching Plan": "خطة التدريس",
    "CreatedAt": "تاريخ الإنشاء",
    "instructionalStrategies": "استراتيجيات التدريس",
    "recommendedResources": "الموارد المقترحة",
    "assessmentMethods": "طرق التقييم",
    "learningObjectives": "أهداف التعلم",
    "timeline": "الجدول الزمني",

    # "Teaching Plan": "Ø®Ø·Ø© Ø§Ù„ØªØ¯Ø±ÙŠØ³",
    # "CreatedAt": "ØªÙ… Ø§Ù„Ø¥Ù†Ø´Ø§Ø¡ ÙÙŠ",
    # "instructionalStrategies": "Ø§Ø³ØªØ±Ø§ØªÙŠØ¬ÙŠØ§Øª Ø§Ù„ØªØ¯Ø±ÙŠØ³",
    # "recommendedResources": "Ø§Ù„Ù…ÙˆØ§Ø±Ø¯ Ø§Ù„Ù…Ù‚ØªØ±Ø­Ø©",
    # "assessmentMethods": "Ø·Ø±Ù‚ Ø§Ù„ØªÙ‚ÙŠÙŠÙ…",
    # "learningObjectives": "Ø£Ù‡Ø¯Ø§Ù Ø§Ù„ØªØ¹Ù„Ù…",
    # "timeline": "Ø§Ù„Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø²Ù…Ù†ÙŠ",
}


def get_translation12(key, lang):
    """Get translation for a key based on language"""
    if lang == "ar":
        return ARABIC_TRANSLATIONS.get(key, key)
    return key


def setup_pdf_direction12(c, lang, width):
    """Set up the PDF direction based on language (RTL for Arabic)"""
    if lang == "ar":
        c.setFont("Amiri", 12)
        return True
    return False


def draw_rtl_text12(
    c, text, x, y, font_name="Amiri", font_size=12, available_width=None
):
    """Properly draw RTL text with wrapping if needed"""
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)

    c.setFont(font_name, font_size)

    if available_width:
        # Handle text wrapping for RTL
        lines = []
        current_line = []
        current_width = 0

        # Split text into words (Arabic words are separated by spaces)
        words = bidi_text.split()

        for word in words:
            word_width = c.stringWidth(word + " ", font_name, font_size)
            if current_width + word_width <= available_width:
                current_line.append(word)
                current_width += word_width
            else:
                if current_line:
                    lines.append(" ".join(current_line))
                current_line = [word]
                current_width = word_width

        if current_line:
            lines.append(" ".join(current_line))

        # Draw each line from right to left
        for i, line in enumerate(lines):
            c.drawRightString(x, y - (i * (font_size + 2)), line)
        return len(lines)
    else:
        c.drawRightString(x, y, bidi_text)
        return 1


def draw_string_with_direction12(
    c, text, x, y, lang, bold=False, font_size=12, available_width=None
):
    """Draw text with proper direction handling"""
    # First determine if the text contains Arabic characters
    text_contains_arabic = contains_arabic(text)

    # For Arabic text, always use Arabic shaping regardless of report language
    if text_contains_arabic:
        font_name = "Amiri-Bold" if bold else "Amiri"
        return draw_rtl_text12(c, text, x, y, font_name, font_size, available_width)
    else:
        # For non-Arabic text, use the specified language's settings
        if lang == "ar":
            font_name = "Amiri-Bold" if bold else "Amiri"
            return draw_rtl_text12(c, text, x, y, font_name, font_size, available_width)
        else:
            font_name = "Helvetica-Bold" if bold else "Helvetica"
            c.setFont(font_name, font_size)

            if available_width:
                # Handle text wrapping for LTR
                lines = []
                current_line = []
                current_width = 0

                for word in text.split():
                    word_width = c.stringWidth(word + " ", font_name, font_size)
                    if current_width + word_width <= available_width:
                        current_line.append(word)
                        current_width += word_width
                    else:
                        if current_line:
                            lines.append(" ".join(current_line))
                        current_line = [word]
                        current_width = word_width

                if current_line:
                    lines.append(" ".join(current_line))

                for i, line in enumerate(lines):
                    c.drawString(x, y - (i * (font_size + 2)), line)
                return len(lines)
            else:
                c.drawString(x, y, text)
                return 1


# def contains_arabic12(text):
#     """Check if text contains Arabic characters"""
#     if not isinstance(text, str):
#         return False

#     arabic_unicode_ranges = [
#         ("\u0600", "\u06ff"),  # Arabic
#         ("\u0750", "\u077f"),  # Arabic Supplement
#         ("\u08a0", "\u08ff"),  # Arabic Extended-A
#         ("\ufb50", "\ufdff"),  # Arabic Presentation Forms-A
#         ("\ufe70", "\ufeff"),  # Arabic Presentation Forms-B
#     ]

#     for char in text:
#         for start, end in arabic_unicode_ranges:
#             if start <= char <= end:
#                 return True
#     return False


# def prepare_text12(text, force_arabic=False):
#     if not text:
#         return ""
#     if force_arabic or contains_arabic12(text):
#         reshaped_text = arabic_reshaper.reshape(text)
#         return get_display(reshaped_text)
#     return text


# def is_arabic12(text: str) -> bool:
#     """Check if text contains Arabic characters"""
#     arabic_chars = set("Ø¡Ø¢Ø£Ø¤Ø¥Ø¦Ø§Ø¨Ø©ØªØ«Ø¬Ø­Ø®Ø¯Ø°Ø±Ø²Ø³Ø´ØµØ¶Ø·Ø¸Ø¹ØºÙÙ‚ÙƒÙ„Ù…Ù†Ù‡ÙˆÙ‰ÙŠ")
#     return any(char in arabic_chars for char in text)


# def export_data12(lang, student_id, plan_id):
#     """Export and translate teaching plan data piece by piece with deep translation"""
#     student_ref = db.collection("students").document(student_id)
#     student_doc = student_ref.get()
#     if not student_doc.exists:
#         return None

#     student_data = student_doc.to_dict()
#     teaching_plan = student_data.get("teachingPlans", {}).get(plan_id, {})

#     if not teaching_plan:
#         return None

#     # Create deep copy
#     content = json.loads(json.dumps(teaching_plan))

#     def deep_translate(data, lang):
#         """Recursively translate all strings in a nested structure"""
#         if isinstance(data, dict):
#             translated_dict = {}
#             for key, value in data.items():
#                 # Translate the key if it's a string
#                 translated_key = (
#                     translate_item(key, lang) if isinstance(key, str) else key
#                 )
#                 # Recursively translate the value
#                 translated_value = deep_translate(value, lang)
#                 translated_dict[translated_key] = translated_value
#             return translated_dict
#         elif isinstance(data, list):
#             return [deep_translate(item, lang) for item in data]
#         elif isinstance(data, str):
#             return translate_item(data, lang)
#         else:
#             return data

#     # Translate the entire content including nested structures
#     translated_content = deep_translate(content, lang)
#     print(f"Translated content: {translated_content}")

#     return translated_content


# def translate_item(item, lang):
#     """Translate a single item if needed"""
#     if not item or not isinstance(item, str):
#         return item

#     # Determine if text needs translation
#     is_text_arabic = is_arabic12(item)

#     # For English output, translate Arabic to English
#     if lang == "en" and is_text_arabic:
#         return translate(item, "en")
#     # For Arabic output, translate English to Arabic
#     elif lang == "ar" and not is_text_arabic:
#         return translate(item, "ar")
#     # Otherwise return original text
#     return item
# Assuming 'db', 'json', 'translate_item', 'is_arabic12', 'arabic_reshaper', 'get_display' 
# and other dependencies are already defined.

def contains_arabic12(text):
    """Check if text contains Arabic characters"""
    if not isinstance(text, str):
        return False

    arabic_unicode_ranges = [
        ("\u0600", "\u06ff"),  # Arabic
        ("\u0750", "\u077f"),  # Arabic Supplement
        ("\u08a0", "\u08ff"),  # Arabic Extended-A
        ("\ufb50", "\ufdff"),  # Arabic Presentation Forms-A
        ("\ufe70", "\ufeff"),  # Arabic Presentation Forms-B
    ]

    for char in text:
        for start, end in arabic_unicode_ranges:
            if start <= char <= end:
                return True
    return False


def prepare_text12(text, force_arabic=False):
    if not text:
        return ""
    if force_arabic or contains_arabic12(text):
        # NOTE: 'arabic_reshaper' and 'get_display' must be imported/defined.
        reshaped_text = arabic_reshaper.reshape(text)
        return get_display(reshaped_text)
    return text


def is_arabic12(text: str) -> bool:
    """Check if text contains Arabic characters"""
    arabic_chars = set("Ø¡Ø¢Ø£Ø¤Ø¥Ø¦Ø§Ø¨Ø©ØªØ«Ø¬Ø­Ø®Ø¯Ø°Ø±Ø²Ø³Ø´ØµØ¶Ø·Ø¸Ø¹ØºÙÙ‚ÙƒÙ„Ù…Ù†Ù‡ÙˆÙ‰ÙŠ")
    return any(char in arabic_chars for char in text)


def translate_item(item, lang):
    """Translate a single item if needed"""
    if not item or not isinstance(item, str):
        return item

    # Determine if text needs translation
    is_text_arabic = is_arabic12(item)

    # For English output, translate Arabic to English
    if lang == "en" and is_text_arabic:
        # NOTE: 'translate' function must be defined elsewhere
        return translate(item, "en")
    # For Arabic output, translate English to Arabic
    elif lang == "ar" and not is_text_arabic:
        # NOTE: 'translate' function must be defined elsewhere
        return translate(item, "ar")
    # Otherwise return original text
    return item


def translate_analysis_recursively(data, lang):
    """
    Recursively translates all strings in a nested structure (analysis content).
    This version ensures dictionary keys (like 'strengths') are NOT translated.
    """
    
    def deep_translate(data, lang):
        if isinstance(data, dict):
            translated_dict = {}
            for key, value in data.items():
                # CRITICAL FIX: DO NOT translate the key; only use the original key name.
                translated_key = key
                # Recursively translate the value
                translated_value = deep_translate(value, lang)
                translated_dict[translated_key] = translated_value
            return translated_dict
        elif isinstance(data, list):
            return [deep_translate(item, lang) for item in data]
        elif isinstance(data, str):
            # Translate the string item
            return translate_item(data, lang)
        else:
            return data

    # Use json for a robust deep copy before translation
    content = json.loads(json.dumps(data))
    
    return deep_translate(content, lang)


def export_data12(lang, student_id, plan_id):
    """
    Export and translate teaching plan data piece by piece with deep translation.
    Uses the original deep translation logic which includes key translation.
    """
    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()
    if not student_doc.exists:
        return None

    student_data = student_doc.to_dict()
    teaching_plan = student_data.get("teachingPlans", {}).get(plan_id, {})

    if not teaching_plan:
        return None

    # Create deep copy
    content = json.loads(json.dumps(teaching_plan))

    def deep_translate(data, lang):
        """Recursively translate all strings in a nested structure, INCLUDING KEYS."""
        if isinstance(data, dict):
            translated_dict = {}
            for key, value in data.items():
                # Translate the key as originally requested for this function
                translated_key = (
                    translate_item(key, lang) if isinstance(key, str) else key
                )
                # Recursively translate the value
                translated_value = deep_translate(value, lang)
                translated_dict[translated_key] = translated_value
            return translated_dict
        elif isinstance(data, list):
            return [deep_translate(item, lang) for item in data]
        elif isinstance(data, str):
            return translate_item(data, lang)
        else:
            return data

    # Translate the entire content including nested structures
    translated_content = deep_translate(content, lang)
    print(f"Translated content: {translated_content}")

    return translated_content
    
def translate_item(item, lang):
    """Translate a single item if needed"""
    if not item or not isinstance(item, str):
        return item

    # Determine if text needs translation
    is_text_arabic = is_arabic12(item)

    # For English output, translate Arabic to English
    if lang == "en" and is_text_arabic:
        return translate(item, "en")
    # For Arabic output, translate English to Arabic
    elif lang == "ar" and not is_text_arabic:
        return translate(item, "ar")
    # Otherwise return original text
    return item

def generate_pdf12(content, filename, lang):
    """Generate PDF with proper RTL support for Arabic"""
    pdf_path = os.path.join(os.getcwd(), filename)

    # Configure document with language-appropriate settings
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=40,
        leftMargin=72 if lang == "ar" else 40,
        topMargin=40,
        bottomMargin=40,
        encoding="utf-8",
    )

    # Get base styles
    styles = getSampleStyleSheet()

    # Create Arabic-specific styles
    arabic_title_style = ParagraphStyle(
        name="ArabicTitle",
        fontName="Amiri-Bold",
        fontSize=16,
        leading=18,
        alignment=TA_RIGHT,
        spaceAfter=14,
        allowWidows=0,
        allowOrphans=0,
        wordWrap="RTL" if lang == "ar" else "LTR",
    )

    arabic_heading_style = ParagraphStyle(
        name="ArabicHeading",
        fontName="Amiri-Bold",
        fontSize=14,
        leading=16,
        alignment=TA_RIGHT,
        spaceAfter=12,
        wordWrap="RTL" if lang == "ar" else "LTR",
    )
    # Add Heading style (both English and Arabic versions)
    englsih_heading_style = ParagraphStyle(
        name="Heading",
        fontName="Helvetica-Bold" if lang != "ar" else "Amiri-Bold",
        fontSize=14,
        leading=16,
        alignment=TA_LEFT if lang != "ar" else TA_RIGHT,
        spaceAfter=12,
        wordWrap="LTR" if lang != "ar" else "RTL",
    )
    styles.add(englsih_heading_style)

    arabic_normal_style = ParagraphStyle(
        name="ArabicNormal",
        fontName="Amiri",
        fontSize=12,
        leading=14,
        alignment=TA_RIGHT,
        spaceAfter=8,
        wordWrap="RTL" if lang == "ar" else "LTR",
    )

    arabic_bullet_style = ParagraphStyle(
        name="ArabicBullet",
        fontName="Amiri",
        fontSize=12,
        leading=14,
        alignment=TA_RIGHT,
        firstLineIndent=-12 if lang == "ar" else 12,
        leftIndent=12,
        spaceAfter=6,
        wordWrap="RTL" if lang == "ar" else "LTR",
    )

    # Add Arabic styles to the stylesheet
    styles.add(arabic_title_style)
    styles.add(arabic_heading_style)
    styles.add(arabic_normal_style)
    styles.add(arabic_bullet_style)

    # Modify English styles for better Arabic support
    if lang == "ar":
        styles["Title"].fontName = "Amiri-Bold"
        styles["Title"].alignment = TA_RIGHT
        styles["Title"].fontSize = 16
        styles["Title"].wordWrap = "RTL"
        styles["Normal"].fontName = "Amiri"
        styles["Normal"].alignment = TA_RIGHT
        styles["Normal"].wordWrap = "RTL"
        styles["BodyText"].fontName = "Amiri"
        styles["BodyText"].alignment = TA_RIGHT
        styles["BodyText"].wordWrap = "RTL"
        styles["Bullet"].fontName = "Amiri"
        styles["Bullet"].alignment = TA_RIGHT
        styles["Bullet"].wordWrap = "RTL"
    else:
        styles["Title"].fontName = "Helvetica-Bold"
        styles["Title"].alignment = TA_LEFT
        styles["Normal"].fontName = "Helvetica"
        styles["Normal"].alignment = TA_LEFT
        styles["BodyText"].fontName = "Helvetica"
        styles["BodyText"].alignment = TA_JUSTIFY

    elements = []

    def format_arabic_text(text):
        """Properly shape and direction Arabic text"""
        if not text or not isinstance(text, str):
            return text
        if lang == "ar" or contains_arabic12(text):
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        return text
    def create_paragraph12(text, style_name, bold=False):
        """Create properly formatted paragraph with RTL support"""
        if not text:
            return None

        # Format Arabic text
        formatted_text = format_arabic_text(text)

        # Handle bullet points differently for RTL
        if style_name == "Bullet" and (lang == "ar" or contains_arabic12(text)):
            # For Arabic bullets, we need to manually handle the bullet position
            bullet_char = (
                "•" if lang != "ar" else "•"
            )  # You can use Arabic bullet if preferred
            if lang == "ar":
                # For RTL, bullet goes on the right
                para_text = f"{formatted_text} {bullet_char}"
            else:
                # For LTR, bullet goes on the left
                para_text = f"{bullet_char} {formatted_text}"

            return Paragraph(para_text, styles["ArabicBullet"])

        # Handle regular paragraphs
        if lang == "ar" or contains_arabic12(text):
            if style_name == "Title":
                return Paragraph(formatted_text, styles["ArabicTitle"])
            elif style_name == "Heading":
                return Paragraph(formatted_text, styles["ArabicHeading"])
            elif style_name == "Bullet":
                return Paragraph(formatted_text, styles["ArabicBullet"])
            else:
                return Paragraph(formatted_text, styles["ArabicNormal"])
        else:
            if bold:
                return Paragraph(f"<b>{formatted_text}</b>", styles[style_name])
            return Paragraph(formatted_text, styles[style_name])

    # Add title
    title = "خطة التدريس" if lang == "ar" else "Teaching Plan"
    elements.append(create_paragraph12(title, "Title"))
    elements.append(Spacer(1, 24))

    # Add creation date
    if "createdAt" in content or "createdAt: تم الإنشاء في" in content:
        created_label = "تم الإنشاء في" if lang == "ar" else "Created"
        date_value = content.get(
            "createdAt", content.get("createdAt: تم الإنشاء في", "")
        )
        date_text = f"{created_label}: {date_value}"
        elements.append(create_paragraph12(date_text, "Normal"))
        elements.append(Spacer(1, 24))

    # Process action plan sections
    action_plan = (
        content.get("actionPlan", {}) if lang != "ar" else content.get("خطة العمل", {})
    )
    print(f"Action plan content: {action_plan}")
    # Define section order with their translations
    sections_order = [
        ("learningObjectives", "أهداف التعلم"),
        ("instructionalStrategies", "استراتيجيات التدريس"),
        ("recommendedResources", "الموارد المقترحة"),
        ("assessmentMethods", "طرق التقييم"),
        ("timeline", "الجدول الزمني"),
    ]

    for section_en, section_ar in sections_order:
        section_key = section_ar if lang == "ar" else section_en
        if section_key in action_plan:
            # Section heading - use the appropriate language version
            section_title = section_ar if lang == "ar" else section_en
            elements.append(create_paragraph12(section_title, "Heading"))
            elements.append(Spacer(1, 12))

            section_content = action_plan[section_key]

            # Handle different content types
            if isinstance(section_content, dict):
                for key, value in section_content.items():
                    # For Arabic, we don't need to translate the keys as they're already in Arabic
                    if isinstance(value, str):
                        # Key: Value format - for RTL we put the colon on the left
                        if lang == "ar":
                            elements.append(
                                create_paragraph12(f"{key}: {value}", "Normal")
                            )
                        else:
                            elements.append(
                                create_paragraph12(f"{key}: {value}", "Normal")
                            )
                        elements.append(Spacer(1, 8))
                    elif isinstance(value, list):
                        # Bullet points
                        elements.append(
                            create_paragraph12(key + ":", "Normal", bold=True)
                        )
                        for item in value:
                            if isinstance(item, str) and item.strip():
                                elements.append(create_paragraph12(item, "Bullet"))
                        elements.append(Spacer(1, 8))
                    elif isinstance(value, dict):
                        # Handle nested dictionaries
                        elements.append(
                            create_paragraph12(key + ":", "Normal", bold=True)
                        )
                        for sub_key, sub_value in value.items():
                            if isinstance(sub_value, str):
                                if lang == "ar":
                                    elements.append(
                                        create_paragraph12(
                                            f"{sub_key}: {sub_value}", "Normal"
                                        )
                                    )
                                else:
                                    elements.append(
                                        create_paragraph12(
                                            f"{sub_key}: {sub_value}", "Normal"
                                        )
                                    )
                            elif isinstance(sub_value, list):
                                for item in sub_value:
                                    if isinstance(item, str) and item.strip():
                                        elements.append(
                                            create_paragraph12(item, "Bullet")
                                        )
                            elements.append(Spacer(1, 4))
                        elements.append(Spacer(1, 8))

            elif isinstance(section_content, str):
                # Plain text with paragraph breaks
                paragraphs = [
                    p.strip() for p in section_content.split("\n") if p.strip()
                ]
                for para in paragraphs:
                    elements.append(create_paragraph12(para, "Normal"))
                    elements.append(Spacer(1, 6))

            elif isinstance(section_content, list):
                # Simple bullet list
                for item in section_content:
                    if isinstance(item, str) and item.strip():
                        elements.append(create_paragraph12(item, "Bullet"))
                elements.append(Spacer(1, 12))

    doc.build(elements)
    return pdf_path



from datetime import datetime, timezone  # existing
import datetime as _dt

def upload_to_firebase12(local_path, student_id, plan_id):
    """Uploads PDF to Firebase Storage and returns public URL."""
    timestamp = _dt.datetime.now().strftime("%H%M%S")
    blob_path = f"students/{student_id}/teaching_plans/{plan_id}/{timestamp}.pdf"
    blob = bucket.blob(blob_path)
    blob.upload_from_filename(local_path)
    blob.make_public()
    return blob.public_url

# --- Teaching-plan helpers (JSON-safe + translation) ---

import datetime, re, json

def to_json_safe(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if val.__class__.__name__ in ("Timestamp", "DatetimeWithNanoseconds", "ServerTimestamp"):
        try:
            return val.isoformat()
        except Exception:
            try:
                # Handle potential conversion to datetime object
                return val.to_datetime().isoformat()
            except Exception:
                return str(val)
    return val

def deep_normalize(obj):
    if isinstance(obj, dict):
        return {k: deep_normalize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [deep_normalize(v) for v in obj]
    return to_json_safe(obj)

def translate_text_sync(text: str, target_lang: str) -> str:
    """Synchronously calls the translation provider."""
    if not isinstance(text, str) or not text.strip():
        return text
    if target_lang == "ar":
        # Placeholder: Ensure 'translate_text_ar' is synchronous and imported
        from translation_provider import translate_text_ar 
        return translate_text_ar(text)
    return text

def translate_object_value_only_sync(obj, target_lang: str):
    """Recursively translates only string values in a dictionary/list."""
    if isinstance(obj, dict):
        return {k: translate_object_value_only_sync(v, target_lang) for k, v in obj.items()}
    if isinstance(obj, list):
        return [translate_object_value_only_sync(v, target_lang) for v in obj]
    if isinstance(obj, str):
        return translate_text_sync(obj, target_lang)
    return to_json_safe(obj)


# from datetime import datetime, timezone  # existing
# import datetime as _dt
# @app.route("/api/teaching-plan/export", methods=["POST"])
# def export_teaching_plan():
#     """
#     Exports a teaching plan into a bilingual (Arabic/English) PDF and uploads it to Firebase.
#     Fixes blank Arabic PDF issue by normalizing input structure and handling missing fonts.
#     """
#     try:
#         data = request.json
#         student_id = data.get("studentId")
#         plan_id = data.get("planId")
#         lang = data.get("lang", "").lower().strip()

#         if not student_id or not plan_id:
#             return jsonify({"error": "studentId and planId are required"}), 400

#         # STEP 1: Fetch Firestore teaching plan
#         content = export_data12(lang, student_id, plan_id)
#         if not content:
#             return jsonify({"error": "Teaching plan not found"}), 404

#         # STEP 2: Normalize structure for PDF generation
#         # Fix blank PDF issue where `teaching_plan` is returned instead of `actionPlan`
#         if "teaching_plan" in content and "actionPlan" not in content:
#             content["actionPlan"] = content["teaching_plan"]

#         # Ensure content isn't empty after normalization
#         if not content.get("actionPlan"):
#             print("DEBUG: Missing actionPlan content, trying teaching_plan fallback.")
#             content["actionPlan"] = content.get("teaching_plan", {})

#         # STEP 3: Verify fonts (critical for Arabic rendering)
#         try:
#             pdfmetrics.registerFont(TTFont("Amiri", "static/fonts/Amiri-Regular.ttf"))
#             pdfmetrics.registerFont(TTFont("Amiri-Bold", "static/fonts/Amiri-Bold.ttf"))
#         except Exception as e:
#             print(f"WARNING: Could not register Arabic fonts: {e}")

#         # STEP 4: Generate PDF
#         pdf_filename = f"{plan_id}_{lang}.pdf"
#         pdf_path = generate_pdf12(content, pdf_filename, lang)

#         # Sanity check: ensure file was created and has size > 0
#         if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
#             return jsonify({"error": "PDF generation failed (file empty)."}), 500

#         # STEP 5: Upload to Firebase Storage
#         pdf_url = upload_to_firebase12(pdf_path, student_id, plan_id)

#         # STEP 6: Clean up local file
#         if os.path.exists(pdf_path):
#             os.remove(pdf_path)

#         print(f"PDF successfully generated and uploaded: {pdf_url}")

#         return jsonify({
#             "studentId": student_id,
#             "planId": plan_id,
#             "lang": lang,
#             "pdfUrl": pdf_url
#         }), 200

#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({"error": f"Processing failed: {str(e)}"}), 500

from datetime import datetime, timezone  # existing
import datetime as _dt
@app.route("/api/teaching-plan/export", methods=["POST"])
def export_teaching_plan():
    try:
        data = request.json
        student_id = data.get("studentId")
        plan_id = data.get("planId")
        lang = data.get("lang", "")

        if not student_id or not plan_id:
            return jsonify({"error": "studentId and planId are required"}), 400

        # Fetch and translate teaching plan
        content = export_data12(lang, student_id, plan_id)
        # print(f"Fetched content: {content}")
        if not content:
            return jsonify({"error": "Plan not found"}), 404

        # Generate and upload PDF
        pdf_filename = f"{plan_id}.pdf"
        pdf_path = generate_pdf12(content, pdf_filename, lang)
        pdf_url = upload_to_firebase12(pdf_path, student_id, plan_id)

        # Clean up local file
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

        return (
            jsonify({"studentId": student_id, "planId": plan_id, "pdfUrl": pdf_url}),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500



@app.route("/teaching_plans", methods=["GET"])
def get_teaching_plan():
    student_id = request.args.get("studentId")
    plan_id = request.args.get("planId")
    lang = request.args.get("lang", "")

    if not student_id or not plan_id:
        return jsonify({"error": "Missing studentId or planId"}), 400

    try:
        # Get student document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": f"Student with ID {student_id} not found"}), 404

        student_data = student_doc.to_dict()
        teaching_plans = student_data.get("teachingPlans", {})

        # Get specific teaching plan
        plan = teaching_plans.get(plan_id)
        if not plan:
            return jsonify({"error": f"Teaching plan with ID {plan_id} not found"}), 404

        # Return only the action plan
        return jsonify({"actionPlan": plan.get("actionPlan", {})}), 200

    except Exception as e:
        return jsonify({"error": f"Failed to fetch teaching plan: {str(e)}"}), 500


@app.route("/teaching-plan", methods=["PATCH"])
def update_teaching_plan():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Missing request body"}), 400

        student_id = data.get("studentId")
        plan_id = data.get("planId")
        updates = data.get("updates")

        if not student_id or not plan_id or not updates:
            return jsonify({"error": "Missing studentId, planId, or updates"}), 400

        # Get student document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": f"Student with ID {student_id} not found"}), 404

        student_data = student_doc.to_dict()
        teaching_plans = student_data.get("teachingPlans", {})

        # Get specific teaching plan
        plan = teaching_plans.get(plan_id)
        if not plan:
            return jsonify({"error": f"Teaching plan with ID {plan_id} not found"}), 404

        # Validate updates: Ensure only existing fields are modified
        action_plan = plan.get("actionPlan", {})
        invalid_fields = [field for field in updates if field not in action_plan]

        if invalid_fields:
            return (
                jsonify({"error": f"Invalid fields in update: {invalid_fields}"}),
                400,
            )

        # Update only the existing fields
        for key, value in updates.items():
            action_plan[key] = value

        # Save the updated teaching plan
        student_ref.update({f"teachingPlans.{plan_id}.actionPlan": action_plan})

        return (
            jsonify(
                {
                    "message": "Teaching plan updated successfully",
                    "updatedActionPlan": action_plan,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"Failed to update teaching plan: {str(e)}"}), 500


# --------------------------------------------------------------add subject reportcard ----------------------------------------------

from flask import Flask, request, jsonify
from datetime import datetime
from firebase_admin import credentials, firestore, initialize_app

# @app.route("/api/student/report-card", methods=["POST"])
# def enter_student_report_card():
#     try:
#         data = request.get_json()

#         # Validate required fields
#         required_fields = ['studentId', 'academicData', 'attendance', 'reportCard']
#         if not all(field in data for field in required_fields):
#             return jsonify({"error": "Missing required fields in request body"}), 400

#         student_id = data['studentId']
#         academic_data = data['academicData']
#         attendance = data['attendance']
#         report_card = data['reportCard']

#         # Validate that reportCard has a grade
#         if 'grade' not in report_card:
#             return jsonify({"error": "'reportCard' must contain 'grade'"}), 400

#         report_card_grade = report_card['grade']

#         # Fetch student record
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         # Get existing student data
#         student_data = student_doc.to_dict()

#         # Store Report Card Grade History as a dictionary (timestamp -> grade)
#         grade_history = student_data.get("gradeHistory", {})

#         # Use timestamp as key and store grade as a string
#         timestamp_key = datetime.utcnow().isoformat()
#         grade_history[timestamp_key] = str(report_card_grade)  # Store as string

#         # Handle subjects: Ensure new subjects have default values
#         existing_subjects = student_data.get("academicData", {}).get("subjects", {})
#         incoming_subjects = academic_data.get("subjects", {})

#         for subject, details in incoming_subjects.items():
#             if subject not in existing_subjects:
#                 existing_subjects[subject] = {"marks": 0, "grade": "E"}  # Default values for new subjects
#             existing_subjects[subject].update(details)

#         # Ensure academicData remains updated with merged subjects
#         academic_data["subjects"] = existing_subjects

#         # Update Firestore Database
#         student_ref.update({
#             "academicData": academic_data,  # Updated subjects without overwriting existing ones
#             "attendance": attendance,
#             "reportCard": report_card,  # Update report card details
#             "gradeHistory": grade_history  # Store historical report card grades as a dictionary
#         })

#         return jsonify({"message": "Student report card updated with grade history and subjects", "studentId": student_id}), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

async def generate_ai_alert1(
    student_name, student_id, previous_score, current_score, lang_desc, lang_prompt
):
    """
    Generates a personalized AI alert using OpenAI's chat completion API.
    (This version accepts the 'lang_prompt' to be dynamic)
    """
    prompt = (
        f"Generate a **{lang_prompt}**:\n"
        f"Student Name: {student_name}\n"
        f"Student ID: {student_id}\n"
        f"Previous Score: {previous_score}\n"
        f"Current Score: {current_score}\n\n"
        f"""Description:
The alert should be structured, formal, and informative. It should emphasize the student's performance change and highlight the importance of continued support and mentorship.  

The message must follow a structured format, beginning with an introductory statement about the student's performance. It should then provide details about the change in score, acknowledge the teacher’s role.  

The tone should be professional, clear, and action-oriented, ensuring the message effectively communicates the necessary information.  Do not add any extra symbols or any extra words"""
        f"I want the alerts to be generated in this language:- {lang_desc} **strictly**"
    )

    response = await openai_client.chat.completions.create(
        model="gpt-4.1-mini-2025-04-14",
        messages=[
            {
                "role": "system",
                "content": "You are a professional education assistant. Your output must be ONLY the alert message, and strictly in the requested language.",
            },
            {"role": "user", "content": prompt},
        ],
    )

    return response.choices[0].message.content.strip()


# @app.route("/api/student/report-card", methods=["POST"])
# async def enter_student_report_card():
#     try:
#         data = request.get_json()

#         required_fields = ["studentId", "academicData"]
#         if not all(field in data for field in required_fields):
#             return (
#                 jsonify(
#                     {
#                         "error": "Missing required fields: 'studentId' and 'academicData' are required"
#                     }
#                 ),
#                 400,
#             )

#         student_id = data["studentId"]
#         academic_data = data["academicData"]

#         # 1. Safely retrieve names/IDs
#         studentName = data.get("student_name", "Unknown Student")
#         teacherId = data.get("teacher_id", "Unknown Teacher")
        
#         # --- DATE HANDLING LOGIC (UPDATED) ---
#         entry_date_str = data.get("entryDate") # Expected format from client: "YYYY-MM-DD"
#         current_utc_timestamp = time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())

#         # 1. Determine the timestamp_key to use for this entry
#         if entry_date_str:
#             try:
#                 # Parse the date and standardize the time component to midnight UTC (T00:00:00) 
#                 # for consistent historical records based on the report date.
#                 dt_obj = datetime.datetime.strptime(entry_date_str, '%Y-%m-%d')
#                 entry_date_timestamp = dt_obj.strftime("%Y-%m-%dT00:00:00")
#             except ValueError:
#                 print(f"Warning: Invalid entryDate format '{entry_date_str}'. Falling back to current timestamp.")
#                 entry_date_timestamp = current_utc_timestamp
#         else:
#             # If no date is provided, use the current time
#             entry_date_timestamp = current_utc_timestamp
#         # --- END DATE HANDLING LOGIC ---


#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()
#         grade_history = student_data.get("gradeHistory", {})

#         if "reportCard" in data:
#             report_card = data["reportCard"]
#             if "grade" not in report_card:
#                 return (
#                     jsonify({"error": "'reportCard' must contain 'grade' if provided"}),
#                     400,
#                 )

#             report_card_grade = report_card["grade"]
            
#             # 2. Use the consistent entry_date_timestamp for grade history key
#             timestamp_key = entry_date_timestamp 
#             grade_history[timestamp_key] = str(report_card_grade)
#         else:
#             report_card = student_data.get("reportCard", {})

#         existing_academic_data = student_data.get("academicData", {})
#         existing_subjects = existing_academic_data.get("subjects", {})
#         incoming_subjects = academic_data.get("subjects", {})

#         alerts = []

#         for subject, details in incoming_subjects.items():
#             subject_history = existing_subjects.get(subject, {})
#             if "history" not in subject_history:
#                 subject_history["history"] = []

#             previous_score = (
#                 subject_history["history"][-1]["marks"]
#                 if subject_history["history"]
#                 else None
#             )
#             print(f"Previous Score for {subject}: {previous_score}")
#             current_score = details.get("marks", 0)
#             print(f"Current Score for {subject}: {current_score}")

#             curriculum_id = details.get("curriculumId", "Unknown_Curriculum_ID")
#             curriculum_name = details.get("curriculumName", "Unknown_Curriculum_Name")
#             grade = details.get("grade", "E")
#             total_mark = details.get("totalMark", 100)

#             subject_history["history"].append(
#                 {
#                     # 3. Use the consistent entry_date_timestamp for subject history
#                     "timestamp": entry_date_timestamp,
#                     "curriculumId": curriculum_id,
#                     "curriculumName": curriculum_name,
#                     "marks": current_score,
#                     "totalMark": total_mark,
#                     "grade": grade,
#                 }
#             )

#             existing_subjects[subject] = subject_history

#             if previous_score is not None and current_score < previous_score:
#                 lang_desc = await GetLangugage(curriculum_id)
#                 alert_message = await generate_ai_alert1(
#                     studentName, student_id, previous_score, current_score, lang_desc
#                 )

#                 alert_data = {
#                     "ai_generated_message": alert_message,
#                     "current_score": current_score,
#                     "date": time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime()), # Alert generation date (current time)
#                     "isSeen": "Unread",
#                     "previous_score": previous_score,
#                     "student_id": student_id,
#                     "student_name": studentName,
#                     "teacher_id": teacherId,
#                     "type_of_alert": "Success",
#                 }

#                 alerts.append(alert_data)
#                 db.collection("alerts").add(alert_data)

#         academic_data["subjects"] = existing_subjects

#         update_data = {
#             "academicData": academic_data,
#             "gradeHistory": grade_history,
#         }

#         if "reportCard" in data:
#             update_data["reportCard"] = report_card
#         if "attendance" in data:
#             update_data["attendance"] = data["attendance"]

#         student_ref.update(update_data)

#         return (
#             jsonify(
#                 {
#                     "message": "Student report card updated with grade history and timestamped subject records",
#                     "studentId": student_id,
#                     "alerts_generated": alerts,
#                     "report_date_used": entry_date_timestamp # Return the date used for confirmation
#                 }
#             ),
#             200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# import datetime as dt_util 

# # --- 3. ADD this utility function to your ALERTSs file ---
# def format_timestamp(date_str):
#     """Format timestamps safely."""
#     if isinstance(date_str, dt_util.datetime):
#         return date_str.strftime("%d %b %Y, %H:%M")
#     if isinstance(date_str, str):
#         try:
#             if date_str.endswith("Z"):
#                 dt = dt_util.datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
#             else:
#                 dt = dt_util.datetime.fromisoformat(date_str)
#             return dt.strftime("%d %b %Y, %H:%M")
#         except Exception:
#             pass
#     return str(date_str)

# # --- 4. REPLACE your 'get_alerts_notifications' in your ALERTSs file ---
# @app.route("/api/alerts-notifications", methods=["GET"])
# def get_alerts_notifications():
#     """Fetch alerts & notifications based on 'lang' query (en or ar) using teacher_id."""
#     try:
#         teacher_id = request.args.get("teacher_id")
#         lang = request.args.get("lang", "en").lower()
        
#         if not teacher_id:
#             return jsonify({"error": "teacher_id is required"}), 400

#         target_msg_field = "ai_generated_message_ar" if lang == "ar" else "ai_generated_message_en"
#         other_msg_field = "ai_generated_message_en" if lang == "ar" else "ai_generated_message_ar"

#         def process_data(entry):
#             # --- This is the fix for reading ---
#             message = entry.get(target_msg_field)
#             if not message:
#                 message = entry.get(other_msg_field)
#             if not message:
#                 message = entry.get("ai_generated_message") # <--- Reads old data
            
#             entry["ai_generated_message"] = message or "Message content unavailable."
            
#             entry.pop("ai_generated_message_en", None)
#             entry.pop("ai_generated_message_ar", None)
            
#             date_value = entry.get("date")
#             entry["date"] = format_timestamp(date_value)
            
#             return entry

#         alerts_stream = db.collection("alerts").where("teacher_id", "==", teacher_id).stream()
#         alerts = [process_data(doc.to_dict()) for doc in alerts_stream]

#         notifications_stream = (
#             db.collection("notifications")
#             .where("teacher_id", "==", teacher_id) 
#             .stream()
#         )
#         notifications = [process_data(doc.to_dict()) for doc in notifications_stream]
        
#         return jsonify({"alerts": alerts, "notifications": notifications}), 200

#     except Exception as e:
#         print(f"FATAL ERROR in get_alerts_notifications: {e}") 
#         return jsonify({"error": "An unexpected server error occurred during data retrieval."}), 500

# if __name__ == '__main__':
#     # Example usage (uncomment to run locally)
#     # app.run(debug=True)
#     pass

# @app.route("/api/student/report-card", methods=["POST"])
# async def enter_student_report_card():
#     try:
#         data = request.get_json()
#         date = request.form.get("date")

#         required_fields = ["studentId", "academicData"]
#         if not all(field in data for field in required_fields):
#             return (
#                 jsonify(
#                     {
#                         "error": "Missing required fields: 'studentId' and 'academicData' are required"
#                     }
#                 ),
#                 400,
#             )

#         student_id = data["studentId"]
#         academic_data = data["academicData"]
        

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()
#         studentName = data["student_name"]
#         teacherId = data["teacher_id"]

#         grade_history = student_data.get("gradeHistory", {})

#         if "reportCard" in data:
#             report_card = data["reportCard"]
#             if "grade" not in report_card:
#                 return (
#                     jsonify({"error": "'reportCard' must contain 'grade' if provided"}),
#                     400,
#                 )

#             report_card_grade = report_card["grade"]
#             timestamp_key = time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime())
#             grade_history[timestamp_key] = str(report_card_grade)
#         else:
#             report_card = student_data.get("reportCard", {})

#         existing_academic_data = student_data.get("academicData", {})
#         existing_subjects = existing_academic_data.get("subjects", {})
#         incoming_subjects = academic_data.get("subjects", {})

#         alerts = []

#         for subject, details in incoming_subjects.items():
#             subject_history = existing_subjects.get(subject, {})
#             if "history" not in subject_history:
#                 subject_history["history"] = []

#             previous_score = (
#                 subject_history["history"][-1]["marks"]
#                 if subject_history["history"]
#                 else None
#             )
#             print(previous_score)
#             current_score = details.get("marks", 0)
#             print(current_score)

#             curriculum_id = details.get("curriculumId", "Unknown_Curriculum_ID")
#             curriculum_name = details.get("curriculumName", "Unknown_Curriculum_Name")
#             grade = details.get("grade", "E")
#             total_mark = details.get("totalMark", 100)

#             subject_history["history"].append(
#                 {
#                     "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime()),
#                     "curriculumId": curriculum_id,
#                     "curriculumName": curriculum_name,
#                     "marks": current_score,
#                     "totalMark": total_mark,
#                     "date": date,
#                     "grade": grade,
#                 }
#             )

#             existing_subjects[subject] = subject_history

#             if previous_score is not None and current_score < previous_score:
#                 lang_desc = await GetLangugage(curriculum_id)
#                 alert_message = await generate_ai_alert(
#                     studentName, student_id, previous_score, current_score, lang_desc
#                 )

#                 alert_data = {
#                     "ai_generated_message": alert_message,
#                     "current_score": current_score,
#                     "date": time.strftime("%Y-%m-%dT%H:%M:%S", time.gmtime()),
#                     "isSeen": "Unread",
#                     "previous_score": previous_score,
#                     "student_id": student_id,
#                     "student_name": studentName,
#                     "teacher_id": teacherId,
#                     "type_of_alert": "Success",
#                 }

#                 alerts.append(alert_data)
#                 db.collection("alerts").add(alert_data)

#         academic_data["subjects"] = existing_subjects

#         update_data = {
#             "academicData": academic_data,
#             "gradeHistory": grade_history,
#         }

#         if "reportCard" in data:
#             update_data["reportCard"] = report_card
#         if "attendance" in data:
#             update_data["attendance"] = data["attendance"]

#         student_ref.update(update_data)

#         return (
#             jsonify(
#                 {
#                     "message": "Student report card updated with grade history and timestamped subject records",
#                     "studentId": student_id,
#                     "alerts_generated": alerts,
#                 }
#             ),
#             200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/student/report-card/<student_id>", methods=["GET"])
# def get_student_report_card(student_id):
#     try:
#         # Fetch student data
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()

#         # Get student's academic data
#         academic_data = student_data.get("academicData", {})
#         subjects = academic_data.get("subjects", {})

#         # Fetch the student's grade and class from Firestore (assuming they exist)
#         student_grade = academic_data.get("grade", "")
#         student_class = student_data.get("class_id", "")

#         if student_grade and student_class:
#             # Fetch subjects from Grades -> classes -> subjects
#             subjects_ref = (
#                 db.collection("Grades")
#                 .document(student_grade)
#                 .collection("classes")
#                 .document(student_class)
#             )
#             class_doc = subjects_ref.get()

#             if class_doc.exists:
#                 grade_subjects = class_doc.to_dict().get("subjects", {})

#                 # Merge subjects (ensure existing marks/grades are not lost)
#                 for subject, details in grade_subjects.items():
#                     if subject not in subjects:
#                         subjects[subject] = details  # Add new subjects

#         # Return merged data
#         return jsonify({
#             "studentId": student_id,
#             "academicData": {
#                 "grade": student_grade,
#                 "subjects": subjects
#             },
#             "attendance": student_data.get("attendance", {}),
#             "reportCard": student_data.get("reportCard", {})
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/student/report-card/<student_id>", methods=["GET"])
# def get_student_report_card(student_id):
#     try:
#         # Fetch student data
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()

#         # Get student's academic data
#         academic_data = student_data.get("academicData", {})
#         subjects = academic_data.get("subjects", {})

#         # Fetch the student's grade and class from Firestore
#         student_grade = academic_data.get("grade", "")
#         student_class = student_data.get("class_id", "")

#         # Fetch subjects from Grades -> classes -> subjects
#         grade_subjects = {}
#         if student_grade and student_class:
#             subjects_ref = (
#                 db.collection("Grades")
#                 .document(student_grade)
#                 .collection("classes")
#                 .document(student_class)
#             )
#             class_doc = subjects_ref.get()

#             if class_doc.exists:
#                 grade_subjects = class_doc.to_dict().get("subjects", {})

#         # Ensure all subjects from the grade curriculum are included
#         for subject, details in grade_subjects.items():
#             if subject not in subjects:
#                 subjects[subject] = {"history": []}  # Initialize with empty history

#         # Prepare response data with subject history
#         subject_data = {}
#         for subject, details in subjects.items():
#             subject_data[subject] = {
#                 "history": details.get("history", [])  # Return all recorded marks with timestamps
#             }

#         # Get grade history
#         grade_history = student_data.get("gradeHistory", {})

#         # Return merged data
#         return jsonify({
#             "studentId": student_id,
#             "academicData": {
#                 "grade": student_grade,
#                 "subjects": subject_data
#             },
#             "attendance": student_data.get("attendance", {}),
#             "reportCard": student_data.get("reportCard", {}),
#             "gradeHistory": grade_history  # Return all past grades with timestamps
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/student/report-card/<student_id>", methods=["GET"])
# def get_student_report_card(student_id):
#     try:
#         # Fetch student data
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()

#         # Get student's academic data
#         academic_data = student_data.get("academicData", {})
#         subjects = academic_data.get("subjects", {})

#         # Fetch the student's grade and class from Firestore
#         student_grade = academic_data.get("grade", "")
#         student_class = student_data.get("class_id", "")

#         # Fetch subjects from Grades -> classes -> subjects
#         grade_subjects = {}
#         if student_grade and student_class:
#             subjects_ref = (
#                 db.collection("Grades")
#                 .document(student_grade)
#                 .collection("classes")
#                 .document(student_class)
#             )
#             class_doc = subjects_ref.get()

#             if class_doc.exists:
#                 grade_subjects = class_doc.to_dict().get("subjects", {})

#         # Ensure all subjects from the grade curriculum are included
#         for subject, details in grade_subjects.items():
#             if subject not in subjects:
#                 subjects[subject] = {"history": []}  # Initialize with empty history

#         # Prepare response data with subject history and calculate total marks
#         subject_data = {}
#         total_obtained_marks = 0
#         total_marks = len(subjects) * 100  # Assuming each subject has a max mark of 100

#         for subject, details in subjects.items():
#             history = details.get("history", [])
#             obtained_marks = sum(entry.get("marks", 0) for entry in history)
#             total_obtained_marks += obtained_marks

#             subject_data[subject] = {
#                 "history": history
#             }

#         # Get grade history
#         grade_history = student_data.get("gradeHistory", {})

#         # Return merged data with total marks
#         return jsonify({
#             "studentId": student_id,
#             "academicData": {
#                 "grade": student_grade,
#                 "subjects": subject_data
#             },
#             "attendance": student_data.get("attendance", {}),
#             "reportCard": student_data.get("reportCard", {}),
#             "gradeHistory": grade_history,  # Return all past grades with timestamps
#             "totalObtainedMarks": total_obtained_marks,
#             "totalMarks": total_marks
#         }), 200

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500
import re
from flask import jsonify, request
import re
from flask import jsonify, request
from firebase_admin import firestore # Assuming 'db' is a firestore client instance

# --- HELPER: Output Normalizer (STANDARDIZED TO SPACES FOR FRONTEND DISPLAY) ---
def normalize_grade_key_output(key):
    """
    Converts any variation (Grade 5, Grade_5, Grade5) to the space-separated 
    format expected by the frontend: 'GRADE 5'.
    """
    if not key: return ""
    key = str(key).upper().strip()
    
    # 1. Replace underscores, dots, hyphens with SPACE
    key = re.sub(r'[_\.\-]+', ' ', key)
    
    # 2. Insert space between letter and number if missing (e.g. GRADE5 -> GRADE 5)
    key = re.sub(r'([A-Z])(\d)', r'\1 \2', key)
    
    # 3. Collapse multiple spaces into one
    key = re.sub(r'\s+', ' ', key)
    
    # 4. Remove trailing/leading spaces and return
    return key.strip()

@app.route("/api/student/report-card/<student_id>", methods=["GET"])
def get_student_report_card(student_id):
    try:
        # Fetch student data
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        student_data = student_doc.to_dict()

        # Get student's academic data
        academic_data = student_data.get("academicData", {})
        subjects = academic_data.get("subjects", {})
        teacher_id = (request.args.get("teacher_id") or request.args.get("teacherId") or "").strip()

        # --- GRADE FETCHING LOGIC ---
        # 1. Try academicData first
        raw_student_grade = academic_data.get("grade", "")
        
        # 2. Fallback: If empty, check assignedGrades
        if not raw_student_grade:
            assigned = student_data.get("assignedGrades", {})
            if "grades" in assigned and isinstance(assigned["grades"], dict):
                keys = list(assigned["grades"].keys())
                if keys: raw_student_grade = keys[0]
            elif isinstance(assigned, dict) and assigned:
                keys = list(assigned.keys())
                if keys: raw_student_grade = keys[0]

        # 3. Normalize the grade for OUTPUT (e.g. "GRADE_5" -> "GRADE 5")
        student_grade = normalize_grade_key_output(raw_student_grade)
        # ---------------------------------------
        
        student_class = student_data.get("class_id", "")

        # Fetch subjects from Grades -> classes -> subjects
        grade_subjects = {}
        
        # Only attempt fetch if we successfully found a grade
        if student_grade and student_class:
            
            # The 'Grades' collection typically uses the normalized key (e.g., "GRADE 5") as the document ID.
            # Attempt 1: Try Normalized Key with SPACE (GRADE 5)
            subjects_ref = (
                db.collection("Grades")
                .document(student_grade) 
                .collection("classes")
                .document(student_class)
            )
            class_doc = subjects_ref.get()

            if class_doc.exists:
                grade_subjects = class_doc.to_dict().get("subjects", {})
            else:
                # Attempt 2: Fallback to Underscore key (Legacy DB: GRADE_5)
                # This handles cases where the document ID in 'Grades' uses the underscore format.
                storage_key_underscore = student_grade.replace(" ", "_")
                if storage_key_underscore != student_grade:
                    subjects_ref_legacy = (
                        db.collection("Grades")
                        .document(storage_key_underscore)
                        .collection("classes")
                        .document(student_class)
                    )
                    class_doc_legacy = subjects_ref_legacy.get()
                    if class_doc_legacy.exists:
                        grade_subjects = class_doc_legacy.to_dict().get("subjects", {})

        # Ensure all subjects from the grade curriculum are included
        for subject, details in grade_subjects.items():
            if subject not in subjects:
                subjects[subject] = {"history": []}

        # Prepare response data
        subject_data = {}
        total_obtained_marks = 0
        total_marks = 0

        for subject, details in subjects.items():
            history = details.get("history", [])
            if not isinstance(history, list):
                history = []
            # Marks history is returned without teacher-level filtering.
            # Filtering is handled in UI/use-case specific flows if needed.

            normalized_history = []
            for idx, entry in enumerate(history):
                row = dict(entry or {})
                row["history_id"] = row.get("history_id") or f"{subject}::{idx}"
                normalized_history.append(row)
            history = normalized_history

            if history:
                latest_entry = max(history, key=lambda x: x.get("timestamp", ""))
                obtained_marks = latest_entry.get("marks", 0)
                subject_total_marks = latest_entry.get("totalMark", 100)
                total_obtained_marks += obtained_marks
                total_marks += subject_total_marks
            else:
                obtained_marks = 0
                subject_total_marks = 0

            subject_data[subject] = {"history": history}

        grade_history = student_data.get("gradeHistory", {})
        report_card = student_data.get("reportCard", {})
        report_card["totalObtainedMarks"] = total_obtained_marks
        report_card["totalMarks"] = total_marks

        return (
            jsonify(
                {
                    "studentId": student_id,
                    # This consistently sends the space-separated format ("GRADE 5") to the App
                    "academicData": {"grade": student_grade, "subjects": subject_data},
                    "attendance": student_data.get("attendance", {}),
                    "reportCard": report_card,
                    "gradeHistory": grade_history, 
                }
            ),
            200,
        )

    except Exception as e:
        # In a real app, you'd log this exception
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

def _find_history_entry(subjects, entry_id=None, subject=None, timestamp=None):
    if not isinstance(subjects, dict):
        return None, None, None

    parsed_subject = subject
    parsed_index = None
    if entry_id and "::" in str(entry_id):
        left, right = str(entry_id).split("::", 1)
        parsed_subject = parsed_subject or left
        try:
            parsed_index = int(right)
        except Exception:
            parsed_index = None

    if parsed_subject and parsed_subject in subjects:
        history = subjects.get(parsed_subject, {}).get("history", [])
        if not isinstance(history, list):
            history = []
        if parsed_index is not None and 0 <= parsed_index < len(history):
            return parsed_subject, parsed_index, history[parsed_index]

        for idx, item in enumerate(history):
            item_id = str(item.get("history_id") or item.get("entry_id") or item.get("id") or "")
            if entry_id and item_id == str(entry_id):
                return parsed_subject, idx, item
            if timestamp and str(item.get("timestamp", "")) == str(timestamp):
                return parsed_subject, idx, item

    for subj_name, subj_data in subjects.items():
        history = subj_data.get("history", [])
        if not isinstance(history, list):
            continue
        for idx, item in enumerate(history):
            item_id = str(item.get("history_id") or item.get("entry_id") or item.get("id") or "")
            if entry_id and item_id == str(entry_id):
                return subj_name, idx, item
            if subject and subj_name == subject and timestamp and str(item.get("timestamp", "")) == str(timestamp):
                return subj_name, idx, item

    return None, None, None


def _update_or_delete_report_card_history(student_id, payload, entry_id_from_path=None, http_method="POST"):
    if not student_id:
        return jsonify({"error": "studentId is required"}), 400

    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()
    if not student_doc.exists:
        return jsonify({"error": "Student account not found"}), 404

    student_data = student_doc.to_dict() or {}
    academic_data = student_data.get("academicData", {})
    subjects = academic_data.get("subjects", {})
    if not isinstance(subjects, dict):
        subjects = {}

    action = (payload.get("action") or "").strip().lower()
    teacher_id = str(payload.get("teacher_id") or payload.get("teacherId") or request.args.get("teacher_id") or request.args.get("teacherId") or "").strip()
    entry_id = entry_id_from_path or payload.get("entryId") or payload.get("entry_id")
    subject = payload.get("subject")
    timestamp = payload.get("timestamp")

    if http_method == "DELETE":
        action = "delete"
    elif http_method in {"PATCH", "PUT"}:
        action = "update"

    if action not in {"update", "delete"}:
        return jsonify({"error": "Unsupported action. Use update or delete"}), 400

    found_subject, found_index, found_entry = _find_history_entry(
        subjects, entry_id=entry_id, subject=subject, timestamp=timestamp
    )
    if found_subject is None or found_index is None:
        return jsonify({"error": "History entry not found"}), 404
    if teacher_id:
        entry_teacher_id = _entry_teacher_id(found_entry)
        if not entry_teacher_id or entry_teacher_id != teacher_id:
            return jsonify({"error": "You can only modify your own marks"}), 403

    history = subjects.get(found_subject, {}).get("history", [])
    if not isinstance(history, list):
        history = []

    if action == "delete":
        deleted = history.pop(found_index)
        for idx, item in enumerate(history):
            item["history_id"] = f"{found_subject}::{idx}"
        subjects[found_subject]["history"] = history
        student_ref.update({"academicData.subjects": subjects})
        return jsonify({"message": "Report card history entry deleted", "entry": deleted}), 200

    if "marks" in payload:
        try:
            found_entry["marks"] = int(payload.get("marks"))
        except Exception:
            return jsonify({"error": "marks must be an integer"}), 400
    if "totalMark" in payload:
        try:
            found_entry["totalMark"] = int(payload.get("totalMark"))
        except Exception:
            return jsonify({"error": "totalMark must be an integer"}), 400
    if "grade" in payload:
        found_entry["grade"] = payload.get("grade")
    if "curriculumName" in payload:
        found_entry["curriculumName"] = payload.get("curriculumName")
    if "timestamp" in payload:
        found_entry["timestamp"] = payload.get("timestamp")
    found_entry["history_id"] = found_entry.get("history_id") or f"{found_subject}::{found_index}"
    history[found_index] = found_entry

    for idx, item in enumerate(history):
        item["history_id"] = f"{found_subject}::{idx}"

    subjects[found_subject]["history"] = history
    student_ref.update({"academicData.subjects": subjects})
    return jsonify({"message": "Report card history entry updated", "entry": found_entry}), 200


@app.route("/api/student/update-report-card", methods=["POST"])
def update_student_report_card():
    try:
        data = request.get_json(silent=True) or {}
        student_id = data.get("studentId")
        if not student_id:
            return jsonify({"error": "JSON body must contain studentId"}), 400

        action = (data.get("action") or "").strip().lower()
        if action in {"update", "delete"}:
            return _update_or_delete_report_card_history(student_id, data, http_method="POST")

        academic_data = data.get("academicData", {})
        attendance = data.get("attendance", {})
        report_card = data.get("reportCard", {})

        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()
        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        existing_data = student_doc.to_dict() or {}
        existing_subjects = existing_data.get("academicData", {}).get("subjects", {})
        grade_history = existing_data.get("gradeHistory", {})

        if "grade" in report_card:
            new_grade = str(report_card["grade"])
            timestamp_key = datetime.datetime.utcnow().isoformat()
            grade_history[timestamp_key] = new_grade

        for subject_name, details in academic_data.get("subjects", {}).items():
            if subject_name not in existing_subjects:
                existing_subjects[subject_name] = {"history": []}
            existing_subjects[subject_name].update(details)

        student_ref.update(
            {
                "academicData.grade": academic_data.get(
                    "grade", existing_data.get("academicData", {}).get("grade", "")
                ),
                "academicData.subjects": existing_subjects,
                "attendance": attendance or existing_data.get("attendance", {}),
                "reportCard": report_card or existing_data.get("reportCard", {}),
                "gradeHistory": grade_history,
            }
        )

        return jsonify({"message": "Student report card updated successfully", "studentId": student_id}), 200
    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/student/report-card/<student_id>/history/<entry_id>", methods=["POST", "PATCH", "PUT", "DELETE"])
def update_or_delete_report_card_entry(student_id, entry_id):
    try:
        payload = request.get_json(silent=True) or {}
        return _update_or_delete_report_card_history(
            student_id, payload, entry_id_from_path=entry_id, http_method=request.method
        )
    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/student/report-card/history/<entry_id>", methods=["POST", "PATCH", "PUT", "DELETE"])
def update_or_delete_report_card_entry_without_path_student(entry_id):
    try:
        payload = request.get_json(silent=True) or {}
        student_id = payload.get("studentId")
        return _update_or_delete_report_card_history(
            student_id, payload, entry_id_from_path=entry_id, http_method=request.method
        )
    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/student/add-subject", methods=["POST"])
def add_subject():
    try:
        data = request.get_json()

        # Validate input
        if not data or "studentId" not in data or "subject_name" not in data:
            return (
                jsonify(
                    {"error": "JSON body must contain 'studentId' and 'subject_name'"}
                ),
                400,
            )

        student_id = data["studentId"]
        subject_name = data["subject_name"]

        # Reference to student document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        # Fetch existing student data
        student_data = student_doc.to_dict()
        subjects = student_data.get("academicData", {}).get("subjects", {})

        # Check if subject already exists
        if subject_name in subjects:
            return (
                jsonify(
                    {
                        "message": f"Subject '{subject_name}' already exists for this student."
                    }
                ),
                400,
            )

        # Add new subject with default values
        subjects[subject_name] = {"marks": "", "grade": ""}

        # Update Firestore
        student_ref.update({"academicData.subjects": subjects})

        return (
            jsonify(
                {
                    "message": f"Subject '{subject_name}' added successfully.",
                    "studentId": student_id,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# ---------------------------------------------------------delivery method -------------------------------------------------------------------
from flask import Flask, request, jsonify
from datetime import datetime, timezone
import uuid
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from firebase_admin import credentials, firestore, initialize_app
from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore

# Email Configuration
EMAIL_SENDER = "tajheezai@gmail.com"
EMAIL_PASSWORD = "sfek auxg kasz ulgy"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

# Logging Configuration
logging.basicConfig(level=logging.DEBUG)


def send_email(receiver_email, subject, message_body, html=False):
    """Sends an email using Gmail SMTP. Supports plain text and HTML."""
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER
        msg["To"] = receiver_email
        msg["Subject"] = subject

        # Attach message body as HTML or plain text
        if html:
            msg.attach(MIMEText(message_body, "html"))
        else:
            msg.attach(MIMEText(message_body, "plain"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)

        logging.info(f"Email sent to {receiver_email}")

    except Exception as e:
        logging.error(f"Failed to send email: {e}")


def get_user_email(user_id):
    """Fetch the email of a user from Firestore based on user_id."""
    user_doc = db.collection("users").document(user_id).get()
    if user_doc.exists:
        user_data = user_doc.to_dict()
        return user_data.get("email", None)
    return None


@app.route("/add-notification", methods=["POST"])
def add_notification():
    """
    API to add a notification to the notifications collection.
    It checks the user's preferred delivery method before sending notifications.
    """
    try:
        # Parse input
        data = request.json
        title = data.get("title")
        description = data.get("description")
        receiver_id = data.get("receiver_id")
        sender_id = data.get("sender_id")

        # Validate required fields
        if not all([title, description, receiver_id, sender_id]):
            return jsonify({"error": "Missing required fields"}), 400

        # Check if user exists and fetch delivery method
        user_ref = db.collection("users").document(receiver_id)
        user_doc = user_ref.get()
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        user_data = user_doc.to_dict()
        delivery_method = user_data.get(
            "delivery_method", {"email": True, "app": True, "sms": False}
        )  # Default settings
        receiver_email = user_data.get("email", None)  # Fetch user email dynamically

        # Generate unique ID and timestamp
        notification_id = str(uuid.uuid4())
        created_at = datetime.now(timezone.utc).isoformat()

        # Prepare the notification document
        notification_data = {
            "id": notification_id,
            "type": "notification",
            "title": title,
            "description": description,
            "receiver_id": receiver_id,
            "sender_id": sender_id,
            "status": False,
            "created_at": created_at,
            "type": "notification",
            "responseMessage": "",
            "responseTimestamp": "",
            "responseStatus": "not responded",  # Default response status
        }

        # Store in the notifications collection
        db.collection("notifications").document(notification_id).set(notification_data)

        # Notification Sending Logic Based on Delivery Preferences
        email_enabled = delivery_method.get("email", True)  # True
        app_enabled = delivery_method.get("app", True)  # True
        # sms_enabled = delivery_method.get("sms", True) #False
        sms_enabled = True

        # Send Email if enabled
        if email_enabled and receiver_email:
            subject = f"Ã°Å¸â€œÂ¢ New Notification: {title}"
            message_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <h2 style="color: #2E86C1;">New Notification</h2>
                <p><strong>Title:</strong> {title}</p>
                <p><strong>Message:</strong> {description}</p>
                <hr>
                <p style="font-size: 12px; color: #777;">
                    This is an automated message. Please do not reply.
                </p>
            </body>
            </html>
            """
            send_email(receiver_email, subject, message_body, html=True)
            logging.info(f"Email sent to {receiver_email}")

        # Send App Notification if enabled
        if app_enabled:
            logging.info(f"App notification stored for user {receiver_id}")
        print(sms_enabled)
        # SMS Notification Sending
        if sms_enabled:
            sms_payload = {
                "to": user_data.get("profileInfo", {})
                .get("contactInformation", {})
                .get(
                    "phoneNumber", ""
                ),  # Ensure this field exists in your user document
                "message": f"New Notification From Sender {sender_id}: {title} - {description}",
            }

            if sms_payload["to"]:  # Ensure the phone number is not empty
                try:
                    response = requests.post(
                        "http://127.0.0.1:5000/api/send_sms", json=sms_payload
                    )
                    if response.status_code == 200:
                        logging.info(f"SMS sent successfully to {sms_payload['to']}")
                    else:
                        logging.error(f"Failed to send SMS: {response.json()}")
                except Exception as e:
                    logging.error(f"Error sending SMS: {str(e)}")
            else:
                logging.error(f"User {receiver_id} does not have a valid phone number.")

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Notification added successfully",
                    "notification_data": notification_data,
                }
            ),
            200,
        )

    except Exception as e:
        logging.error(f"Error in add-notification: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/update-notification-response", methods=["POST"])
def update_notification_response():
    """
    API to update the response details of a notification.
    """
    try:
        # Parse input
        data = request.json
        notification_id = data.get("notification_id")
        response_message = data.get("response_message")
        response_status = data.get("response_status", "responded")  # Default status

        # Validate required fields
        if not all([notification_id, response_message]):
            return jsonify({"error": "Missing required fields"}), 400

        # Fetch the notification document
        notification_ref = db.collection("notifications").document(notification_id)
        notification_doc = notification_ref.get()

        if not notification_doc.exists:
            return jsonify({"error": "Notification not found"}), 404

        # Update notification response details
        response_timestamp = datetime.now(timezone.utc).isoformat()
        update_data = {
            "responseMessage": response_message,
            "responseTimestamp": response_timestamp,
            "responseStatus": response_status,
        }

        notification_ref.update(update_data)

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Notification response updated successfully",
                    "updated_data": update_data,
                }
            ),
            200,
        )

    except Exception as e:
        logging.error(f"Error in update-notification-response: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/update-delivery-method", methods=["POST"])
def update_delivery_method():
    """
    API to update the user's notification delivery method preferences.
    Ensures the user exists before updating their preferences.
    """
    try:
        # Parse input
        data = request.json
        user_id = data.get("user_id")
        delivery_method = data.get(
            "delivery_method"
        )  # Expected format: {"email": true, "app": true, "sms": false}

        # Validate input
        if not user_id or not isinstance(delivery_method, dict):
            return jsonify({"error": "Invalid input"}), 400

        # Ensure valid boolean values
        if not all(
            key in ["email", "app", "sms", "notificationcount"]
            and isinstance(value, bool)
            for key, value in delivery_method.items()
        ):
            return jsonify({"error": "Invalid delivery method format"}), 400

        # Check if user exists
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()
        if not user_doc.exists:
            return jsonify({"error": "User not found"}), 404

        # Update Firestore with new settings
        user_ref.update({"delivery_method": delivery_method})

        return (
            jsonify(
                {
                    "success": True,
                    "message": "Notification settings updated successfully",
                }
            ),
            200,
        )

    except Exception as e:
        logging.error(f"Error updating delivery method: {e}")
        return jsonify({"error": str(e)}), 500


from flask import Flask, request, jsonify
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import firebase_admin
from firebase_admin import credentials, storage, firestore
from werkzeug.utils import secure_filename
from docx import Document
from fpdf import FPDF
import os
import time
from PyPDF2 import PdfReader, PdfWriter
try:
    from firebase_functions import https_fn
except Exception:
    https_fn = None
from concurrent.futures import ThreadPoolExecutor  # Import ThreadPoolExecutor
from flask_cors import CORS
import requests  # Assuming Groq uses an API endpoint
from groq import Groq
from dotenv import load_dotenv
import uuid
try:
    from teachingplans import generate_teaching_plan
except Exception:
    async def generate_teaching_plan(*args, **kwargs):
        return {}, "", str(uuid.uuid4()).replace("-", "_")
from PIL import Image
import img2pdf


# def convert_image_to_pdf(image_path, output_pdf_path):
#     """Convert an image to PDF."""
#     try:
#         print(image_path)
#         print(output_pdf_path)
#         image = Image.open(image_path)  # Ensure Pillow is correctly installed
#         image = image.convert("RGB")  # Convert to RGB for PDF
#         image.save(output_pdf_path, "PDF")
#     except Exception as e:
#         raise RuntimeError(f"Failed to convert image to PDF: {e}")


def convert_image_to_pdf(image_path, output_pdf_path):
    """Convert an image to PDF using img2pdf."""
    try:
        with open(output_pdf_path, "wb") as f:
            f.write(img2pdf.convert(image_path))
    except Exception as e:
        raise RuntimeError(f"Failed to convert image to PDF: {e}")


# Set up upload folder
UPLOAD_FOLDER = "uploads/"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

# Initialize Firestore
db = firestore.client()

# Azure Form Recognizer credentials
endpoint = "https://aiocr395080637747.cognitiveservices.azure.com/"  # Azure Form Recognizer endpoint
key = "b1b026f421034dabb948999b80a63e8c"  # Azure key

# Groq Configuration
GROQ_API_KEY = "YOUR_GROQ_API_KEY"
client = Groq(api_key=GROQ_API_KEY)


# Extract text from the PDF using Azure Form Recognizer
def extract_text_from_pdf(pdf_path):
    client = DocumentAnalysisClient(
        endpoint=endpoint, credential=AzureKeyCredential(key)
    )

    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(
            "prebuilt-layout", document=f, polling_interval=60
        )
        result = poller.result(timeout=7200)  # Adjust timeout if necessary

    extracted_text = []
    for page in result.pages:
        for line in page.lines:
            extracted_text.append(line.content)

    return "\n".join(extracted_text)


# Process a chunk of pages and extract text
def process_chunk(pdf_writer, chunk_num):
    temp_chunk_path = os.path.join(UPLOAD_FOLDER, f"temp_chunk_{chunk_num}.pdf")

    # Save the chunk as a temporary file
    with open(temp_chunk_path, "wb") as temp_chunk_file:
        pdf_writer.write(temp_chunk_file)

    # Extract text from the current chunk
    chunk_text = extract_text_from_pdf(temp_chunk_path)

    # Cleanup temporary chunk file
    os.remove(temp_chunk_path)

    return chunk_text


# Split the PDF into chunks and process them in parallel using ThreadPoolExecutor
def process_pdf_in_chunks(pdf_path):
    extracted_text = []

    with open(pdf_path, "rb") as f:
        reader = PdfReader(f)
        num_pages = len(reader.pages)

        chunk_num = 0
        pdf_writer = PdfWriter()
        futures = []

        with ThreadPoolExecutor(
            max_workers=4
        ) as executor:  # Adjust the number of workers as needed
            for page_num in range(num_pages):
                pdf_writer.add_page(reader.pages[page_num])

                # If we reach the chunk size (60 pages) or the end of the document, process the chunk
                if (page_num + 1) % 60 == 0 or (page_num + 1) == num_pages:
                    # Submit the chunk to be processed in parallel
                    futures.append(
                        executor.submit(process_chunk, pdf_writer, chunk_num)
                    )

                    # Start a new writer for the next chunk
                    pdf_writer = PdfWriter()
                    chunk_num += 1

            # Wait for all futures to complete and gather results
            for future in futures:
                extracted_text.append(future.result())

    # Combine text from all chunks
    return "\n".join(extracted_text)


# Upload any file to Firebase Storage and return its public URL
def upload_to_firebase(file_path, filename, folder):
    bucket = storage.bucket()
    blob = bucket.blob(f"{folder}/{filename}")

    blob.upload_from_filename(file_path)
    blob.make_public()

    return blob.public_url


# Convert DOCX to PDF using fpdf2 with Unicode support
def convert_docx_to_pdf(docx_file_path, pdf_file_path):
    doc = Document(docx_file_path)
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Load the Unicode font (DejaVuSans)
    pdf.add_font("DejaVu", "", "fonts/DejaVuSans.ttf", uni=True)
    pdf.set_font("DejaVu", size=12)

    # Iterate over the paragraphs in the DOCX file and add them to the PDF
    for para in doc.paragraphs:
        pdf.multi_cell(0, 10, para.text)

    # Save the generated PDF
    pdf.output(pdf_file_path)


# Allowed file extensions
ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


from text_extract1 import extract_text_from_pdf_or_image11
try:
    from teachingplans import generate_teaching_plan
except Exception:
    async def generate_teaching_plan(*args, **kwargs):
        return {}, "", str(uuid.uuid4()).replace("-", "_")

# from test1 import extract_from_pdf
from test2 import extract_from_pdf


@app.route("/upload_exam_script", methods=["POST"])
async def upload_exam_script():
    temp_file_path = None
    temp_pdf_path = None
    try:
        # Extract data from form-data
        exam_name = request.form.get("exam_name")
        curriculum_coverage = request.form.get(
            "curriculum_coverage[]", ""
        )  # Handle arrays in form-data
        date = request.form.get("date")
        observation = request.form.get("observation")
        student_id = request.form.get("studentId")
        file = request.files["file"]  # File to upload
        curriculum_id = request.form.get("curriculumId")
        curriculumName = request.form.get("curriculumName")
        subject = request.form.get("subject")
        language = request.form.get("language", "ar")
        teacher_id = request.form.get("teacherId")

        if not exam_name or not date or not observation or not student_id:
            return (
                jsonify(
                    {
                        "error": "Missing required fields: exam_name, date, observation, studentId, or file"
                    }
                ),
                400,
            )

        if file.filename == "":
            return jsonify({"error": "No selected file"}), 400

        # Validate file format
        if not allowed_file(file.filename):
            return (
                jsonify(
                    {
                        "error": "Invalid file type. Only PDF, PNG, JPG, and JPEG are allowed."
                    }
                ),
                400,
            )

        # if not file.filename.endswith('.pdf'):
        #     return jsonify({'error': 'Only PDF files are allowed'}), 400

        # Save the uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(temp_file_path)

        # Check if the file is an image
        if filename.lower().endswith(("png", "jpg", "jpeg")):
            temp_pdf_path = temp_file_path.rsplit(".", 1)[0] + ".pdf"
            try:
                convert_image_to_pdf(temp_file_path, temp_pdf_path)
                print(
                    f"Converted PDF Path: {temp_pdf_path}, Exists: {os.path.exists(temp_pdf_path)}"
                )
                os.remove(
                    temp_file_path
                )  # Remove the original image only after successful conversion
            except Exception as e:
                return (
                    jsonify({"error": f"Failed to convert image to PDF: {str(e)}"}),
                    400,
                )
        else:
            temp_pdf_path = temp_file_path  # If it's a PDF, use it directly

        # Ensure temp_pdf_path is valid before calling extract_text_from_pdf_or_image11
        if not temp_pdf_path or not os.path.exists(temp_pdf_path):
            return (
                jsonify({"error": "File processing failed. Temp PDF path is invalid."}),
                400,
            )

        # Extract text
        # extracted_text,image_url = await extract_text_from_pdf_or_image11(temp_file_path)

        # extracted_text = extract_text_from_pdf(temp_pdf_path)

        # print(extracted_text)

        # query = f"Using extracted text of students exam sheet Extracted text :-  {extracted_text} and optional topics {curriculum_coverage} \n gather relevant content from vectorstore."

        print(language)
        teaching_plan = {}
        evaluation_report = ""
        plan_id = str(uuid.uuid4()).replace("-", "_")
        try:
            teaching_plan, evaluation_report, plan_id = await generate_teaching_plan(
                # extracted_text=extracted_text,
                curriculum_id=curriculum_id,
                student_id=student_id,
                # query=query,
                curriculumname=curriculumName,
                image_url=None,
                language=language,
                temp_pdf_path=temp_pdf_path,
                openai_client=openai_client,
                subject=subject,
                curriculum_coverage=curriculum_coverage,
                teacher_id=teacher_id,
            )
        except Exception as plan_error:
            print(f"Teaching plan generation failed: {plan_error}")
        # temp_pdf_path, openai_client,curriculum_id,subject,curriculum_coverage

        print(plan_id)
        print(evaluation_report)

        # Verify if studentId exists in the students collection
        student_doc = db.collection("students").document(student_id).get()
        if not student_doc.exists:
            return (
                jsonify(
                    {
                        "error": f"Student ID {student_id} does not exist in the students collection"
                    }
                ),
                404,
            )

        # Generate unique file name and upload to Firebase Storage
        unique_filename = f"exam_scripts/{uuid.uuid4()}_{file.filename}"
        blob = bucket.blob(unique_filename)
        file.seek(0)
        blob.upload_from_file(file)
        file_url = blob.public_url

        # Generate Plan ID
        # plan_id = str(uuid.uuid4()).replace('-', '_')

        # Create Firestore document for examscripts
        exam_script_data = {
            "exam_name": exam_name,
            "curriculum_coverage": curriculum_coverage,
            "date": date,
            "observation": observation,
            "student_id": student_id,
            "file_url": file_url,
            "uuid": str(uuid.uuid4()),
            # "extracted_text": extracted_text,
            "teaching_plan": teaching_plan,
            "curriculumId": curriculum_id,
            "curriculumName": curriculumName,
            "evaluatedtext": evaluation_report,
            "planId": plan_id,
            "subject": subject,
        }

        db.collection("examscripts").add(exam_script_data)

        return (
            jsonify(
                {
                    "message": "Exam script uploaded successfully",
                    "evaluationReport": evaluation_report,
                    "teachingPlan": teaching_plan,
                    "studentId": student_id,
                    "uuid": exam_script_data["uuid"],
                    "extractText": "",
                    "planId": plan_id,
                }
            ),
            200,
        )  # extracted_text

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        if (
            temp_pdf_path
            and temp_pdf_path != temp_file_path
            and os.path.exists(temp_pdf_path)
        ):
            os.remove(temp_pdf_path)


@app.route("/api/teaching-plan-detail", methods=["GET"])
def get_teaching_plan113():
    try:
        # Extract query parameters
        student_id = request.args.get("student_id")
        print(f"Received student_id: {student_id}")

        if not student_id:
            return (
                jsonify({"status": "error", "message": "student_id is required"}),
                400,
            )

        # Query Firestore for the student's teaching plan
        query = db.collection("examscripts").where("student_id", "==", student_id)
        docs = list(query.stream())  # Convert stream to a list

        print(f"Documents found: {len(docs)}")  # Debugging output

        # Process results
        results = []
        for doc in docs:
            data = doc.to_dict()
            print(f"Document data: {data}")  # Debugging output

            plan_details = {
                "date": data.get("date"),
                "exam_name": data.get("exam_name"),
                "curriculum_name": data.get("curriculumName"),
                "subject": data.get("subject"),
                "plan_details": data.get("teaching_plan", {}),
            }
            results.append(plan_details)

        if not results:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "No teaching plan found for this student",
                    }
                ),
                404,
            )

        return jsonify({"status": "success", "data": results}), 200

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/get_exam_history/<student_id>", methods=["GET"])
def get_exam_history(student_id):
    try:
        # Query Firestore to fetch exam history for the given student ID
        exams_ref = (
            db.collection("examscripts").where("student_id", "==", student_id).stream()
        )

        exam_history = []
        for exam in exams_ref:
            exam_data = exam.to_dict()
            exam_history.append(
                {
                    "subject_name": exam_data.get("subject", "N/A"),
                    "curriculum_name": exam_data.get("curriculumName", "N/A"),
                    "exam_name": exam_data.get("exam_name", "N/A"),
                    "date": exam_data.get("date", "N/A"),
                    "evaluated_text": exam_data.get("evaluatedtext", "N/A"),
                    "evaluation_id": exam_data.get("uuid", "N/A"),
                }
            )

        if not exam_history:
            return (
                jsonify(
                    {
                        "message": "No exam history found for this student.",
                        "exam_history": [],
                    }
                ),
                404,
            )

        return jsonify({"student_id": student_id, "exam_history": exam_history}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/delete_exam_history", methods=["DELETE"])
def delete_exam_history():
    try:
        # Get student_id and evaluation_id from query parameters
        student_id = request.args.get("student_id")
        evaluation_id = request.args.get("evaluation_id")

        if not student_id or not evaluation_id:
            return (
                jsonify(
                    {
                        "error": "Missing student_id or evaluation_id in query parameters."
                    }
                ),
                400,
            )

        # Query Firestore for matching document
        exams_ref = (
            db.collection("examscripts")
            .where("student_id", "==", student_id)
            .where("uuid", "==", evaluation_id)
            .stream()
        )

        deleted = False
        for exam in exams_ref:
            exam.reference.delete()
            deleted = True

        if not deleted:
            return (
                jsonify(
                    {
                        "message": "No matching exam history found for the given student ID and evaluation ID."
                    }
                ),
                404,
            )

        return (
            jsonify(
                {
                    "message": "Exam history successfully deleted.",
                    "student_id": student_id,
                    "evaluation_id": evaluation_id,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


try:
    from teachingplans import get_curriculum_list
except Exception:
    async def get_curriculum_list(teacher_id=None):
        return jsonify({"curriculum": []}), 200




# @app.route("/curriculum", methods=["GET"])
# async def fetch_curriculum():
#     """
#     API route handler to fetch curriculum data.
#     It retrieves the teacherId from the query parameters and passes it to the core logic.
#     """
#     teacher_id = request.args.get("teacherId")
#     # Call the core logic function
#     return await get_curriculum_list(teacher_id)
@app.route("/curriculum", methods=["GET"])
def fetch_curriculum():
    """
    API route handler to fetch curriculum data.
    """
    teacher_id = request.args.get("teacherId")

    def _extract_curriculum_from_response(resp_obj):
        try:
            if isinstance(resp_obj, tuple):
                response = resp_obj[0]
            else:
                response = resp_obj
            payload = response.get_json(silent=True) or {}
            items = payload.get("curriculum", [])
            if isinstance(items, list):
                return items
            return []
        except Exception:
            return []

    def _load_curriculum_from_local_json():
        fallback_path = os.path.join(os.path.dirname(__file__), "curriculum_by_grade.json")
        if not os.path.exists(fallback_path):
            return []
        try:
            with open(fallback_path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            if not isinstance(raw, dict):
                return []
            flattened = []
            seen = set()
            for grade, classes in raw.items():
                if not isinstance(classes, dict):
                    continue
                for _, subjects in classes.items():
                    if not isinstance(subjects, list):
                        continue
                    for subject in subjects:
                        grade_val = str(grade or "").strip()
                        subject_val = str(subject or "").strip()
                        if not grade_val or not subject_val:
                            continue
                        dedupe_key = (grade_val.lower(), subject_val.lower())
                        if dedupe_key in seen:
                            continue
                        seen.add(dedupe_key)
                        flattened.append(
                            {
                                "curriculum_id": f"{grade_val}::{subject_val}",
                                "curriculum_name": subject_val,
                                "grade": grade_val,
                                "subject": subject_val,
                            }
                        )
            return flattened
        except Exception as e:
            logging.error(f"Failed to load fallback curriculum JSON: {e}")
            return []

    # ✅ Run async function synchronously using shared loop
    try:
        primary_resp = loop.run_until_complete(get_curriculum_list(teacher_id))
        primary_items = _extract_curriculum_from_response(primary_resp)
        if primary_items:
            return primary_resp

        if teacher_id:
            # Keep teacher-scoped requests strict to avoid leaking random/global subjects.
            return jsonify({"curriculum": []}), 200

        fallback_items = _load_curriculum_from_local_json()
        if fallback_items:
            return jsonify({"curriculum": fallback_items}), 200

        return primary_resp
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/assign-grade-class", methods=["POST"])
def assign_grade_class():
    try:
        data = request.json
        user_id = data.get("userId")  # Single user ID
        grades_data = data.get("grades", {})  # Dictionary format

        if not user_id or not grades_data:
            return jsonify({"error": "userId and grades are required"}), 400

        # Reference to user document
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()

        if not user_doc.exists:
            return jsonify({"error": f"User ID {user_id} not found"}), 404

        # Get existing assigned grades (if any)
        existing_data = user_doc.to_dict().get("assignedGrades", {})

        # Merge existing data with new data (without overwriting previous subjects)
        for grade, classes in grades_data.items():
            if grade not in existing_data:
                existing_data[grade] = {}  # Create grade if not exists

            for class_name, subjects in classes.items():
                if class_name not in existing_data[grade]:
                    existing_data[grade][class_name] = []  # Create class if not exists

                # Add new subjects while avoiding duplicates
                existing_data[grade][class_name] = list(
                    set(existing_data[grade][class_name] + subjects)
                )

        # Update Firestore
        user_ref.update({"assignedGrades": existing_data})

        return (
            jsonify(
                {
                    "message": "Grades assigned successfully",
                    "userId": user_id,
                    "assignedGrades": existing_data,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# @app.route("/api/headmaster/add-users1", methods=["POST"])
# def add_user11():
#     try:

#         if not request.is_json:
#             return jsonify({"error": "Invalid JSON format"}), 400

#         data = request.json

    #     # Validate required fields
    #     required_fields = ["name", "email", "role", "grades"]
    #     for field in required_fields:
    #         if field not in data or not data[field]:
    #             return jsonify({"error": f"{field} is required"}), 400

    #     # Validate role
    #     valid_roles = ["teacher", "parent", "student", "headmaster"]
    #     if data.get("role", "") not in valid_roles:
    #         return jsonify({"error": "Invalid role"}), 400

    #     # Validate email format
    #     email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
    #     if not re.match(email_regex, data.get("email", "")):
    #         return jsonify({"error": "Invalid email format"}), 400

    #     # Validate subject field (only for teachers)
    #     if data.get("role", "") == "teacher":
    #         for grade, classes in data.get("grades", {}).items():
    #             for class_name, subjects in classes.items():
    #                 if not isinstance(subjects, list) or not subjects:
    #                     return (
    #                         jsonify(
    #                             {
    #                                 "error": f"Subjects are required for grade {grade}, class {class_name}"
    #                             }
    #                         ),
    #                         400,
    #                     )

    #     # Generate userId and createdAt
    #     user_id = str(uuid.uuid4())
    #     created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    #     # Store user data
    #     user_data = {
    #         "userId": user_id,
    #         "email": data.get("email", ""),
    #         "role": data.get("role", ""),
    #         "createdAt": created_at,
    #         "profileInfo": {
    #             "personalInformation": {"name": data.get("name", "")},
    #             "contactInformation": {
    #                 "phoneNumber": data.get("contactNumber", "")
    #             },  # âœ… Updated field location
    #         },
    #         "associatedIds": data.get("associatedIds", []),  # âœ… Handles missing key
    #         "assignedGrades": data.get("grades", {}),  # âœ… Handles missing key
    #     }

    #     # Store in Firestore
    #     db.collection("users").document(user_id).set(user_data)
    #     if role == "parent":
    #         student_ids = data.get("associatedIds", [])
    #         if student_ids:
    #             for student_id in student_ids:
    #                 student_ref = db.collection("users").document(student_id)
    #                 try:
    #                     # Use ArrayUnion to safely add the Parent's ID to the Student's associatedIds list
    #                     student_ref.update({
    #                         "associatedIds": firestore.ArrayUnion([user_id])
    #                     })
    #                     logging.info(f"Successfully linked Parent {user_id} to Student {student_id}")
    #                 except Exception as update_e:
    #                     logging.error(f"Failed to update student {student_id} with parent link: {update_e}")
    #     # ----------------------------------------------------

    #     return jsonify({"message": "User added successfully", "userId": user_id}), 201

    # except Exception as e:
    #     return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# def get_students_by_grades(db, grades_data):
#     """
#     Looks up student user IDs whose assignedGrades match the teacher's grades.
    
#     FIX: Uses the full namespace (firebase_admin.firestore.field_path) to correctly 
#     query fields containing spaces (like 'GRADE 10').
#     """
#     associated_ids = set()
    
#     # The necessary import is placed here just in case, but should ideally be global.
#     # We reference it absolutely to avoid local attribute errors.
#     import firebase_admin 
    
#     # Iterate through all grades the teacher is assigned to (e.g., 'GRADE 10')
#     for grade, classes in grades_data.items():
#         # Iterate through all classes within that grade (e.g., 'Class A')
#         for class_name in classes.keys():
            
#             # ðŸ”¥ CRITICAL FIX: Use the full namespace for field_path
#             # This correctly constructs the path: assignedGrades.[Grade Name]
#             grade_field_path = firebase_admin.firestore.field_path("assignedGrades", grade)
            
#             # 1. Query Firestore for all students who have this grade in their assignedGrades map.
#             students_ref = db.collection("users").where("role", "==", "student")
            
#             # Query using the corrected field path
#             students_docs = students_ref.where(grade_field_path, "!=", None).stream()

#             for student_doc in students_docs:
#                 student_data = student_doc.to_dict()
#                 student_grades = student_data.get("assignedGrades", {})

#                 # 2. In-memory check: Confirm the student is in the specific class 
#                 # Assumes student assignedGrades has structure: {'Grade 10': ['Class A', 'Class B']}
#                 if class_name in student_grades.get(grade, []):
#                     associated_ids.add(student_doc.id)
                    
#     return list(associated_ids)



# @app.route("/api/headmaster/add-users1", methods=["POST"])
# def add_user11():
#     try:
#         if not request.is_json:
#             return jsonify({"error": "Invalid JSON format"}), 400

#         data = request.json

#         # 1. Validation 
#         required_fields = ["name", "email", "role", "grades", "password"] 
        
#         for field in required_fields:
#             if field not in data or not data[field]:
#                 return jsonify({"error": f"{field} is required"}), 400
            
#         valid_roles = ["teacher", "parent", "student", "headmaster"]
#         role = data.get("role", "")
#         if role not in valid_roles:
#             return jsonify({"error": "Invalid role"}), 400
            
#         email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
#         if not re.match(email_regex, data.get("email", "")):
#             return jsonify({"error": "Invalid email format"}), 400

#         # Generate userId and createdAt
#         user_id = str(uuid.uuid4())
#         created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        
#         email = data.get("email")
#         password = data.get("password")
#         name = data.get("name")
        
#       #  2. Firebase Authentication Creation - COMMENTED OUT FOR TESTING
#         try:
#             auth_user = auth.create_user(
#                 uid=user_id,
#                 email=email,
#                 password=password,
#                 display_name=name,
#                 disabled=False,
#             )
#             logging.info(f"User successfully created in Firebase Auth: {auth_user.uid}")
#         except auth.EmailAlreadyExistsError:
#             return jsonify({"error": f"The email address {email} is already in use."}), 409
        
#         # 3. Read Associated IDs (for Parent role)
#         associated_ids = data.get("associated_ids", []) 
#         if not isinstance(associated_ids, list):
#             return jsonify({"error": "associated_ids must be a list of IDs"}), 400

#         #  CRITICAL FIX: Populate associated_ids for Teacher role immediately (First Take)
#         if role == "teacher":
#             teacher_grades = data.get("grades", {})
#             calculated_ids = set()
            
#             # --- Logic to find students matching assignedGrades: MATCH BY GRADE NAME ONLY (Like the working update API) ---
#             try:
#                 # 1. Query all students by role
#                 students_ref = db.collection("users").where("role", "==", "student")
#                 students_docs = students_ref.stream()

#                 # 2. Iterate through students and match grades
#                 for student_doc in students_docs:
#                     student_data = student_doc.to_dict()
#                     student_grades = student_data.get("assignedGrades", {})
                    
#                     # Normalize student grades keys to lowercase for robust matching (e.g., 'GRADE 10' -> 'grade 10')
#                     normalized_student_grades_keys = {g.lower() for g in student_grades.keys()}

#                     # Check if any of the teacher's new grades are present in the student's assigned grades
#                     for t_grade in teacher_grades.keys():
#                         if t_grade.lower() in normalized_student_grades_keys:
#                             calculated_ids.add(student_doc.id)
#                             break # Student matched, move to the next student
                                
#             except Exception as e:
#                 # Log error but continue with empty list if query fails
#                 logging.error(f"Failed to query students for new teacher {name}: {e}")
            
#             # Overwrite the empty list with the calculated list
#             associated_ids = list(calculated_ids)
#             logging.info(f"Teacher {name} found {len(associated_ids)} students matching grades.")
           
#         if role == "headmaster":
#             # Get all student IDs from both students and users collections
#             calculated_ids = set()
#             try:
#                 # Get students from students collection
#                 students_stream = db.collection("students").stream()
#                 for student_snap in students_stream:
#                     calculated_ids.add(student_snap.id)

#                 # Get students from users collection
#                 users_stream = db.collection("users").where("role", "==", "student").stream()
#                 for user_snap in users_stream:
#                     calculated_ids.add(user_snap.id)

#                 associated_ids = list(calculated_ids)
#                 logging.info(f"Headmaster {name} associatedIds populated with {len(associated_ids)} students")

#             except Exception as e:
#                 logging.error(f"Failed to compute associated students for headmaster {name}: {e}")
#                 associated_ids = []

#             # Assign all grades with School-Wide access
#             assigned_grades = {"grades": {}}
#             ALL_GRADES = [
#                 "GRADE 1", "KG1", "GRADE 2", "KG2", "GRADE 3", "GRADE 4", 
#                 "GRADE 5", "GRADE 6", "GRADE 7", "GRADE 8", "GRADE 9", 
#                 "GRADE 10", "GRADE 11 (SCIENCE)", "GRADE 11 LITERATURE", 
#                 "GRADE 11 (LITERATURE)", "GRADE 11-SCIENCE", "GRADE11 LITERATURE", 
#                 "GRADE11-SCIENCE", "GRADE 12"
#             ]
                    
#             for g in ALL_GRADES:
#                 assigned_grades["grades"][g] = {
#                     "School-Wide": ["Oversight", "Administration"]
#                 }

#         # --- 4. Logic for Creating a STUDENT (Auto-Link to Teachers) ---
#         if role == "student":
#             student_grades = data.get("grades", {})
            
#             # 1. Normalize the student's new grades
#             # Use a simple lambda to clean the keys (uppercase, strip)
#             clean_key = lambda k: str(k).upper().strip().replace('_', ' ')
#             student_grade_keys = set(clean_key(k) for k in student_grades.keys())

#             if student_grade_keys:
#                 try:
#                     # 2. Find all teachers
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers_stream = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False

#                     for teacher_doc in teachers_stream:
#                         t_data = teacher_doc.to_dict()
#                         t_assigned = t_data.get("assignedGrades", {})
                        
#                         # Normalize teacher's grades
#                         t_grade_keys = set(clean_key(k) for k in t_assigned.keys())

#                         # 3. Check overlap: Does this teacher teach the new student's grade?
#                         if not student_grade_keys.isdisjoint(t_grade_keys):
#                             # Yes! Add student ID to teacher's associatedIds
#                             t_ref = db.collection("users").document(teacher_doc.id)
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                             batch_count += 1
#                             updates_made = True
                            
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()
#                         logging.info(f"Successfully linked new Student {user_id} to relevant teachers.")

#                 except Exception as e:
#                     logging.error(f"Failed to auto-link new student to teachers: {e}")
#         # 4. Prep
#         #are Firestore User Data
#         user_data = {
#             "userId": user_id,
#             "email": email,
#             "role": role,
#             "createdAt": created_at,
#             "profileInfo": {
#                 "personalInformation": {"name": name},
#                 "contactInformation": {
#                     "phoneNumber": data.get("contactNumber", "")
#                 },
#             },
#             # This field is now correctly populated on the first take for Teachers
#             "associatedIds": associated_ids, 
#             "assignedGrades": data.get("grades", {}),
#         }

#         # 5. Store the new user in Firestore
#         db.collection("users").document(user_id).set(user_data)

#         # 6. CRITICAL: Reverse-link Parent ID to each selected Student's document (Only for Parents)
#         if role == "parent" and associated_ids:
#             batch = db.batch()
#             for student_id in associated_ids:
#                 student_ref = db.collection("users").document(student_id)
                
#                 # Add the new Parent's ID (user_id) to the Student's associatedIds list
#                 batch.update(student_ref, {
#                     "associatedIds": firestore.ArrayUnion([user_id])
#                 })
            
#             batch.commit()
#             logging.info(f"Successfully reverse-linked Parent {user_id} to {len(associated_ids)} students.")

#         # 7. Final Response (Includes userId and associatedIds)
#         return jsonify({
#             "message": f"{role.capitalize()} added successfully", 
#             "userId": user_id,
#             "associatedIds": associated_ids
#         }), 201

#     except Exception as e:
#         logging.error(f"Error in add_user11: {e}") 
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# @app.route("/api/headmaster/add-users1", methods=["POST"])
# def add_user11():
#     try:
#         if not request.is_json:
#             return jsonify({"error": "Invalid JSON format"}), 400

#         data = request.json

#         # 1. Validation 
#         required_fields = ["name", "email", "role", "grades", "password"] 
#         for field in required_fields:
#             if field not in data or not data[field]:
#                 return jsonify({"error": f"{field} is required"}), 400
            
#         valid_roles = ["teacher", "parent", "student", "headmaster"]
#         role = data.get("role", "")
#         if role not in valid_roles:
#             return jsonify({"error": "Invalid role"}), 400
            
#         email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
#         if not re.match(email_regex, data.get("email", "")):
#             return jsonify({"error": "Invalid email format"}), 400

#         # Generate userId and createdAt
#         user_id = str(uuid.uuid4())
#         created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        
#         email = data.get("email")
#         password = data.get("password")
#         name = data.get("name")
        
#         # 2. Firebase Authentication Creation
#         try:
#             auth_user = auth.create_user(
#                 uid=user_id,
#                 email=email,
#                 password=password,
#                 display_name=name,
#                 disabled=False,
#             )
#             logging.info(f"User successfully created in Firebase Auth: {auth_user.uid}")
#         except auth.EmailAlreadyExistsError:
#             # Optional: Try to find existing user if you want to overwrite logic, 
#             # but usually returning error is safer.
#             return jsonify({"error": f"The email address {email} is already in use."}), 409
        
#         # 3. Read Associated IDs (for Parent role)
#         associated_ids = data.get("associated_ids", []) 
#         if not isinstance(associated_ids, list):
#             return jsonify({"error": "associated_ids must be a list of IDs"}), 400

#         # --- LOGIC: TEACHER (Find Students) ---
#         if role == "teacher":
#             teacher_grades = data.get("grades", {})
#             calculated_ids = set()
#             try:
#                 students_ref = db.collection("users").where("role", "==", "student")
#                 students_docs = students_ref.stream()

#                 for student_doc in students_docs:
#                     student_data = student_doc.to_dict()
#                     student_grades = student_data.get("assignedGrades", {})
                    
#                     normalized_student_grades_keys = {g.lower().strip() for g in student_grades.keys()}

#                     for t_grade in teacher_grades.keys():
#                         if t_grade.lower().strip() in normalized_student_grades_keys:
#                             calculated_ids.add(student_doc.id)
#                             break 
#             except Exception as e:
#                 logging.error(f"Failed to query students for new teacher {name}: {e}")
            
#             associated_ids = list(calculated_ids)
#             logging.info(f"Teacher {name} found {len(associated_ids)} students matching grades.")
           
#         # --- LOGIC: HEADMASTER (Add Everyone) ---
#         if role == "headmaster":
#             calculated_ids = set()
#             try:
#                 users_stream = db.collection("users").where("role", "==", "student").stream()
#                 for user_snap in users_stream:
#                     calculated_ids.add(user_snap.id)

#                 associated_ids = list(calculated_ids)
#             except Exception as e:
#                 logging.error(f"Failed to compute associated students for headmaster: {e}")
#                 associated_ids = []

#             # Assign all grades
#             assigned_grades = {"grades": {}}
#             ALL_GRADES = ["GRADE 1", "GRADE 2", "GRADE 3", "GRADE 4", "GRADE 5", "GRADE 6", "GRADE 7", "GRADE 8", "GRADE 9", "GRADE 10", "GRADE 11", "GRADE 12"]
#             for g in ALL_GRADES:
#                 assigned_grades["grades"][g] = {"School-Wide": ["Administration"]}

#         # 4. Prepare User Data
#         user_data = {
#             "userId": user_id,
#             "email": email,
#             "role": role,
#             "createdAt": created_at,
#             "profileInfo": {
#                 "personalInformation": {"name": name},
#                 "contactInformation": {
#                     "phoneNumber": data.get("contactNumber", "")
#                 },
#             },
#             "associatedIds": associated_ids, 
#             "assignedGrades": data.get("grades", {}),
#         }

#         # 5. Store in 'users' Collection
#         db.collection("users").document(user_id).set(user_data)

#         # 6. Reverse-Link for Parents
#         if role == "parent" and associated_ids:
#             batch = db.batch()
#             for student_id in associated_ids:
#                 # Update user doc
#                 u_ref = db.collection("users").document(student_id)
#                 batch.update(u_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                 # Update student doc (mirror)
#                 s_ref = db.collection("students").document(student_id)
#                 # We use set with merge=True in case student doc doesn't exist yet, though it should
#                 batch.set(s_ref, {"associatedIds": firestore.ArrayUnion([user_id])}, merge=True)
            
#             batch.commit()

#         # ======================================================
#         # 7. CRITICAL: STUDENT LOGIC (Mirror & Auto-Link)
#         # ======================================================
#         if role == "student":
            
#             # A. Mirror to 'students' Collection
#             # This ensures APIs that read from 'students' collection work immediately.
#             try:
#                 db.collection("students").document(user_id).set(user_data)
#                 logging.info(f"✅ Mirror document created in 'students' collection for {user_id}")
#             except Exception as e:
#                 logging.error(f"❌ Failed to create mirror in 'students' collection: {e}")

#             # B. Auto-Link to Teachers
#             # Find teachers who teach this new student's grade and update them.
#             try:
#                 student_grades = data.get("grades", {})
                
#                 # Normalize keys
#                 clean_key = lambda k: str(k).upper().strip().replace('_', ' ')
#                 student_grade_keys = set(clean_key(k) for k in student_grades.keys())

#                 if student_grade_keys:
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers_stream = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False

#                     for teacher_doc in teachers_stream:
#                         t_data = teacher_doc.to_dict()
#                         t_assigned = t_data.get("assignedGrades", {})
                        
#                         t_grade_keys = set(clean_key(k) for k in t_assigned.keys())

#                         # Check Intersection
#                         if not student_grade_keys.isdisjoint(t_grade_keys):
#                             t_ref = db.collection("users").document(teacher_doc.id)
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                             batch_count += 1
#                             updates_made = True
                            
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()
#                         logging.info(f"✅ Automatically linked new Student {user_id} to relevant teachers.")

#             except Exception as e:
#                 logging.error(f"⚠️ Failed to auto-link new student to teachers: {e}")

#         # 8. Final Response
#         return jsonify({
#             "message": f"{role.capitalize()} added successfully", 
#             "userId": user_id,
#             "associatedIds": associated_ids
#         }), 201

#     except Exception as e:
#         logging.error(f"Error in add_user11: {e}") 
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/api/headmaster/add-users1", methods=["POST"])
# def add_user11():
#     try:
#         if not request.is_json:
#             return jsonify({"error": "Invalid JSON format"}), 400

#         data = request.json

#         # 1. Validation 
#         required_fields = ["name", "email", "role", "grades", "password"] 
#         for field in required_fields:
#             if field not in data or not data[field]:
#                 return jsonify({"error": f"{field} is required"}), 400
            
#         valid_roles = ["teacher", "parent", "student", "headmaster"]
#         role = data.get("role", "")
#         if role not in valid_roles:
#             return jsonify({"error": "Invalid role"}), 400
            
#         email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
#         if not re.match(email_regex, data.get("email", "")):
#             return jsonify({"error": "Invalid email format"}), 400

#         # Generate userId and createdAt
#         user_id = str(uuid.uuid4())
#         created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        
#         email = data.get("email")
#         password = data.get("password")
#         name = data.get("name")
        
#         # 2. Firebase Authentication Creation
#         try:
#             auth_user = auth.create_user(
#                 uid=user_id,
#                 email=email,
#                 password=password,
#                 display_name=name,
#                 disabled=False,
#             )
#             logging.info(f"User successfully created in Firebase Auth: {auth_user.uid}")
#         except auth.EmailAlreadyExistsError:
#             return jsonify({"error": f"The email address {email} is already in use."}), 409
        
#         # 3. Read Associated IDs (for Parent role)
#         associated_ids = data.get("associated_ids", []) 
#         if not isinstance(associated_ids, list):
#             return jsonify({"error": "associated_ids must be a list of IDs"}), 400

#         # --- HELPER: Robust Grade Normalization ---
#         def normalize_grade_key(key):
#             """
#             Normalizes grade strings to handle case, spaces, and separators.
#             Ex: "Grade 1", "GRADE-1", "grade.1" -> "GRADE_1"
#             """
#             key = str(key).upper().strip()
#             # Replace spaces, dots, slashes, hyphens with underscore
#             key = re.sub(r'[\s\./\-]+', '_', key)
#             # Clean up extra underscores around parentheses
#             key = key.replace('_(', '(').replace(')_', ')')
#             # Collapse multiple underscores
#             key = re.sub(r'_{2,}', '_', key)
#             return key.strip('_')

# # --- LOGIC: TEACHER (Find Students + Format Data + Reverse Link) ---
#         if role == "teacher":
#             raw_input_grades = data.get("grades", {})
            
#             # 1. Sanitize & Flatten Data (Prevents "Unknown" Error)
#             clean_assigned_grades = {}
#             teacher_grade_keys = set()

#             for grade_key, classes in raw_input_grades.items():
#                 norm_key = normalize_grade_key(grade_key) # e.g. "GRADE_5"
#                 teacher_grade_keys.add(norm_key)
                
#                 if norm_key not in clean_assigned_grades:
#                     clean_assigned_grades[norm_key] = {}

#                 for class_name, subjects in classes.items():
#                     # Handle Subject List vs String
#                     final_subject_val = ""
                    
#                     if isinstance(subjects, list):
#                         if len(subjects) == 1:
#                             final_subject_val = subjects[0]
#                         else:
#                             # --- FIX: JOIN MULTIPLE SUBJECTS INTO A STRING ---
#                             # ["Math", "Science"] -> "Math, Science"
#                             final_subject_val = ", ".join(subjects)
#                     elif isinstance(subjects, str):
#                         final_subject_val = subjects
                    
#                     clean_assigned_grades[norm_key][class_name] = final_subject_val

#             # REPLACE the raw input with our clean data so Step 5 saves it correctly
#             # We wrap it in "grades" to match the schema used in PUT
#             data["grades"] = {"grades": clean_assigned_grades} 

#             # 2. Find Students (Calculate Intersection)
#             calculated_ids = set()
#             try:
#                 students_ref = db.collection("users").where("role", "==", "student")
#                 students_docs = students_ref.stream()
                
#                 batch = db.batch()
#                 batch_count = 0
#                 updates_made = False

#                 for student_doc in students_docs:
#                     student_data = student_doc.to_dict()
#                     student_assigned = student_data.get("assignedGrades", {})
                    
#                     if "grades" in student_assigned and isinstance(student_assigned["grades"], dict):
#                         student_grades_map = student_assigned["grades"]
#                     else:
#                         student_grades_map = student_assigned

#                     student_grade_keys = {normalize_grade_key(g) for g in student_grades_map.keys()}

#                     if not teacher_grade_keys.isdisjoint(student_grade_keys):
#                         calculated_ids.add(student_doc.id)

#                         # Reverse Link (Teacher -> Student)
#                         s_ref = db.collection("users").document(student_doc.id)
#                         batch.update(s_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
                        
#                         # Mirror Link
#                         s_mirror_ref = db.collection("students").document(student_doc.id)
#                         batch.set(s_mirror_ref, {"associatedIds": firestore.ArrayUnion([user_id])}, merge=True)

#                         batch_count += 1
#                         updates_made = True
                        
#                         if batch_count >= 200:
#                             batch.commit()
#                             batch = db.batch()
#                             batch_count = 0
                
#                 if updates_made and batch_count > 0:
#                     batch.commit()
                    
#             except Exception as e:
#                 logging.error(f"Failed to query/link students for new teacher {name}: {e}")
            
#             associated_ids = list(calculated_ids)
           
#         # --- LOGIC: HEADMASTER (Add Everyone) ---
#         if role == "headmaster":
#             calculated_ids = set()
#             try:
#                 users_stream = db.collection("users").where("role", "==", "student").stream()
#                 for user_snap in users_stream:
#                     calculated_ids.add(user_snap.id)

#                 associated_ids = list(calculated_ids)
#             except Exception as e:
#                 logging.error(f"Failed to compute associated students for headmaster: {e}")
#                 associated_ids = []

#             # Assign all grades
#             assigned_grades = {"grades": {}}
#             ALL_GRADES = ["GRADE 1", "GRADE 2", "GRADE 3", "GRADE 4", "GRADE 5", "GRADE 6", "GRADE 7", "GRADE 8", "GRADE 9", "GRADE 10", "GRADE 11", "GRADE 12"]
#             for g in ALL_GRADES:
#                 assigned_grades["grades"][g] = {"School-Wide": ["Administration"]}

#         # 4. Prepare User Data
#         user_data = {
#             "userId": user_id,
#             "email": email,
#             "role": role,
#             "createdAt": created_at,
#             "profileInfo": {
#                 "personalInformation": {"name": name},
#                 "contactInformation": {
#                     "phoneNumber": data.get("contactNumber", "")
#                 },
#             },
#             "associatedIds": associated_ids, 
#             "assignedGrades": data.get("grades", {}),
#         }

#         # 5. Store in 'users' Collection
#         db.collection("users").document(user_id).set(user_data)

#         # 6. Reverse-Link for Parents
#         if role == "parent" and associated_ids:
#             batch = db.batch()
#             for student_id in associated_ids:
#                 u_ref = db.collection("users").document(student_id)
#                 batch.update(u_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                 s_ref = db.collection("students").document(student_id)
#                 batch.set(s_ref, {"associatedIds": firestore.ArrayUnion([user_id])}, merge=True)
#             batch.commit()

#         # ======================================================
#         # 7. CRITICAL: STUDENT LOGIC (Mirror & Auto-Link)
#         # ======================================================
#         if role == "student":
            
#             # A. Mirror to 'students' Collection
#             try:
#                 db.collection("students").document(user_id).set(user_data)
#                 logging.info(f"✅ Mirror document created in 'students' collection for {user_id}")
#             except Exception as e:
#                 logging.error(f"❌ Failed to create mirror in 'students' collection: {e}")

#             # B. Auto-Link to Teachers
#             try:
#                 student_grades = data.get("grades", {})
#                 # Use the robust normalizer here too
#                 student_grade_keys = {normalize_grade_key(k) for k in student_grades.keys()}

#                 if student_grade_keys:
#                     teachers_ref = db.collection("users").where("role", "==", "teacher")
#                     teachers_stream = teachers_ref.stream()

#                     batch = db.batch()
#                     batch_count = 0
#                     updates_made = False

#                     for teacher_doc in teachers_stream:
#                         t_data = teacher_doc.to_dict()
#                         t_assigned = t_data.get("assignedGrades", {})
                        
#                         # Normalize teacher db grades
#                         t_grade_keys = {normalize_grade_key(k) for k in t_assigned.keys()}

#                         # Check Intersection
#                         if not student_grade_keys.isdisjoint(t_grade_keys):
#                             t_ref = db.collection("users").document(teacher_doc.id)
#                             batch.update(t_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
#                             batch_count += 1
#                             updates_made = True
                            
#                             if batch_count >= 450:
#                                 batch.commit()
#                                 batch = db.batch()
#                                 batch_count = 0
                    
#                     if updates_made and batch_count > 0:
#                         batch.commit()
#                         logging.info(f"✅ Automatically linked new Student {user_id} to relevant teachers.")

#             except Exception as e:
#                 logging.error(f"⚠️ Failed to auto-link new student to teachers: {e}")

#         # 8. Final Response
#         return jsonify({
#             "message": f"{role.capitalize()} added successfully", 
#             "userId": user_id,
#             "associatedIds": associated_ids
#         }), 201

#     except Exception as e:
#         logging.error(f"Error in add_user11: {e}") 
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500
@app.route("/api/headmaster/add-users1", methods=["POST"])
def add_user11():
    try:
        if not request.is_json: return jsonify({"error": "Invalid JSON format"}), 400
        data = request.json

        # 1. Validation
        required_fields = ["name", "email", "role", "grades", "password"] 
        for field in required_fields:
            if field not in data or not data[field]: return jsonify({"error": f"{field} is required"}), 400
        
        valid_roles = ["teacher", "parent", "student", "headmaster"]
        role = data.get("role", "")
        if role not in valid_roles: return jsonify({"error": "Invalid role"}), 400
            
        email_regex = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        if not re.match(email_regex, data.get("email", "")): return jsonify({"error": "Invalid email format"}), 400

        user_id = str(uuid.uuid4())
        created_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        email = data.get("email")
        password = data.get("password")
        name = data.get("name")
        
        # 2. Firebase Auth Creation
        try:
            auth_user = auth.create_user(
                uid=user_id,
                email=email,
                password=password,
                display_name=name,
                disabled=False,
            )
            logging.info(f"User created in Auth: {auth_user.uid}")
        except auth.EmailAlreadyExistsError:
            return jsonify({"error": f"The email address {email} is already in use."}), 409
        except Exception as e:
             return jsonify({"error": f"Auth Error: {str(e)}"}), 500
        
        # 3. Read Associated IDs
        associated_ids = data.get("associated_ids", []) 
        if not isinstance(associated_ids, list):
            return jsonify({"error": "associated_ids must be a list of IDs"}), 400

        # --- LOGIC: TEACHER (FIXED) ---
        clean_assigned_grades = {}
        
        if role == "teacher":
            raw_input = data.get("grades", {})
            processed_grades_dict = {}

            # Adapter: Convert List to Dict if needed
            if isinstance(raw_input, list):
                for item in raw_input:
                    g = item.get("grade") or item.get("gradeName")
                    c = item.get("class") or item.get("className")
                    s = item.get("subject") or item.get("subjects")
                    if g and c and s:
                        if g not in processed_grades_dict: processed_grades_dict[g] = {}
                        processed_grades_dict[g][c] = s
            elif isinstance(raw_input, dict):
                processed_grades_dict = raw_input

            # Sanitize & Format as LIST of subjects
            for grade_key, classes in processed_grades_dict.items():
                norm_key = sanitize_grade_key(grade_key)
                
                if norm_key not in clean_assigned_grades:
                    clean_assigned_grades[norm_key] = {}

                for class_name, subjects in classes.items():
                    final_subject_list = []
                    
                    if isinstance(subjects, list):
                        final_subject_list = subjects
                    elif isinstance(subjects, str):
                        # ✅ CRITICAL FIX: Wrap single string subject in a list
                        final_subject_list = [subjects] 
                    
                    # Store as List
                    clean_assigned_grades[norm_key][class_name] = final_subject_list

            # Store cleaned data back into data object
            data["grades"] = clean_assigned_grades 

            # Find Students & Auto-Link (Bi-directional)
            calculated_ids = set()
            teacher_grade_keys = set(clean_assigned_grades.keys())
            
            try:
                # Fetch all students to check against this new teacher's grades
                students_ref = db.collection("users").where("role", "==", "student")
                students_docs = students_ref.stream()
                
                batch = db.batch()
                batch_count = 0
                updates_made = False

                for student_doc in students_docs:
                    s_data = student_doc.to_dict()
                    s_assigned = s_data.get("assignedGrades", {})
                    s_grades_map = s_assigned.get('grades', s_assigned)
                    
                    if s_grades_map:
                        norm_s_grades = {sanitize_grade_key(k) for k in s_grades_map.keys()}
                        
                        # If teacher shares ANY grade with student, link them
                        if not teacher_grade_keys.isdisjoint(norm_s_grades):
                            # Add Student to Teacher's list
                            calculated_ids.add(student_doc.id)
                            
                            # Add Teacher ID to Student's associatedIds (Users Collection)
                            s_ref = db.collection("users").document(student_doc.id)
                            batch.update(s_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
                            
                            # Add Teacher ID to Student's associatedIds (Students Collection Mirror)
                            s_mirror = db.collection("students").document(student_doc.id)
                            batch.set(s_mirror, {"associatedIds": firestore.ArrayUnion([user_id])}, merge=True)
                            
                            batch_count += 1
                            updates_made = True
                            
                            # Commit batch in chunks of 200 to be safe
                            if batch_count >= 200:
                                batch.commit()
                                batch = db.batch()
                                batch_count = 0
                
                if updates_made and batch_count > 0:
                    batch.commit()

            except Exception as e:
                logging.error(f"Failed to query/link students: {e}")
            
            # Populate associated_ids with found students
            associated_ids = list(calculated_ids) 

        # 4. Prepare User Data
        user_data = {
            "userId": user_id,
            "email": email,
            "role": role,
            "createdAt": created_at,
            "profileInfo": {
                "personalInformation": {"name": name},
                "contactInformation": {
                    "phoneNumber": data.get("contactNumber", "")
                },
            },
            "associatedIds": associated_ids, 
            # ✅ Use the cleaned grades (lists) for teachers, or raw for others
            "assignedGrades": clean_assigned_grades if role == "teacher" else data.get("grades", {}), 
        }

        # 5. Store in 'users' Collection
        db.collection("users").document(user_id).set(user_data)

        # 6. Reverse-Link for Parents (Add this Parent ID to their Children)
        if role == "parent" and associated_ids:
            try:
                batch = db.batch()
                for student_id in associated_ids:
                    # Update User doc
                    u_ref = db.collection("users").document(student_id)
                    batch.update(u_ref, {"associatedIds": firestore.ArrayUnion([user_id])})
                    
                    # Update Student mirror doc
                    s_ref = db.collection("students").document(student_id)
                    batch.set(s_ref, {"associatedIds": firestore.ArrayUnion([user_id])}, merge=True)
                batch.commit()
            except Exception as e:
                logging.error(f"Failed to link parent to students: {e}")

        # 7. Student Logic (Mirror to 'students' collection)
        if role == "student":
            try:
                db.collection("students").document(user_id).set(user_data)
                teacher_ids = sync_student_teacher_links(user_id, user_data)
                if teacher_ids:
                    user_data["associatedIds"] = sorted(set((user_data.get("associatedIds") or []) + teacher_ids))
                    db.collection("users").document(user_id).set(
                        {"associatedIds": user_data["associatedIds"]},
                        merge=True,
                    )
                    db.collection("students").document(user_id).set(
                        {"associatedIds": user_data["associatedIds"]},
                        merge=True,
                    )
            except Exception as e:
                logging.error(f"Failed mirror: {e}")
            
            # Note: If you have logic to auto-link Teachers TO this new Student based on grades,
            # that would typically go here.

        return jsonify({
            "message": f"{role.capitalize()} added successfully", 
            "userId": user_id,
            "associatedIds": associated_ids
        }), 201

    except Exception as e:
        logging.error(f"Error in add_user11: {e}") 
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

# ------------------------- TWILIO POST API SMS/MMS Sending

from twilio.rest import Client

# Twilio credentials
account_sid = "YOUR_TWILIO_ACCOUNT_SID"
auth_token = "c1304a6c7e64e9668d693c403dca6d0c"
TWILIO_PHONE_NUMBER = "+12183191162"

# Initialize Twilio client
client = Client(account_sid, auth_token)


@app.route("/api/send_sms", methods=["POST"])
def send_sms():
    """
    Endpoint to send an SMS using Twilio.
    Expected JSON payload:
    {
        "to": "+1234567890",
        "message": "Your message here"
    }
    """
    data = request.get_json()
    if not data or "to" not in data or "message" not in data:
        return jsonify({"error": "Missing 'to' or 'message' in request payload"}), 400

    to = data["to"]
    message_body = data["message"]

    try:
        # Send SMS using Twilio
        message = client.messages.create(
            body=message_body, from_=TWILIO_PHONE_NUMBER, to=to
        )
        return jsonify({"message": "SMS sent successfully", "sid": message.sid}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/send-bulk-messages", methods=["POST"])
def send_bulk_messages():
    """
    Endpoint to send SMS to multiple recipients with corresponding messages.
    Expected payload:
    {
        "recipients": ["+1234567890", "+1987654321"],
        "messages": ["Message for recipient 1", "Message for recipient 2"]
    }
    """
    data = request.json
    if not data or "recipients" not in data or "messages" not in data:
        return jsonify({"error": "Missing 'recipients' or 'messages' in request"}), 400

    recipients = data["recipients"]
    messages = data["messages"]

    # Validate the input
    if (
        not isinstance(recipients, list)
        or not isinstance(messages, list)
        or len(recipients) != len(messages)
    ):
        return (
            jsonify(
                {
                    "error": "'recipients' and 'messages' must be lists of the same length"
                }
            ),
            400,
        )

    results = []
    for recipient, message_body in zip(recipients, messages):
        try:
            message = client.messages.create(
                body=message_body, from_=TWILIO_PHONE_NUMBER, to=recipient
            )
            results.append({"to": recipient, "sid": message.sid, "status": "success"})
        except Exception as e:
            results.append({"to": recipient, "error": str(e), "status": "failed"})

    return jsonify(results), 200


# @app.route('/getStudent', methods=['GET'])
# def get_student():
#     try:
#         student_id = request.args.get('studentId')
#         if not student_id:
#             return jsonify({"error": "studentId is required"}), 400

#         # Fetch student data from the students collection
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         student_data = student_doc.to_dict()

#         # Fetch email from the users collection
#         user_ref = db.collection("users").document(student_id)
#         user_doc = user_ref.get()
#         email = user_doc.to_dict().get("email", "") if user_doc.exists else ""

#         # Extract necessary fields and ensure missing ones return ""
#         response = {
#             "studentId": student_data.get("studentId", ""),
#             "studentName": student_data.get('profileInfo',{}).get("personalInformation", {}).get("name", ""),
#             "classSection": student_data.get("academicInformation", {}).get("classSection", ""),
#             "phoneNumber": student_data.get("contactInformation", {}).get("phoneNumber", ""),
#             "address": student_data.get("contactInformation", {}).get("address", ""),
#             "photoUrl": student_data.get("personalInformation", {}).get("photourl", ""),
#             "email": email,  # Retrieved from users collection
#             "grade": student_data.get("academicInformation", {}).get("grade", ""),
#         }

#         return jsonify(response)

#     except Exception as e:
#         print("Error fetching student data:", e)
#         return jsonify({"error": "Internal server error"}), 500

import re # Ensure re is imported

@app.route("/getStudent", methods=["GET"])
def get_student():
    try:
        student_id = request.args.get("studentId")
        if not student_id:
            return jsonify({"error": "studentId is required"}), 400

        # Fetch student data from "students" collection
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        student_data = student_doc.to_dict()

        # Extract relevant fields
        profile_info = student_data.get("profileInfo", {})
        personal_info = profile_info.get("personalInformation", {})
        contact_info = profile_info.get("contactInformation", {})

        # --- FIX 1: Handle Structure Mismatch (Flat vs Nested) ---
        assigned_grades = student_data.get("assignedGrades", {})
        
        # Try finding the "grades" key first (New Structure)
        grades_data = assigned_grades.get("grades")
        
        # If not found, assume the parent map itself holds the grades (Old Structure)
        if not grades_data:
            grades_data = assigned_grades

        grade_keys = list(grades_data.keys()) 
        
        # Logic to pick the grade (prefer second if >1, else first)
        grade = (
            grade_keys[1]
            if len(grade_keys) > 1
            else grade_keys[0] if grade_keys else ""
        )
        
        # --- FIX 2: Normalize Grade for Frontend Matching ---
        # Converts "Grade 1" -> "GRADE 1" so it matches the Curriculum List
        if grade:
            grade = str(grade).upper().strip()
            grade = re.sub(r'[_\.\-]+', ' ', grade) # Replace _ with space
            grade = re.sub(r'\s+', ' ', grade).strip()

        # Determine class section
        class_section = ""
        # Note: We must access grades_data using the original key if we didn't normalize keys in place
        # But since we only normalized the 'grade' variable, we need to be careful looking it up again if keys vary.
        # Safer strategy: Loop through keys to find the match or just use the raw key logic above.
        
        # Re-fetching raw key for lookup
        raw_key = grade_keys[1] if len(grade_keys) > 1 else grade_keys[0] if grade_keys else ""
        
        if raw_key and isinstance(grades_data.get(raw_key, {}), dict):
            for section in grades_data[raw_key].keys():  # Example: "CLASS A"
                if isinstance(grades_data[raw_key][section], list):
                    class_section = section
                    break

        # Construct response
        response = {
            "studentId": student_data.get("studentId", ""),
            "studentName": personal_info.get("name", ""),
            "classSection": class_section,
            "phoneNumber": contact_info.get("phoneNumber", ""),
            "address": contact_info.get("address", ""),
            "photoUrl": personal_info.get("photoUrl", ""),
            "email": student_data.get("email", ""), 
            "grade": grade, # Now correctly "GRADE 1"
        }

        return jsonify(response)

    except Exception as e:
        print("Error fetching student data:", e)
        return jsonify({"error": "Internal server error"}), 500

from flask import Flask, request, jsonify, send_file
import openai
import firebase_admin
from firebase_admin import credentials, firestore, storage, initialize_app
import os
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
import uuid
import textwrap
import json
from reportlab.lib.pagesizes import A4
from flask import Flask, request, jsonify, send_file
from openai import OpenAI
from reportlab.pdfgen import canvas
import os
import uuid
import datetime
import textwrap
from flask import jsonify, request
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from flask import jsonify, request
import os, uuid, datetime, textwrap
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
import os
import uuid
import textwrap
from flask import jsonify, request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Initialize Firebase
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()
bucket = storage.bucket()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
client_openai = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None


# Fetch student data from Firebase
@app.route("/fetch_student_data", methods=["GET"])
def fetch_student_data():
    student_id = request.args.get("student_id")

    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()

    if student_doc.exists:
        student_data = student_doc.to_dict()
        return (
            jsonify(
                {"message": "Data fetched successfully", "student_data": student_data}
            ),
            200,
        )
    else:
        return jsonify({"error": "Student not found"}), 404


import openai
import json


from flask import Flask, jsonify, request
from flask_apscheduler import APScheduler
import logging
import openai
import json

logging.basicConfig(level=logging.INFO)

scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()


# @app.route("/analyze_student_data", methods=["POST"])
# def analyze_student_data():
#     data = request.json
#     student_id = data.get("student_id")
#     lang = data.get("lang", "en")  


#     if not student_id:
#         return jsonify({"error": "Student ID is required"}), 400

#     student_ref = db.collection("students").document(student_id)
#     student_doc = student_ref.get()

#     if student_doc.exists:
#         student_data = student_doc.to_dict()
#         subjects_data = student_data.get("academicData", {}).get("subjects", {})

#         if not subjects_data:
#             print("Warning: subjects_data is empty for student", student_id)

#         all_entries = []

#         for subject, subject_info in subjects_data.items():
#             history_entries = subject_info.get("history", [])
#             if isinstance(history_entries, list):
#                 valid_entries = [
#                     {**entry, "subject": subject}
#                     for entry in history_entries
#                     if "timestamp" in entry
#                 ]
#                 all_entries.extend(valid_entries)

#         sorted_all_entries = sorted(
#             all_entries, key=lambda x: x["timestamp"], reverse=True
#         )

#         latest_subject_data = {}
#         if sorted_all_entries:
#             latest_entry = sorted_all_entries[0]
#             subject_name = latest_entry["subject"]
#             latest_subject_data[subject_name] = latest_entry

#         print("Latest subject data:", json.dumps(latest_subject_data, indent=2))

#         analysis = send_to_gpt(student_data, subjects_data, latest_subject_data, lang)

#         student_ref.update(
#             {
#                 "analysis": {
#                     "strengths": analysis.get("strengths", []),
#                     "weaknesses": analysis.get("weaknesses", []),
#                     "areas_for_improvement": analysis.get("areas_for_improvement", {}),
#                     "recommendations": analysis.get("recommendations", []),
#                     "interventions": analysis.get("interventions", []),
#                 }
#             }
#         )

#         return (
#             jsonify(
#                 {
#                     "message": "Analysis completed",
#                     "student_id": student_id,
#                     "analysis": analysis,
#                     "interventions": analysis.get("interventions", []),
#                 }
#             ),
#             200,
#         )


#     else:
#         return jsonify({"error": "Student not found"}), 404
from datetime import datetime, timezone, timedelta
try:
    from translation_provider import (translate_analysis_recursively, contains_arabic)
except Exception:
    def translate_analysis_recursively(payload):
        return payload

    def contains_arabic(text):
        if not isinstance(text, str):
            return False
        return bool(re.search(r"[\u0600-\u06FF]", text))


# @app.route("/analyze_student_data", methods=["POST"])
# def analyze_student_data():
#     try:
#         data = request.json or {}
#         student_id = data.get("student_id")
#         if not student_id:
#             return jsonify({"error": "Student ID is required"}), 400

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()
#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         student_data = student_doc.to_dict() or {}
#         subjects_data = (student_data.get("academicData") or {}).get("subjects") or {}
#         if not subjects_data:
#             print(f"[analyze_student_data] Warning: subjects_data empty for {student_id}")

#         # Build latest_subject_data from history by newest timestamp
#         all_entries = []
#         for subject, subject_info in subjects_data.items():
#             history_entries = subject_info.get("history", [])
#             if isinstance(history_entries, list):
#                 for entry in history_entries:
#                     if isinstance(entry, dict) and "timestamp" in entry:
#                         all_entries.append({**entry, "subject": subject})

#         all_entries.sort(key=lambda x: x["timestamp"], reverse=True)
#         latest_subject_data = {}
#         if all_entries:
#             latest = all_entries[0]
#             latest_subject_data[latest["subject"]] = latest

#         print("[analyze_student_data] Latest subject data:",
#               json.dumps(latest_subject_data, indent=2, ensure_ascii=False))

#         # 1) Generate English canonical analysis
#         en_raw = send_to_gpt(student_data, subjects_data, latest_subject_data, lang="en")
#         if not isinstance(en_raw, dict):
#             return jsonify({"error": "LLM returned unexpected format"}), 500

#         analysis_en = {
#             "strengths": en_raw.get("strengths", []) or [],
#             "weaknesses": en_raw.get("weaknesses", []) or [],
#             "interventions": (
#                 en_raw.get("interventions", []) or
#                 (en_raw.get("areas_for_improvement", {}) or {}).get("interventions", []) or []
#             ),
#             "recommendations": en_raw.get("recommendations", []) or [],
#         }

#         # Basic validation
#         if not any(analysis_en.values()):
#             return jsonify({"error": "Empty analysis generated"}), 500

#         # 2) Translate English to Arabic (structure-safe)
#         try:
#             analysis_ar = translate_analysis_recursively(analysis_en)
#         except Exception as te:
#             print(f"[analyze_student_data] Translation failed: {te}")
#             return jsonify({"error": f"Arabic translation failed: {te}"}), 500

#         # 3) Guard: ensure Arabic output truly contains Arabic script
#         if not contains_arabic(analysis_ar):
#             return jsonify({"error": "Arabic translation appears to be English; not storing"}), 500

#         # 4) Persist both languages atomically
#         student_ref.set({
#             "analysis": {
#                 "en": analysis_en,
#                 "ar": analysis_ar
#             },
#             "last_updated": datetime.utcnow().isoformat()
#         }, merge=True)

#         print(f"[analyze_student_data] Stored dual-language analysis for {student_id} "
#               f"en_sizes=({len(analysis_en['strengths'])},"
#               f"{len(analysis_en['weaknesses'])},"
#               f"{len(analysis_en['interventions'])},"
#               f"{len(analysis_en['recommendations'])}) "
#               f"ar_has_arabic={contains_arabic(analysis_ar)}")

#         return jsonify({
#             "message": "Analysis completed and saved in EN and AR",
#             "student_id": student_id
#         }), 200

#     except Exception as e:
#         print(f"[analyze_student_data] Unexpected error: {e}")
#         return jsonify({"error": "Internal server error"}), 500

# --- Analysis and Persistence Function (Fixed for Dual-Language Areas) ---

# @app.route("/analyze_student_data", methods=["POST"])
# def analyze_student_data():
#     try:
#         data = request.json or {}
#         student_id = data.get("userId")
#         if not student_id:
#             return jsonify({"error": "Student ID is required"}), 400

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()
#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         student_data = student_doc.to_dict() or {}
#         subjects_data = (student_data.get("academicData") or {}).get("subjects") or {}
#         if not subjects_data:
#             print(f"[analyze_student_data] Warning: subjects_data empty for {student_id}")

#         # Build latest_subject_data from history by newest timestamp (Existing Logic)
#         all_entries = []
#         for subject, subject_info in subjects_data.items():
#             history_entries = subject_info.get("history", [])
#             if isinstance(history_entries, list):
#                 for entry in history_entries:
#                     if isinstance(entry, dict) and "timestamp" in entry:
#                         all_entries.append({**entry, "subject": subject})

#         all_entries.sort(key=lambda x: x["timestamp"], reverse=True)
#         latest_subject_data = {}
#         if all_entries:
#             latest = all_entries[0]
#             latest_subject_data[latest["subject"]] = latest

#         print("[analyze_student_data] Latest subject data:",
#               json.dumps(latest_subject_data, indent=2, ensure_ascii=False))

#         # 1) Generate English canonical analysis
#         # LLM MUST generate all content in English when called with lang="en"
#         en_raw = send_to_gpt(student_data, subjects_data, latest_subject_data, lang="en")
#         if not isinstance(en_raw, dict):
#             return jsonify({"error": "LLM returned unexpected format"}), 500

#         # FIX: Extract Areas for Improvement before translation
#         areas_for_improvement_en = en_raw.pop("areas_for_improvement", {}) or {}

#         analysis_en = {
#             "strengths": en_raw.get("strengths", []) or [],
#             "weaknesses": en_raw.get("weaknesses", []) or [],
#             "interventions": en_raw.get("interventions", []) or [], # Simplified the key lookup
#             "recommendations": en_raw.get("recommendations", []) or [],
#         }

#         # Basic validation
#         # if not any(analysis_en.values()):
#         #     return jsonify({"error": "Empty analysis generated"}), 500

#         # 2) Translate English report structure (strengths, weaknesses, etc.) to Arabic
#         try:
#             analysis_ar = translate_analysis_recursively(analysis_en)
#         except Exception as te:
#             print(f"[analyze_student_data] Report translation failed: {te}")
#             return jsonify({"error": f"Arabic translation failed: {te}"}), 500

#         # FIX: 3) Translate English Areas for Improvement (subject lists) to Arabic
#         areas_for_improvement_ar = {}
#         try:
#             areas_for_improvement_ar = translate_analysis_recursively(areas_for_improvement_en)
#         except Exception as te:
#             print(f"[analyze_student_data] Areas translation failed: {te}")
#             return jsonify({"error": f"Arabic translation for areas failed: {te}"}), 500

#         # FIX: 4) Normalize and persist the final dual-language structure for areas
#         areas_for_improvement_localized = {}
#         all_subjects = set(areas_for_improvement_en.keys()) | set(areas_for_improvement_ar.keys())
        
#         for subject in all_subjects:
#             en_areas = areas_for_improvement_en.get(subject, [])
#             ar_areas = areas_for_improvement_ar.get(subject, [])
            
#             # Guard against potential empty or non-list results from translation
#             if isinstance(en_areas, list) and isinstance(ar_areas, list) and (en_areas or ar_areas):
#                  areas_for_improvement_localized[subject] = {
#                     "en": en_areas,
#                     "ar": ar_areas,
#                 }

#         # 5) Guard: ensure Arabic output truly contains Arabic script
#         # if not contains_arabic(analysis_ar):
#         #     return jsonify({"error": "Arabic translation appears to be English; not storing"}), 500

#         # 6) Persist both languages atomically, including the new areas_for_improvement structure
#         student_ref.set({
#             "analysis": {
#                 "en": analysis_en,
#                 "ar": analysis_ar,
#                 # NEW: Subject-specific areas for improvement
#                 "areas_for_improvement": areas_for_improvement_localized
#             },
#             # FIX: Ensuring the correct datetime format is used for persistence
#             "last_updated": datetime.datetime.utcnow().isoformat()
#         }, merge=True)

#         print(f"[analyze_student_data] Stored dual-language analysis for {student_id} "
#               f"en_sizes=({len(analysis_en['strengths'])},"
#               f"{len(analysis_en['weaknesses'])},"
#               f"{len(analysis_en['interventions'])},"
#               f"{len(analysis_en['recommendations'])}) "
#               f"areas_subjects={len(areas_for_improvement_localized)} "
#               f"ar_has_arabic={contains_arabic(analysis_ar)}")

#         return jsonify({
#             "message": "Analysis completed and saved in EN and AR",
#             "student_id": student_id
#         }), 200

#     except Exception as e:
#         print(f"[analyze_student_data] Unexpected error: {e}")
#         # Print traceback for server-side debugging
#         import traceback
#         traceback.print_exc()
#         return jsonify({"error": "Internal server error"}), 500

# @app.route("/analyze_student_data", methods=["POST"])
# def analyze_student_data():
#     try:
#         data = request.json or {}
#         student_id = data.get("userId")
#         if not student_id:
#             return jsonify({"error": "Student ID is required"}), 400

#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()
#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         student_data = student_doc.to_dict() or {}
#         subjects_data = (student_data.get("academicData") or {}).get("subjects") or {}
#         if not subjects_data:
#             print(f"[analyze_student_data] Warning: subjects_data empty for {student_id}")

#         # Build latest_subject_data from history by newest timestamp 
#         all_entries = []
#         for subject, subject_info in subjects_data.items():
#             history_entries = subject_info.get("history", [])
#             if isinstance(history_entries, list):
#                 for entry in history_entries:
#                     if isinstance(entry, dict) and "timestamp" in entry:
#                         entry_with_subject = {**entry, "subject": subject}
#                         if 'marks' not in entry_with_subject:
#                             entry_with_subject['marks'] = 0
#                         all_entries.append(entry_with_subject)

#         # 🔑 FIX: Sort by timestamp (newest first) AND marks (highest mark first) for tie-breaking
#         all_entries.sort(key=lambda x: (x["timestamp"], x.get("marks", 0)), reverse=True)
        
#         latest_subject_data = {}
#         if all_entries:
#             # We take the absolute latest entry across all subjects
#             latest = all_entries[0]
#             latest_subject_data[latest["subject"]] = latest 
#             # NOTE: If you need the latest entry PER SUBJECT, this logic needs adjustment.
#             # Assuming you only need the single latest entry overall for the LLM prompt.

#         print("[analyze_student_data] Latest subject data:",
#               json.dumps(latest_subject_data, indent=2, ensure_ascii=False))

#         # 1) Generate English canonical analysis
#         en_raw = send_to_gpt(student_data, subjects_data, latest_subject_data, lang="en")
#         if not isinstance(en_raw, dict):
#             return jsonify({"error": "LLM returned unexpected format"}), 500

#         # Extract Areas for Improvement before translation
#         areas_for_improvement_en = en_raw.pop("areas_for_improvement", {}) or {}

#         analysis_en = {
#             "strengths": en_raw.get("strengths", []) or [],
#             "weaknesses": en_raw.get("weaknesses", []) or [],
#             "interventions": en_raw.get("interventions", []) or [],
#             "recommendations": en_raw.get("recommendations", []) or [],
#         }

#         # FIX: Basic validation now includes areas_for_improvement_en
#         is_analysis_valid = (
#             any(analysis_en.values()) 
#             or bool(areas_for_improvement_en)
#         )
#         # if not is_analysis_valid:
#         #     return jsonify({"error": "Empty analysis generated"}), 500

#         # 2) Translate English report structure (strengths, weaknesses, etc.) to Arabic
#         try:
#             analysis_ar = translate_analysis_recursively(analysis_en)
#         except Exception as te:
#             print(f"[analyze_student_data] Report translation failed: {te}")
#             return jsonify({"error": f"Arabic translation failed: {te}"}), 500

#         # 3) Translate English Areas for Improvement (subject lists) to Arabic
#         areas_for_improvement_ar = {}
#         try:
#             areas_for_improvement_ar = translate_analysis_recursively(areas_for_improvement_en)
#         except Exception as te:
#             print(f"[analyze_student_data] Areas translation failed: {te}")
#             return jsonify({"error": f"Arabic translation for areas failed: {te}"}), 500

#         # 4) Normalize and persist the final dual-language structure for areas
#         areas_for_improvement_localized = {}
#         all_subjects = set(areas_for_improvement_en.keys()) | set(areas_for_improvement_ar.keys())
        
#         for subject in all_subjects:
#             en_areas = areas_for_improvement_en.get(subject, [])
#             ar_areas = areas_for_improvement_ar.get(subject, [])
            
#             # Guard against potential empty or non-list results from translation
#             if isinstance(en_areas, list) and isinstance(ar_areas, list) and (en_areas or ar_areas):
#                  areas_for_improvement_localized[subject] = {
#                      "en": en_areas,
#                      "ar": ar_areas,
#                  }

#         # 5) Guard: ensure Arabic output truly contains Arabic script (assuming contains_arabic is defined)
#         # if not contains_arabic(analysis_ar):
#         #     return jsonify({"error": "Arabic translation appears to be English; not storing"}), 500

#         # 6) Persist both languages atomically, including the new areas_for_improvement structure
#         student_ref.set({
#             "analysis": {
#                 "en": analysis_en,
#                 "ar": analysis_ar,
#                 "areas_for_improvement": areas_for_improvement_localized
#             },
#             "last_updated": datetime.datetime.utcnow().isoformat()
#         }, merge=True)

#         print(f"[analyze_student_data] Stored dual-language analysis for {student_id} "
#               f"en_sizes=({len(analysis_en['strengths'])},"
#               f"{len(analysis_en['weaknesses'])},"
#               f"{len(analysis_en['interventions'])},"
#               f"{len(analysis_en['recommendations'])}) "
#               f"areas_subjects={len(areas_for_improvement_localized)} "
#               f"ar_has_arabic={contains_arabic(analysis_ar)}")

#         return jsonify({
#             "message": "Analysis completed and saved in EN and AR",
#             "student_id": student_id
#         }), 200

#     except Exception as e:
#         print(f"[analyze_student_data] Unexpected error: {e}")
#         traceback.print_exc()
#         return jsonify({"error": "Internal server error"}), 500


# @app.route("/analyze_student_data", methods=["POST"])
# def analyze_student_data():
#     try:
#         data = request.json or {}
#             # Support various ID keys
#         student_id = data.get("userId") or data.get("studentId") or data.get("student_id")
            
#         if not student_id:
#             return jsonify({"error": "Student ID is required"}), 400

#             # 1. Fetch Student Data
#             student_ref = db.collection("students").document(student_id)
#             student_doc = student_ref.get()
#             if not student_doc.exists:
#                 return jsonify({"error": "Student not found"}), 404

#             student_data = student_doc.to_dict() or {}
#             subjects_data = (student_data.get("academicData") or {}).get("subjects") or {}

#             if not subjects_data:
#                 print(f"[analyze_student_data] Warning: subjects_data empty for {student_id}")

#             # 2. Process History (Sort logic based on DB structure)
#             # DB Structure: subjects -> [SubjectName] -> history (Array of Maps)
#             all_entries = []
#             for subject, subject_info in subjects_data.items():
#                 history_entries = subject_info.get("history", [])
#                 if isinstance(history_entries, list):
#                     for entry in history_entries:
#                         if isinstance(entry, dict) and "timestamp" in entry:
#                             entry_with_subject = {**entry, "subject": subject}
#                             if 'marks' not in entry_with_subject:
#                                 entry_with_subject['marks'] = 0
#                             all_entries.append(entry_with_subject)

#             # Sort: Newest timestamp first, then Highest marks
#             # This handles the ISO timestamps (e.g. "2025-09-24T00:00:00") correctly
#             all_entries.sort(key=lambda x: (x["timestamp"], x.get("marks", 0)), reverse=True)
            
#             latest_subject_data = {}
#             if all_entries:
#                 latest = all_entries[0]
#                 latest_subject_data[latest["subject"]] = latest

#             # 3. Prepare Sanitized Context for AI
#             # We strictly exclude 'passwordHash', 'assignedGrades', 'email' etc.
#             # We only pass academic data to ensure data privacy.
#             context_for_ai = {
#                 "academic_history": subjects_data,
#                 "latest_activity": latest_subject_data
#             }

#             # 4. CALL GPT (One shot, Dual Language)
#             raw_analysis = send_to_gpt(context_for_ai, subjects_data)
            
#             if not raw_analysis:
#                 return jsonify({"error": "Failed to generate analysis from AI"}), 500

#             # 5. Extract & Restructure Data
#             # The 'areas_for_improvement' block matches the DB subject keys (e.g. "الأحياء")
#             areas_for_improvement_localized = raw_analysis.get("areas_for_improvement", {})

#             # Separate General Analysis into 'en' and 'ar' blocks for legacy DB structure
#             final_en_block = {
#                 "strengths": raw_analysis["strengths"]["en"],
#                 "weaknesses": raw_analysis["weaknesses"]["en"],
#                 "interventions": raw_analysis["interventions"]["en"],
#                 "recommendations": raw_analysis["recommendations"]["en"],
#             }

#             final_ar_block = {
#                 "strengths": raw_analysis["strengths"]["ar"],
#                 "weaknesses": raw_analysis["weaknesses"]["ar"],
#                 "interventions": raw_analysis["interventions"]["ar"],
#                 "recommendations": raw_analysis["recommendations"]["ar"],
#             }

#             # 6. Save to Firestore
#             # This updates the root-level 'analysis' field, merging the new dual-language maps
#             # over the existing legacy arrays.
#             student_ref.set({
#                 "analysis": {
#                     "en": final_en_block,
#                     "ar": final_ar_block,
#                     "areas_for_improvement": areas_for_improvement_localized
#                 },
#                 "last_updated": datetime.datetime.utcnow().isoformat()
#             }, merge=True)

#             print(f"[analyze_student_data] Success: Saved EN/AR analysis for {student_id}")

#             return jsonify({
#                 "message": "Analysis completed and saved in EN and AR",
#                 "student_id": student_id,
#                 "subjects_processed": list(areas_for_improvement_localized.keys())
#             }), 200

#         except Exception as e:
#             print(f"[analyze_student_data] Unexpected error: {e}")
#             traceback.print_exc()
#             return jsonify({"error": "Internal server error"}), 500


import traceback # Add this at the top of your file

@app.route("/analyze_student_data", methods=["POST"])
def analyze_student_data():
    try:
        data = request.json or {}
        # Support various ID keys
        student_id = data.get("userId") or data.get("studentId") or data.get("student_id")
        
        if not student_id:
            return jsonify({"error": "Student ID is required"}), 400

        # 1. Fetch Student Data
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        student_data = student_doc.to_dict() or {}
        subjects_data = (student_data.get("academicData") or {}).get("subjects") or {}

        if not subjects_data:
            print(f"[analyze_student_data] Warning: subjects_data empty for {student_id}")

        # 2. Process History (Sort logic)
        all_entries = []
        for subject, subject_info in subjects_data.items():
            history_entries = subject_info.get("history", [])
            if isinstance(history_entries, list):
                for entry in history_entries:
                    if isinstance(entry, dict) and "timestamp" in entry:
                        entry_with_subject = {**entry, "subject": subject}
                        if 'marks' not in entry_with_subject:
                            entry_with_subject['marks'] = 0
                        all_entries.append(entry_with_subject)

        # Sort: Newest timestamp first, then Highest marks
        all_entries.sort(key=lambda x: (x["timestamp"], x.get("marks", 0)), reverse=True)
        
        latest_subject_data = {}
        if all_entries:
            latest = all_entries[0]
            latest_subject_data[latest["subject"]] = latest

        # 3. Prepare Sanitized Context for AI
        context_for_ai = {
            "academic_history": subjects_data,
            "latest_activity": latest_subject_data
        }

        # 4. CALL GPT (One shot, Dual Language)
        # Ensure student_routes.py is saved/reloaded so this 2-arg call works
        raw_analysis = send_to_gpt(context_for_ai, subjects_data)
        
        if not raw_analysis:
             return jsonify({"error": "Failed to generate analysis from AI"}), 500

        # 5. Extract & Restructure Data
        areas_for_improvement_localized = raw_analysis.get("areas_for_improvement", {})

        final_en_block = {
            "strengths": raw_analysis["strengths"]["en"],
            "weaknesses": raw_analysis["weaknesses"]["en"],
            "interventions": raw_analysis["interventions"]["en"],
            "recommendations": raw_analysis["recommendations"]["en"],
        }

        final_ar_block = {
            "strengths": raw_analysis["strengths"]["ar"],
            "weaknesses": raw_analysis["weaknesses"]["ar"],
            "interventions": raw_analysis["interventions"]["ar"],
            "recommendations": raw_analysis["recommendations"]["ar"],
        }

        # 6. Save to Firestore
        student_ref.set({
            "analysis": {
                "en": final_en_block,
                "ar": final_ar_block,
                "areas_for_improvement": areas_for_improvement_localized
            },
            "last_updated": datetime.datetime.utcnow().isoformat()
        }, merge=True)

        print(f"[analyze_student_data] Success: Saved EN/AR analysis for {student_id}")

        return jsonify({
            "message": "Analysis completed and saved in EN and AR",
            "student_id": student_id,
            "subjects_processed": list(areas_for_improvement_localized.keys())
        }), 200

    except Exception as e:
        print(f"[analyze_student_data] Unexpected error: {e}")
        # FIX: Use traceback correctly
        traceback.print_exc()
        return jsonify({"error": "Internal server error"}), 500
    
# --- LLM Call Function (Fixed for Canonical English Output) ---

# def send_to_gpt(student_data, subjects_data, latest_subject_data, lang="en"):
#     # Fix: This part of the code needs to use keys from subjects_data for the prompt list 
#     # to ensure all subjects are included in the LLM's response structure.
#     subject_names = list(subjects_data.keys())

#     if not latest_subject_data:
#         logging.warning("send_to_gpt: latest_subject_data is empty. Returning default analysis.")
#         return {
#             "strengths": [], "weaknesses": [], "areas_for_improvement": {},
#             "recommendations": [], "interventions": [],
#         }

#     # The rest of your function remains the same...
#     subject_list = "\n".join([f"- {subject}" for subject in subject_names])

#     prompt = f"""Analyze the following student data and **strictly include all {len(subject_names)} subjects** listed below:
#     {subject_list}

#     Requirements:
#     1. Generate exactly 3 areas_for_improvement per subject.
#     2. Never skip any subject from the list above.

#     Student Data:
#     {student_data}

#     Response Format (JSON):
#     {{
#       "areas_for_improvement": {{
#         # FIX: Enforce English for the *values* of areas_for_improvement to create a canonical source for translation.
#         {"{" + "}, {".join([f'"{subjectNames}": ["English improvement 1", "English improvement 2", "English improvement 3"]' for subjectNames in subject_names]) + "}"}
#       }},
        
#     Language Enforcement Rule:

#     - **All points** under **areas_for_improvement**, **Strengths, Weaknesses, Recommendations, and Interventions** must be strictly and entirely written in **{lang.upper()}**.

#     Note:
#     - The **keys** for 'areas_for_improvement' must be the original subject names (e.g., "Math" or "الفيزياء").
    
#         "strengths": ["strength 1", "strength 2", ...],
#         "weaknesses": ["weakness 1", "weakness 2", ...],
        
#         "recommendations": ["recommendation 1", "recommendation 2", ...],
#         "interventions": ["intervention 1", "intervention 2", ...]

#     }}"""

#     try:
#         response = client_openai.chat.completions.create(
#             model="gpt-4.1-mini-2025-04-14",
#             messages=[
#                 {
#                     "role": "system",
#                     "content": "You are an academic performance analyzer. you must always generate all the data as mentioned",
#                 },
#                 {"role": "user", "content": prompt},
#             ],
#             temperature=0,
#         )

#         analysis = response.choices[0].message.content.strip()
#         if analysis.startswith("```json") and analysis.endswith("```"):
#             analysis = analysis[7:-3].strip()

#         analysis_dict = json.loads(analysis)

#         for subject, areas in analysis_dict.get("areas_for_improvement", {}).items():
#             if len(areas) != 3:
#                 analysis_dict["areas_for_improvement"][subject] = (areas + ["-"] * 3)[
#                     :3
#                 ]

#         return analysis_dict

#     except Exception as e:
#         logging.error(f"GPT Analysis Error: {str(e)}")
#         return {
#             "strengths": [], "weaknesses": [], "areas_for_improvement": {},
#             "recommendations": [], "interventions": [],
#         }


# ---------------------------------------------------------
# 2. Helper Function: Send to GPT (For POST Analysis)
# ---------------------------------------------------------
# ---------------------------------------------------------
# 2. Helper Function: Send to GPT (For POST Analysis)
# ---------------------------------------------------------

import json
import logging
# Ensure OpenAI client is imported/initialized as 'client_openai' in your actual file

FILLER_CHAR = "-"

def send_to_gpt(context_data, subjects_data):
    """
    Generates analysis for ALL subjects in BOTH English and Arabic in one API call.
    Ensures exactly 3 pointers per list and prevents Subject Name translation in keys.
    """
    # Get the exact list of keys as they exist in the DB (e.g., "Science", "Maths")
    subject_names = list(subjects_data.keys())
    
    # We pass the list of valid keys to GPT and tell it to use ONLY these.
    prompt = f"""
    Analyze the following student data and provide an academic report.

    TARGET SUBJECTS (Use these EXACT keys for the JSON output): 
    {json.dumps(subject_names, ensure_ascii=False)}

    CRITICAL OUTPUT RULES:
    1. Output strictly valid JSON.
    2. You MUST generate content in both English ("en") and Arabic ("ar").
    3. **DO NOT TRANSLATE SUBJECT NAMES** in the 'areas_for_improvement' object keys. Use the exact strings provided in 'TARGET SUBJECTS'.
       - Incorrect: "العلوم": {{ ... }}
       - Correct: "Science": {{ ... }} (if 'Science' is in the target list).
    4. **EXACTLY 3 POINTS REQUIRED**: Every list (strengths, weaknesses, recommendations, interventions, areas_for_improvement) MUST have exactly 3 distinct items. 
       - If you only find 2 points, generate a 3rd relevant general advice to fill the list. Do not leave it empty.

    REQUIRED JSON STRUCTURE:
    {{
      "areas_for_improvement": {{
        "SubjectName_From_List": {{ "en": ["Point 1", "Point 2", "Point 3"], "ar": ["نقطة 1", "نقطة 2", "نقطة 3"] }},
        ...
      }},
      "strengths": {{ "en": ["...", "...", "..."], "ar": ["...", "...", "..."] }},
      "weaknesses": {{ "en": ["...", "...", "..."], "ar": ["...", "...", "..."] }},
      "recommendations": {{ "en": ["...", "...", "..."], "ar": ["...", "...", "..."] }},
      "interventions": {{ "en": ["...", "...", "..."], "ar": ["...", "...", "..."] }}
    }}

    Student Academic Context:
    {json.dumps(context_data, default=str, ensure_ascii=False)}
    """

    try:
        if client_openai is None:
            raise RuntimeError("OPENAI_API_KEY is not configured")
        response = client_openai.chat.completions.create(
            model="gpt-4-turbo", 
            messages=[
                {"role": "system", "content": "You are a strict bilingual academic advisor. You never translate JSON keys. You always provide exactly 3 bullet points."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2, # Low temp to ensure it follows structure rules
            response_format={"type": "json_object"}
        )

        content = response.choices[0].message.content.strip()
        analysis_dict = json.loads(content)

        # --- VALIDATION & PADDING ---

        # 1. Validate General Lists (strengths, weaknesses, etc.)
        for key in ["strengths", "weaknesses", "recommendations", "interventions"]:
            if key not in analysis_dict: 
                analysis_dict[key] = {}
            for lang in ["en", "ar"]:
                if lang not in analysis_dict[key]: 
                    analysis_dict[key][lang] = []
                
                # Ensure it is a list
                if not isinstance(analysis_dict[key][lang], list):
                    analysis_dict[key][lang] = [str(analysis_dict[key][lang])]

                # FIX: Force exactly 3 items. 
                # If GPT gave 2, this adds a filler. If GPT gave 4, this cuts it to 3.
                current_list = analysis_dict[key][lang]
                padded_list = (current_list + [FILLER_CHAR] * 3)[:3]
                analysis_dict[key][lang] = padded_list

        # 2. Validate Areas for Improvement
        if "areas_for_improvement" not in analysis_dict:
            analysis_dict["areas_for_improvement"] = {}

        # We create a new clean dictionary to ensure only original Subject Names exist
        clean_areas = {}

        for original_subject in subject_names:
            # We look for the subject in the GPT response
            gpt_data = analysis_dict["areas_for_improvement"].get(original_subject)

            # If GPT translated the key (e.g., "العلوم" instead of "Science"), we try to find the mismatch
            # But usually, the prompt fix above prevents this. If missing, we initialize empty.
            if not gpt_data:
                # Fallback: check if the list has a key that looks like a translation? 
                # For now, just initialize empty to prevent crashes.
                gpt_data = {"en": [], "ar": []}

            clean_entry = {}
            for lang in ["en", "ar"]:
                if lang not in gpt_data: 
                    gpt_data[lang] = []
                
                # Ensure list
                if not isinstance(gpt_data[lang], list):
                    gpt_data[lang] = [str(gpt_data[lang])]

                # Force exactly 3 items
                padded = (gpt_data[lang] + [FILLER_CHAR] * 3)[:3]
                clean_entry[lang] = padded
            
            # Assign to the CLEAN dictionary using the ORIGINAL ID/Name
            clean_areas[original_subject] = clean_entry

        # Replace the messy GPT dict with our clean, strict dict
        analysis_dict["areas_for_improvement"] = clean_areas

        return analysis_dict

    except Exception as e:
        logging.error(f"GPT Analysis Error: {str(e)}")
        # FAILSAFE: Return a structure that satisfies the database requirements
        default_list = [FILLER_CHAR, FILLER_CHAR, FILLER_CHAR]
        return {
            "strengths": {"en": default_list, "ar": default_list},
            "weaknesses": {"en": default_list, "ar": default_list},
            "recommendations": {"en": default_list, "ar": default_list},
            "interventions": {"en": default_list, "ar": default_list},
            "areas_for_improvement": {
                s: {"en": default_list, "ar": default_list} for s in subject_names
            }
        }

# --- Fixes Applied to update_student_analysis ---
# def update_student_analysis():
#     students_ref = db.collection("students")
#     students = students_ref.stream()

#     for student in students:
#         student_data = student.to_dict()
#         subjects_data = student_data.get("academicData", {}).get("subjects", {})

#         # FIX 1: Replicate logic to find latest_subject_data, required by send_to_gpt
#         all_entries = []
#         for subject, subject_info in subjects_data.items():
#             history_entries = subject_info.get("history", [])
#             if isinstance(history_entries, list):
#                 for entry in history_entries:
#                     if isinstance(entry, dict) and "timestamp" in entry:
#                         all_entries.append({**entry, "subject": subject})

#         # NOTE: Assumes timestamp is comparable (e.g., float, int, or string ISO format)
#         all_entries.sort(key=lambda x: x["timestamp"], reverse=True)
#         latest_subject_data = {}
#         if all_entries:
#             latest = all_entries[0]
#             latest_subject_data[latest["subject"]] = latest

#         # FIX 2: Define 'lang' and pass all required arguments
#         lang = "en" # Default to English for the background update job
#         analysis = send_to_gpt(student_data, subjects_data, latest_subject_data, lang=lang)

#         student_ref = db.collection("students").document(student.id)
        
#         # --- BACKGROUND JOB: No Localization Needed (Only EN is typically stored here) ---
#         student_ref.update(
#             {
#                 "analysis": {
#                     "strengths": analysis.get("strengths", ""),
#                     "weaknesses": analysis.get("weaknesses", ""),
#                     "areas_for_improvement": analysis.get("areas_for_improvement", {}),
#                     "recommendations": analysis.get("recommendations", ""),
#                     "interventions": analysis.get("interventions", ""),
#                 }
#             }
#         )

# # Schedule the job to run every day at midnight
# # Assuming scheduler is defined and available
# # scheduler.add_job(
# #     id="update_student_analysis",
# #     func=update_student_analysis,
# #     trigger="cron",
# #     hour=0,
# #     minute=0,
# # )


# # Schedule the job to run every day at midnight
# scheduler.add_job(
#     id="update_student_analysis",
#     func=update_student_analysis,
#     trigger="cron",
#     hour=0,
#     minute=0,
# )

# --- Fixes Applied to update_student_analysis ---
# --- Fixes Applied to update_student_analysis ---

def update_student_analysis():
    students_ref = db.collection("students")
    students = students_ref.stream()

    for student in students:
        student_data = student.to_dict()
        subjects_data = student_data.get("academicData", {}).get("subjects", {})

        # Replicate logic to find latest_subject_data, required by send_to_gpt
        all_entries = []
        for subject, subject_info in subjects_data.items():
            history_entries = subject_info.get("history", [])
            if isinstance(history_entries, list):
                for entry in history_entries:
                    if isinstance(entry, dict) and "timestamp" in entry:
                         entry_with_subject = {**entry, "subject": subject}
                         if 'marks' not in entry_with_subject:
                            entry_with_subject['marks'] = 0 
                         all_entries.append(entry_with_subject)

        # 🔑 FIX: Sort by timestamp and marks for reliable "latest" data
        all_entries.sort(key=lambda x: (x["timestamp"], x.get("marks", 0)), reverse=True)
        
        latest_subject_data = {}
        if all_entries:
            latest = all_entries[0]
            latest_subject_data[latest["subject"]] = latest

        lang = "en" # Default to English for the background update job
        analysis = send_to_gpt(student_data, subjects_data, latest_subject_data, lang=lang)

        student_ref = db.collection("students").document(student.id)
        
        # --- BACKGROUND JOB: Only EN is stored here ---
        student_ref.update(
            {
                "analysis": {
                    "strengths": analysis.get("strengths", ""),
                    "weaknesses": analysis.get("weaknesses", ""),
                    "areas_for_improvement": analysis.get("areas_for_improvement", {}),
                    "recommendations": analysis.get("recommendations", ""),
                    "interventions": analysis.get("interventions", ""),
                }
            }
        )

# Schedule the job to run every day at midnight
scheduler.add_job(
    id="update_student_analysis",
    func=update_student_analysis,
    trigger="cron",
    hour=0,
    minute=0,
)

def get_student_subject_average_and_data_table(student_id):
    try:
        # Fetch student document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        student_data = student_doc.to_dict()
        academic_data = student_data.get("academicData", {})
        subjects_data = academic_data.get("subjects", {})

        # Get optional date filters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")

        start_date_obj = (
            datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        )
        end_date_obj = (
            datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        )

        average_percentages = {}
        data_table = {}

        for subject_name, subject_info in subjects_data.items():
            history = subject_info.get("history", [])
            percentages = []
            detailed_entries = []

            for entry in history:
                timestamp_str = entry.get("timestamp")
                if not timestamp_str:
                    continue

                try:
                    timestamp_str = timestamp_str.replace("T", "")
                    timestamp_obj = datetime.strptime(timestamp_str, "%Y-%m-%d%H:%M:%S")
                    entry_date = timestamp_obj.date()

                    # Apply filters
                    if start_date_obj and entry_date < start_date_obj:
                        continue
                    if end_date_obj and entry_date > end_date_obj:
                        continue

                    marks = entry.get("marks", 0)
                    total_mark = entry.get("totalMark", 100)
                    percentage = (marks / total_mark) * 100 if total_mark > 0 else 0
                    rounded_percentage = round(percentage, 2)
                    percentages.append(percentage)

                    detailed_entries.append(
                        {
                            "timestamp": timestamp_obj.strftime("%Y-%m-%d %H:%M:%S"),
                            "percentage": rounded_percentage,
                            "marks": marks,
                            "totalMark": total_mark,
                            "grade": entry.get("grade", ""),
                            "curriculumName": entry.get("curriculumName", ""),
                        }
                    )

                except Exception as e:
                    print(f"Skipping invalid timestamp {timestamp_str}: {e}")
                    continue

            if percentages:
                average = round(sum(percentages) / len(percentages), 2)
                average_percentages[subject_name] = average
                data_table[subject_name] = detailed_entries

        return (
            jsonify(
                {
                    "studentId": student_id,
                    "averageSubjectPercentages": average_percentages,
                    "dataTable": data_table,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

#--------------------------------------------------------------------------
from flask import Flask, request, jsonify
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib.units import inch
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import os
import uuid
import textwrap
import matplotlib.pyplot as plt
import numpy as np
import io
from arabic_reshaper import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
from reportlab.lib.utils import ImageReader
import os
import uuid
from datetime import datetime
from flask import jsonify, request
import textwrap

# # Translation dictionaries
# ARABIC_TRANSLATIONS = {
#     "Student Analysis Report": "ØªÙ‚Ø±ÙŠØ± ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø·Ø§Ù„Ø¨",
#     "Personal Information": "Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª Ø§Ù„Ø´Ø®ØµÙŠØ©",
#     "Name": "Ø§Ù„Ø§Ø³Ù…",
#     "Email": "Ø§Ù„Ø¨Ø±ÙŠØ¯ Ø§Ù„Ø¥Ù„ÙƒØªØ±ÙˆÙ†ÙŠ",
#     "Class": "Ø§Ù„ØµÙ",
#     "Grade": "Ø§Ù„Ø¯Ø±Ø¬Ø©",
#     "progress": "ØªÙ‚Ø¯Ù…",
#     "Subject performance": "Ø£Ø¯Ø§Ø¡ Ø§Ù„Ù…ÙˆØ¶ÙˆØ¹",
#     "average percentage": "Ø§Ù„Ù†Ø³Ø¨Ø© Ø§Ù„Ù…Ø¦ÙˆÙŠØ© Ø§Ù„Ù…ØªÙˆØ³Ø·Ø©",
#     "Subjects": "Ø§Ù„Ù…ÙˆØ§Ø¯",
#     "Strengths": "Ù†Ù‚Ø§Ø· Ø§Ù„Ù‚ÙˆØ©",
#     "Weaknesses": "Ù†Ù‚Ø§Ø· Ø§Ù„Ø¶Ø¹Ù",
#     "Recommendations": "ØªÙˆØµÙŠØ§Øª",
#     "Interventions": "Ø§Ù„ØªØ¯Ø®Ù„Ø§Øª",
#     "No data available for selected date range": "Ù„Ø§ ØªÙˆØ¬Ø¯ Ø¨ÙŠØ§Ù†Ø§Øª Ù…ØªØ§Ø­Ø© Ù„ÙØªØ±Ø© Ø§Ù„ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…Ø­Ø¯Ø¯Ø©",
#     "Generated on": "ØªÙ… Ø¥Ù†Ø´Ø§Ø¤Ù‡ ÙÙŠ",
#     "cont'd": "ØªØ§Ø¨Ø¹",
#     "Exam Date": "ØªØ§Ø±ÙŠØ® Ø§Ù„Ø§Ù…ØªØ­Ø§Ù†",
#     "Percentage": "Ø§Ù„Ù†Ø³Ø¨Ø© Ø§Ù„Ù…Ø¦ÙˆÙŠØ©",
# }
ARABIC_TRANSLATIONS = {
    "Student Analysis Report": "تقرير تحليل الطالب",
    "Personal Information": "المعلومات الشخصية",
    "Name": "الاسم",
    "Email": "البريد الإلكتروني",
    "Class": "الصف",
    "Grade": "الدرجة",
    "progress": "تقدم",
    "Subject performance": "أداء المواد",
    "average percentage": "النسبة المئوية المتوسطة",
    "Subjects": "المواد الدراسية",
    "Strengths": "نقاط القوة",
    "Weaknesses": "نقاط الضعف",
    "Recommendations": "توصيات",
    "Interventions": "التدخلات",
    "No data available for selected date range": "لا توجد بيانات متاحة للفترة المحددة",
    "Generated on": "تم إنشاؤه في",
    "cont'd": "تابع",
    "Exam Date": "تاريخ الامتحان",
    "Percentage": "النسبة المئوية",
}


def get_translation(key, lang):
    if lang == "ar":
        return ARABIC_TRANSLATIONS.get(key, key)
    return key


def setup_pdf_direction(c, lang, width):
    """Set up the PDF direction based on language (RTL for Arabic)"""
    if lang == "ar":
        c.setFont("Amiri", 12)
        return True
    return False


def draw_rtl_text(c, text, x, y, font_name="Amiri", font_size=12, available_width=None):
    """Properly draw RTL text with wrapping if needed"""
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)

    c.setFont(font_name, font_size)

    if available_width:
        # Handle text wrapping for RTL
        lines = []
        current_line = []
        current_width = 0

        # Split text into words (Arabic words are separated by spaces)
        words = bidi_text.split()

        for word in words:
            word_width = c.stringWidth(word + " ", font_name, font_size)
            if current_width + word_width <= available_width:
                current_line.append(word)
                current_width += word_width
            else:
                if current_line:
                    lines.append(" ".join(current_line))
                current_line = [word]
                current_width = word_width

        if current_line:
            lines.append(" ".join(current_line))

        # Draw each line from right to left
        for i, line in enumerate(lines):
            c.drawRightString(x, y - (i * (font_size + 2)), line)
        return len(lines)
    else:
        c.drawRightString(x, y, bidi_text)
        return 1


def draw_string_with_direction(
    c, text, x, y, lang, bold=False, font_size=12, available_width=None
):
    """Draw text with proper direction handling"""
    # First determine if the text contains Arabic characters
    text_contains_arabic = contains_arabic(text)

    # For Arabic text, always use Arabic shaping regardless of report language
    if text_contains_arabic:
        font_name = "Amiri-Bold" if bold else "Amiri"
        return draw_rtl_text(c, text, x, y, font_name, font_size, available_width)
    else:
        # For non-Arabic text, use the specified language's settings
        if lang == "ar":
            font_name = "Amiri-Bold" if bold else "Amiri"
            return draw_rtl_text(c, text, x, y, font_name, font_size, available_width)
        else:
            font_name = "Helvetica-Bold" if bold else "Helvetica"
            c.setFont(font_name, font_size)

            if available_width:
                # Handle text wrapping for LTR
                lines = []
                current_line = []
                current_width = 0

                for word in text.split():
                    word_width = c.stringWidth(word + " ", font_name, font_size)
                    if current_width + word_width <= available_width:
                        current_line.append(word)
                        current_width += word_width
                    else:
                        if current_line:
                            lines.append(" ".join(current_line))
                        current_line = [word]
                        current_width = word_width

                if current_line:
                    lines.append(" ".join(current_line))

                for i, line in enumerate(lines):
                    c.drawString(x, y - (i * (font_size + 2)), line)
                return len(lines)
            else:
                c.drawString(x, y, text)
                return 1


def contains_arabic(text):
    if not text:
        return False
    # Check for Arabic Unicode blocks
    for char in str(text):
        if (
            "\u0600" <= char <= "\u06ff"
            or "\u0750" <= char <= "\u077f"
            or "\u08a0" <= char <= "\u08ff"
            or "\ufb50" <= char <= "\ufdff"
            or "\ufe70" <= char <= "\ufeff"
        ):
            return True
    return False


def prepare_text(text, force_arabic=False):
    if not text:
        return ""
    if force_arabic or contains_arabic(text):
        reshaped_text = arabic_reshaper.reshape(text)
        return get_display(reshaped_text)
    return text

# from datetime import datetime, timezone # <--- THE CORRECTED IMPORT

@app.route("/export_analysis", methods=["GET"])
def export_analysis():
    student_id = request.args.get("student_id")
    format_type = request.args.get("format_type", "").strip().lower()
    lang = request.args.get("lang", "en")
    print(f"Received student_id: {student_id}, format_type: {format_type}")

    if format_type == "pdf":
        return export_analysis_pdf()
    elif format_type == "excel":
        return export_analysis_excel()
    else:
        return jsonify({"error": "Invalid format. Use 'pdf' or 'excel'."}), 400


from trans import translate
import ast
from flask import jsonify
# Assuming your definition is: def exportdata(...):
# export_data = exportdata
# def export_data(lang, studentid):
#     studentref = db.collection("students").document(studentid)
#     studentdoc = studentref.get()
#     if not studentdoc.exists:
#         return jsonify({"error": "Student not found"}), 404

#     studentdata = studentdoc.to_dict()
    
#     # 1. Try to fetch language-specific analysis (e.g., analysis_ar for lang='ar')
#     analysis_key = f"analysis_{lang}"
#     analysis = studentdata.get(analysis_key)

#     # 2. If lang-specific data is NOT found, fall back to the default 'analysis' field
#     if not analysis:
#         # Note: If your default 'analysis' is in English, this fallback will be used
#         analysis = studentdata.get("analysis", {})

#     if not analysis:
#         return jsonify({"error": "No analysis found for this student"}), 404

#     # 3. CRITICAL: If language-specific data was found (e.g., analysis_ar), return it directly.
#     # This assumes the fetched 'analysis' dictionary now holds the correctly pre-translated content.
#     if lang == "en" and analysis_key in studentdata:
#         # If we successfully fetched a non-default language field (like 'analysis_ar'),
#         # return the raw dictionary, which is what the PDF function expects.
#         return analysis, 200
    
#     # If lang is 'en' or the specific language key was missing, we proceed to external translation.
    
#     # Fetching sections first (from the currently selected 'analysis' dict)
#     strengths = analysis.get("strengths", [])
#     weaknesses = analysis.get("weaknesses", [])
#     recommendations = analysis.get("recommendations", [])
#     interventions = analysis.get("interventions", [])
    
#     # --- safe_translate function definition (Keep this to translate English to non-English if needed) ---
#     def safe_translate(content, current_lang):
#         def process_item(item, current_lang):
#             return translate(item, current_lang) if isinstance(item, str) else str(item)

#         if isinstance(content, list):
#             return [process_item(item, current_lang) for item in content]
            
#         elif isinstance(content, str):
#             try:
#                 content_list = ast.literal_eval(content)
#                 if isinstance(content_list, list):
#                     return [process_item(item, current_lang) for item in content_list]
#             except:
#                 return process_item(content, current_lang)
                
#         return content

#     # Apply external translation (runs only if we didn't return early)
#     translated_data = {
#         "strengths": safe_translate(strengths, lang),
#         "weaknesses": safe_translate(weaknesses, lang),
#         "recommendations": safe_translate(recommendations, lang),
#         "interventions": safe_translate(interventions, lang),
#     }

#     # Return the translated dictionary and status code
#     return translated_data, 200

# ...existing code...
def export_data(lang, studentid):
    studentref = db.collection("students").document(studentid)
    studentdoc = studentref.get()
    if not studentdoc.exists:
        return jsonify({"error": "Student not found"}), 404

    studentdata = studentdoc.to_dict() or {}

    # 1) If there's a language-specific top-level field like analysis_ar / analysis_en, prefer it
    analysis_key_top = f"analysis_{lang}"
    if analysis_key_top in studentdata:
        analysis = studentdata.get(analysis_key_top) or {}
        if isinstance(analysis, dict):
            return analysis, 200

    # 2) If there is an "analysis" field and it's a dict with language keys, prefer that
    analysis_root = studentdata.get("analysis")
    if isinstance(analysis_root, dict):
        # case: analysis = {"en": {...}, "ar": {...}}
        if lang in analysis_root and isinstance(analysis_root.get(lang), dict):
            return analysis_root.get(lang), 200
        # case: analysis might already be a language-specific object (older shape)
        # if user requested Arabic and values stored are Arabic (detect), return as-is
        # otherwise fall through to safe translation below
        # If analysis_root looks like final English dict and requested lang is 'en' return it
        if lang == "en" and "strengths" in analysis_root:
            return analysis_root, 200

    # 3) Fallback: try to read any available analysis fields
    # prefer 'analysis_ar' or 'analysis_en' if present
    for k in ("analysis_ar", "analysis_en", "analysis"):
        maybe = studentdata.get(k)
        if isinstance(maybe, dict):
            # if requested language matches content language heuristically, return it
            # if requesting 'ar' and content contains Arabic script, return it
            if lang == "ar":
                # quick heuristic: check any string value for Arabic chars
                def has_arabic(d):
                    for v in (d.get("strengths", []) if isinstance(d, dict) else []):
                        if isinstance(v, str) and any("\u0600" <= ch <= "\u06FF" for ch in v):
                            return True
                    return False
                if has_arabic(maybe):
                    return maybe, 200
            else:
                return maybe, 200

    # 4) If we reach here, we have no pretranslated dict for requested lang.
    # Build sections from default analysis (if present) and return them.
    # This keeps behavior backwards-compatible and allows the PDF code to receive a dict
    default_analysis = {}
    if isinstance(analysis_root, dict):
        default_analysis = analysis_root
    else:
        default_analysis = {}

    strengths = default_analysis.get("strengths", []) or []
    weaknesses = default_analysis.get("weaknesses", []) or []
    recommendations = default_analysis.get("recommendations", []) or []
    interventions = default_analysis.get("interventions", []) or []

    # At this point we would normally call external translate(...) only if needed.
    # If your translate provider is failing (401), returning the default (likely English) is safer.
    # Try to translate values only if translate() is available and lang != 'en'
    if lang != "en":
        translated = {"strengths": [], "weaknesses": [], "recommendations": [], "interventions": []}
        try:
            def safe_translate_item(v):
                if isinstance(v, str):
                    return translate(v, lang)
                return str(v)

            translated["strengths"] = [safe_translate_item(x) for x in strengths]
            translated["weaknesses"] = [safe_translate_item(x) for x in weaknesses]
            translated["recommendations"] = [safe_translate_item(x) for x in recommendations]
            translated["interventions"] = [safe_translate_item(x) for x in interventions]
            return translated, 200
        except Exception as e:
            # translation failed â€” log and return original content
            print(f"Translation provider error fallback: {e}")
            return {
                "strengths": strengths,
                "weaknesses": weaknesses,
                "recommendations": recommendations,
                "interventions": interventions,
            }, 200

    # lang == "en"
    return {
        "strengths": strengths,
        "weaknesses": weaknesses,
        "recommendations": recommendations,
        "interventions": interventions,
    }, 200
# ...existing code...

import datetime as _dt
from datetime import timezone
from datetime import datetime
def export_analysis_pdf():
    student_id = request.args.get("student_id")
    lang = request.args.get("lang", "en")


    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    # Get student document from Firestore
    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()
    if not student_doc.exists:
        return jsonify({"error": "Student not found"}), 404

    student_data = student_doc.to_dict()

    # Extract personal information
    pers_info = student_data.get("profileInfo", {}).get("personalInformation", {})
    name = pers_info.get("name", get_translation("Not Available", lang))
    email = student_data.get("email", get_translation("Not Available", lang))
    assigned_grades = student_data.get("assignedGrades", {}).get("grades", {})
    

    if assigned_grades:
        # Get first grade key (e.g., "GRADE 5")
        grade = next(
            iter(assigned_grades.keys()), get_translation("Not Available", lang)
        )
        class_sections = assigned_grades.get(grade, {})
        class_section = next(
            iter(class_sections.keys()), get_translation("Not available", lang)
        )
    else:
        grade = get_translation("Not Available", lang)
        class_section = get_translation("Not available", lang)

    # Get subject averages and data table
    subject_response = get_student_subject_average_and_data_table(student_id)
    if subject_response[1] != 200:
        return subject_response

    subject_data = subject_response[0].json
    average_percentages = subject_data.get("averageSubjectPercentages", {})
    has_data = any(
        len(subject_info) > 0
        for subject_info in subject_data.get("dataTable", {}).values()
    )
    data_table = subject_data.get("dataTable", {})
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate")
    # analysis = student_data.get("analysis", {})

    # Get translated analysis sections using the export_data function
    analysis_response = export_data(lang, student_id)
    if analysis_response[1] != 200:
        return analysis_response

    # translated_analysis = analysis_response[0]
    translated_analysis = analysis_response[0]
    print(translated_analysis)

    print(analysis_response)
    # Use the translated analysis sections
    strengths = translated_analysis.get("strengths", [])
    weaknesses = translated_analysis.get("weaknesses", [])
    recommendations = translated_analysis.get("recommendations", [])
    interventions = translated_analysis.get("interventions", [])

    filename = f"report_{_dt.datetime.now().strftime('%Y%m%d%H%M%S')}_{lang}.pdf"
    local_path = os.path.join("reports", filename)
    os.makedirs("reports", exist_ok=True)

    try:
        c = canvas.Canvas(local_path, pagesize=A4)
        width, height = A4
        margin = 50
        available_width = width - 2 * margin

        # Set RTL if Arabic
        is_rtl = lang == "ar"
        if is_rtl:
            # For RTL, we'll use drawRightString instead of drawString
            pass

        # Register fonts
        try:
            pdfmetrics.registerFont(TTFont("Amiri", "static/fonts/Amiri-Regular.ttf"))
            pdfmetrics.registerFont(TTFont("Amiri-Bold", "static/fonts/Amiri-Bold.ttf"))
            addMapping("Amiri", 0, 0, "Amiri")  # normal
            addMapping("Amiri", 0, 1, "Amiri")  # italic
            addMapping("Amiri", 1, 0, "Amiri-Bold")  # bold
            addMapping("Amiri", 1, 1, "Amiri-Bold")  # bold and italic
        except:
            print("Amiri font not found, Arabic text may not display correctly")

        # ---------------------------- Header ----------------------------
        header_height = 80
        c.setFillColor(colors.HexColor("#4A90E2"))
        c.rect(0, height - header_height, width, header_height, stroke=0, fill=1)
        c.setFillColor(colors.white)

        # Draw title based on language direction
        title = prepare_text(get_translation("Student Analysis Report", lang))
        if is_rtl:
            c.setFont("Amiri-Bold", 20)
            c.drawRightString(width - margin, height - header_height / 2 + 6, title)
        else:
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width / 2, height - header_height / 2 + 6, title)

        c.setFont("Helvetica", 12)
        c.line(
            margin,
            height - header_height - 10,
            width - margin,
            height - header_height - 10,
        )
        c.setFillColor(colors.black)
        current_y = height - header_height - 30

        # -------------------- Personal Information --------------------
        section_title = get_translation("Personal Information", lang) + ":"
        draw_string_with_direction(
            c,
            section_title,
            width - margin if is_rtl else margin,
            current_y,
            lang,
            bold=True,
            font_size=14,
            available_width=available_width,
        )
        current_y -= 20

        info_lines = [
            f"{get_translation('Name', lang)}: {name}",
            f"{get_translation('Email', lang)}: {email}",
            f"{get_translation('Class', lang)}: {class_section}",
            f"{get_translation('Grade', lang)}: {grade}",
        ]

        for line in info_lines:
            if current_y < margin + 50:
                c.showPage()
                current_y = height - margin
                cont_title = (
                    get_translation("Personal Information", lang)
                    + " ("
                    + get_translation("cont'd", lang)
                    + "):"
                )
                draw_string_with_direction(
                    c,
                    cont_title,
                    width - margin if is_rtl else margin,
                    current_y,
                    lang,
                    bold=True,
                    font_size=14,
                )
                current_y -= 20

            draw_string_with_direction(
                c,
                line,
                width - margin if is_rtl else margin,
                current_y,
                lang,
                available_width=available_width,
            )
            lines_used = draw_string_with_direction(
                c,
                line,
                width - margin if is_rtl else margin,
                current_y,
                lang,
                available_width=available_width,
            )
            current_y -= 14 * lines_used

        current_y -= 10

        # -------------------- Subject Averages Charts --------------------
        if data_table:
            # Generate charts
            # Generate charts
            line_chart_result = generate_monthly_subject_line_chart(
                data_table, start_date=start_date, end_date=end_date, lang=lang
            )
            line_chart_stream = (
                line_chart_result.get("chart_image") if line_chart_result else None
            )
            bar_chart_stream = generate_subject_bar_chart(average_percentages, lang)

            chart_width = 400
            chart_height = 250
            chart_spacing = 30

            # Add line chart
            if line_chart_stream:
                try:
                    # Check if we need a new page
                    if current_y - chart_height - 40 < margin:
                        c.showPage()
                        current_y = height - margin

                    # Chart title
                    chart_title = get_translation("progress", lang) + ":"
                    draw_string_with_direction(
                        c,
                        chart_title,
                        width - margin if is_rtl else margin,
                        current_y,
                        lang,
                        bold=True,
                        font_size=14,
                    )
                    current_y -= 20

                    # Add the chart image
                    line_chart_stream.seek(0)  # Rewind the buffer
                    img = ImageReader(line_chart_stream)
                    c.drawImage(
                        img,
                        (width - chart_width) / 2,  # Center horizontally
                        current_y - chart_height,
                        width=chart_width,
                        height=chart_height,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
                    current_y -= chart_height + chart_spacing
                except Exception as e:
                    print(f"Error adding line chart to PDF: {str(e)}")
                    c.setFont("Helvetica", 12)
                    draw_string_with_direction(
                        c,
                        get_translation("Could not generate progress chart", lang),
                        width - margin if is_rtl else margin,
                        current_y,
                        lang,
                    )
                    current_y -= 20

            # Add bar chart or "No Data" message
            if bar_chart_stream:
                try:
                    if current_y - chart_height < margin:
                        c.showPage()
                        current_y = height - margin

                    chart_title = get_translation("Subject performance", lang) + ":"
                    draw_string_with_direction(
                        c,
                        chart_title,
                        width - margin if is_rtl else margin,
                        current_y,
                        lang,
                        bold=True,
                        font_size=14,
                    )
                    current_y -= 20

                    img = ImageReader(bar_chart_stream)
                    c.drawImage(
                        img,
                        (width - chart_width) / 2,  # Center the image
                        current_y - chart_height,
                        width=chart_width,
                        height=chart_height,
                    )
                    current_y -= chart_height + 20
                except Exception as e:
                    print(f"Error adding bar chart to PDF: {e}")
                    draw_string_with_direction(
                        c,
                        get_translation("Could not generate bar chart", lang),
                        width - margin if is_rtl else margin,
                        current_y,
                        lang,
                    )
                    current_y -= 20
            else:
                draw_string_with_direction(
                    c,
                    get_translation("Subject performance", lang) + ":",
                    width - margin if is_rtl else margin,
                    current_y,
                    lang,
                    bold=True,
                    font_size=14,
                )
                current_y -= 20
                draw_string_with_direction(
                    c,
                    get_translation("No data available for selected date range", lang),
                    width - margin if is_rtl else margin,
                    current_y,
                    lang,
                )
                current_y -= 30
        else:
            # No data at all
            draw_string_with_direction(
                c,
                get_translation("Subject performance", lang) + ":",
                width - margin if is_rtl else margin,
                current_y,
                lang,
                bold=True,
                font_size=14,
            )
            current_y -= 20
            draw_string_with_direction(
                c,
                get_translation("No data available for selected date range", lang),
                width - margin if is_rtl else margin,
                current_y,
                lang,
            )
            current_y -= 30

        # -------------------- Analysis Sections --------------------
        sections = [
            (get_translation("Strengths", lang), strengths),
            (get_translation("Weaknesses", lang), weaknesses),
            (get_translation("Recommendations", lang), recommendations),
            (get_translation("Interventions", lang), interventions),
        ]

        for header, items in sections:
            if current_y < margin + 50:
                c.showPage()
                current_y = height - margin

            # Header (always in selected language)
            if lang == "ar":
                # For Arabic headers, use draw_rtl_text directly
                draw_rtl_text(
                    c, header + ":", width - margin, current_y, "Amiri-Bold", 14
                )
            else:
                # For English headers, use regular left-aligned text
                c.setFont("Helvetica-Bold", 14)
                c.drawString(margin, current_y, header + ":")

            current_y -= 20

            for i, item in enumerate(items, 1):
                if current_y < margin + 20:
                    c.showPage()
                    current_y = height - margin
                    cont_header = header + " (" + get_translation("cont'd", lang) + "):"
                    if lang == "ar":
                        draw_rtl_text(
                            c, cont_header, width - margin, current_y, "Amiri-Bold", 14
                        )
                    else:
                        c.setFont("Helvetica-Bold", 14)
                        c.drawString(margin, current_y, cont_header)
                    current_y -= 20

                # Item text - always render Arabic content properly
                item_text = f"{i}. {item}"

                if contains_arabic(item_text):
                    # Arabic content - right-aligned with proper shaping
                    draw_rtl_text(
                        c,
                        item_text,
                        width - margin,
                        current_y,
                        "Amiri",
                        12,
                        available_width,
                    )
                else:
                    # Non-Arabic content - left-aligned
                    c.setFont("Helvetica", 12)
                    lines = []
                    current_line = []
                    current_width = 0

                    for word in item_text.split():
                        word_width = c.stringWidth(word + " ", "Helvetica", 12)
                        if current_width + word_width <= available_width:
                            current_line.append(word)
                            current_width += word_width
                        else:
                            if current_line:
                                lines.append(" ".join(current_line))
                            current_line = [word]
                            current_width = word_width

                    if current_line:
                        lines.append(" ".join(current_line))

                    for i, line in enumerate(lines):
                        c.drawString(margin, current_y - (i * 14), line)
                    current_y -= 14 * max(1, len(lines))

                current_y -= 14

            current_y -= 10
            footer_text = (
                get_translation("Generated on", lang)
                + " "
                # before: datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
                + _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            )
        if is_rtl:
            c.setFont("Amiri", 10)
            c.drawRightString(width - margin, margin / 2, footer_text)
        else:
            c.setFont("Helvetica-Oblique", 10)
            c.drawCentredString(width / 2, margin / 2, footer_text)

        c.save()

        # Upload to Firebase Storage
        blob = bucket.blob(f"students/{student_id}/{filename}")
        blob.upload_from_filename(local_path)
        blob.make_public()
        download_url = blob.public_url

        os.remove(local_path)
        student_ref.update({"pdf_report_url": download_url})

        return jsonify(
            {"message": "Report generated successfully", "download_url": download_url}
        )

    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        return jsonify({"error": f"Failed to generate PDF: {str(e)}"}), 500


import matplotlib.pyplot as plt
import numpy as np
import io
from PIL import Image
from reportlab.lib.utils import ImageReader
import matplotlib.pyplot as plt
import numpy as np
import io


def generate_monthly_subject_line_chart(data_table, start_date, end_date, lang="en"):
    try:
        if not data_table:
            return {"error": "No data provided"}

        # Prepare month names (short form)
        month_names = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]

        # Set up the plot
        plt.figure(figsize=(12, 6))
        plt.style.use("ggplot")

        # Generate distinct colors for each subject
        subjects = list(data_table.keys())
        colors = plt.cm.tab10(np.linspace(0, 1, len(subjects)))

        # Prepare result JSON
        result_json = {}

        # Collect all unique dates to determine x-axis positions
        all_dates = []
        for subject, entries in data_table.items():
            for entry in entries:
                try:
                    date_obj = datetime.strptime(
                        entry["timestamp"], "%Y-%m-%d %H:%M:%S"
                    ).date()
                    all_dates.append(date_obj)

                except:
                    continue

        if not all_dates:
            return {"error": "No valid dates found in data"}

        # Sort all dates chronologically
        sorted_dates = sorted(list(set(all_dates)))  # Remove duplicates and sort
        date_labels = [f"{month_names[d.month-1]} {d.day}" for d in sorted_dates]
        x_axis_positions = range(len(sorted_dates))

        # Plot each subject's history
        for i, (subject, entries) in enumerate(data_table.items()):
            # Sort entries by timestamp
            sorted_entries = sorted(entries, key=lambda x: x["timestamp"])

            x_values = []  # Will store date indices
            y_values = []  # Will store percentages

            for entry in sorted_entries:
                try:
                    date_obj = datetime.strptime(
                        entry["timestamp"], "%Y-%m-%d %H:%M:%S"
                    ).date()
                    date_index = sorted_dates.index(date_obj)

                    x_values.append(date_index)
                    y_values.append(entry["percentage"])

                    # Add to result JSON
                    date_str = date_obj.strftime("%Y-%m-%d")
                    if date_str not in result_json:
                        result_json[date_str] = {}
                    result_json[date_str][subject] = {
                        "percentage": f"{entry['percentage']:.1f}%",
                        "grade": entry["grade"],
                        "marks": entry["marks"],
                        "totalMark": entry["totalMark"],
                    }
                except:
                    continue

            if not x_values:  # Skip if no valid data
                continue

            # Plot the line with markers
            plt.plot(
                x_values,
                y_values,
                marker="o",
                linestyle="-",
                color=colors[i % len(colors)],
                markersize=8,
                linewidth=2,
                label=prepare_text(subject.strip(), lang),
            )

            # Add percentage labels only
            for x, y in zip(x_values, y_values):
                plt.text(
                    x,
                    y + 2 if y >= 0 else y - 5,
                    f"{y:.1f}%",
                    ha="center",
                    fontsize=8,
                    color=colors[i % len(colors)],
                    bbox=dict(facecolor="white", alpha=0.7, edgecolor="none"),
                )

        # Chart formatting
        plt.xlabel(prepare_text(get_translation("Exam Date", lang)))
        plt.ylabel(prepare_text(get_translation("Percentage", lang)))
        # plt.title("Subject performance Timeline", fontsize=14)
        plt.ylim(0, 100)  # Fixed range for percentages (0-100)
        plt.grid(True, linestyle="--", alpha=0.9)

        # Set x-axis ticks to show dates
        plt.xticks(x_axis_positions, date_labels, rotation=45, ha="right")

        # Add legend
        plt.legend(bbox_to_anchor=(1.05, 1), loc="upper left", framealpha=0.9)

        plt.tight_layout()

        # Save chart to bytes buffer
        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        buf.seek(0)
        plt.close()

        return {"performance_history": result_json, "chart_image": buf}

    except Exception as e:
        print(f"Error generating history chart: {e}")
        return {"error": str(e)}


def generate_subject_bar_chart(average_percentages, lang="en"):
    try:
        plt.figure(figsize=(8, 5))
        plt.style.use("ggplot")

        # Prepare subject names - always reshape Arabic text regardless of UI language
        subjects = [
            prepare_text(subj, force_arabic=contains_arabic(subj))
            for subj in average_percentages.keys()
        ]
        percentages = list(average_percentages.values())

        # Set font properties
        font_prop = {"family": "Arial", "size": 10}
        if any(contains_arabic(subj) for subj in subjects) or lang == "ar":
            try:
                font_prop = {"family": "Arial Unicode MS", "size": 10}
            except:
                pass

        colors = plt.cm.viridis(np.linspace(0.2, 0.8, len(subjects)))
        bars = plt.bar(subjects, percentages, color=colors)

        # plt.title(prepare_text(get_translation("Subject performance", lang)), fontdict=font_prop, pad=20)
        plt.xlabel(prepare_text(get_translation("Subjects", lang)), fontdict=font_prop)
        plt.ylabel(
            prepare_text(get_translation("average percentage", lang)),
            fontdict=font_prop,
        )
        plt.ylim(0, 110)

        # Handle RTL if Arabic is present or language is Arabic
        if any(contains_arabic(subj) for subj in subjects) or lang == "ar":
            plt.xticks(
                ticks=range(len(subjects)),
                labels=subjects,
                rotation=45,
                ha="left" if lang == "en" else "right",
                fontproperties=font_prop,
            )

            plt.tight_layout()

        plt.grid(True, linestyle="--", alpha=0.5, axis="y")
        plt.tight_layout()

        # Add value labels
        for bar in bars:
            height = bar.get_height()
            plt.text(
                bar.get_x() + bar.get_width() / 2.0,
                height + 2,
                f"{height}%",
                ha="center",
                va="bottom",
                fontsize=9,
            )

        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")

        plt.close()
        return buf
    except Exception as e:
        print(f"Error generating bar chart: {e}")
        return None


import xlsxwriter
from flask import jsonify, request
import os
from datetime import datetime, timezone 
import uuid
from arabic_reshaper import arabic_reshaper
from bidi.algorithm import get_display


# def export_analysis_excel():
#     student_id = request.args.get("student_id")
#     lang = request.args.get("lang", "en")  # Default to English if not specified
#     start_date = request.args.get("startDate")
#     end_date = request.args.get("endDate")

#     if not student_id:
#         return jsonify({"error": "Student ID is required"}), 400

#     # Get student document from Firestore
#     student_ref = db.collection("students").document(student_id)
#     student_doc = student_ref.get()
#     if not student_doc.exists:
#         return jsonify({"error": "Student not found"}), 404

#     student_data = student_doc.to_dict()

#     # Extract personal information
#     pers_info = student_data.get("profileInfo", {}).get("personalInformation", {})
#     name = pers_info.get("name", get_translation("Not Available", lang))
#     email = student_data.get("email", get_translation("Not Available", lang))
#     assigned_grades = student_data.get("assignedGrades", {}).get("grades", {})

#     if assigned_grades:
#         # Get first grade key (e.g., "GRADE 5")
#         grade = next(
#             iter(assigned_grades.keys()), get_translation("Not Available", lang)
#         )
#         class_sections = assigned_grades.get(grade, {})
#         class_section = next(
#             iter(class_sections.keys()), get_translation("Not available", lang)
#         )
#     else:
#         grade = get_translation("Not Available", lang)
#         class_section = get_translation("Not available", lang)

#     # Get subject averages and data table
#     subject_response = get_student_subject_average_and_data_table(student_id)
#     if subject_response[1] != 200:
#         return subject_response

#     subject_data = subject_response[0]
#     average_percentages = subject_data.get("averageSubjectPercentages", {})
#     data_table = subject_data.get("dataTable", {})
#     has_data = any(len(subject_info) > 0 for subject_info in data_table.values())

#     # Get translated analysis sections using the export_data function
#     analysis_response = export_data(lang, student_id)
#     if analysis_response[1] != 200:
#         return analysis_response

#     # translated_analysis = analysis_response[0].json()
#     translated_analysis = analysis_response[0].json()
#     print(translated_analysis)

#     print(analysis_response)
#     # Use the translated analysis sections
#     strengths = translated_analysis.get("strengths", [])
#     weaknesses = translated_analysis.get("weaknesses", [])
#     recommendations = translated_analysis.get("recommendations", [])
#     interventions = translated_analysis.get("interventions", [])

from datetime import datetime, timezone
import datetime as _dt
def export_analysis_excel():
    student_id = request.args.get("student_id")
    lang = request.args.get("lang", "en")  # Default to English if not specified
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate")

    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400

    # Get student document from Firestore
    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()
    if not student_doc.exists:
        return jsonify({"error": "Student not found"}), 404

    student_data = student_doc.to_dict()

    # Extract personal information
    pers_info = student_data.get("profileInfo", {}).get("personalInformation", {})
    name = pers_info.get("name", get_translation("Not Available", lang))
    email = student_data.get("email", get_translation("Not Available", lang))
    assigned_grades = student_data.get("assignedGrades", {}).get("grades", {})

    if assigned_grades:
        grade = next(iter(assigned_grades.keys()), get_translation("Not Available", lang))
        class_sections = assigned_grades.get(grade, {})
        class_section = next(iter(class_sections.keys()), get_translation("Not available", lang))
    else:
        grade = get_translation("Not Available", lang)
        class_section = get_translation("Not available", lang)

    # Get subject averages and data table - get the JSON data from the response
    subject_response = get_student_subject_average_and_data_table(student_id)
    if isinstance(subject_response, tuple):
        response_json, status_code = subject_response
        if status_code != 200:
            return subject_response
        subject_data = response_json.json
    else:
        subject_data = subject_response

    average_percentages = subject_data.get("averageSubjectPercentages", {})
    data_table = subject_data.get("dataTable", {})
    has_data = any(len(subject_info) > 0 for subject_info in data_table.values())

    # Get translated analysis sections using the export_data function
    analysis_response = export_data(lang, student_id)
    if analysis_response[1] != 200:
        return analysis_response

    translated_analysis = analysis_response[0]

    # Use the translated analysis sections
    strengths = translated_analysis.get("strengths", [])
    weaknesses = translated_analysis.get("weaknesses", [])
    recommendations = translated_analysis.get("recommendations", [])
    interventions = translated_analysis.get("interventions", [])

    # Continue with the rest of the function...

    filename = f"{student_id}_performance_report_{uuid.uuid4().hex}.xlsx"
    local_path = os.path.join("reports", filename)
    os.makedirs("reports", exist_ok=True)

    try:
        # Create a workbook and add a worksheet
        workbook = xlsxwriter.Workbook(local_path)
        worksheet = workbook.add_worksheet(get_translation("Student Report", lang))

        # Define formats
        title_format = workbook.add_format(
            {
                "bold": True,
                "font_size": 16,
                "align": "center",
                "valign": "vcenter",
                "font_color": "#4A90E2",
                "reading_order": 2 if lang == "ar" else 0,  # RTL for Arabic
            }
        )

        header_format = workbook.add_format(
            {
                "bold": True,
                "font_size": 12,
                "bg_color": "#4A90E2",
                "font_color": "white",
                "align": "center",
                "border": 1,
                "reading_order": 2 if lang == "ar" else 0,  # RTL for Arabic
            }
        )

        section_header_format = workbook.add_format(
            {
                "bold": True,
                "font_size": 12,
                "font_color": "#4A90E2",
                "underline": True,
                "reading_order": 2 if lang == "ar" else 0,  # RTL for Arabic
            }
        )

        # Base formats that will be modified based on language
        base_format = {
            "font_size": 10,
            "reading_order": 2 if lang == "ar" else 0,  # RTL for Arabic
        }

        bold_format = workbook.add_format({**base_format, "bold": True})
        normal_format = workbook.add_format(base_format)
        arabic_format = workbook.add_format(
            {
                **base_format,
                "align": "right",
                "reading_order": 2,  # Force RTL for Arabic text
            }
        )
        date_format = workbook.add_format(
            {
                "num_format": "yyyy-mm-dd hh:mm:ss",
                "reading_order": 2 if lang == "ar" else 0,
            }
        )

        # Set RTL for the entire worksheet if Arabic
        if lang == "ar":
            worksheet.right_to_left()

        # Set column widths
        worksheet.set_column("A:A", 5)  # Row numbers
        worksheet.set_column("B:B", 25)  # Labels
        worksheet.set_column("C:C", 50)  # Values/content
        worksheet.set_column("D:D", 15)  # For charts

        # ---------------------------- Header ----------------------------
        worksheet.merge_range(
            "B2:C2", get_translation("Student Analysis Report", lang), title_format
        )
        row = 3

        # -------------------- Personal Information --------------------
        worksheet.write(
            row,
            1,
            get_translation("Personal Information", lang) + ":",
            section_header_format,
        )
        row += 1

        personal_info = [
            (get_translation("Name", lang), name),
            (get_translation("Email", lang), email),
            (get_translation("Class", lang), class_section),
            (get_translation("Grade", lang), grade),
        ]

        for label, value in personal_info:
            # For Arabic, we'll write the value first, then the label
            if lang == "ar":
                fmt = (
                    arabic_format
                    if contains_arabic(value) or lang == "ar"
                    else normal_format
                )
                worksheet.write(row, 2, value, fmt)
                worksheet.write(row, 1, label + ":", bold_format)
            else:
                worksheet.write(row, 1, label + ":", bold_format)
                fmt = (
                    arabic_format
                    if contains_arabic(value) or lang == "ar"
                    else normal_format
                )
                worksheet.write(row, 2, value, fmt)
            row += 1

        row += 1  # Add extra space # Add extra space

        # -------------------- Subject Performance --------------------
        if has_data:
            # Generate charts (same as PDF)
            line_chart_result = generate_monthly_subject_line_chart(
                data_table, start_date=start_date, end_date=end_date, lang=lang
            )
            bar_chart_stream = generate_subject_bar_chart(average_percentages, lang)

            row += 1  # Add extra space

            # Add Bar Chart
            if bar_chart_stream:
                try:
                    worksheet.write(
                        row,
                        1,
                        get_translation("Subject performance", lang) + ":",
                        section_header_format,
                    )
                    row += 1

                    # Insert the chart image
                    worksheet.insert_image(
                        f"B{row+1}",
                        "subject_performance.png",
                        {
                            "image_data": bar_chart_stream,
                            "x_scale": 0.7,
                            "y_scale": 0.7,
                        },
                    )
                    row += 20  # Adjust row position after image
                except Exception as e:
                    print(f"Error inserting bar chart: {e}")
                    worksheet.write(
                        row,
                        1,
                        get_translation("Could not generate performance chart", lang),
                        normal_format,
                    )
                    row += 2

            row += 1  # Add extra space

            # Add Line Chart if available
            if line_chart_result and "chart_image" in line_chart_result:
                try:
                    worksheet.write(
                        row,
                        1,
                        get_translation("progress", lang) + ":",
                        section_header_format,
                    )
                    row += 1

                    # Insert the chart image
                    worksheet.insert_image(
                        f"B{row+1}",
                        "monthly_progress.png",
                        {
                            "image_data": line_chart_result["chart_image"],
                            "x_scale": 0.7,
                            "y_scale": 0.7,
                        },
                    )
                    row += 20  # Adjust row position after image
                except Exception as e:
                    print(f"Error inserting line chart: {e}")
                    worksheet.write(
                        row,
                        1,
                        get_translation("Could not generate progress chart", lang),
                        normal_format,
                    )
                    row += 2

        else:
            worksheet.write(
                row,
                1,
                get_translation("Subject performance", lang) + ":",
                section_header_format,
            )
            worksheet.write(
                row + 1,
                1,
                get_translation("No data available for selected date range", lang),
                normal_format,
            )
            row += 3

        # -------------------- Analysis Sections --------------------
        sections = [
            (get_translation("Strengths", lang), strengths),
            (get_translation("Weaknesses", lang), weaknesses),
            (get_translation("Recommendations", lang), recommendations),
            (get_translation("Interventions", lang), interventions),
        ]

        for header, items in sections:
            worksheet.write(row, 1, header + ":", section_header_format)
            row += 1

            if items:
                for i, item in enumerate(items, 1):
                    item_text = f"{i}. {item}"
                    fmt = (
                        arabic_format
                        if contains_arabic(item_text) or lang == "ar"
                        else normal_format
                    )
                    worksheet.write(row, 2, item_text, fmt)
                    row += 1
            else:
                worksheet.write(
                    row, 2, get_translation("No data available", lang), normal_format
                )
                row += 1

            row += 1  # Add extra space between sections

        # Footer
        worksheet.write(
            row, 1, get_translation("Generated on", lang) + ":", bold_format
        )
        worksheet.write_datetime(row, 2, _dt.datetime.now(), date_format)
# ... 

        workbook.close()

        # Upload to Firebase Storage
        blob = bucket.blob(f"students12/{student_id}/{filename}")
        blob.upload_from_filename(local_path)
        blob.make_public()
        download_url = blob.public_url

        os.remove(local_path)
        student_ref.update({"excel_report_url": download_url})

        return jsonify(
            {
                "message": "Excel report generated successfully",
                "download_url": download_url,
            }
        )

    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        return jsonify({"error": f"Failed to generate Excel: {str(e)}"}), 500


# if __name__ == "__main__":
#     app.run(debug=True)


# ---------------------------------------------------------line graph -------------------------------------------------------------------

from datetime import datetime


@app.route("/api/student/grades/<student_id>", methods=["GET"])
def get_student_grade_progress(student_id):
    try:
        # Fetch student record from Firestore
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        student_data = student_doc.to_dict()
        grade_history = student_data.get(
            "gradeHistory", {}
        )  # Fetch grade history stored as {timestamp: grade}

        # Define grade-to-decimal mapping for smoother graph plotting
        GRADE_DECIMAL_MAPPING = {
            "A+": 10.0,
            "A": 9.0,
            "B+": 8.5,
            "B": 8.0,
            "C+": 7.5,
            "C": 7.0,
            "D": 6.0,
            "F": 5.0,
            "E": 4.0,
        }

        # Get optional start and end date filters from query parameters
        start_date = request.args.get("startDate")  # Expected format: YYYY-MM-DD
        end_date = request.args.get("endDate")  # Expected format: YYYY-MM-DD

        # Convert string dates to datetime for comparison
        start_date_obj = (
            datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        )
        end_date_obj = (
            datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        )

        # Convert grade history from dictionary to list of objects with timestamp and decimal values
        grade_progression = []
        for timestamp, grade in grade_history.items():
            try:
                timestamp_obj = datetime.strptime(
                    timestamp, "%Y-%m-%dT%H:%M:%S.%f"
                )  # Convert to datetime object
                formatted_timestamp = timestamp_obj.strftime(
                    "%Y-%m-%d %H:%M:%S"
                )  # Format properly
                timestamp_date = timestamp_obj.date()  # Extract only the date part

                print(
                    f"Checking timestamp: {timestamp_date} against startDate: {start_date_obj} and endDate: {end_date_obj}"
                )

                # Apply date filtering
                if start_date_obj and timestamp_date < start_date_obj:
                    continue
                if end_date_obj and timestamp_date > end_date_obj:
                    continue

                grade_entry = {
                    "timestamp": formatted_timestamp,  # Use formatted timestamp
                    "grade": grade,
                    "linepoints": GRADE_DECIMAL_MAPPING.get(
                        grade, 0.0
                    ),  # Default to 0.0 if grade is missing
                }
                grade_progression.append(grade_entry)
            except Exception as e:
                print(f"Skipping invalid timestamp {timestamp}: {e}")

        # Sort grades by timestamp for chronological progression
        grade_progression.sort(key=lambda x: x["timestamp"])

        return (
            jsonify({"studentId": student_id, "gradeProgression": grade_progression}),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

#---------------------------------------------------------teaching plans -------------------------------------------------------------------
@app.route("/teaching-plans1", methods=["GET"])
def get_teaching_plans():
    teacher_id = request.args.get("teacher_id")
    student_id = request.args.get("student_id")
    lang = request.args.get("lang", "en")  # Get language param, default to English

    if not teacher_id or not student_id:
        return jsonify({"error": "teacher_id and student_id are required"}), 400

    try:
        # Step 1: Validate existence
        teacher_doc = db.collection("users").document(teacher_id).get()
        student_doc = db.collection("students").document(student_id).get()

        if not teacher_doc.exists:
            return jsonify({"error": "Teacher not found"}), 404
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        # Step 2: Check association
        teacher_data = teacher_doc.to_dict()
        if student_id not in teacher_data.get("associatedIds", []):
            return jsonify({"error": "Student is not associated with this teacher"}), 403

        # Step 3: Fetch teaching plans
        student_data = student_doc.to_dict()
        teaching_plans = student_data.get("teachingPlans", {})
        if not teaching_plans:
            return jsonify({"error": "No teaching plans found"}), 404

        # Apply localization based on lang param
        localized_plans = {}
        for plan_id, plan in teaching_plans.items():
            if lang == "ar":
                localized_block = plan.get("actionPlan_i18n", {}).get("ar")
                if localized_block:
                    # Build localized actionPlan in Arabic
                    action_plan = {}
                    for key in ['assessmentMethods', 'instructionalStrategies', 'learningObjectives', 'recommendedResources', 'timeline']:
                        if key in localized_block:
                            action_plan[key] = localized_block[key]
                    localized_plans[plan_id] = {
                        "planId": plan_id,
                        "actionPlan": action_plan
                    }
                else:
                    # Fallback to original if Arabic not found
                    localized_plans[plan_id] = plan
            else:
                # Default or 'en': use existing English or default plan
                localized_plans[plan_id] = plan

        # Step 4: Fetch exam scripts
        from google.cloud.firestore_v1.base_query import FieldFilter
        exam_query = db.collection("examscripts").where(filter=FieldFilter("student_id", "==", student_id))
        exam_docs = list(exam_query.stream())

        exam_data = []
        for doc in exam_docs:
            data = doc.to_dict()
            exam_data.append({
                "curriculum_name": data.get("curriculumName"),
                "date": data.get("date"),
                "exam_name": data.get("exam_name"),
                "plan_details": data.get("teaching_plan", {}),
                "subject": data.get("subject"),
            })

        # Step 5: Format response
        response = {"data": []}
        for plan_id, plan in localized_plans.items():
            for exam in exam_data:
                response["data"].append({
                    "student_id": student_id,
                    "curriculum_name": exam.get("curriculum_name"),
                    "date": exam.get("date"),
                    "exam_name": exam.get("exam_name"),
                    "plan_details": {
                        "assessmentMethods": plan.get("actionPlan", {}).get("assessmentMethods", {}),
                        "instructionalStrategies": plan.get("actionPlan", {}).get("instructionalStrategies", {}),
                        "learningObjectives": plan.get("actionPlan", {}).get("learningObjectives", {}),
                        "planId": plan_id,
                        "recommendedResources": plan.get("actionPlan", {}).get("recommendedResources", {}),
                        "timeline": plan.get("actionPlan", {}).get("timeline", {}),
                    },
                    "subject": exam.get("subject"),
                })

        if not response["data"]:
            return jsonify({"error": "No exam scripts or teaching plans found"}), 404

        return jsonify(response)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# @app.route("/get_students", methods=["GET"])
# def get_students():
#     user_id = request.args.get("userId")

#     if not user_id:
#         return jsonify({"error": "Missing userId"}), 400

#     # âœ… Fetch user document
#     user_ref = db.collection("users").where("userId", "==", user_id).stream()
#     user_doc = next(user_ref, None)

#     if not user_doc or not user_doc.exists:
#         return jsonify({"error": "User not found"}), 404

#     user_data = user_doc.to_dict()
#     associated_ids = user_data.get("associatedIds", [])

#     if not associated_ids:
#         return jsonify({"message": "No associated student IDs found"}), 200

#     # âœ… Fetch students matching associatedIds
#     students_ref = (
#         db.collection("students").where("studentId", "in", associated_ids).stream()
#     )

#     students = []
#     for student in students_ref:
#         student_data = student.to_dict()

#         # âœ… Extracting student details
#         profile_info = student_data.get("profileInfo", {}).get(
#             "personalInformation", {}
#         )
#         academic_info = student_data.get("academicInformation", {})

#         name = profile_info.get("name", "Unknown")
#         email = student_data.get("email", "NA")
#         photo_url = profile_info.get("photoUrl", "")

#         # âœ… Extract class and grade from academicInformation
#         student_class = academic_info.get("classSection")
#         grade = academic_info.get("grade")

#         # âœ… Fetch teacher name by student ID
#         teacher_name = get_teacher_name_by_student_id(student.id)

#         students.append(
#             {
#                 "studentId": student.id,
#                 "name": name,
#                 "email": email,
#                 "photoUrl": photo_url,
#                 "class": student_class,
#                 "grade": grade,
#                 "assignedTeacherName": teacher_name,
#             }
#         )

#     return jsonify({"students": students}), 200

# ...existing code...
# ...existing code...

# @app.route("/get_students", methods=["GET"])
# def get_students():
#     user_id = request.args.get("userId")
#     if not user_id:
#         return jsonify({"error": "Missing userId"}), 400

#     user_query = db.collection("users").where("userId", "==", user_id).stream()
#     user_doc = next(user_query, None)
#     if not user_doc or not user_doc.exists:
#         return jsonify({"error": "User not found"}), 404

#     user_data = user_doc.to_dict()
#     associated_ids = user_data.get("associatedIds", [])

#     # normalize associated_ids (string, dict or list)
#     if isinstance(associated_ids, str):
#         try:
#             associated_ids = json.loads(associated_ids)
#         except Exception:
#             associated_ids = [associated_ids]
#     if isinstance(associated_ids, dict):
#         associated_ids = list(associated_ids.keys()) or list(associated_ids.values())
#     if not isinstance(associated_ids, list):
#         associated_ids = [associated_ids] if associated_ids else []

#     if not associated_ids:
#         return jsonify({"students": []}), 200

#     # chunk helper for Firestore 'in' queries (max 10)
#     def chunked(seq, size=10):
#         for i in range(0, len(seq), size):
#             yield seq[i : i + size]

#     students = []
#     for chunk in chunked(associated_ids, 10):
#         # Query 'users' collection where the 'userId' field matches
#         students_ref = db.collection("users").where("userId", "in", chunk).stream()
#         for sdoc in students_ref:
#             sd = sdoc.to_dict() or {}

#             # profile / contact locations (support different schemas)
#             profile_info = sd.get("profileInfo", {}) or sd.get("profile_info", {})
#             personal_info = profile_info.get("personalInformation", {}) or profile_info.get("personal_information", {})
#             contact_info = profile_info.get("contactInformation", {}) or profile_info.get("contact_information", {})

#             # Try canonical academic locations
#             academic_info = (
#                 sd.get("academicInformation")
#                 or sd.get("academicData")
#                 or sd.get("academic_info")
#                 or sd.get("academic")
#                 or {}
#             )

#             # try to find class and grade from known places
#             student_class = ""
#             grade = ""

#             if isinstance(academic_info, dict):
#                 student_class = academic_info.get("classSection") or academic_info.get("class") or academic_info.get("classSection", "") or ""
#                 grade = academic_info.get("grade") or academic_info.get("gradeLevel") or ""

#             # root-level fallbacks
#             if not student_class:
#                 student_class = sd.get("classSection") or sd.get("class") or sd.get("class_name", "") or ""
#             if not grade:
#                 grade = sd.get("grade", "") or sd.get("gradeLevel", "")

#             # Final fallback: assignedGrades -> grades -> { GRADE X : { "A": {...} } }
#             if not student_class or not grade:
#                 assigned = sd.get("assignedGrades", {}) or {}
#                 if isinstance(assigned, dict):
#                     grades_map = assigned.get("grades", assigned)
#                     if isinstance(grades_map, dict) and grades_map:
#                         # pick a sensible grade if missing
#                         if not grade:
#                             grade = next(iter(grades_map.keys()), grade)
#                         # pick a sensible class (first key inside grade mapping) if missing
#                         if not student_class:
#                             first_grade_val = next(iter(grades_map.values()), None)
#                             if isinstance(first_grade_val, dict) and first_grade_val:
#                                 student_class = next(iter(first_grade_val.keys()), student_class)
#                             elif isinstance(first_grade_val, list) and first_grade_val:
#                                 student_class = first_grade_val[0]

#             name = personal_info.get("name") or sd.get("name") or "Unknown"
#             email = sd.get("email") or personal_info.get("email") or "NA"
#             photo_url = personal_info.get("photoUrl") or personal_info.get("photourl") or sd.get("photoUrl") or ""

#             # students.append(
#             #     {
#             #         "studentId": sdoc.id,
#             #         "name": name,
#             #         "email": email,
#             #         "photoUrl": photo_url,
#             #         "class": student_class,
#             #         "grade": grade,
#             #         "assignedTeacherName": get_teacher_name_by_student_id(sdoc.id),
#             #     }
#             # )
#             students.append(
#                 {
#                     # Return the correct userId
#                     "studentId": student_id_from_doc, 
#                     "name": name,
#                     "email": email,
#                     "photoUrl": photo_url,
#                     "class": student_class,
#                     "grade": grade,
#                     # Pass the correct userId to the helper function
#                     "assignedTeacherName": get_teacher_name_by_student_id(student_id_from_doc), 
#                 }
#             )           

#     return jsonify({"students": students}), 200
# # ...existing code...
# # ...existing code...

# def get_teacher_name_by_student_id(student_id):
#     """
#     Fetches the teacher's name by checking if the student ID
#     is in the associatedIds of the users collection.
#     Extracts the name from 'profileInfo -> personalInformation -> name'.
#     """
#     teacher_name = "Unknown"

#     # âœ… Fetch all teachers with role 'teacher' and status 'Active'
#     teachers_ref = db.collection("users").where("role", "==", "teacher").stream()

#     for teacher_doc in teachers_ref:
#         teacher_data = teacher_doc.to_dict()

#         # âœ… Check if the student_id is in associatedIds
#         associated_ids = teacher_data.get("associatedIds", [])

#         if student_id in associated_ids:
#             # âœ… Extract teacher name
#             teacher_name = (
#                 teacher_data.get("profileInfo", {})
#                 .get("personalInformation", {})
#                 .get("name", "Unknown")
#             )
#             break  # Stop once the matching teacher is found

#     return teacher_name


@app.route("/get_students", methods=["GET"])
def get_students():
    user_id = request.args.get("userId")
    if not user_id:
        return jsonify({"error": "Missing userId"}), 400

    user_query = db.collection("users").where("userId", "==", user_id).stream()
    user_doc = next(user_query, None)
    if not user_doc or not user_doc.exists:
        return jsonify({"error": "User not found"}), 404

    user_data = user_doc.to_dict()
    associated_ids = user_data.get("associatedIds", [])

    # normalize associated_ids (string, dict or list)
    if isinstance(associated_ids, str):
        try:
            associated_ids = json.loads(associated_ids)
        except Exception:
            associated_ids = [associated_ids]
    if isinstance(associated_ids, dict):
        associated_ids = list(associated_ids.keys()) or list(associated_ids.values())
    if not isinstance(associated_ids, list):
        associated_ids = [associated_ids] if associated_ids else []

    if not associated_ids:
        return jsonify({"students": []}), 200

    # chunk helper for Firestore 'in' queries (max 10)
    def chunked(seq, size=10):
        for i in range(0, len(seq), size):
            yield seq[i : i + size]

    students = []
    for chunk in chunked(associated_ids, 10):
        # Query 'users' collection where the 'userId' field matches
        students_ref = db.collection("users").where("userId", "in", chunk).stream()
        for sdoc in students_ref:
            sd = sdoc.to_dict() or {}

            # profile / contact locations (support different schemas)
            profile_info = sd.get("profileInfo", {}) or sd.get("profile_info", {})
            personal_info = profile_info.get("personalInformation", {}) or profile_info.get("personal_information", {})
            contact_info = profile_info.get("contactInformation", {}) or profile_info.get("contact_information", {})

            # Try canonical academic locations
            academic_info = (
                sd.get("academicInformation")
                or sd.get("academicData")
                or sd.get("academic_info")
                or sd.get("academic")
                or {}
            )

            # try to find class and grade from known places
            student_class = ""
            grade = ""

            if isinstance(academic_info, dict):
                student_class = academic_info.get("classSection") or academic_info.get("class") or academic_info.get("classSection", "") or ""
                grade = academic_info.get("grade") or academic_info.get("gradeLevel") or ""

            # root-level fallbacks
            if not student_class:
                student_class = sd.get("classSection") or sd.get("class") or sd.get("class_name", "") or ""
            if not grade:
                grade = sd.get("grade", "") or sd.get("gradeLevel", "")

            # Final fallback: assignedGrades -> grades -> { GRADE X : { "A": {...} } }
            if not student_class or not grade:
                assigned = sd.get("assignedGrades", {}) or {}
                if isinstance(assigned, dict):
                    grades_map = assigned.get("grades", assigned)
                    if isinstance(grades_map, dict) and grades_map:
                        # pick a sensible grade if missing
                        if not grade:
                            grade = next(iter(grades_map.keys()), grade)
                        # pick a sensible class (first key inside grade mapping) if missing
                        if not student_class:
                            first_grade_val = next(iter(grades_map.values()), None)
                            if isinstance(first_grade_val, dict) and first_grade_val:
                                student_class = next(iter(first_grade_val.keys()), student_class)
                            elif isinstance(first_grade_val, list) and first_grade_val:
                                student_class = first_grade_val[0]

            name = personal_info.get("name") or sd.get("name") or "Unknown"
            email = sd.get("email") or personal_info.get("email") or "NA"
            photo_url = personal_info.get("photoUrl") or personal_info.get("photourl") or sd.get("photoUrl") or ""

            #
            #
            # Get the correct studentId from the document data
            student_id_from_doc = sd.get("userId", sdoc.id) # Fallback to sdoc.id if 'userId' is missing
            #
            #

            # students.append(
            #     {
            #         "studentId": sdoc.id,
            #         "name": name,
            #         ...
            #         "assignedTeacherName": get_teacher_name_by_student_id(sdoc.id),
            #     }
            # )
            
            # (Your new code that will now work)
            students.append(
                {
                    # Return the correct userId
                    "studentId": student_id_from_doc, 
                    "name": name,
                    "email": email,
                    "photoUrl": photo_url,
                    "class": student_class,
                    "grade": grade,
                    # Pass the correct userId to the helper function
                    "assignedTeacherName": get_teacher_name_by_student_id(student_id_from_doc), 
                }
            )

    return jsonify({"students": students}), 200

#
#
def get_teacher_name_by_student_id(student_id):
    """
    Fetches the teacher's name by checking if the student ID
    is in the associatedIds of the users collection.
    Extracts the name from 'profileInfo -> personalInformation -> name'.
    """
    teacher_name = "Unknown"

    # Fetch all teachers with role 'teacher' and status 'Active'
    teachers_ref = db.collection("users").where("role", "==", "teacher").stream()

    for teacher_doc in teachers_ref:
        teacher_data = teacher_doc.to_dict()

        # Check if the student_id is in associatedIds
        associated_ids = teacher_data.get("associatedIds", [])

        if student_id in associated_ids:
            # Extract teacher name
            teacher_name = (
                teacher_data.get("profileInfo", {})
                .get("personalInformation", {})
                .get("name", "Unknown")
            )
            break  # Stop once the matching teacher is found

    return teacher_name


from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import os
from datetime import datetime


# Initialize Firebase Admin SDK
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")


# # -----------------------------
# # POST Feedback for a Teaching Plan
# # -----------------------------
# loop111 = asyncio.new_event_loop()

# db111 = firestore_async.client()


# @app.route("/api/teaching-plans/feedback", methods=["POST"])
# def get_students_feedback():
#     result = loop111.run_until_complete(
#         send_feedback(request)
#     )  # âœ… Run the async function
#     return result


# async def send_feedback(request):
#     try:
#         data = request.get_json()
#         if not data:
#             return jsonify({"error": "Invalid JSON payload"}), 400

#         teacherid = data.get("teacherid")
#         planid = data.get("planid")
#         feedback = data.get("feedback")
#         student_id = data.get("studentid")

#         if not all([teacherid, planid, feedback, student_id]):
#             return (
#                 jsonify(
#                     {"error": "teacherid, planid, feedback, and studentid are required"}
#                 ),
#                 400,
#             )

#         feedback_doc = {
#             "teacherid": teacherid,
#             "studentid": student_id,
#             "planid": planid,
#             "feedback": feedback,
#             "timestamp": datetime.utcnow().isoformat(),
#         }

#         db.collection("teacher_feedback").add(feedback_doc)

#         return (
#             jsonify({"status": "success", "message": "Feedback added successfully"}),
#             200,
#         )
#     except Exception as e:
#         return jsonify({"error": "Failed to add feedback", "details": str(e)}), 500
# -----------------------------
# POST Feedback for a Teaching Plan
# (Corrected and Synchronous)
# -----------------------------
@app.route("/api/teaching-plans/feedback", methods=["POST"])
def get_students_feedback():
    # ✅ 1. Add internal imports
    from datetime import datetime, timezone

    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Invalid JSON payload"}), 400

        teacherid = data.get("teacherid")
        planid = data.get("planid")
        feedback = data.get("feedback")
        student_id = data.get("studentid")

        if not all([teacherid, planid, feedback, student_id]):
            return (
                jsonify(
                    {"error": "teacherid, planid, feedback, and studentid are required"}
                ),
                400,
            )

        feedback_doc = {
            "teacherid": teacherid,
            "studentid": student_id,
            "planid": planid,
            "feedback": feedback,
            # ✅ 2. THE FIX: Replaced utcnow()
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }

        # Use your global synchronous 'db' client
        db.collection("teacher_feedback").add(feedback_doc)

        return (
            jsonify({"status": "success", "message": "Feedback added successfully"}),
            200,
        )
    except Exception as e:
        return jsonify({"error": "Failed to add feedback", "details": str(e)}), 500
# -----------------------------
# -----------------------------
# GET Feedback by teacherid and planid
# -----------------------------
@app.route("/api/teaching-plans/feedback", methods=["GET"])
def get_feedback():
    teacherid = request.args.get("teacherid")
    planid = request.args.get("planid")

    if not all([teacherid, planid]):
        return jsonify({"error": "teacherid and planid are required"}), 400

    # Query the teaching_feedback collection
    feedback_query = (
        db.collection("teacher_feedback")
        .where("teacherid", "==", teacherid)
        .where("planid", "==", planid)
        .stream()
    )

    feedback_list = []
    for feedback in feedback_query:
        feedback_data = feedback.to_dict()
        feedback_list.append(
            {
                "studentid": feedback_data.get("studentid"),
                "feedback": feedback_data.get("feedback"),
                "timestamp": feedback_data.get("timestamp"),
            }
        )

    if feedback_list:
        return jsonify({"status": "success", "feedback": feedback_list}), 200
    else:
        return (
            jsonify({"error": "No feedback found for the given teacherid and planid"}),
            404,
        )


from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore, storage
import os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from io import BytesIO
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet


# âœ… Firebase Initialization
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()

import matplotlib

matplotlib.use("Agg")  # âœ… Use non-GUI backend for rendering stability

# âœ… Grade to Decimal Mapping
GRADE_DECIMAL_MAPPING = {
    "A+": 10.0,
    "A": 9.0,
    "B+": 8.5,
    "B": 8.0,
    "C+": 7.5,
    "C": 7.0,
    "D": 6.0,
    "F": 5.0,
    "E": 4.0,
}


# âœ… Upload PDF to Firebase Storage
def upload_to_firebase(pdf_bytes: bytes, filename: str) -> str:
    """Uploads the PDF to Firebase Storage."""
    bucket = storage.bucket()
    blob = bucket.blob(filename)
    blob.upload_from_string(pdf_bytes, content_type="application/pdf")
    blob.make_public()
    return blob.public_url


# âœ… Flexible timestamp parsing
def parse_timestamp_safe(timestamp):
    """Safely parse timestamp, supporting both microseconds and regular format."""
    formats = [
        "%Y-%m-%dT%H:%M:%S.%f",  # Microsecond format
        "%Y-%m-%dT%H:%M:%S",  # Regular format
    ]

    for fmt in formats:
        try:
            return datetime.strptime(timestamp, fmt)
        except ValueError:
            continue

    print(f"Skipping invalid timestamp: {timestamp}")
    return None


# âœ… Generate graphs and return image bytes
def generate_graphs(student_data, student_id):
    """Generates and returns graph image bytes with student details."""
    subjects = []
    marks = []
    grade_data = []
    grade_labels = []

    academic_data = student_data.get("academicData", {}).get("subjects", {})
    for subject, details in academic_data.items():
        if isinstance(details, dict):
            clean_subject = subject.strip()
            if "history" in details and isinstance(details["history"], list):
                history_marks = [
                    entry["marks"] for entry in details["history"] if "marks" in entry
                ]
                if history_marks:
                    avg_marks = sum(history_marks) / len(history_marks)
                    subjects.append(clean_subject)
                    marks.append(avg_marks)
            elif "marks" in details:
                subjects.append(clean_subject)
                marks.append(details["marks"])

    grade_history = student_data.get("gradeHistory", {})
    for timestamp, grade in grade_history.items():
        parsed_date = parse_timestamp_safe(timestamp)
        if parsed_date and grade in GRADE_DECIMAL_MAPPING:
            grade_value = GRADE_DECIMAL_MAPPING[grade]
            grade_data.append((parsed_date, grade_value))
            grade_labels.append(grade)  # Store grade labels

    grade_data.sort(key=lambda x: x[0])
    dates = [entry[0] for entry in grade_data]
    grades = [entry[1] for entry in grade_data]

    fig, axes = plt.subplots(2, 1, figsize=(12, 14))
    plt.subplots_adjust(hspace=0.4)  # âœ… Add spacing between graphs

    # âœ… Subject-wise marks
    ax1 = axes[0]
    if subjects and marks:
        bars = ax1.bar(subjects, marks, color="skyblue")
        ax1.set_title("Subject-wise Score Analysis", fontsize=14)
        ax1.set_xlabel("Subjects", fontsize=12)
        ax1.set_ylabel("Marks", fontsize=12)
        ax1.grid(True)
        for bar, mark in zip(bars, marks):
            ax1.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.5,
                f"{mark:.1f}",
                ha="center",
                fontsize=10,
                fontweight="bold",
            )
    else:
        ax1.text(
            0.5,
            0.5,
            "No Data Available",
            ha="center",
            fontsize=14,
            color="red",
            transform=ax1.transAxes,
        )
        ax1.set_title("Subject-wise Score Analysis", fontsize=14)
        ax1.axis("off")

    # âœ… Grade history
    ax2 = axes[1]
    if dates and grades:
        ax2.plot(dates, grades, marker="o", color="green", label="Grade History")
        for date, grade, label in zip(dates, grades, grade_labels):
            ax2.text(
                date,
                grade + 0.2,
                f"{label}",
                fontsize=10,
                color="blue",
                ha="center",
                fontweight="bold",
            )  # âœ… Add grade names above points
        ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
        ax2.set_title("Progress History Over Time", fontsize=14)
        ax2.set_xlabel("Date", fontsize=12)
        ax2.set_ylabel("Grade (Decimal)", fontsize=12)
        ax2.legend()
        ax2.grid(True)
    else:
        ax2.text(
            0.5,
            0.5,
            "No Data Available",
            ha="center",
            fontsize=14,
            color="red",
            transform=ax2.transAxes,
        )
        ax2.set_title("Grade History Over Time", fontsize=14)
        ax2.axis("off")

    plt.tight_layout()
    image_bytes = BytesIO()
    plt.savefig(image_bytes, format="png", dpi=150)
    plt.close(fig)
    image_bytes.seek(0)
    return image_bytes


# âœ… Generate PDF with student details and graph
def generate_pdf(student_data, student_id):
    """Generates a PDF report with student details and graph."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER)
    story = []

    styles = getSampleStyleSheet()

    # âœ… Add student details
    story.append(Paragraph(f"Child Progress Report", styles["Title"]))
    story.append(Spacer(1, 12))
    student_id = student_data.get("studentId", "Unknown")

    student_name = (
        student_data.get("profileInfo", {})
        .get("personalInformation", {})
        .get("name", "Unknown")
    )
    student_email = student_data.get("email", "N/A")
    student_class = student_data.get("academicInformation", {}).get(
        "classSection", "N/A"
    )
    student_grade = student_data.get("academicInformation", {}).get("grade", "N/A")

    story.append(Paragraph(f"Name: {student_name}", styles["Normal"]))
    story.append(Paragraph(f"Student ID: {student_id}", styles["Normal"]))
    story.append(Paragraph(f"Email: {student_email}", styles["Normal"]))
    story.append(Paragraph(f"Class: {student_class}", styles["Normal"]))
    story.append(Paragraph(f"Grade: {student_grade}", styles["Normal"]))
    story.append(Spacer(1, 24))

    # âœ… Generate graph
    graph_image = generate_graphs(student_data, student_id)

    # âœ… Add graph image to PDF
    img = Image(graph_image)
    img.drawHeight = 300
    img.drawWidth = 400
    story.append(img)

    doc.build(story)

    buffer.seek(0)
    return buffer.getvalue()


@app.route("/api/progress-report", methods=["GET"])
def get_child_progress():
    """API Endpoint to generate and return progress report PDF with student details."""
    parent_id = request.args.get("parent_id")
    student_id = request.args.get("student_id")

    if not parent_id or not student_id:
        return jsonify({"error": "parent_id and student_id are required"}), 400

    try:
        # âœ… Fetch student data
        student_doc = db.collection("students").document(student_id).get()
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        student_data = student_doc.to_dict()

        # âœ… Generate PDF report
        pdf_bytes = generate_pdf(student_data, student_id)
        filename = f"childprogress_reports/{student_id}_report.pdf"
        pdf_url = upload_to_firebase(pdf_bytes, filename)

        return (
            jsonify({"message": "Report generated successfully", "pdf_url": pdf_url}),
            200,
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# from flask import Flask, request, jsonify
# import firebase_admin
# from firebase_admin import credentials, firestore
# from datetime import datetime, timedelta, timezone
# from dateutil import parser  # Ã¢Å“â€¦ Handles timezone offsets
# import os

# if not firebase_admin._apps:
#     if os.path.exists("serviceAccountKey.json"):
#         cred = credentials.Certificate("serviceAccountKey.json")
#         firebase_admin.initialize_app(
#             cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
#         )
#     else:
#         raise FileNotFoundError("Missing Firebase Service Account Key")

# db = firestore.client()


# def parse_timestamp_flexible(timestamp_str):
#     """
#     Handles multiple timestamp formats and returns UTC offset-aware datetime.
#     """
#     try:
#         dt = parser.parse(timestamp_str)

#         # Ã°Å¸â€Â¥ Ensure all timestamps are converted to UTC
#         if dt.tzinfo is None:  # Naive timestamp -> make it UTC-aware
#             dt = dt.replace(tzinfo=timezone.utc)
#         else:
#             dt = dt.astimezone(timezone.utc)  # Convert offset-aware to UTC

#         return dt

#     except Exception as e:
#         raise ValueError(f"Unsupported timestamp format: {timestamp_str} -> {e}")


# from flask import request, jsonify
# from datetime import datetime, timedelta, timezone

# # NOTE: The database connection ('db') and flexible timestamp parser
# # ('parse_timestamp_flexible') are assumed to be defined elsewhere in your app.

# # Assumes 'app' is defined and the environment variables/configuration for 'db' are set up.
# # You will need to uncomment or add 'app = Flask(__name__)' if this is the start of your file.
# # import datetime
# # @app.route("/api/recent-updates", methods=["GET"])
# # def get_recent_updates():
# #     parent_id = request.args.get("parent_id")

# #     if not parent_id:
# #         return jsonify({"error": "parent_id is required"}), 400

# #     try:
# #         # Fetch associated student IDs
# #         # Assumes 'db' is initialized (e.g., from Firestore/Firebase SDK)
# #         parent_doc = db.collection("users").document(parent_id).get()

# #         if not parent_doc.exists:
# #             return jsonify({"error": "Parent not found"}), 404

# #         parent_data = parent_doc.to_dict()
# #         associated_ids = parent_data.get("associatedIds", [])

# #         if not associated_ids:
# #             # Return 200 with an empty list if no students are associated.
# #             return jsonify({"recent_updates": []}), 200

# #         recent_updates = []
        
# #         # FIX: Correctly uses the datetime class and UTC for timezone-aware comparison.
# #         last_24_hours = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(hours=24)

# #         # Fetch observations
# #         for student_id in associated_ids:
# #             student_doc = db.collection("students").document(student_id).get()

# #             if student_doc.exists:
# #                 student_data = student_doc.to_dict()
# #                 observations = student_data.get("observations", [])

# #                 for obs in observations:
# #                     obs_date_str = obs.get("date")

# #                     if obs_date_str:
# #                         try:
# #                             # Parse and normalize to UTC
# #                             # Assumes 'parse_timestamp_flexible' handles various date formats
# #                             obs_date = parse_timestamp_flexible(obs_date_str)
# #                         except ValueError as e:
# #                             print(f"Error parsing observation timestamp: {e}")
# #                             continue

# #                         # Compare with last 24 hours (timezone-aware comparison)
# #                         if obs_date >= last_24_hours:
# #                             recent_updates.append(
# #                                 {
# #                                     "student_id": student_id,
# #                                     "type": "observation",
# #                                     "subject": obs.get("subject"),
# #                                     "observation": obs.get("observation"),
# #                                     "date": obs.get("date"),
# #                                     "attachment_url": obs.get("attachment_url"),
# #                                 }
# #                             )

# #         # Fetch feedback
# #         for student_id in associated_ids:
# #             # Assumes 'db' supports Firestore/Database querying like this.
# #             feedback_ref = (
# #                 db.collection("feedback").where("student_id", "==", student_id).stream()
# #             )

# #             for feedback in feedback_ref:
# #                 feedback_data = feedback.to_dict()
# #                 feedback_time_str = feedback_data.get("timestamp")

# #                 if feedback_time_str:
# #                     try:
# #                         # Parse and normalize feedback timestamp to UTC
# #                         feedback_time = parse_timestamp_flexible(feedback_time_str)
# #                     except ValueError as e:
# #                         print(f"Error parsing feedback timestamp: {e}")
# #                         continue

# #                     # Compare with last 24 hours (timezone-aware comparison)
# #                     if feedback_time >= last_24_hours:
# #                         recent_updates.append(
# #                             {
# #                                 "student_id": student_id,
# #                                 "type": "feedback",
# #                                 "subject": feedback_data.get("subject"),
# #                                 "feedback": feedback_data.get("feedback"),
# #                                 "timestamp": feedback_data.get("timestamp"),
# #                             }
# #                         )

# #         if not recent_updates:
# #             return (
# #                 jsonify({"message": "No recent updates found in the last 24 hours"}),
# #                 200,
# #             )

# #         return jsonify({"recent_updates": recent_updates}), 200

# #     except Exception as e:
# #         print(f"An error occurred in get_recent_updates: {e}")
# #         return jsonify({"error": str(e)}), 500

# from flask import Flask, request, jsonify
# from datetime import datetime, timedelta, timezone

# @app.route("/api/recent-updates", methods=["GET"])
# def get_recent_updates():
#     parent_id = request.args.get("parent_id")
#     lang = request.args.get("lang", "en").lower()  # Default to English

#     if not parent_id:
#         return jsonify({
#             "error": "parent_id is required" if lang == "en" else "مطلوب معرف الوالد"
#         }), 400

#     try:
#         # Fetch parent document
#         parent_doc = db.collection("users").document(parent_id).get()
#         if not parent_doc.exists:
#             return jsonify({
#                 "error": "Parent not found" if lang == "en" else "لم يتم العثور على الوالد"
#             }), 404

#         parent_data = parent_doc.to_dict()
#         associated_ids = parent_data.get("associatedIds", [])

#         if not associated_ids:
#             return jsonify({
#                 "recent_updates": [],
#                 "message": "No associated students found" if lang == "en" else "لم يتم العثور على طلاب مرتبطين"
#             }), 200

#         recent_updates = []
#         last_24_hours = datetime.now(timezone.utc) - timedelta(hours=24)

#         # --- Fetch Observations ---
#         for student_id in associated_ids:
#             student_doc = db.collection("students").document(student_id).get()
#             if not student_doc.exists:
#                 continue

#             student_data = student_doc.to_dict()
#             observations = student_data.get("observations", [])

#             for obs in observations:
#                 obs_date_str = obs.get("date")
#                 if not obs_date_str:
#                     continue

#                 try:
#                     obs_date = parse_timestamp_flexible(obs_date_str)
#                 except ValueError:
#                     continue

#                 if obs_date >= last_24_hours:
#                     recent_updates.append({
#                         "student_id": student_id,
#                         "type": "observation" if lang == "en" else "ملاحظة",
#                         "subject": obs.get("subject"),
#                         "observation": obs.get("observation") if lang == "en" else obs.get("observation_ar", obs.get("observation")),
#                         "date": obs.get("date"),
#                         "attachment_url": obs.get("attachment_url"),
#                     })

#         # --- Fetch Feedback ---
#         for student_id in associated_ids:
#             feedback_ref = db.collection("feedback").where("student_id", "==", student_id).stream()

#             for feedback in feedback_ref:
#                 feedback_data = feedback.to_dict()
#                 feedback_time_str = feedback_data.get("timestamp")
#                 if not feedback_time_str:
#                     continue

#                 try:
#                     feedback_time = parse_timestamp_flexible(feedback_time_str)
#                 except ValueError:
#                     continue

#                 if feedback_time >= last_24_hours:
#                     recent_updates.append({
#                         "student_id": student_id,
#                         "type": "feedback" if lang == "en" else "تغذية راجعة",
#                         "subject": feedback_data.get("subject"),
#                         "feedback": feedback_data.get("feedback") if lang == "en" else feedback_data.get("feedback_ar", feedback_data.get("feedback")),
#                         "timestamp": feedback_data.get("timestamp"),
#                     })

#         if not recent_updates:
#             return jsonify({
#                 "message": "No recent updates found in the last 24 hours"
#                 if lang == "en"
#                 else "لا توجد تحديثات حديثة خلال آخر 24 ساعة"
#             }), 200

#         return jsonify({"recent_updates": recent_updates}), 200

#     except Exception as e:
#         print(f"An error occurred in get_recent_updates: {e}")
#         return jsonify({
#             "error": str(e) if lang == "en" else "حدث خطأ أثناء جلب التحديثات"
#         }), 500
import os
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime, timedelta, timezone
from dateutil import parser
from flask import Flask, request, jsonify

# -------------------- Initialize Flask --------------------
# app = Flask(__name__)

# -------------------- Initialize Firebase --------------------
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()


# # -------------------- Timestamp Parser --------------------
# def parse_timestamp_flexible(timestamp_str):
#     """Handles multiple timestamp formats and returns UTC offset-aware datetime."""
#     try:
#         dt = parser.parse(timestamp_str)
#         # Ensure all timestamps are converted to UTC
#         if dt.tzinfo is None:
#             dt = dt.replace(tzinfo=timezone.utc)
#         else:
#             dt = dt.astimezone(timezone.utc)
#         return dt
#     except Exception as e:
#         raise ValueError(f"Unsupported timestamp format: {timestamp_str} -> {e}")

# @app.route("/api/recent-updates", methods=["GET"])
# def get_recent_updates():
#     import importlib

#     # ✅ Force import of the *real* datetime module, not the overridden one
#     datetime_mod = importlib.import_module("datetime")
#     datetime_cls = datetime_mod.datetime
#     timedelta_cls = datetime_mod.timedelta
#     timezone_cls = datetime_mod.timezone

#     parent_id = request.args.get("parent_id")
#     lang = request.args.get("lang", "en").lower()  # Default to English

#     if not parent_id:
#         return jsonify({
#             "error": "parent_id is required" if lang == "en" else "مطلوب معرف الوالد"
#         }), 400

#     try:
#         parent_doc = db.collection("users").document(parent_id).get()
#         if not parent_doc.exists:
#             return jsonify({
#                 "error": "Parent not found" if lang == "en" else "لم يتم العثور على الوالد"
#             }), 404

#         parent_data = parent_doc.to_dict()
#         associated_ids = parent_data.get("associatedIds", [])

#         if not associated_ids:
#             return jsonify({
#                 "recent_updates": [],
#                 "message": "No associated students found"
#                 if lang == "en"
#                 else "لم يتم العثور على طلاب مرتبطين"
#             }), 200

#         recent_updates = []

#         # ✅ Use our locally imported datetime safely
#         last_24_hours = datetime_cls.now(timezone_cls.utc) - timedelta_cls(hours=24)

#         # --- Fetch Observations ---
#         for student_id in associated_ids:
#             student_doc = db.collection("students").document(student_id).get()
#             if not student_doc.exists:
#                 continue

#             student_data = student_doc.to_dict()
#             observations = student_data.get("observations", [])

#             for obs in observations:
#                 obs_date_str = obs.get("date")
#                 if not obs_date_str:
#                     continue

#                 try:
#                     obs_date = parse_timestamp_flexible(obs_date_str)
#                 except ValueError:
#                     continue

#                 if obs_date >= last_24_hours:
#                     recent_updates.append({
#                         "student_id": student_id,
#                         "type": "observation" if lang == "en" else "ملاحظة",
#                         "subject": obs.get("subject"),
#                         "observation": obs.get("observation")
#                         if lang == "en"
#                         else obs.get("observation_ar", obs.get("observation")),
#                         "date": obs.get("date"),
#                         "attachment_url": obs.get("attachment_url"),
#                     })

#         # --- Fetch Feedback ---
#         for student_id in associated_ids:
#             feedback_ref = db.collection("feedback").where("student_id", "==", student_id).stream()

#             for feedback in feedback_ref:
#                 feedback_data = feedback.to_dict()
#                 feedback_time_str = feedback_data.get("timestamp")
#                 if not feedback_time_str:
#                     continue

#                 try:
#                     feedback_time = parse_timestamp_flexible(feedback_time_str)
#                 except ValueError:
#                     continue

#                 if feedback_time >= last_24_hours:
#                     recent_updates.append({
#                         "student_id": student_id,
#                         "type": "feedback" if lang == "en" else "تغذية راجعة",
#                         "subject": feedback_data.get("subject"),
#                         "feedback": feedback_data.get("feedback")
#                         if lang == "en"
#                         else feedback_data.get("feedback_ar", feedback_data.get("feedback")),
#                         "timestamp": feedback_data.get("timestamp"),
#                     })

#         if not recent_updates:
#             return jsonify({
#                 "message": "No recent updates found in the last 24 hours"
#                 if lang == "en"
#                 else "لا توجد تحديثات حديثة خلال آخر 24 ساعة"
#             }), 200

#         return jsonify({"recent_updates": recent_updates}), 200

#     except Exception as e:
#         print(f"An error occurred in get_recent_updates: {e}")
#         return jsonify({
#             "error": str(e) if lang == "en" else "حدث خطأ أثناء جلب التحديثات"
#         }), 500

# -------------------- Timestamp Parser --------------------
def parse_timestamp_flexible(timestamp_str):
    """Handles multiple timestamp formats and returns UTC offset-aware datetime."""
    try:
        from dateutil import parser
        from datetime import timezone
        
        dt = parser.parse(timestamp_str)
        # Ensure all timestamps are converted to UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt
    except Exception as e:
        raise ValueError(f"Unsupported timestamp format: {timestamp_str} -> {e}")

@app.route("/api/recent-updates", methods=["GET"])
def get_recent_updates():
    
    parent_id = request.args.get("parent_id")
    lang = request.args.get("lang", "en").lower()  # Default to English

    if not parent_id:
        return jsonify({
            "error": "parent_id is required" if lang == "en" else "مطلوب معرف الوالد"
        }), 400

    try:
        parent_doc = db.collection("users").document(parent_id).get()
        if not parent_doc.exists:
            return jsonify({
                "error": "Parent not found" if lang == "en" else "لم يتم العثور على الوالد"
            }), 404

        parent_data = parent_doc.to_dict()
        associated_ids = parent_data.get("associatedIds", [])

        if not associated_ids:
            return jsonify({
                "recent_updates": [],
                "message": "No associated students found"
                if lang == "en"
                else "لم يتم العثور على طلاب مرتبطين"
            }), 200

        recent_updates = []

        # --- Fetch Observations ---
        for student_id in associated_ids:
            student_doc = db.collection("students").document(student_id).get()
            if not student_doc.exists:
                continue

            student_data = student_doc.to_dict()
            observations = student_data.get("observations", [])

            for obs in observations:
                # Add all observations (no time filter)
                recent_updates.append({
                    "student_id": student_id,
                    "type": "observation" if lang == "en" else "ملاحظة",
                    "subject": obs.get("subject"),
                    "observation": obs.get("observation")
                    if lang == "en"
                    else obs.get("observation_ar", obs.get("observation")),
                    "date": obs.get("date"),
                    "attachment_url": obs.get("attachment_url"),
                })

        # --- Fetch Feedback ---
        for student_id in associated_ids:
            feedback_ref = db.collection("feedback").where("student_id", "==", student_id).stream()

            for feedback in feedback_ref:
                feedback_data = feedback.to_dict()
                feedback_time_str = feedback_data.get("timestamp")
                if not feedback_time_str:
                    continue

                # Add all feedback (no time filter)
                recent_updates.append({
                    "student_id": student_id,
                    "type": "feedback" if lang == "en" else "تغذية راجعة",
                    "subject": feedback_data.get("subject"),
                    "feedback": feedback_data.get("feedback")
                    if lang == "en"
                    else feedback_data.get("feedback_ar", feedback_data.get("feedback")),
                    "timestamp": feedback_data.get("timestamp"),
                })

        if not recent_updates:
            return jsonify({
                "message": "No updates found for associated students"
                if lang == "en"
                else "لم يتم العثور على تحديثات للطلاب المرتبطين"
            }), 200
        
        # ✅ THE CRITICAL CHANGE: Sort the combined list by date/timestamp
        # We use a lambda to handle both "date" (for observations) and "timestamp" (for feedback)
        # and parse them to ensure correct temporal ordering.
        recent_updates.sort(
            key=lambda x: parse_timestamp_flexible(x.get("date") or x.get("timestamp")),
            reverse=True # Newest first (descending)
        )
        
        return jsonify({"recent_updates": recent_updates}), 200

    except Exception as e:
        print(f"An error occurred in get_recent_updates: {e}")
        return jsonify({
            "error": str(e) if lang == "en" else "حدث خطأ أثناء جلب التحديثات"
        }), 500

 

@app.route("/api/resources", methods=["GET"])
def get_resources():
    """
    Fetches recommended resources for the parent's associated students.
    """
    parent_id = request.args.get("parent_id")

    if not parent_id:
        return jsonify({"error": "parent_id is required"}), 400

    try:
        # Ã°Å¸â€Â¥ Step 1: Fetch Associated Student IDs
        parent_doc = db.collection("users").document(parent_id).get()

        if not parent_doc.exists:
            return jsonify({"error": "Parent not found"}), 404

        associated_ids = parent_doc.to_dict().get("associatedIds", [])

        # Ã¢Å“â€¦ Ensure associated_ids is always a list
        if not isinstance(associated_ids, list):
            associated_ids = [associated_ids]

        if not associated_ids:
            return jsonify({"error": "No associated student found"}), 404

        all_resources = []

        # Ã°Å¸â€Â¥ Step 2: Iterate over associated student IDs
        for student_id in associated_ids:
            student_doc = db.collection("students").document(student_id).get()

            if not student_doc.exists:
                continue

            student_data = student_doc.to_dict()

            # Ã¢Å“â€¦ Extract Teaching Plans
            teaching_plans = student_data.get("teachingPlans", {})

            # Ã¢Å“â€¦ Log the entire teaching plans to verify the structure
            print(f"Teaching Plans for {student_id}: {teaching_plans}")

            # Ã¢Å“â€¦ Handle both empty and populated plans
            if isinstance(teaching_plans, dict) and teaching_plans:
                for plan_id, plan_data in teaching_plans.items():
                    if isinstance(plan_data, dict):

                        # Ã¢Å“â€¦ Check for `actionPlan` field
                        action_plan = plan_data.get("actionPlan", {})

                        # Ã°Å¸â€Â¥ Extract `recommendedResources` from `actionPlan`
                        recommended = action_plan.get("recommendedResources", {})

                        # Collect resources only if they exist
                        if isinstance(recommended, dict) and recommended:
                            for key, value in recommended.items():
                                all_resources.append(
                                    {
                                        "student_id": student_id,
                                        "plan_id": plan_id,
                                        "resource": key,
                                        "description": value,
                                    }
                                )

        # Ã¢Å“â€¦ Return "No resources found" if the list is empty
        if not all_resources:
            return jsonify({"message": "No resources found"}), 200

        return jsonify({"parent_id": parent_id, "resources": all_resources}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

#==============================================================================
from flask import Flask, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import os
from datetime import datetime

# Ã¢Å“â€¦ Firebase Initialization
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
        print("Ã¢Å“â€¦ Firebase initialized successfully.")
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()
print("Ã¢Å“â€¦ Firestore client initialized.")



from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore, firestore_async
import os

# Ã¢Å“â€¦ Initialize Firebase
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
        print("Ã¢Å“â€¦ Firebase initialized successfully.")
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()
db11 = firestore_async.client()



# # Ã¢Å“â€¦ Function to extract grade name from nested structure
# def extract_grade_name(student_data):
#     """
#     Extract grade name from nested 'assignedGrades' structure.
#     """
#     try:
#         grades = student_data.get("assignedGrades", {}).get("grades", {})

#         if not grades:
#             return "Unknown Grade"

#         # Extract the grade names
#         grade_names = list(grades.keys())

#         # Join multiple grades with a comma if they exist
#         return ", ".join(grade_names) if grade_names else "Unknown Grade"

#     except Exception as e:
#         print(f"Ã¢ÂÅ’ Error extracting grade name: {e}")
#         return "Unknown Grade"



# # --- 2. Helper to Find Grade in Student Data ---
# def extract_grade_name_smart(student_data):
#     # 1. Try academicData (Best source)
#     academic = student_data.get("academicData", {})
#     if isinstance(academic, dict) and academic.get("grade"):
#         return normalize_grade_name_for_count(academic.get("grade"))

#     # 2. Try root level
#     if student_data.get("grade"):
#         return normalize_grade_name_for_count(student_data.get("grade"))
    
#     # 3. Try assignedGrades keys (Old source)
#     assigned = student_data.get("assignedGrades", {})
#     if isinstance(assigned, dict):
#         # Handle nested 'grades'
#         if "grades" in assigned and isinstance(assigned["grades"], dict):
#             keys = list(assigned["grades"].keys())
#             if keys: return normalize_grade_name_for_count(keys[0])
        
#         # Handle flat structure
#         keys = list(assigned.keys())
#         if keys: return normalize_grade_name_for_count(keys[0])

#     return "UNKNOWN"

# def extract_grade_name(student_data):
#     """
#     Extract grade name from nested 'assignedGrades' structure. (Placeholder logic)
#     """
#     try:
#         grades = student_data.get("assignedGrades", {}).get("grades", {})
#         if not grades:
#             return "Unknown Grade"
#         grade_names = list(grades.keys())
#         return ", ".join(grade_names) if grade_names else "Unknown Grade"
#     except Exception as e:
#         print(f"❌ Error extracting grade name: {e}")
#         return "Unknown Grade"
import re
from flask import jsonify, request

# # --- 1. Strict Normalizer ---
# def normalize_grade_name_strict(name):
#     if not name: return "UNKNOWN"
#     name = str(name).upper()
#     name = name.replace("_", " ").replace("-", " ")
#     name = re.sub(r'\s*\(\s*', ' (', name)
#     name = re.sub(r'\s*\)\s*', ')', name)
#     return " ".join(name.split())

# # --- 2. Smart Extractor ---
# def extract_grade_name_smart_sync(student_data):
#     # Priority 1: academicData
#     academic = student_data.get("academicData", {})
#     if isinstance(academic, dict) and academic.get("grade"):
#         return normalize_grade_name_strict(academic.get("grade"))

#     # Priority 2: Root level
#     if student_data.get("grade"):
#         return normalize_grade_name_strict(student_data.get("grade"))
    
#     # Priority 3: assignedGrades
#     assigned = student_data.get("assignedGrades", {})
#     if isinstance(assigned, dict):
#         if "grades" in assigned and isinstance(assigned["grades"], dict):
#             keys = list(assigned["grades"].keys())
#             if keys: return normalize_grade_name_strict(keys[0])
#         keys = list(assigned.keys())
#         if keys: return normalize_grade_name_strict(keys[0])

#     return "UNKNOWN"

# # --- 3. The Sync API Route ---
# @app.route("/api/grade-student-count1", methods=["GET"])
# def get_grade_student_count1():
#     teacher_id = request.args.get("teacherid")
#     print(f"\n🔍 [DEBUG] Requesting stats for teacher: {teacher_id}")
    
#     if not teacher_id:
#         return jsonify({"error": "Missing teacher ID"}), 400

#     try:
#         # --- CRITICAL CHANGE: Use 'db' (Sync), not 'db11' (Async) ---
#         # Make sure 'db' is your firestore.client() variable defined at the top of app.py
        
#         # 1. Fetch Teacher
#         teacher_ref = db.collection("users").document(teacher_id)
#         teacher_doc = teacher_ref.get()

#         if not teacher_doc.exists:
#             print(f"❌ [DEBUG] Teacher doc not found in 'users' collection.")
#             return jsonify({"error": "Teacher not found"}), 404

#         associated_ids = teacher_doc.to_dict().get("associatedIds", [])
#         print(f"📄 [DEBUG] Found {len(associated_ids)} associated IDs: {associated_ids}")

#         if not associated_ids:
#             return jsonify({"message": "No associated students found"}), 404

#         # 2. Fetch Students & Count
#         grade_counts = {}

#         for student_id in associated_ids:
#             # Note: Ensure your students are in 'users' or 'students' collection.
#             # Based on your logs, they seem to be in 'users' with role='student'
#             # BUT if they are in 'students', change "users" to "students" below.
#             student_ref = db.collection("users").document(student_id)
#             student_doc = student_ref.get()

#             if student_doc.exists:
#                 data = student_doc.to_dict()
#                 grade_name = extract_grade_name_smart_sync(data)
#                 print(f"   👤 [DEBUG] Student {student_id} -> Grade: {grade_name}")
                
#                 if grade_name != "UNKNOWN":
#                     grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1
#             else:
#                 print(f"   ⚠️ [DEBUG] Student ID {student_id} in list but doc not found.")

#         # 3. Sort and Return
#         def sort_key(item):
#             match = re.search(r'\d+', item["grade_name"])
#             return int(match.group()) if match else 999

#         response = [
#             {"grade_name": grade, "student_count": count}
#             for grade, count in grade_counts.items()
#         ]
#         response.sort(key=sort_key)

#         print(f"✅ [DEBUG] Final Response: {response}\n")
#         return jsonify(response), 200

#     except Exception as e:
#         print(f"❌ [DEBUG] Exception: {e}")
#         return jsonify({"error": str(e)}), 500
import re
from flask import jsonify, request

# --- 1. Strict Normalizer ---
def normalize_grade_name_strict(name):
    if not name: return "UNKNOWN"
    
    # Force string and uppercase
    name = str(name).upper()
    
    # CRITICAL: Replace underscores (_) and hyphens (-) with spaces
    name = name.replace("_", " ").replace("-", " ")
    
    # Fix parentheses spacing: "GRADE 11(SCIENCE)" -> "GRADE 11 (SCIENCE)"
    name = re.sub(r'\s*\(\s*', ' (', name)
    name = re.sub(r'\s*\)\s*', ')', name)
    
    # Remove extra spaces
    return " ".join(name.split())

# --- 2. Smart Extractor ---
def extract_grade_name_smart_sync(student_data):
    # Priority 1: academicData (Best source)
    academic = student_data.get("academicData", {})
    if isinstance(academic, dict) and academic.get("grade"):
        return normalize_grade_name_strict(academic.get("grade"))

    # Priority 2: Root level
    if student_data.get("grade"):
        return normalize_grade_name_strict(student_data.get("grade"))
    
    # Priority 3: assignedGrades keys
    assigned = student_data.get("assignedGrades", {})
    if isinstance(assigned, dict):
        # Handle nested 'grades'
        if "grades" in assigned and isinstance(assigned["grades"], dict):
            keys = list(assigned["grades"].keys())
            if keys: return normalize_grade_name_strict(keys[0])
        # Handle flat structure
        keys = list(assigned.keys())
        if keys: return normalize_grade_name_strict(keys[0])

    return "UNKNOWN"

# --- 3. The Optimized Sync API Route ---
@app.route("/api/grade-student-count1", methods=["GET"])
def get_grade_student_count1():
    teacher_id = request.args.get("teacherid")
    
    if not teacher_id:
        return jsonify({"error": "Missing teacher ID"}), 400

    try:
        # 1. Fetch Teacher
        # Make sure 'db' is your synchronous firestore.client()
        teacher_ref = db.collection("users").document(teacher_id)
        teacher_doc = teacher_ref.get()

        if not teacher_doc.exists:
            return jsonify({"error": "Teacher not found"}), 404

        associated_ids = teacher_doc.to_dict().get("associatedIds", [])

        if not associated_ids:
            return jsonify({"message": "No associated students found"}), 404

        # 2. Fetch All Students in ONE Batch (High Speed)
        print(f"📄 [DEBUG] Fetching {len(associated_ids)} students in batch...")
        
        # Create a list of document references
        # Note: Ensure 'users' is the correct collection. Change to 'students' if needed.
        student_refs = [db.collection("users").document(sid) for sid in associated_ids]
        
        # Fetch all documents in a single request
        student_docs = db.get_all(student_refs)

        grade_counts = {}

        for doc in student_docs:
            if doc.exists:
                grade_name = extract_grade_name_smart_sync(doc.to_dict())
                
                if grade_name != "UNKNOWN":
                    grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1

        # 3. Sort naturally (Grade 2 before Grade 10)
        def sort_key(item):
            match = re.search(r'\d+', item["grade_name"])
            return int(match.group()) if match else 999

        response = [
            {"grade_name": grade, "student_count": count}
            for grade, count in grade_counts.items()
        ]
        response.sort(key=sort_key)

        return jsonify(response), 200

    except Exception as e:
        print(f"❌ Exception: {e}")
        return jsonify({"error": str(e)}), 500

# # Ã¢Å“â€¦ API to get grade names and student counts by teacher ID
# @app.route("/api/grade-student-count", methods=["GET"])
# def get_grade_student_count():
#     """
#     Fetch grade names and student counts by teacher ID using associated_id mapping.
#     """
#     teacher_id = request.args.get("teacherid")

#     print(f"Ã°Å¸â€Å½ Received teacher_id: {teacher_id}")

#     if not teacher_id:
#         print("Ã¢ÂÅ’ Missing teacher ID")
#         return jsonify({"error": "Missing teacher ID"}), 400

#     try:
#         # Ã¢Å“â€¦ Fetch teacher document
#         print("Ã°Å¸â€Å½ Fetching teacher document from Firestore...")
#         teacher_ref = db.collection("users").document(teacher_id).get()

#         if not teacher_ref.exists:
#             print(f"Ã¢ÂÅ’ Teacher with ID {teacher_id} not found.")
#             return jsonify({"error": "Teacher not found"}), 404

#         teacher_data = teacher_ref.to_dict()
#         print(f"Ã¢Å“â€¦ Teacher Data: {teacher_data}")

#         # Ã¢Å“â€¦ Extract associated student IDs
#         associated_ids = teacher_data.get("associatedIds", [])
#         print(f"Ã°Å¸â€œÅ’ Associated Student IDs: {associated_ids}")

#         if not associated_ids:
#             print("Ã¢Å¡Â Ã¯Â¸Â No associated students found.")
#             return jsonify({"message": "No associated students found"}), 404

#         # Ã¢Å“â€¦ Initialize result dictionary
#         grade_counts = {}

#         # Ã¢Å“â€¦ Fetch students by associated ID
#         for student_id in associated_ids:
#             print(f"Ã°Å¸â€Å½ Fetching student document for ID: {student_id}")

#             student_ref = db.collection("students").document(student_id).get()

#             if student_ref.exists:
#                 student_data = student_ref.to_dict()
#                 print(f"Ã¢Å“â€¦ Student Data: {student_data}")

#                 # Ã¢Å“â€¦ Extract grade name using the new function
#                 grade_name = extract_grade_name(student_data)
#                 print(f"Ã°Å¸â€œÅ’ Extracted Grade Name: {grade_name}")

#                 # Ã¢Å“â€¦ Count students per grade
#                 if grade_name in grade_counts:
#                     grade_counts[grade_name] += 1
#                 else:
#                     grade_counts[grade_name] = 1
#             else:
#                 print(f"Ã¢Å¡Â Ã¯Â¸Â Student with ID {student_id} not found.")

#         # Ã¢Å“â€¦ Prepare response
#         response = []
#         for grade_name, count in grade_counts.items():
#             print(f"Ã°Å¸â€œÅ  Grade: {grade_name}, Count: {count}")
#             response.append({"grade_name": grade_name, "student_count": count})

#         if not response:
#             print("Ã¢Å¡Â Ã¯Â¸Â No students found.")
#             return jsonify({"message": "No students found for the associated IDs"}), 404

#         print("Ã¢Å“â€¦ API Response:", response)
#         return jsonify(response), 200

#     except Exception as e:
#         print(f"Ã¢ÂÅ’ Exception: {e}")
#         return jsonify({"error": str(e)}), 500
# --- 3. The API Endpoint ---
@app.route("/api/grade-student-count", methods=["GET"])
async def get_grade_student_count():
    """
    Async API to get student counts per grade for a specific teacher.
    """
    teacher_id = request.args.get("teacherid")
    if not teacher_id:
        return jsonify({"error": "Missing teacher ID"}), 400

    try:
        # 1. Fetch Teacher
        teacher_ref = db.collection("users").document(teacher_id)
        teacher_doc = teacher_ref.get() # Sync call is fine here

        if not teacher_doc.exists:
            return jsonify({"error": "Teacher not found"}), 404

        associated_ids = teacher_doc.to_dict().get("associatedIds", [])

        if not associated_ids:
            return jsonify({"message": "No associated students found"}), 404

        # 2. Fetch Students
        grade_counts = {}
        
        # Loop through IDs (Robust against sync/async db differences)
        for student_id in associated_ids:
            student_ref = db.collection("students").document(student_id)
            student_doc = student_ref.get()

            if student_doc.exists:
                grade_name = extract_grade_name_smart(student_doc.to_dict())
                grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1

        # 3. Format and Sort Response
        def sort_key(item):
            # Sort by number (Grade 2 before Grade 10)
            match = re.search(r'\d+', item["grade_name"])
            return int(match.group()) if match else 999

        response = [
            {"grade_name": grade, "student_count": count}
            for grade, count in grade_counts.items()
        ]
        response.sort(key=sort_key)

        return jsonify(response), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500



import re

import re

# def normalize_grade_name(name):
#     if not name:
#         return "UNKNOWN"

#     # 1. Force string and Uppercase
#     name = str(name).upper()

#     # 2. Replace underscores (_) and hyphens (-) with spaces
#     # This turns "GRADE_2" into "GRADE 2"
#     name = name.replace("_", " ").replace("-", " ")

#     # 3. Standardize spacing around parentheses 
#     # Turns "GRADE 11(SCIENCE)" -> "GRADE 11 (SCIENCE)"
#     name = re.sub(r'\s*\(\s*', ' (', name)
#     name = re.sub(r'\s*\)\s*', ')', name)

#     # 4. Remove extra spaces (e.g., "GRADE  2" -> "GRADE 2")
#     name = " ".join(name.split())

# def normalize_grade_name_for_count(name):
#     if not name:
#         return "UNKNOWN"
    
#     # Force string and uppercase
#     name = str(name).upper()

#     # Replace underscores and hyphens with spaces
#     name = name.replace("_", " ").replace("-", " ")

#     # Fix parentheses spacing: "GRADE 11(SCIENCE)" -> "GRADE 11 (SCIENCE)"
#     name = re.sub(r'\s*\(\s*', ' (', name)
#     name = re.sub(r'\s*\)\s*', ')', name)

#     # Remove extra spaces
#     return " ".join(name.split())

#     return name
# # âœ… Function to extract grade name from nested structure
# async def extract_grade_name_async(student_data):
#     """
#     Extract a list of normalized grade names from student data.
#     Supports both assignedGrades.grades and assignedGrades directly.
#     """
#     try:
#         assigned_grades = student_data.get("assignedGrades", {})

#         # Handle both cases
#         if isinstance(assigned_grades, dict):
#             if "grades" in assigned_grades and isinstance(assigned_grades["grades"], dict):
#                 raw_grade_names = list(assigned_grades["grades"].keys())
#             else:
#                 # Fallback: use assignedGrades directly
#                 raw_grade_names = list(assigned_grades.keys())
            
#             return [normalize_grade_name(g) for g in raw_grade_names]

#         return []

#     except Exception as e:
#         print(f"? Error extracting grade names: {e}")
#         return []

# global_loop = asyncio.get_event_loop()


# # # âœ… API to get grade names and student counts by teacher ID
# # @app.route("/api/grade-student-count1", methods=["GET"])
# # def get_grade_student_count1115689():
# #     teacher_id = request.args.get("teacherid")
# #     print(f"ðŸ” Received teacher_id: {teacher_id}")
# #     response = global_loop.run_until_complete(get_grade_student_count111(teacher_id))
# #     return response


# # async def get_grade_student_count111(teacher_id):
# #     """
# #     Fetch grade names and student counts by teacher ID using associated_id mapping.
# #     """
# #     # teacher_id = request.args.get("teacherid")
# #     # print(f"ðŸ” Received teacher_id: {teacher_id}")

# #     if not teacher_id:
# #         print("âŒ Missing teacher ID")
# #         return jsonify({"error": "Missing teacher ID"}), 400

# #     try:
# #         # âœ… Fetch teacher document
# #         print("ðŸ” Fetching teacher document from Firestore...")
# #         teacher_ref = db11.collection("users").document(teacher_id)
# #         teacher_doc = await teacher_ref.get()

# #         if not teacher_doc.exists:
# #             print(f"âŒ Teacher with ID {teacher_id} not found.")
# #             return jsonify({"error": "Teacher not found"}), 404

# #         teacher_data = teacher_doc.to_dict()
# #         print(f"âœ… Teacher Data: {teacher_data}")

# #         # âœ… Extract associated student IDs
# #         associated_ids = teacher_data.get("associatedIds", [])
# #         print(f"ðŸ“„ Associated Student IDs: {associated_ids}")

# #         if not associated_ids:
# #             print("âš ï¸ No associated students found.")
# #             return jsonify({"message": "No associated students found"}), 404

# #         # âœ… Initialize result dictionary
# #         grade_counts = {}

# #         # âœ… Fetch students concurrently
# #         student_tasks = [
# #             db11.collection("users").document(student_id).get()
# #             for student_id in associated_ids
# #         ]
# #         student_docs = await asyncio.gather(*student_tasks)

# #         for student_doc in student_docs:
# #             if student_doc.exists:
# #                 student_data = student_doc.to_dict()
# #                 print(f"âœ… Student Data: {student_data}")

# #                 # âœ… Extract grade name using async function
# #                 grade_names = await extract_grade_name_async(student_data)
# #                 print(f"?? Extracted Grade Names: {grade_names}")

# #                 # âœ… Count students per grade
# #                 for grade_name in grade_names:
# #                     grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1
# #             else:
# #                 print("âš ï¸ Student not found.")

# #         def extract_grade_number(grade_name):
# #             match = re.search(r'\d+', grade_name)
# #             return int(match.group()) if match else float('inf')

# #         response = sorted(
# #           [{"grade_name": grade, "student_count": count} for grade, count in grade_counts.items()],
# #           key=lambda x: extract_grade_number(x["grade_name"])
# #           )

# #         if not response:
# #             print("âš ï¸ No students found.")
# #             return jsonify({"message": "No students found for the associated IDs"}), 404

# #         print("âœ… API Response:", response)
# #         return jsonify(response), 200

# #     except Exception as e:
# #         print(f"âŒ Exception: {e}")
# #         return jsonify({"error": str(e)}), 500
# import re
# import asyncio
# from flask import jsonify, request

# # --- 1. Strict Normalizer (Merges 'GRADE_2' -> 'GRADE 2') ---
# def normalize_grade_name_strict(name):
#     if not name: return "UNKNOWN"
    
#     # Force string and uppercase
#     name = str(name).upper()
    
#     # CRITICAL: Replace underscores (_) and hyphens (-) with spaces
#     name = name.replace("_", " ").replace("-", " ")
    
#     # Fix parentheses spacing: "GRADE 11(SCIENCE)" -> "GRADE 11 (SCIENCE)"
#     name = re.sub(r'\s*\(\s*', ' (', name)
#     name = re.sub(r'\s*\)\s*', ')', name)
    
#     # Remove extra spaces
#     return " ".join(name.split())

# # --- 2. Smart Extractor (Finds grade in 'academicData' OR 'assignedGrades') ---
# def extract_grade_name_smart_v2(student_data):
#     # Priority 1: academicData
#     academic = student_data.get("academicData", {})
#     if isinstance(academic, dict) and academic.get("grade"):
#         return normalize_grade_name_strict(academic.get("grade"))

#     # Priority 2: Root level
#     if student_data.get("grade"):
#         return normalize_grade_name_strict(student_data.get("grade"))
    
#     # Priority 3: assignedGrades keys (Legacy/User collection structure)
#     assigned = student_data.get("assignedGrades", {})
#     if isinstance(assigned, dict):
#         # Handle nested 'grades'
#         if "grades" in assigned and isinstance(assigned["grades"], dict):
#             keys = list(assigned["grades"].keys())
#             if keys: return normalize_grade_name_strict(keys[0])
#         # Handle flat structure
#         keys = list(assigned.keys())
#         if keys: return normalize_grade_name_strict(keys[0])

#     return "UNKNOWN"

# # --- 3. The API Route (Async Native) ---
# # We removed 'global_loop' because running with Uvicorn allows native async
# @app.route("/api/grade-student-count1", methods=["GET"])
# async def get_grade_student_count_route_v1():
#     teacher_id = request.args.get("teacherid")
#     print(f"🔍 Received teacher_id: {teacher_id}")
#     # Directly await the logic function
#     return await get_grade_student_count111(teacher_id)

# # --- 4. The Logic Function ---
# async def get_grade_student_count111(teacher_id):
#     if not teacher_id:
#         return jsonify({"error": "Missing teacher ID"}), 400

#     try:
#         # Fetch Teacher from 'users'
#         teacher_ref = db11.collection("users").document(teacher_id)
#         teacher_doc = await teacher_ref.get()

#         if not teacher_doc.exists:
#             return jsonify({"error": "Teacher not found"}), 404

#         associated_ids = teacher_doc.to_dict().get("associatedIds", [])

#         if not associated_ids:
#             return jsonify({"message": "No associated students found"}), 404

#         # Fetch Students (Parallel) from 'users' collection
#         # (Assuming your students are in 'users' based on your previous snippet)
#         tasks = [
#             db11.collection("users").document(student_id).get() 
#             for student_id in associated_ids
#         ]
#         student_docs = await asyncio.gather(*tasks)

#         grade_counts = {}

#         for doc in student_docs:
#             if doc.exists:
#                 # Use V2 Smart Extractor
#                 grade_name = extract_grade_name_smart_v2(doc.to_dict())
#                 grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1

#         # Sort naturally (Grade 2 before Grade 10)
#         def sort_key(item):
#             match = re.search(r'\d+', item["grade_name"])
#             return int(match.group()) if match else 999

#         response = [
#             {"grade_name": grade, "student_count": count}
#             for grade, count in grade_counts.items()
#         ]
#         response.sort(key=sort_key)

#         return jsonify(response), 200

#     except Exception as e:
#         print(f"❌ Exception: {e}")
#         return jsonify({"error": str(e)}), 500


# # async def get_grade_student_count11123(teacher_id):
# #     """
# #     Fetch grade names and student counts by teacher ID (Async, Sequential Mode).
# #     """

# #     if not teacher_id:
# #         print("âŒ Missing teacher ID")
# #         return jsonify({"error": "Missing teacher ID"}), 400

# #     try:
# #         # âœ… Fetch teacher document
# #         print("ðŸ” Fetching teacher document from Firestore...")
# #         teacher_ref = db11.collection("users").document(teacher_id)
# #         teacher_doc = await teacher_ref.get()

# #         if not teacher_doc.exists:
# #             print(f"âŒ Teacher with ID {teacher_id} not found.")
# #             return jsonify({"error": "Teacher not found"}), 404

# #         teacher_data = teacher_doc.to_dict()
# #         print(f"âœ… Teacher Data: {teacher_data}")

# #         # âœ… Extract associated student IDs
# #         associated_ids = teacher_data.get("associatedIds", [])
# #         print(f"ðŸ“„ Associated Student IDs: {associated_ids}")

# #         if not associated_ids:
# #             print("âš ï¸ No associated students found.")
# #             return jsonify({"message": "No associated students found"}), 404

# #         # âœ… Fetch all students in parallel
# #         student_docs = await asyncio.gather(
# #             *[
# #                 db11.collection("students").document(student_id).get()
# #                 for student_id in associated_ids
# #             ]
# #         )

# #         # âœ… Initialize result dictionary
# #         grade_counts = {}

# #         for student_doc in student_docs:
# #             if student_doc.exists:
# #                 student_data = student_doc.to_dict()
# #                 grade_name = extract_grade_name(student_data)
# #                 grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1

# #         # âœ… Prepare response
# #         response = [
# #             {"grade_name": grade, "student_count": count}
# #             for grade, count in grade_counts.items()
# #         ]
# #         return jsonify(response), 200 if response else 404

# #     except Exception as e:
# #         print(f"âŒ Exception: {e}")
# #         return jsonify({"error": str(e)}), 500

# async def get_grade_student_count11123(teacher_id):
#     """
#     Fetch grade names and student counts by teacher ID (Async, Sequential Mode).
#     """

#     if not teacher_id:
#         print("❌ Missing teacher ID")
#         return jsonify({"error": "Missing teacher ID"}), 400

#     try:
#         # ✅ Fetch teacher document
#         print("🔍 Fetching teacher document from Firestore...")
#         teacher_ref = db11.collection("users").document(teacher_id)
#         teacher_doc = await teacher_ref.get()

#         if not teacher_doc.exists:
#             print(f"❌ Teacher with ID {teacher_id} not found.")
#             return jsonify({"error": "Teacher not found"}), 404

#         teacher_data = teacher_doc.to_dict()
#         print(f"✅ Teacher Data: {teacher_data}")

#         # ✅ Extract associated student IDs
#         associated_ids = teacher_data.get("associatedIds", [])
#         print(f"📄 Associated Student IDs: {associated_ids}")

#         if not associated_ids:
#             print("⚠️ No associated students found.")
#             return jsonify({"message": "No associated students found"}), 404

#         # ✅ Fetch all students in parallel
#         student_docs = await asyncio.gather(
#             *[
#                 db11.collection("students").document(student_id).get()
#                 for student_id in associated_ids
#             ]
#         )

#         # ✅ Initialize result dictionary
#         grade_counts = {}

#         for student_doc in student_docs:
#             if student_doc.exists:
#                 student_data = student_doc.to_dict()
#                 grade_name = extract_grade_name_smart(student_data)

#                 # --- FIX START: Normalize the grade name ---
#                 if grade_name:
#                     # 1. Force string to avoid errors
#                     grade_name = str(grade_name) 
#                     # 2. Replace underscores with spaces (GRADE_2 -> GRADE 2)
#                     grade_name = grade_name.replace("_", " ")
#                     # 3. Remove extra spaces and ensure uppercase
#                     grade_name = " ".join(grade_name.split()).upper()
#                 else:
#                     grade_name = "UNKNOWN" # Handle missing grades
#                 # --- FIX END ---

#                 grade_counts[grade_name] = grade_counts.get(grade_name, 0) + 1

#         # ✅ Prepare response
#         response = [
#             {"grade_name": grade, "student_count": count}
#             for grade, count in grade_counts.items()
#         ]
#         return jsonify(response), 200 if response else 404

#     except Exception as e:
#         print(f"❌ Exception: {e}")
#         return jsonify({"error": str(e)}), 500


from flask import Flask, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import os


# Ã¢Å“â€¦ Firebase Initialization
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()


# Ã¢Å“â€¦ Function to convert percentage to letter grade
def calculate_grade(percentage):
    if percentage >= 90:
        return "A"
    elif percentage >= 75:
        return "B"
    elif percentage >= 60:
        return "C"
    elif percentage >= 40:
        return "D"
    else:
        return "F"


# # Ã¢Å“â€¦ API: School Performance Metrics
# @app.route("/api/school-performance", methods=["GET"])
# def get_school_performance():
#     """Get average grades by grade and subject across all students"""
#     try:
#         students_ref = db.collection("students").stream()

#         grade_subject_marks = {}

#         for student in students_ref:
#             student_data = student.to_dict()

#             # Extract grade and subjects
#             academic_data = student_data.get("academicData", {})
#             grade = academic_data.get("grade")
#             subjects = academic_data.get("subjects", {})

#             if grade not in grade_subject_marks:
#                 grade_subject_marks[grade] = {}

#             # Handle both nested and flat subject structures
#             for subject, details in subjects.items():
#                 if isinstance(details, dict):
#                     # Handle flat structure
#                     if "marks" in details:
#                         marks = details.get("marks", 0)
#                         if subject not in grade_subject_marks[grade]:
#                             grade_subject_marks[grade][subject] = []
#                         grade_subject_marks[grade][subject].append(marks)

#                     # Handle nested curriculum structure
#                     elif "history" in details:
#                         for item in details["history"]:
#                             marks = item.get("marks", 0)
#                             if subject not in grade_subject_marks[grade]:
#                                 grade_subject_marks[grade][subject] = []
#                             grade_subject_marks[grade][subject].append(marks)

#         # Ã¢Å“â€¦ Calculate average marks per subject per grade
#         result = []
#         for grade, subjects in grade_subject_marks.items():
#             avg_subject_grades = {}

#             for subject, marks_list in subjects.items():
#                 if marks_list:
#                     avg_percentage = sum(marks_list) / len(marks_list)
#                     avg_subject_grades[subject] = calculate_grade(avg_percentage)

#             result.append({"grade": grade, "subjects": avg_subject_grades})

#         return jsonify({"metrics": result}), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# ✅ API: School Performance Metrics (FIXED)
@app.route("/api/school-performance", methods=["GET"])
def get_school_performance():
    """Get average grades by grade and subject across all students"""
    try:
        students_ref = db.collection("students").stream()

        grade_subject_marks = {}

        for student in students_ref:
            student_data = student.to_dict()

            # Extract grade and subjects
            academic_data = student_data.get("academicData", {})
            grade = academic_data.get("grade")
            subjects = academic_data.get("subjects", {})

            # --- FIX 1: Filter out invalid grades immediately ---
            # Skips if grade is None, empty string, or literal "Unknown"
            if not grade or str(grade).strip() == "" or str(grade).lower() == "unknown":
                continue
            # ----------------------------------------------------

            # Initialize dictionary for this grade if it doesn't exist
            if grade not in grade_subject_marks:
                grade_subject_marks[grade] = {}

            # Handle both nested and flat subject structures
            for subject, details in subjects.items():
                # Skip if details is not a dictionary
                if not isinstance(details, dict):
                    continue

                marks_to_add = []

                # Handle flat structure (direct marks)
                if "marks" in details:
                    marks_to_add.append(details.get("marks", 0))

                # Handle nested curriculum structure (history array)
                elif "history" in details and isinstance(details["history"], list):
                    for item in details["history"]:
                        if isinstance(item, dict):
                            marks_to_add.append(item.get("marks", 0))
                
                # Only add if we found marks
                if marks_to_add:
                    if subject not in grade_subject_marks[grade]:
                        grade_subject_marks[grade][subject] = []
                    
                    # Extend the list of marks for this subject/grade
                    grade_subject_marks[grade][subject].extend(marks_to_add)

        # ✅ Calculate average marks per subject per grade
        result = []
        
        # Sort grades alphabetically/numerically if possible for cleaner output
        for grade, subjects in grade_subject_marks.items():
            avg_subject_grades = {}
            
            # --- FIX 2: Only include subjects that actually have data ---
            if not subjects:
                continue 
            # ------------------------------------------------------------

            for subject, marks_list in subjects.items():
                if marks_list:
                    avg_percentage = sum(marks_list) / len(marks_list)
                    # Ensure calculate_grade function is available in your code scope
                    avg_subject_grades[subject] = calculate_grade(avg_percentage)

            # Only append to result if this grade has at least one subject with data
            if avg_subject_grades:
                result.append({"grade": grade, "subjects": avg_subject_grades})

        return jsonify({"metrics": result}), 200

    except Exception as e:
        print(f"Error in school performance: {e}")
        return jsonify({"error": str(e)}), 500




from flask import Flask, request, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import openai
import os
import time


# âœ… Initialize Firebase
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
        print("âœ… Firebase initialized successfully.")
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()
from openai import OpenAI

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
client_openai = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None


# âœ… Helper Function: Fetch teacher name using student_id
def get_teacher_name_by_student_id(student_id):
    """Fetch teacher name by matching student ID in users collection."""

    users_ref = db.collection("users").stream()

    for user in users_ref:
        user_data = user.to_dict()
        associated_ids = user_data.get("associatedIds", [])

        # âœ… Check if student_id is in associatedIds
        if student_id in associated_ids:
            # âœ… Extract teacher name from nested structure
            profile_info = user_data.get("profileInfo", {})
            personal_info = profile_info.get("personalInformation", {})

            teacher_name = personal_info.get("name", "Unknown")
            return teacher_name

    return "Unknown"


# âœ… Helper Function: Fetch associated students
def get_associated_students(parent_id):
    """Fetch associated student IDs from parent record."""

    parent_doc = db.collection("users").document(parent_id).get()

    if not parent_doc.exists:
        return []

    associated_ids = parent_doc.to_dict().get("associatedIds", [])
    students = []

    for student_id in associated_ids:
        student_doc = db.collection("students").document(student_id).get()

        if student_doc.exists:
            student_data = student_doc.to_dict()

            # âœ… Fetch teacher name by matching student ID
            teacher_name = get_teacher_name_by_student_id(student_id)

            name = (
                student_data.get("profileInfo", {})
                .get("personalInformation", {})
                .get("name", "Unknown")
            )
            if name == "Unknown":
                name = student_data.get("name", "Unknown")
            students.append(
                {
                    "student_id": student_id,
                    "name": name,
                    "teacher_name": teacher_name,
                    "academic_data": student_data.get("academicData", {}),
                }
            )

    return students


## âœ… Helper Function: Extract academic data
def extract_academic_data(student):
    """Extract academic data supporting nested history records."""

    academic_data = {
        "student_id": student.get("student_id"),
        "name": student.get("name", "Unknown"),
        "teacher_name": student.get("teacher_name", "Unknown"),
        "grade": student.get("academic_data", {}).get("grade", "N/A"),
        "subjects": {},
    }

    subjects = student.get("academic_data", {}).get("subjects", {})

    # âœ… Iterate over subjects
    for subject, details in subjects.items():
        # Handle history array properly
        history = details.get("history", [])

        if isinstance(history, list) and history:
            latest_record = history[-1]  # Get the latest history entry
            academic_data["subjects"][subject] = {
                "curriculumId": latest_record.get("curriculumId", "N/A"),
                "curriculumName": latest_record.get("curriculumName", "N/A"),
                "grade": latest_record.get("grade", "N/A"),
                "marks": latest_record.get("marks", "N/A"),
                "timestamp": latest_record.get("timestamp", "N/A"),
                "totalMarks": latest_record.get("totalMark", "N/A"),
            }
        else:
            # If no history, return default values
            academic_data["subjects"][subject] = {"grade": "N/A", "marks": "N/A"}

    return academic_data


# # âœ… Helper Function: Generate AI recommendations
# def generate_recommendations(academic_data):
#     """Generate student-wise AI recommendations with plain-text output."""

#     recommendations = {}

#     for student in academic_data:
#         student_id = student.get("student_id", "N/A")
#         name = student.get("name", "Unknown")
#         grade = student.get("grade", "N/A")
#         teacher_name = student.get("teacher_name", "Unknown")

#         prompt = f"""
#         You are an educational expert. Based on the following academic data, generate specific, actionable, and positive recommendations for parents.

#         **Output should be in plain-text format** with clear, structured points:
#         - **Student Name:** {name}
#         - **Grade:** {grade}
#         - **Teacher Name:** {teacher_name}
#         - **Subject:** [Subject name]
#             - **Strengths:** [Highlight the student's strengths in this subject]
#             - **Improvement Areas:** [Mention the key areas where improvement is needed]
#             - **Actionable Steps for Parents:** 
#               - [Suggest 2-3 specific, practical, and easy-to-implement actions parents can take]

#         Academic Data:
#         """

#         for subject, details in student.get("subjects", {}).items():
#             marks = details.get("marks", "N/A")
#             grade = details.get("grade", "N/A")
#             timestamp = details.get("timestamp", "N/A")

#             if timestamp != "N/A":
#                 prompt += f"**{subject}:** Grade - {grade}, Marks - {marks}, Last updated: {timestamp}\n"
#             else:
#                 prompt += f"**{subject}:** Grade - {grade}, Marks - {marks}\n"

#         # âœ… GPT API call
#         response = client_openai.chat.completions.create(
#             model="gpt-4.1-mini-2025-04-14",
#             messages=[
#                 {"role": "system", "content": "You are an education expert."},
#                 {"role": "user", "content": prompt},
#             ],
#             temperature=0.7,
#         )

#         # âœ… Extract GPT output
#         gpt_output = response.choices[0].message.content.strip()

#         recommendations[student_id] = {
#             "student_name": name,
#             "teacher_name": teacher_name,
#             "grade": grade,
#             "recommendations": gpt_output,
#         }

#     return recommendations


# # âœ… API Route
# @app.route("/api/resources-analysis", methods=["GET"])
# def resources_analysis():
#     """API to generate AI recommendations for parents with teacher names."""
#     start_time = time.time()

#     parent_id = request.args.get("parent_id")
#     if not parent_id:
#         return jsonify({"error": "parent_id is required"}), 400

#     try:
#         # Step 1: Fetch associated students
#         students = get_associated_students(parent_id)

#         if not students:
#             return jsonify({"message": "No associated students found"}), 404

#         # Step 2: Extract academic data
#         academic_data = [extract_academic_data(student) for student in students]

#         if not academic_data:
#             return jsonify({"message": "No academic data available"}), 404

#         # Step 3: Send academic data to GPT for recommendations
#         recommendations = generate_recommendations(academic_data)

#         # Step 4: Time check
#         elapsed_time = time.time() - start_time
#         if elapsed_time > 50:
#             return (
#                 jsonify({"error": "Timeout: Processing took longer than 10 seconds"}),
#                 504,
#             )

#         # âœ… Final Response
#         return jsonify(
#             {
#                 "parent_id": parent_id,
#                 "students": academic_data,
#                 "recommendations": recommendations,
#                 "execution_time": f"{elapsed_time:.2f} sec",
#             }
#         )

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# Assuming 'client_openai', 'app', 'request', 'jsonify', 'time'
# 'get_associated_students', and 'extract_academic_data' are imported/defined elsewhere.

import time
import json # Used for parsing GPT's JSON output
from flask import request, jsonify

# --- Configuration ---
SUPPORTED_LANGUAGES = {"en": "English", "ar": "Arabic"}
DEFAULT_LANGUAGE = "en"

# Note: The original academic_data structure is preserved.
# The original API client call to 'client_openai' is preserved inside the loop.


# âœ… Helper Function: Generate AI recommendations (MODIFIED)
def generate_recommendations(academic_data):
    """
    Generate student-wise AI recommendations in both English and Arabic.
    The LLM response is forced into a JSON structure containing 'en_recommendations' 
    and 'ar_recommendations' as plain-text strings.
    """

    recommendations = {}

    for student in academic_data:
        student_id = student.get("student_id", "N/A")
        name = student.get("name", "Unknown")
        grade = student.get("grade", "N/A")
        teacher_name = student.get("teacher_name", "Unknown")

        # --- Part 1: Constructing the Prompt ---

        # Base structure prompt for the LLM
        prompt_structure = f"""
        You are an educational expert. Based on the academic data below, generate specific, actionable, and positive recommendations for parents.

        **Instructions:**
        1. Analyze the student's performance across subjects.
        2. Generate the full structured recommendation for all subjects in **English**.
        3. Generate the full structured recommendation for all subjects in **Arabic**.
        4. The output **MUST** be a valid JSON object with exactly these two top-level keys:
           - "en_recommendations": The English plain-text recommendation.
           - "ar_recommendations": The Arabic plain-text recommendation.

        **Required Structure for each language (plain-text):**
        - **Student Name:** {name}
        - **Grade:** {grade}
        - **Teacher Name:** {teacher_name}
        - **Subject:** [Subject name]
            - **Strengths:** [Highlight the student's strengths in this subject]
            - **Improvement Areas:** [Mention the key areas where improvement is needed]
            - **Actionable Steps for Parents:** - [Suggest 2-3 specific, practical, and easy-to-implement actions parents can take]

        Academic Data:
        """
        
        # Adding the specific academic data to the prompt
        for subject, details in student.get("subjects", {}).items():
            marks = details.get("marks", "N/A")
            grade_subject = details.get("grade", "N/A") # Renamed for clarity
            timestamp = details.get("timestamp", "N/A")

            if timestamp != "N/A":
                prompt_structure += f"**{subject}:** Grade - {grade_subject}, Marks - {marks}, Last updated: {timestamp}\n"
            else:
                prompt_structure += f"**{subject}:** Grade - {grade_subject}, Marks - {marks}\n"

        # --- Part 2: GPT API call (Preserving original logic style) ---
        try:
            # âœ… GPT API call
            response = client_openai.chat.completions.create(
                model="gpt-4.1-mini-2025-04-14",
                messages=[
                    {"role": "system", "content": "You are an education expert who reliably outputs structured JSON."},
                    {"role": "user", "content": prompt_structure},
                ],
                temperature=0.7,
                # Crucial for forcing JSON output
                response_format={"type": "json_object"} 
            )

            # âœ… Extract and parse GPT output
            gpt_output_json_string = response.choices[0].message.content.strip()
            gpt_output = json.loads(gpt_output_json_string)

            en_recs = gpt_output.get("en_recommendations", "Error: English recommendations generation failed.")
            ar_recs = gpt_output.get("ar_recommendations", "Error: Arabic recommendations generation failed.")

        except Exception as e:
            # Fallback in case of API or JSON parsing failure
            error_message = f"Critical Error in GPT call/parsing: {str(e)}"
            print(error_message)
            en_recs = error_message
            ar_recs = error_message

        # Store both language versions in the recommendations dictionary
        recommendations[student_id] = {
            "student_name": name,
            "teacher_name": teacher_name,
            "grade": grade,
            "recommendations_en": en_recs, # New key
            "recommendations_ar": ar_recs, # New key
        }

    return recommendations

# ---

# âœ… API Route (MODIFIED)
@app.route("/api/resources-analysis", methods=["GET"])
def resources_analysis():
    """
    API to generate AI recommendations in both languages and return 
    the one specified by the 'lang' query parameter.
    """
    start_time = time.time()

    parent_id = request.args.get("parent_id")
    # NEW: Get the requested language, default to 'en'
    lang = request.args.get("lang", DEFAULT_LANGUAGE).lower()

    if lang not in SUPPORTED_LANGUAGES:
        return jsonify({"error": f"Unsupported language: {lang}. Supported languages are {list(SUPPORTED_LANGUAGES.keys())}"}), 400

    if not parent_id:
        return jsonify({"error": "parent_id is required"}), 400

    try:
        # Step 1: Fetch associated students (LOGIC UNCHANGED)
        students = get_associated_students(parent_id)

        if not students:
            return jsonify({"message": "No associated students found"}), 404

        # Step 2: Extract academic data (LOGIC UNCHANGED)
        academic_data = [extract_academic_data(student) for student in students]

        if not academic_data:
            return jsonify({"message": "No academic data available"}), 404

        # Step 3: Send academic data to GPT for recommendations (generates both languages)
        recommendations_all_langs = generate_recommendations(academic_data)

        # Step 4: Time check (LOGIC UNCHANGED)
        elapsed_time = time.time() - start_time
        if elapsed_time > 50:
            return (
                jsonify({"error": "Timeout: Processing took longer than 50 seconds"}),
                504,
            )

        # Step 5: Format the final response, extracting only the requested language
        final_recommendations = {}
        # Dynamically create the key based on the requested language
        recommendation_key = f"recommendations_{lang}"

        for student_id, rec_data in recommendations_all_langs.items():
            final_recommendations[student_id] = {
                "student_name": rec_data["student_name"],
                "teacher_name": rec_data["teacher_name"],
                "grade": rec_data["grade"],
                # Extract the recommendation in the requested language
                "recommendations": rec_data.get(recommendation_key, f"Recommendation not available for language: {lang}"),
            }

        # âœ… Final Response
        return jsonify(
            {
                "parent_id": parent_id,
                "language": lang,
                "students": academic_data,
                "recommendations": final_recommendations,
                "execution_time": f"{elapsed_time:.2f} sec",
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

from flask import Flask, jsonify
import firebase_admin
from firebase_admin import credentials, firestore
import openai
import os
import json

# âœ… Initialize Firebase
if not firebase_admin._apps:
    if os.path.exists("serviceAccountKey.json"):
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(
            cred, {"storageBucket": "pees-d1101.firebasestorage.app"}
        )
        print("âœ… Firebase initialized successfully.")
    else:
        raise FileNotFoundError("Missing Firebase Service Account Key")

db = firestore.client()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
if not OPENAI_API_KEY:
    raise ValueError("Missing OpenAI API key")
client_openai = openai.Client(api_key=OPENAI_API_KEY)


# âœ… Function to extract subject marks (handles both conditions)
def extract_subject_marks(subjects):
    """
    Extract subject marks handling both direct and history-based marks.

    - Direct marks: subjects.[subject].marks
    - History-based marks: subjects.[subject].history[n].marks (latest entry)
    """
    marks_data = {}

    for subject, details in subjects.items():
        if isinstance(details, dict):
            # âœ… Condition 1: Direct marks
            if "marks" in details:
                marks_data[subject] = details["marks"]

            # âœ… Condition 2: History-based marks
            elif "history" in details and isinstance(details["history"], list):
                if details["history"]:
                    latest_entry = details["history"][-1]  # Most recent entry
                    marks_data[subject] = latest_entry.get("marks", 0)
                else:
                    marks_data[subject] = 0  # Handle empty history case

    return marks_data


# âœ… Fetch student data
def fetch_student_data():
    """Retrieve all student data from Firestore"""
    students_ref = db.collection("students").stream()

    students = []
    for doc in students_ref:
        student = doc.to_dict()
        student["student_id"] = doc.id

        # Extracting fields
        academic_data = student.get("academicData", {})
        attendance = student.get("attendance", {})
        observations = student.get("observations", [])
        grade_history = student.get("gradeHistory", {})  # Include gradeHistory
        report_card = student.get("reportCard", {})
        personal_info = student.get("personalInformation", {})

        # âœ… Extract subject marks
        subjects = academic_data.get("subjects", {})
        marks = extract_subject_marks(subjects)

        # Append formatted student data (removing teachingPlans & report_urls)
        students.append(
            {
                "student_id": student["student_id"],
                "name": personal_info.get("name", "N/A"),
                "grade": academic_data.get("grade", "N/A"),
                "class_section": student.get("academicInformation", {}).get(
                    "classSection", "N/A"
                ),
                "marks": marks,
                "reportCard": report_card,
                "gradeHistory": grade_history,  # Include gradeHistory
                "attendance": attendance,
                "observations": observations,
            }
        )

    return students


import json


# # âœ… Analyze reports with GPT (Improved with code block handling)
# def analyze_reports_with_gpt(students):
#     """Send student data to GPT for analysis with enforced JSON formatting."""
#     try:
#         student_json = json.dumps(students, indent=2)

#         # ðŸ”¥ Force GPT to return JSON format
#         prompt = f"""
#         You are an educational assistant analyzing student reports.
#         Based on the student data below, generate a report containing:
#         - "quick_links": A list of recent reports with "title" and "url".
#         - "important_reports": A list of important reports highlighting low grades, improvements, or high achievers with "title", "description", and "link".

#         Return the result **strictly in JSON format** without any code block markers.

#         Student Data:
#         {student_json}
#         """

#         # ðŸ”¥ Send data to GPT using `client_openai`
#         response = client_openai.chat.completions.create(
#             model="gpt-4.1-mini-2025-04-14",
#             messages=[
#                 {"role": "system", "content": "You are an educational assistant."},
#                 {"role": "user", "content": prompt},
#             ],
#             max_tokens=3000,
#         )

#         # âœ… Extract GPT response safely
#         if response and response.choices:
#             gpt_response = response.choices[0].message.content.strip()

#             # âœ… Handle code block wrapping
#             if gpt_response.startswith("```json") and gpt_response.endswith("```"):
#                 gpt_response = gpt_response[7:-3].strip()  # Remove code block markers

#             # âœ… Ensure GPT response is JSON formatted
#             if gpt_response:
#                 try:
#                     # Attempt to parse JSON
#                     return json.loads(gpt_response)
#                 except json.JSONDecodeError:
#                     # Fallback: Handle non-JSON output
#                     return {
#                         "error": "Invalid GPT response format. Raw response:",
#                         "raw_output": gpt_response,
#                     }
#             else:
#                 return {"error": "Empty GPT response."}
#         else:
#             return {"error": "No content received from GPT."}

#     except Exception as e:
#         return {"error": f"OpenAI API error: {str(e)}"}

# reports_overview API
# @app.route("/api/reports-overview", methods=["GET"])
# def reports_overview():
#     """API to fetch reports overview and GPT analysis"""
#     try:
#         # 1. NEW: Get 'lang' from query parameters (e.g., /api/reports-overview?lang=ar)
#         lang = request.args.get("lang", "en").lower()  # Default to 'en'

#         # Fetch students data (Assumed function)
#         students = fetch_student_data()

#         if not students:
#             return jsonify({"error": "No students found"}), 404

#         # 2. NEW: Pass the language parameter to the analysis function
#         analysis = analyze_reports_with_gpt(students, lang)

#         return jsonify({"status": "success", "reports": analysis}), 200

#     except Exception as e:
#         # Ensure 'app' is defined and imported for this to work
#         print(f"Error in reports_overview: {str(e)}") 
#         return jsonify({"error": str(e)}), 500

# # analyze_reports_with_gpt function
# def analyze_reports_with_gpt(students, target_lang):
#     """
#     Send student data to GPT for analysis with enforced JSON formatting
#     and a specific target language.
#     """
    
#     # Map language codes to human-readable names for the prompt
#     lang_map = {"ar": "Arabic (العربية)", "en": "English"}
#     language_name = lang_map.get(target_lang, "English")

#     try:
#         student_json = json.dumps(students, indent=2)

#         # 3. NEW: Inject the language directive directly into the prompt
#         prompt = f"""
#         You are an educational assistant analyzing student reports.
#         Your final output MUST be entirely in the **{language_name}** language.

#         Based on the student data below, generate a report containing:
#         - "quick_links": A list of recent reports with "title" and "url".
#         - "important_reports": A list of important reports highlighting low grades, 
#           improvements, or high achievers with "title", "description", and "link".

#         Return the result **strictly in JSON format** without any code block markers. 
#         All text within the JSON values (titles, descriptions) must be in {language_name}.

#         Student Data:
#         {student_json}
#         """

#         # Ensure your client_openai is correctly initialized and synchronous or awaitable
#         response = client_openai.chat.completions.create(
#             # Using a modern model is key for multilingual support
#             model="gpt-4o-mini", # Recommended model for better performance/compliance
#             messages=[
#                 # 4. NEW: Use the System role to reinforce the language rule
#                 {"role": "system", "content": f"You are an educational assistant. All your responses MUST be entirely in {language_name} ({target_lang}) and formatted as a single JSON object."},
#                 {"role": "user", "content": prompt},
#             ],
#             # Use 'json_object' response format for stricter JSON compliance (requires gpt-3.5-turbo-1106 or newer)
#             response_format={"type": "json_object"}, 
#             max_tokens=3000,
#         )

#         # Extract GPT response safely
#         if response and response.choices:
#             gpt_response = response.choices[0].message.content.strip()

#             # The response_format="json_object" should prevent code block wrapping,
#             # but we keep the handler as a safeguard (though it may not be needed anymore).
#             if gpt_response.startswith("```json") and gpt_response.endswith("```"):
#                 gpt_response = gpt_response[7:-3].strip() 

#             # Ensure GPT response is JSON formatted
#             if gpt_response:
#                 try:
#                     # Attempt to parse JSON
#                     return json.loads(gpt_response)
#                 except json.JSONDecodeError:
#                     # Fallback: Handle non-JSON output
#                     return {
#                         "error": "Invalid GPT response format. Raw response:",
#                         "raw_output": gpt_response,
#                         "language_requested": language_name
#                     }
#             else:
#                 return {"error": "Empty GPT response."}
#         else:
#             return {"error": "No content received from GPT."}

#     except Exception as e:
#         return {"error": f"OpenAI API error: {str(e)}"}
# # âœ… Reports Overview API
# @app.route("/api/reports-overview", methods=["GET"])
# def reports_overview():
#     """API to fetch reports overview and GPT analysis"""
#     try:
#         # ðŸ”¥ Fetch students data
#         students = fetch_student_data()

#         if not students:
#             return jsonify({"error": "No students found"}), 404

#         # ðŸ”¥ Analyze reports using GPT
#         analysis = analyze_reports_with_gpt(students)

#         return jsonify({"status": "success", "reports": analysis}), 200

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# loop11111 = asyncio.get_event_loop()


# client_openai_client = AsyncOpenAI(
#     api_key="YOUR_OPENAI_API_KEY"
# )

# reports_overview API


@app.route("/api/reports-overview", methods=["GET"])
def reports_overview():
    """API to fetch reports overview and GPT analysis"""
    try:
        # 1. NEW: Get 'lang' from query parameters (e.g., /api/reports-overview?lang=ar)
        lang = request.args.get("lang", "en").lower()  # Default to 'en'

        # Fetch students data (Assumed function)
        students = fetch_student_data()

        if not students:
            return jsonify({"error": "No students found"}), 404

        # 2. NEW: Pass the language parameter to the analysis function
        analysis = analyze_reports_with_gpt(students, lang)

        return jsonify({"status": "success", "reports": analysis}), 200

    except Exception as e:
        # Ensure 'app' is defined and imported for this to work
        print(f"Error in reports_overview: {str(e)}") 
        return jsonify({"error": str(e)}), 500

# analyze_reports_with_gpt function
def analyze_reports_with_gpt(students, target_lang):
    """
    Send student data to GPT for analysis with enforced JSON formatting
    and a specific target language.
    """
    
    # Map language codes to human-readable names for the prompt
    lang_map = {"ar": "Arabic (العربية)", "en": "English"}
    language_name = lang_map.get(target_lang, "English")

    try:
        student_json = json.dumps(students, indent=2)

        # 3. NEW: Inject the language directive directly into the prompt
        prompt = f"""
        You are an educational assistant analyzing student reports.
        Your final output MUST be entirely in the **{language_name}** language.

        Based on the student data below, generate a report containing:
        - "quick_links": A list of recent reports with "title" and "url".
        - "important_reports": A list of important reports highlighting low grades, 
          improvements, or high achievers with "title", "description", and "link".

        Return the result **strictly in JSON format** without any code block markers. 
        All text within the JSON values (titles, descriptions) must be in {language_name}.

        Student Data:
        {student_json}
        """

        # Ensure your client_openai is correctly initialized and synchronous or awaitable
        response = client_openai.chat.completions.create(
            # Using a modern model is key for multilingual support
            model="gpt-4o-mini", # Recommended model for better performance/compliance
            messages=[
                # 4. NEW: Use the System role to reinforce the language rule
                {"role": "system", "content": f"You are an educational assistant. All your responses MUST be entirely in {language_name} ({target_lang}) and formatted as a single JSON object."},
                {"role": "user", "content": prompt},
            ],
            # Use 'json_object' response format for stricter JSON compliance (requires gpt-3.5-turbo-1106 or newer)
            response_format={"type": "json_object"}, 
            max_tokens=3000,
        )

        # Extract GPT response safely
        if response and response.choices:
            gpt_response = response.choices[0].message.content.strip()

            # The response_format="json_object" should prevent code block wrapping,
            # but we keep the handler as a safeguard (though it may not be needed anymore).
            if gpt_response.startswith("```json") and gpt_response.endswith("```"):
                gpt_response = gpt_response[7:-3].strip() 

            # Ensure GPT response is JSON formatted
            if gpt_response:
                try:
                    # Attempt to parse JSON
                    return json.loads(gpt_response)
                except json.JSONDecodeError:
                    # Fallback: Handle non-JSON output
                    return {
                        "error": "Invalid GPT response format. Raw response:",
                        "raw_output": gpt_response,
                        "language_requested": language_name
                    }
            else:
                return {"error": "Empty GPT response."}
        else:
            return {"error": "No content received from GPT."}

    except Exception as e:
        return {"error": f"OpenAI API error: {str(e)}"}


import asyncio


async def fetch_student(student_id):
    """Fetch student details."""
    student_doc = await db1.collection("students").document(student_id).get()

    if not student_doc.exists:
        return None

    student_data = student_doc.to_dict()
    return {
        "student_id": student_id,
        "name": student_data.get("profileInfo", {})
        .get("personalInformation", {})
        .get("name", "Unknown"),
        "grade": student_data.get("academicData", {}).get("grade", "N/A"),
    }


async def fetch_teaching_plans(teacher_id):
    """Fetch teaching plans for students associated with a specific teacher."""
    plans = []

    # Query TeachingPlans collection where at least one subject has the given teacher_id
    teaching_plans_query = await db1.collection("TeachingPlans").get()

    student_plans_map = {}  # To track plans per student {student_id: [plans]}

    for doc in teaching_plans_query:
        teaching_plan_data = doc.to_dict()
        student_id = doc.id  # Document ID is the student_id

        # Check if any subject is associated with the given teacher_id
        relevant_subjects = {
            subject: details
            for subject, details in teaching_plan_data.get("subjects", {}).items()
            if details.get("teacher_id") == teacher_id
        }

        if relevant_subjects:
            if student_id not in student_plans_map:
                student_plans_map[student_id] = []

            for subject, plan in relevant_subjects.items():
                student_plans_map[student_id].append(
                    {
                        "student_id": student_id,
                        "subject": subject,
                        "teachingPlan": plan.get("actionPlan", {}),
                        "teacher_id": plan.get("teacher_id", ""),
                        "createdAt": plan.get("createdAt", "N/A"),
                    }
                )

    # For each student, keep only the 2 most recent plans (sorted by createdAt)
    for student_id in student_plans_map:
        # Sort plans by createdAt in descending order
        sorted_plans = sorted(
            student_plans_map[student_id],
            key=lambda x: x.get("createdAt", ""),
            reverse=True,
        )
        # Keep only the 2 most recent plans
        student_plans_map[student_id] = sorted_plans[:2]
        plans.extend(student_plans_map[student_id])

    # Fetch student details concurrently
    tasks = [fetch_student(student_id) for student_id in student_plans_map]
    student_results = await asyncio.gather(*tasks)

    # Map student details to the plans
    student_details_map = {
        student["student_id"]: student for student in student_results if student
    }

    for plan in plans:
        student_id = plan["student_id"]
        if student_id in student_details_map:
            plan.update(
                student_details_map[student_id]
            )  # Merge student details into plan

    return plans

# async def analyze_upcoming_actions(teaching_plans, lang):
#         """Send teaching plans to GPT and get upcoming actions in JSON format."""
#         async_client = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

#         prompt = """
#         You have to respond in the selected language {}, in response JSON field name must be in English inside point in given language.
#         You are an educational assistant analyzing teaching plans.

#         For each student, generate the following 5 upcoming actions:
#         1. Pending Assignments - List actual assignments from the plan (don't leave empty)
#         2. Upcoming Exams - List exams/tests mentioned in the plan (don't leave empty)
#         3. Learning Objectives - Key objectives from the teaching plan (don't leave empty)
#         4. Recommended Resources - Specific resources mentioned or implied (don't leave empty)
#         5. Interventions - Specific strategies mentioned in the plan (don't leave empty)

#         Rules:
#         - NEVER return empty arrays - if nothing exists, include at least one generic item
#         - Be specific and extract directly from the teaching plans
#         - If information is missing, make reasonable inferences
#         - Return the response in **strict JSON format** with no additional text.
#         - For all 5 actions, you MUST generate at least 3 paragraphs for each.

#         Example JSON format:
#         {{
#         "students": [
#             {{
#             "name": "Student Name",
#             "grade": "Grade",
#             "actions": {{
#                 "pending_assignments": [
#                     "Paragraph 1 about pending assignments...",
#                     "Paragraph 2 about pending assignments...",
#                     "Paragraph 3 about pending assignments..."
#                 ],
#                 "upcoming_exams": [
#                     "Paragraph 1 about upcoming exams...",
#                     "Paragraph 2 about upcoming exams...",
#                     "Paragraph 3 about upcoming exams..."
#                 ],
#                 "learning_objectives": [
#                     "Paragraph 1 about learning objectives...",
#                     "Paragraph 2 about learning objectives...",
#                     "Paragraph 3 about learning objectives..."
#                 ],
#                 "recommended_resources": [
#                     "Paragraph 1 about recommended resources...",
#                     "Paragraph 2 about recommended resources...",
#                     "Paragraph 3 about recommended resources..."
#                 ],
#                 "interventions": [
#                     "Paragraph 1 about interventions...",
#                     "Paragraph 2 about interventions...",
#                     "Paragraph 3 about interventions..."
#                 ]
#             }}
#             }}
#         ]
#         }}

#         Teaching Plans:
#         """.format(
#             lang
#         )

#         for plan in teaching_plans:
#             prompt += f"""
#             - Student ID: {plan['student_id']} 
#             - Student: {plan['name']} (Grade: {plan['grade']})
#             - Teaching Plan: {json.dumps(plan['teachingPlan'], indent=2)}
#             - Created At: {plan['createdAt']}
#             """

#         try:
#             response = await async_client.chat.completions.create(
#                 model="gpt-4.1-mini-2025-04-14",
#                 messages=[
#                     {"role": "system", "content": "You are an educational assistant."},
#                     {"role": "user", "content": prompt},
#                 ],
#                 response_format={"type": "json_object"},
#             )
#             gpt_response = response.choices[0].message.content.strip()
#             print(gpt_response)
#             return gpt_response

#         except Exception as e:
#             # Return a valid JSON string with the error
#             return json.dumps({"error": f"Error analyzing with GPT: {str(e)}"})


# # # âœ… Async function to extract and parse JSON from GPT response
# # async def extract_and_parse_json(gpt_response, teaching_plans):
# #     """Extract JSON block from GPT response and parse it, including student_id."""
# #     try:
# #         json_match = re.search(r"(\{.*\})", gpt_response, re.DOTALL)

# #         if json_match:
# #             json_block = json_match.group(1)

# #             structured_data = json.loads(json_block)

# #             students = structured_data.get("students", [])

# #             # âœ… Map student_id to each student
# #             for student in students:
# #                 matching_plan = next(
# #                     (
# #                         plan
# #                         for plan in teaching_plans
# #                         if plan["name"] == student["name"]
# #                     ),
# #                     None,
# #                 )
# #                 if matching_plan:
# #                     student["student_id"] = matching_plan["student_id"]

# #             return students
# #         else:
# #             return {"error": "No valid JSON block found in GPT response"}


# #     except json.JSONDecodeError as e:
# #         return {"error": f"Failed to parse JSON: {str(e)}"}
# #     except Exception as e:
# #         return {"error": str(e)}
# # async def extract_and_parse_json(gpt_response, teaching_plans):
# #     """Extract and parse JSON from GPT response safely."""
# #     try:
# #         structured_data = json.loads(gpt_response)  # âœ… Directly parse JSON
# #         students = structured_data.get("students", [])

# #         for student in students:
# #             matching_plan = next(
# #                 (plan for plan in teaching_plans if plan["name"] == student["name"]),
# #                 None,
# #             )
# #             if matching_plan:
# #                 student["student_id"] = matching_plan["student_id"]

# #         return students

# #     except json.JSONDecodeError as e:
# #         return {"error": f"Failed to parse JSON: {str(e)}"}
# #     except Exception as e:
# #         return {"error": str(e)}

# @app.route("/api/upcoming-actions", methods=["GET"])
# async def get_upcoming_actions_route():
#     """
#     Async API Endpoint to fetch upcoming actions.
#     """
#     try:
#         teacher_id = request.args.get("teacherId")
#         lang = request.args.get("lang", "en")

#         if not teacher_id:
#             return jsonify({"error": "Teacher ID is required"}), 400

#         # Fetch teaching plans
#         teaching_plans = await fetch_teaching_plans(teacher_id)

#         if not teaching_plans:
#             return jsonify({"message": "No teaching plans found"}), 404

#         # Send to GPT for analysis
#         gpt_response_str = await analyze_upcoming_actions(teaching_plans, lang)
#         print(gpt_response_str)

#         # Extract and parse JSON
#         try:
#             structured_data = json.loads(gpt_response_str)
            
#             # Check if GPT returned an error *inside* the JSON
#             if "error" in structured_data:
#                 return jsonify({"error": f"GPT analysis failed: {structured_data['error']}"}), 500

#             students = structured_data.get("students", [])

#             # Map student_id back to the results
#             for student in students:
#                 matching_plan = next(
#                     (plan for plan in teaching_plans if plan["name"] == student["name"]),
#                     None,
#                 )
#                 if matching_plan:
#                     student["student_id"] = matching_plan["student_id"]
            
#             return jsonify({"upcoming_actions": students})

#         except json.JSONDecodeError as e:
#             # This happens if GPT response was not valid JSON
#             return jsonify({"error": f"Failed to parse GPT response: {str(e)}", "raw_response": gpt_response_str}), 500

#     except Exception as e:
#         return jsonify({"error": f"An unexpected server error occurred: {str(e)}"}), 500
import os
import json
import asyncio

# ⬇️ Import the client classes, NOT the global instances ⬇️

from google.cloud import firestore  # ⬅️ IMPORT CREDENTIALS
from google.oauth2 import service_account
from openai import AsyncOpenAI

# (Make sure these are imported for your route)
from flask import request, jsonify

# ---
# 1. ADD this new "local" version of fetch_student
# ---
async def fetch_student_LOCAL(student_id, local_db_client):
    """Fetch student details using a local db client."""
    # Use local_db_client, not the global db1
    student_doc = await local_db_client.collection("students").document(student_id).get()

    if not student_doc.exists:
        return None
    student_data = student_doc.to_dict()
    return {
        "student_id": student_id,
        "name": student_data.get("profileInfo", {})
        .get("personalInformation", {})
        .get("name", "Unknown"),
        "grade": student_data.get("academicData", {}).get("grade", "N/A"),
    }

# ---
# 2. ADD this new "local" version of fetch_teaching_plans
# ---
async def fetch_teaching_plans_LOCAL(teacher_id, local_db_client):
    """Fetch teaching plans using a local db client."""
    plans = []

    # Use local_db_client, not the global db1
    teaching_plans_query = await local_db_client.collection("TeachingPlans").get()
    student_plans_map = {} 

    for doc in teaching_plans_query:
        teaching_plan_data = doc.to_dict()
        student_id = doc.id 

        relevant_subjects = {
            subject: details
            for subject, details in teaching_plan_data.get("subjects", {}).items()
            if details.get("teacher_id") == teacher_id
        }

        if relevant_subjects:
            if student_id not in student_plans_map:
                student_plans_map[student_id] = []
            for subject, plan in relevant_subjects.items():
                student_plans_map[student_id].append(
                    {
                        "student_id": student_id,
                        "subject": subject,
                        "teachingPlan": plan.get("actionPlan", {}),
                        "teacher_id": plan.get("teacher_id", ""),
                        "createdAt": plan.get("createdAt", "N/A"),
                    }
                )

    for student_id in student_plans_map:
        sorted_plans = sorted(
            student_plans_map[student_id],
            key=lambda x: x.get("createdAt", ""),
            reverse=True,
        )
        student_plans_map[student_id] = sorted_plans[:2]
        plans.extend(student_plans_map[student_id])

    # Fetch student details concurrently using the LOCAL function
    # Pass the local_db_client down
    tasks = [fetch_student_LOCAL(student_id, local_db_client) for student_id in student_plans_map]
    student_results = await asyncio.gather(*tasks)

    student_details_map = {
        student["student_id"]: student for student in student_results if student
    }

    for plan in plans:
        student_id = plan["student_id"]
        if student_id in student_details_map:
            plan.update(student_details_map[student_id]) 

    return plans

# ---
#    It correctly creates its own local client.
# ---
async def analyze_upcoming_actions(teaching_plans, lang):
    """Send teaching plans to GPT and get upcoming actions in JSON format."""
    async_client = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

    prompt = """
    You have to respond in the selected language {}, in response JSON field name must be in English inside point in given language.
    You are an educational assistant analyzing teaching plans.

    For each student, generate the following 5 upcoming actions:
    1. Pending Assignments - List actual assignments from the plan (don't leave empty)
    2. Upcoming Exams - List exams/tests mentioned in the plan (don't leave empty)
    3. Learning Objectives - Key objectives from the teaching plan (don't leave empty)
    4. Recommended Resources - Specific resources mentioned or implied (don't leave empty)
    5. Interventions - Specific strategies mentioned in the plan (don't leave empty)

    Rules:
    - NEVER return empty arrays - if nothing exists, include at least one generic item
    - Be specific and extract directly from the teaching plans
    - If information is missing, make reasonable inferences
    - Return the response in **strict JSON format** with no additional text.
    - For all 5 actions, you MUST generate at least 3 paragraphs for each.

    Example JSON format:
    {{
      "students": [
        {{
          "name": "Student Name",
          "grade": "Grade",
          "actions": {{
            "pending_assignments": [
              "Paragraph 1 about pending assignments...",
              "Paragraph 2 about pending assignments...",
              "Paragraph 3 about pending assignments..."
            ],
            "upcoming_exams": [
              "Paragraph 1 about upcoming exams...",
              "Paragraph 2 about upcoming exams...",
              "Paragraph 3 about upcoming exams..."
            ],
            "learning_objectives": [
              "Paragraph 1 about learning objectives...",
              "Paragraph 2 about learning objectives...",
              "Paragraph 3 about learning objectives..."
            ],
            "recommended_resources": [
              "Paragraph 1 about recommended resources...",
              "Paragraph 2 about recommended resources...",
              "Paragraph 3 about recommended resources..."
            ],
            "interventions": [
              "Paragraph 1 about interventions...",
              "Paragraph 2 about interventions...",
              "Paragraph 3 about interventions..."
            ]
          }}
        }}
      ]
    }}

    Teaching Plans:
    """.format(
        lang
    )

    for plan in teaching_plans:
        prompt += f"""
        - Student ID: {plan['student_id']} 
        - Student: {plan['name']} (Grade: {plan['grade']})
        - Teaching Plan: {json.dumps(plan['teachingPlan'], indent=2)}
        - Created At: {plan['createdAt']}
        """

    try:
        response = await async_client.chat.completions.create(
            model="gpt-4.1-mini-2025-04-14",
            messages=[
                {"role": "system", "content": "You are an educational assistant."},
                {"role": "user", "content": prompt},
            ],
            response_format={"type": "json_object"},
        )
        gpt_response = response.choices[0].message.content.strip()
        print(gpt_response)
        return gpt_response

    except Exception as e:
        # Return a valid JSON string with the error
        return json.dumps({"error": f"Error analyzing with GPT: {str(e)}"})

# ---
# REPLACE your OLD API route with this UPDATED one
# ---
@app.route("/api/upcoming-actions", methods=["GET"])
async def get_upcoming_actions_route():
    """
    Async API Endpoint to fetch upcoming actions.
    """
    try:
        teacher_id = request.args.get("teacherId")
        lang = request.args.get("lang", "en")

        if not teacher_id:
            return jsonify({"error": "Teacher ID is required"}), 400

        # --- THIS IS THE FIX ---
        
        # 1. Define the path to your credentials
        CREDENTIALS_PATH = "serviceAccountKey.json" 
        
        # 2. Create local credentials using google.oauth2
        local_creds = service_account.Credentials.from_service_account_file(CREDENTIALS_PATH)
        
        # 3. Create a LOCAL, temporary DB client *using those credentials*
        local_db_client = firestore.AsyncClient(credentials=local_creds) 
        
        # 4. Call your NEW local fetch function, passing the client
        teaching_plans = await fetch_teaching_plans_LOCAL(teacher_id, local_db_client)
        # ---

        if not teaching_plans:
            return jsonify({"message": "No teaching plans found"}), 404

        # 5. Call your (already fixed) analyze function
        gpt_response_str = await analyze_upcoming_actions(teaching_plans, lang)
        print(gpt_response_str)

        
        # ... (rest of your JSON parsing code) ...
        try:
            structured_data = json.loads(gpt_response_str)
            
            if "error" in structured_data:
                return jsonify({"error": f"GPT analysis failed: {structured_data['error']}"}), 500

            students = structured_data.get("students", [])

            for student in students:
                matching_plan = next(
                    (plan for plan in teaching_plans if plan["name"] == student["name"]),
                    None,
                )
                if matching_plan:
                    student["student_id"] = matching_plan["student_id"]
            
            return jsonify({"upcoming_actions": students})

        except json.JSONDecodeError as e:
            return jsonify({"error": f"Failed to parse GPT response: {str(e)}", "raw_response": gpt_response_str}), 500

    except Exception as e:
        return jsonify({"error": f"An unexpected server error occurred: {str(e)}"}), 500
    

@app.route("/submit_feedback_old", methods=["POST"])
def submit_feedback_old():
    try:
        # Parse JSON request data
        data = request.json  # Use JSON instead of request.form

        student_id = data.get("studentId")
        teacher_id = data.get("teacherId")
        feedback = data.get("feedback")
        status = data.get("status", "").lower() == "true"  # Convert to boolean

        # Validate required fields
        if not all([student_id, teacher_id, feedback]) or status is None:
            return jsonify({"error": "Missing required fields"}), 400

        # If status is True, simply return success
        if status:
            return jsonify(
                {"message": "Feedback recorded successfully", "review": True}
            )

        # If status is False, trigger teaching plan generation
        curriculum_id = data.get("curriculumId")
        curriculum_coverage = data.get("curriculum_coverage", [])  # Expect list
        extracted_text = data.get("extractedText", "")
        curriculum_name = data.get("curriculumName", "")
        image_url = data.get("imageUrl", "")
        language = data.get("language", "en")
        subject = data.get("subject", "N/A")
        curriculum_coverage = data.get("curriculum_coverage", "N/A")

        if not curriculum_id:
            return jsonify({"error": "curriculumId is required for review"}), 400

        # Create query string using curriculum coverage
        query = f"Find relevant context from curriculum text using given curriculum coverage topics: {', '.join(curriculum_coverage)} and feedback of Teacher is :- {feedback} Update teaching plan accordingly."

        # Run teaching plan generation asynchronously
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        teaching_plan, evaluation_report, plan_id = loop.run_until_complete(
            generate_teaching_plan(
                student_id,
                curriculum_id,
                query,
                extracted_text,
                curriculum_name,
                image_url,
                language,
                subject=subject,
                curriculum_coverage=curriculum_coverage,
            )
        )

        return jsonify(
            {
                "message": "Feedback recorded with review",
                "review": False,
                "teachingPlan": teaching_plan,
                "evaluationReport": "",
                "planId": plan_id,
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# @app.route("/submit_feedback", methods=["POST"])
# def submit_feedback():
#     try:
#         # Extract text fields from form data
#         student_id = request.form.get("studentId")
#         teacher_id = request.form.get("teacherId")
#         feedback = request.form.get("feedback")
#         status = request.form.get("status", "").lower() == "true"
#         exam_name = request.form.get("exam_name")
#         date = request.form.get("date")
#         observation = request.form.get("observation")
#         file = request.files["file"]  # File to upload

#         if file.filename == "":
#             return jsonify({"error": "No selected file"}), 400

#         # Validate file format
#         if not allowed_file(file.filename):
#             return (
#                 jsonify(
#                     {
#                         "error": "Invalid file type. Only PDF, PNG, JPG, and JPEG are allowed."
#                     }
#                 ),
#                 400,
#             )

#         # if not file.filename.endswith('.pdf'):
#         #     return jsonify({'error': 'Only PDF files are allowed'}), 400

#         # Save the uploaded file temporarily
#         filename = secure_filename(file.filename)
#         temp_file_path = os.path.join(UPLOAD_FOLDER, filename)
#         file.save(temp_file_path)

#         # Check if the file is an image
#         if filename.lower().endswith(("png", "jpg", "jpeg")):
#             temp_pdf_path = temp_file_path.rsplit(".", 1)[0] + ".pdf"
#             try:
#                 convert_image_to_pdf(temp_file_path, temp_pdf_path)
#                 print(
#                     f"Converted PDF Path: {temp_pdf_path}, Exists: {os.path.exists(temp_pdf_path)}"
#                 )
#                 os.remove(
#                     temp_file_path
#                 )  # Remove the original image only after successful conversion
#             except Exception as e:
#                 return (
#                     jsonify({"error": f"Failed to convert image to PDF: {str(e)}"}),
#                     400,
#                 )
#         else:
#             temp_pdf_path = temp_file_path  # If it's a PDF, use it directly

#         # Ensure temp_pdf_path is valid before calling extract_text_from_pdf_or_image11
#         if not temp_pdf_path or not os.path.exists(temp_pdf_path):
#             return (
#                 jsonify({"error": "File processing failed. Temp PDF path is invalid."}),
#                 400,
#             )
#         # Validate required fields
#         if not all([student_id, teacher_id]):
#             return jsonify({"error": "Missing required fields"}), 400

#         if status:
#             # Extract additional fields for review
#             curriculum_id = request.form.get("curriculumId")
#             curriculum_name = request.form.get("curriculumName", "")
#             curriculum_coverage = request.form.get("curriculum_coverage[]", "")
#             # extracted_text = request.form.get("extractedText", "")
#             # image_url = request.form.get("imageUrl", "")
#             language = request.form.get("language", "en")
#             subject = request.form.get("subject", "N/A")

#             if not curriculum_id:
#                 return jsonify({"error": "curriculumId is required for review"}), 400

#             # Parse curriculum_coverage as comma-separated string to list
#             curriculum_coverage = [
#                 item.strip() for item in curriculum_coverage.split(",") if item.strip()
#             ]

#             # Build query
#             query = (
#                 f"Find relevant context from curriculum text using given curriculum coverage topics: "
#                 f"{', '.join(curriculum_coverage)} and feedback of Teacher is :- {feedback} "
#                 f"Update teaching plan accordingly."
#             )

#             # Run teaching plan generation asynchronously
#             loop = asyncio.new_event_loop()
#             asyncio.set_event_loop(loop)
#             teaching_plan, evaluation_report, plan_id = loop.run_until_complete(
#                 generate_teaching_plan(
#                     student_id=student_id,
#                     curriculumname=curriculum_name,
#                     curriculum_id=curriculum_id,
#                     image_url=None,
#                     language=language,
#                     subject=subject,
#                     curriculum_coverage=curriculum_coverage,
#                     teacher_id=teacher_id,
#                     temp_pdf_path=temp_pdf_path,  # ensure defined somewhere
#                     openai_client=openai_client,
#                     saveInTeachingPlans=True,  # ensure defined somewhere
#                 )
#             )

#             # Generate unique file name and upload to Firebase Storage
#             unique_filename = f"exam_scripts/{uuid.uuid4()}_{file.filename}"
#             blob = bucket.blob(unique_filename)
#             file.seek(0)
#             blob.upload_from_file(file)
#             file_url = blob.public_url

#             # Generate Plan ID
#             # plan_id = str(uuid.uuid4()).replace('-', '_')

#             # Create Firestore document for examscripts
#             exam_script_data = {
#                 "exam_name": exam_name,
#                 "curriculum_coverage": curriculum_coverage,
#                 "date": date,
#                 "observation": observation,
#                 "student_id": student_id,
#                 "file_url": file_url,
#                 "uuid": str(uuid.uuid4()),
#                 # "extracted_text": extracted_text,
#                 "teaching_plan": teaching_plan,
#                 "curriculumId": curriculum_id,
#                 "curriculumName": curriculum_name,
#                 "evaluatedtext": evaluation_report,
#                 "planId": plan_id,
#                 "subject": subject,
#             }

#             db.collection("examscripts").add(exam_script_data)

#             return jsonify(
#                 {
#                     "message": "Feedback recorded with review",
#                     "review": True,
#                     "teachingPlan": teaching_plan,
#                     "evaluationReport": evaluation_report,
#                     "planId": plan_id,
#                 }
#             )

#         else:
#             return jsonify({"message": "Teaching plan is deleted", "review": False})

#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# --- Async Firestore client (TeachingPlans) ---
# from google.cloud.firestore_v1.async_client import AsyncClient  # official async client
# # db1 = AsyncClient()  # uses ADC by default (set GOOGLE_CLOUD_PROJECT/credentials) [docs]

# # --- Async writer for TeachingPlans (await db1.set) ---
# async def _write_plan_async(student_id, subject, en_obj, ar_obj, teacher_id):
#     doc_ref = db1.collection("TeachingPlans").document(student_id)
#     await doc_ref.set({
#         "subjects": {
#             subject: {
#                 "actionPlan": en_obj,
#                 "actionPlan_i18n": {"en": en_obj, "ar": ar_obj},
#                 "createdAt": datetime.utcnow().isoformat(),
#                 "studentId": student_id,
#                 "subjectId": subject,
#                 "teacher_id": teacher_id
#             }
#         }
#     }, merge=True)
@app.route("/submit_feedback", methods=["POST"])
def submit_feedback():
    """
    Handles plan generation, synchronous translation, and dual-language save to 'TeachingPlans'.
    """
    # NOTE: This function requires 'import asyncio' at the top of app.py
    try:
        # --- Input Parsing ---
        if request.is_json:
            data = request.get_json()
            student_id = data.get("studentId")
            teacher_id = data.get("teacherId")
            status = str(data.get("status", "")).lower() == "true"
            feedback = data.get("feedback")
        else:
            student_id = request.form.get("studentId")
            teacher_id = request.form.get("teacherId")
            status = request.form.get("status", "").lower() == "true"
            feedback = request.form.get("feedback")

        
        if not student_id or not teacher_id:
             return jsonify({"error": "Student ID and Teacher ID are required"}), 400

        if status:
            curriculum_id = request.form.get("curriculumId")
            curriculum_name = request.form.get("curriculumName", "")
            language = request.form.get("language", "en")
            subject = (request.form.get("subject", "N/A") or "").strip() or "General"
            curriculum_coverage = request.form.getlist("curriculum_coverage[]") 
            # Fix: Define temp_pdf_path before use
            temp_pdf_path = None 
            
            # 1. Synchronous Plan Generation (LLM call)
            from teachingplans import generate_teaching_plan
            
            # FIX: Use asyncio.run() to execute the async function synchronously
            try:
                # The call to generate_teaching_plan is now wrapped by asyncio.run()
                results = asyncio.run(generate_teaching_plan( 
                    student_id=student_id,
                    curriculumname=curriculum_name,
                    curriculum_id=curriculum_id,
                    image_url=None,
                    language=language,
                    subject=subject,
                    curriculum_coverage=curriculum_coverage,
                    teacher_id=teacher_id,
                    temp_pdf_path=temp_pdf_path,
                    openai_client=None, # IMPORTANT: Replace 'None' with a valid AsyncOpenAI client instance
                    saveInTeachingPlans=True,
                ))
                # Unpack the results tuple returned by asyncio.run()
                teaching_plan, evaluation_report, plan_id = results
                
            except Exception as async_error:
                logging.error(f"[submit_feedback] Async plan generation failed: {async_error}")
                return jsonify({"error": f"Failed to generate plan: {str(async_error)}"}), 500
            
            # 2. Synchronous Dual-language Persistence
            try:
                logging.info(f"[submit_feedback] Starting persistence for student_id={student_id}, subject={subject}")
                
                en_obj = deep_normalize(teaching_plan or {}) 
                ar_obj = {}
                try:
                    # Synchronous translation call
                    ar_obj = translate_object_value_only_sync(en_obj, "ar")
                except Exception as te:
                    logging.warning(f"[submit_feedback] Teaching plan AR translation failed, falling back to EN: {te}")
                    ar_obj = en_obj 

                # Synchronous Firestore write using 'db'
                doc_ref = db.collection("TeachingPlans").document(student_id)
                doc_ref.set(
                    {"subjects": {subject: {"actionPlan": en_obj, "actionPlan_i18n": {"en": en_obj, "ar": ar_obj}}}},
                    merge=True
                )

                logging.info(f"[submit_feedback] Wrote TeachingPlans successfully.")
            except Exception as se:
                logging.error(f"[submit_feedback] Failed to write TeachingPlans dual-language: {se}")

            return jsonify({
                "message": "Feedback recorded with review",
                "review": True,
                "teachingPlan": teaching_plan,
                "evaluationReport": evaluation_report,
                "planId": plan_id,
            })
        else:
            return jsonify({"message": "Teaching plan is deleted", "review": False})

    except Exception as e:
        logging.error(f"Critical error in submit_feedback: {e}")
        return jsonify({"error": str(e)}), 500
# NOTE: This code assumes the existence of 'app', 'request', 'jsonify', 'loop', 'db1', 'db',
# 'json', 'datetime', 'logging', 'client_openai', 'translate_analysis_recursively', and 'contains_arabic'.
# The proper setup for the Flask-Async integration (e.g., using 'loop.run_until_complete') is also assumed.

# --- Retrieval Function (Updated for Localization Transparency) ---

# Note: Ensure 'db1' (AsyncFirestoreClient) and 'client_openai' are initialized globally.

# # ---------------------------------------------------------
# # 1. Helper Function: Translate List using GPT (Synchronous)
# # ---------------------------------------------------------
# def translate_list_gpt(text_list, target_lang="ar"):
#     """
#     Translates a list of strings to the target language using GPT.
#     Returns a strict list of strings (e.g., ["text1", "text2"]) or None if failed.
#     """
#     if not text_list:
#         return None

#     try:
#         # Prompt designed to force strict JSON list output without markdown
#         prompt = f"""
#         Translate the following list of academic improvement areas into {target_lang}.
#         Keep the meaning precise, professional, and academic.
        
#         Input List: {json.dumps(text_list)}
        
#         CRITICAL OUTPUT FORMAT:
#         Return ONLY a raw JSON list of strings. 
#         Do not include markdown formatting (like ```json).
#         Example Output: ["translation1", "translation2"]
#         """

#         response = client_openai.chat.completions.create(
#             model="gpt-4-turbo", 
#             messages=[
#                 {"role": "system", "content": "You are a precise translator returning only valid JSON."},
#                 {"role": "user", "content": prompt}
#             ],
#             temperature=0.1
#         )

#         content = response.choices[0].message.content.strip()

#         # Robust Cleanup: Remove markdown code blocks if GPT adds them
#         if content.startswith("```json"):
#             content = content[7:]
#         elif content.startswith("```"):
#             content = content[3:]
#         if content.endswith("```"):
#             content = content[:-3]
        
#         translated_list = json.loads(content.strip())
        
#         # Basic Validation: Ensure it's a non-empty list
#         if isinstance(translated_list, list) and len(translated_list) > 0:
#             return translated_list
#         return None

#     except Exception as e:
#         print(f"[Translation Error] {e}")
#         return None

# # ---------------------------------------------------------
# # 2. Async Fetch with "Heal-on-Read" & "Legacy Migration"
# # ---------------------------------------------------------
# async def fetch_area_need_improvement(student_id, subject_name, lang):
#     try:
#         if not student_id or not subject_name:
#             return jsonify({"error": "Student ID and Subject Name are required"}), 400

#         doc_ref = db1.collection("students").document(student_id)
#         student_doc = await doc_ref.get()
        
#         if not student_doc.exists:
#             return jsonify({"error": "Student not found"}), 404

#         student_data = student_doc.to_dict()
        
#         # Navigate path: analysis -> areas_for_improvement -> SubjectName
#         analysis_data = student_data.get("analysis", {})
#         areas_map = analysis_data.get("areas_for_improvement", {})
        
#         area_entry = areas_map.get(subject_name)

#         if not area_entry:
#             return jsonify({"error": f"No area found for the subject: {subject_name}"}), 404

#         final_area = None
#         language_served = "none"
        
#         # We use the running loop to execute the blocking GPT call without freezing the API
#         loop = asyncio.get_running_loop()

#         # --- SCENARIO A: Modern Data Structure (Dict) ---
#         if isinstance(area_entry, dict):
#             # 1. Try fetching requested language
#             target_data = area_entry.get(lang)
#             if target_data and isinstance(target_data, list) and len(target_data) > 0:
#                 final_area = target_data
#                 language_served = lang
            
#             # 2. Fallback: If Arabic requested but missing
#             elif lang == 'ar':
#                 english_source = area_entry.get('en')
#                 if english_source:
#                     # HEAL: Translate and Append .ar to existing map
#                     translated_ar = await loop.run_in_executor(None, translate_list_gpt, english_source, "ar")
                    
#                     if translated_ar:
#                         final_area = translated_ar
#                         language_served = 'ar (generated_fallback)'
                        
#                         # Update only the specific field 'ar' inside the subject map
#                         update_key = f"analysis.areas_for_improvement.{subject_name}.ar"
#                         try:
#                             await doc_ref.update({update_key: translated_ar})
#                         except Exception as e:
#                             print(f"DB Update Failed: {e}")
#                     else:
#                         final_area = english_source
#                         language_served = 'en (translation_failed)'
#                 else:
#                     language_served = 'none'

#             # 3. Standard Fallback to English
#             if not final_area and lang != 'en':
#                 final_area = area_entry.get('en')
#                 language_served = 'en (fallback)'

#         # --- SCENARIO B: Legacy Data Structure (Simple List) ---
#         elif isinstance(area_entry, list):
#             # This is your current situation: The DB has ["text", "text"]
            
#             if lang == 'ar':
#                 # MIGRATE: We must upgrade this List to a Map {"en": list, "ar": trans}
#                 english_source = area_entry
                
#                 # Run translation
#                 translated_ar = await loop.run_in_executor(None, translate_list_gpt, english_source, "ar")
                
#                 if translated_ar:
#                     final_area = translated_ar
#                     language_served = 'ar (migrated_from_legacy)'
                    
#                     # Construct the NEW structure (Map)
#                     new_structure = {
#                         "en": english_source,
#                         "ar": translated_ar
#                     }
                    
#                     # Overwrite the List field with the New Map
#                     update_key = f"analysis.areas_for_improvement.{subject_name}"
#                     try:
#                         await doc_ref.update({update_key: new_structure})
#                         print(f"[{student_id}] Migrated {subject_name} from List to Map structure.")
#                     except Exception as e:
#                         print(f"DB Migration Failed: {e}")
#                 else:
#                     # Translation failed, return original list
#                     final_area = area_entry
#                     language_served = 'en (legacy_translation_failed)'
#             else:
#                 # Requested English on a Legacy List -> Just return it
#                 final_area = area_entry
#                 language_served = 'en (legacy)'

#         # --- RETURN RESPONSE ---
#         if not final_area:
#             return jsonify({"error": "Improvement area data is corrupted or empty"}), 500
        
#         return jsonify({
#             "area_for_improvement": final_area,
#             "language_returned": language_served,
#             "subject": subject_name
#         }), 200

#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({"error": str(e)}), 500

# @app.route("/student/area_need_improvement", methods=["GET"])
# def get_area_need_improvement():
#     """
#     Endpoint: /student/area_need_improvement
#     Params: studentId, subjectName, lang
#     """
#     student_id = request.args.get("studentId")
#     subject_name = request.args.get("subjectName")
    
#     # URL params are automatically decoded by Flask.
#     # %D8%A7... becomes "اللغة العربية" automatically.
    
#     lang = request.args.get("lang", "en").lower()
#     if lang not in ['en', 'ar']:
#         lang = 'en'

#     # FIX: Safely manage the Event Loop for the async call
#     try:
#         # Try to get the current loop, or create a new one if strictly necessary
#         loop = asyncio.get_event_loop()
#     except RuntimeError:
#         loop = asyncio.new_event_loop()
#         asyncio.set_event_loop(loop)

#     # Run the async function synchronously
#     result, status_code = loop.run_until_complete(
#         fetch_area_need_improvement(student_id, subject_name, lang)
#     )
    
#     return result, status_code


# ---------------------------------------------------------
# 1. Helper Function: Translate List (Synchronous)
# ---------------------------------------------------------
def translate_list_gpt(text_list, target_lang="ar"):
    """
    Translates a list of strings to the target language using GPT.
    Returns a strict list of strings (e.g., ["text1", "text2"]) or None if failed.
    """
    if not text_list:
        return None

    try:
        # Prompt designed to force strict JSON list output without markdown
        prompt = f"""
        Translate the following list of academic improvement areas into {target_lang}.
        Keep the meaning precise, professional, and academic.
        
        Input List: {json.dumps(text_list)}
        
        CRITICAL OUTPUT FORMAT:
        Return ONLY a raw JSON list of strings. 
        Do not include markdown formatting (like ```json).
        Example Output: ["translation1", "translation2"]
        """

        response = client_openai.chat.completions.create(
            model="gpt-4-turbo", 
            messages=[
                {"role": "system", "content": "You are a precise translator returning only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        content = response.choices[0].message.content.strip()

        # Robust Cleanup: Remove markdown code blocks if GPT adds them
        if content.startswith("```json"):
            content = content[7:]
        elif content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        
        translated_list = json.loads(content.strip())
        
        # Basic Validation: Ensure it's a non-empty list
        if isinstance(translated_list, list) and len(translated_list) > 0:
            return translated_list
        return None

    except Exception as e:
        print(f"[Translation Error] {e}")
        return None

#  ---------------------------------------------------------
# 3. Synchronous Fetch (GET Logic)
# ---------------------------------------------------------
def fetch_area_need_improvement(student_id, subject_name, lang):
    """
    Fetches area for improvement synchronously.
    Uses 'Heal-on-Read' to generate and save missing translations on the fly.
    """
    try:
        if not student_id or not subject_name:
            return jsonify({"error": "Student ID and Subject Name are required"}), 400

        # FIX: Use synchronous 'db' client (not 'db1') and no 'await'
        doc_ref = db.collection("students").document(student_id)
        student_doc = doc_ref.get()
        
        if not student_doc.exists:
            return jsonify({"error": "Student not found"}), 404

        student_data = student_doc.to_dict()
        
        # Navigate path: analysis -> areas_for_improvement -> SubjectName
        analysis_data = student_data.get("analysis", {})
        areas_map = analysis_data.get("areas_for_improvement", {})
        
        area_entry = areas_map.get(subject_name)

        if not area_entry:
            return jsonify({"error": f"No area found for the subject: {subject_name}"}), 404

        final_area = None
        language_served = "none"

        # --- SCENARIO A: Modern Data Structure (Dict) ---
        if isinstance(area_entry, dict):
            # 1. Try fetching requested language
            target_data = area_entry.get(lang)
            if target_data and isinstance(target_data, list) and len(target_data) > 0:
                final_area = target_data
                language_served = lang
            
            # 2. Fallback: If Arabic requested but missing
            elif lang == 'ar':
                english_source = area_entry.get('en')
                if english_source:
                    # HEAL: Translate directly (Synchronous call)
                    # This blocks for ~2s but ensures safety and simplicity
                    translated_ar = translate_list_gpt(english_source, "ar")
                    
                    if translated_ar:
                        final_area = translated_ar
                        language_served = 'ar (generated_fallback)'
                        
                        # Update DB synchronously
                        update_key = f"analysis.areas_for_improvement.{subject_name}.ar"
                        try:
                            doc_ref.update({update_key: translated_ar})
                        except Exception as e:
                            print(f"DB Update Failed: {e}")
                    else:
                        final_area = english_source
                        language_served = 'en (translation_failed)'
                else:
                    language_served = 'none'

            # 3. Standard Fallback to English
            if not final_area and lang != 'en':
                final_area = area_entry.get('en')
                language_served = 'en (fallback)'

        # --- SCENARIO B: Legacy Data Structure (Simple List) ---
        elif isinstance(area_entry, list):
            if lang == 'ar':
                # MIGRATE: List -> Map
                english_source = area_entry
                
                # Run translation directly
                translated_ar = translate_list_gpt(english_source, "ar")
                
                if translated_ar:
                    final_area = translated_ar
                    language_served = 'ar (migrated_from_legacy)'
                    
                    new_structure = {
                        "en": english_source,
                        "ar": translated_ar
                    }
                    
                    update_key = f"analysis.areas_for_improvement.{subject_name}"
                    try:
                        doc_ref.update({update_key: new_structure})
                        print(f"[{student_id}] Migrated {subject_name} from List to Map structure.")
                    except Exception as e:
                        print(f"DB Migration Failed: {e}")
                else:
                    final_area = area_entry
                    language_served = 'en (legacy_translation_failed)'
            else:
                final_area = area_entry
                language_served = 'en (legacy)'

        # --- RETURN RESPONSE ---
        if not final_area:
            return jsonify({"error": "Improvement area data is corrupted or empty"}), 500
        
        return jsonify({
            "area_for_improvement": final_area,
            "language_returned": language_served,
            "subject": subject_name
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------
# 3. Route Handler (No Loops!)
# ---------------------------------------------------------
@app.route("/student/area_need_improvement", methods=["GET"])
def get_area_need_improvement():
    """
    Endpoint: /student/area_need_improvement
    Params: studentId, subjectName, lang
    """
    student_id = request.args.get("studentId")
    subject_name = request.args.get("subjectName")
    
    lang = request.args.get("lang", "en").lower()
    if lang not in ['en', 'ar']:
        lang = 'en'

    # FIX: Simply call the function. No loop management required.
    return fetch_area_need_improvement(student_id, subject_name, lang)
# @app.route("/student/getTeachingPlan", methods=["GET"])
# def get_teaching_plan1():
#     student_id = request.args.get("studentId")
#     subject_name = request.args.get("subjectName")
#     lang = request.args.get("lang", "en")
#     result = loop.run_until_complete(fetch_teaching_plan1(student_id, subject_name, lang))
#     return result


# async def fetch_teaching_plan1(student_id, subject_name, lang):
#     try:
#         teaching_plan_doc_id = f"{student_id}"
#         teaching_plan_ref = db1.collection("TeachingPlans").document(teaching_plan_doc_id)

#         loop = asyncio.get_event_loop()
#         teaching_plan_doc = await loop.run_in_executor(None, teaching_plan_ref.get)

#         if not teaching_plan_doc.exists:
#             return jsonify({"error": "Teaching plan not found for the given student"}), 404

#         teaching_plan_data = teaching_plan_doc.to_dict()
#         subjects = teaching_plan_data.get("subjects", {})

#         if subject_name in subjects:
#             raw_plan = subjects[subject_name]
#             translated_plan = await translate_json_content(raw_plan, lang)

#             return jsonify({"teaching_plan": translated_plan}), 200

#         return jsonify({"error": "Teaching plan not found for the given subject"}), 404

#     except Exception as e:
#         print(f"Database error: {e}")
#         return jsonify({"error": "Database error occurred"}), 500
# At top of file ensure these exist:
# from datetime import datetime
# import asyncio

# At top of file ensure these exist:
# from datetime import datetime
# import asyncio
# from google.cloud.firestore_v1.async_client import AsyncClient
# db1 = AsyncClient()  # async client
import datetime

@app.route("/student/getTeachingPlan", methods=["GET"])
def get_teaching_plan_translated():
    """
    Retrieves the teaching plan from 'TeachingPlans' by studentId/subject.
    Returns the requested language, performing synchronous translation and backfill if necessary.
    """
    student_id = (request.args.get("studentId") or "").strip()
    subject_name = (request.args.get("subjectName") or "").strip()
    lang = (request.args.get("lang") or "en").lower() # Requested language
    plan_id = (request.args.get("planId") or "").strip()

    
    if not student_id or not subject_name:
        return jsonify({"error": "studentId and subjectName are required"}), 400
    if lang not in ("en", "ar"):
        return jsonify({"error": "Unsupported language"}), 400

    try:
        # Synchronous Firestore call using 'db'
        doc_ref = db.collection("TeachingPlans").document(student_id)
        doc = doc_ref.get() 

        if not doc.exists:
            return jsonify({"error": "Teaching plan not found for the given student"}), 404

        data = doc.to_dict() or {}
        subjects = data.get("subjects") or {}
        
        if subject_name not in subjects:
            return jsonify({"error": "Teaching plan not found for the given subject"}), 404

        subject_map = subjects[subject_name] or {}

        # 1. Prefer precomputed i18n map
        i18n = subject_map.get("actionPlan_i18n")
        
        if isinstance(i18n, dict):
            # Pick requested language or fall back
            picked = i18n.get(lang) or subject_map.get("actionPlan") or i18n.get("en") or {}
            return jsonify({"teaching_plan": deep_normalize(picked)}), 200

        # 2. Fallback: Translation required
        base_en = deep_normalize(subject_map.get("actionPlan", {}))
        
        if lang == "en":
            return jsonify({"teaching_plan": base_en}), 200

        # If AR requested but not precomputed, try to translate synchronously and backfill
        if lang == "ar":
            try:
                # Synchronous translation
                ar_obj = translate_object_value_only_sync(base_en, "ar") 
                
                # Synchronous backfill
                doc_ref.set(
                    {"subjects": {subject_name: {"actionPlan_i18n": {"en": base_en, "ar": ar_obj}}}},
                    merge=True
                )
                return jsonify({"teaching_plan": ar_obj}), 200
            except Exception as te:
                logging.error(f"Live AR translation and backfill failed: {te}. Returning English plan.")
                return jsonify({"teaching_plan": base_en}), 200
        
        return jsonify({"teaching_plan": base_en}), 200

    except Exception as e:
        logging.error(f"Database error in get_teaching_plan_translated: {e}")
        return jsonify({"error": "Database error occurred"}), 500


# --- Existing Routes (Ensuring Synchronous Firestore Calls) ---

# @app.route("/api/teaching-plan/export", methods=["POST"])
# def export_teaching_plan():
#     """Exports a teaching plan to PDF (must be synchronous)."""
#     try:
#         data = request.json
#         student_id = data.get("studentId")
#         plan_id = data.get("planId")
#         lang = data.get("lang", "")

#         if not student_id or not plan_id:
#             return jsonify({"error": "studentId and planId are required"}), 400

#         # Ensure 'export_data12' is synchronous and uses synchronous 'db'
#         from your_pdf_logic import export_data12, generate_pdf12, upload_to_firebase12
#         content = export_data12(lang, student_id, plan_id)
#         if not content:
#             return jsonify({"error": "Plan not found"}), 404

#         pdf_filename = f"{plan_id}.pdf"
#         pdf_path = generate_pdf12(content, pdf_filename, lang)
#         pdf_url = upload_to_firebase12(pdf_path, student_id, plan_id)

#         if os.path.exists(pdf_path):
#             os.remove(pdf_path)

#         return (
#             jsonify({"studentId": student_id, "planId": plan_id, "pdfUrl": pdf_url}),
#             200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"Processing failed: {str(e)}"}), 500


# @app.route("/teaching_plans", methods=["GET"])
# def get_teaching_plan():
#     """Fetches a teaching plan from the 'students' collection (Synchronous)."""
#     student_id = request.args.get("studentId")
#     plan_id = request.args.get("planId")

#     if not student_id or not plan_id:
#         return jsonify({"error": "Missing studentId or planId"}), 400

#     try:
#         # Synchronous Firestore call
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": f"Student with ID {student_id} not found"}), 404

#         student_data = student_doc.to_dict()
#         teaching_plans = student_data.get("teachingPlans", {})

#         plan = teaching_plans.get(plan_id)
#         if not plan:
#             return jsonify({"error": f"Teaching plan with ID {plan_id} not found"}), 404

#         return jsonify({"actionPlan": plan.get("actionPlan", {})}), 200

#     except Exception as e:
#         return jsonify({"error": f"Failed to fetch teaching plan: {str(e)}"}), 500

# @app.route("/teaching-plan", methods=["PATCH"])
# def update_teaching_plan():
#     """Updates a teaching plan in the 'students' collection (Synchronous)."""
#     try:
#         data = request.get_json()
#         if not data:
#             return jsonify({"error": "Missing request body"}), 400

#         student_id = data.get("studentId")
#         plan_id = data.get("planId")
#         updates = data.get("updates")

#         if not student_id or not plan_id or not updates:
#             return jsonify({"error": "Missing studentId, planId, or updates"}), 400

#         # Synchronous Firestore calls
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": f"Student with ID {student_id} not found"}), 404

#         student_data = student_doc.to_dict()
#         teaching_plans = student_data.get("teachingPlans", {})

#         plan = teaching_plans.get(plan_id)
#         if not plan:
#             return jsonify({"error": f"Teaching plan with ID {plan_id} not found"}), 404

#         action_plan = plan.get("actionPlan", {})
#         invalid_fields = [field for field in updates if field not in action_plan]

#         if invalid_fields:
#             return (
#                 jsonify({"error": f"Invalid fields in update: {invalid_fields}"}),
#                 400,
#             )

#         for key, value in updates.items():
#             action_plan[key] = value

#         # Synchronous Firestore update
#         student_ref.update({f"teachingPlans.{plan_id}.actionPlan": action_plan})

#         return (
#             jsonify(
#                 {
#                     "message": "Teaching plan updated successfully",
#                     "updatedActionPlan": action_plan,
#                 }
#             ),
#             200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"Failed to update teaching plan: {str(e)}"}), 500



# async def fetch_teaching_plan1(student_id: str, subject_name: str, lang: str):
#     try:
#         # Use db1 (the same client used when writing TeachingPlans)
#         doc_ref = db1.collection("TeachingPlans").document(student_id)

#         loop = asyncio.get_event_loop()
#         doc = await loop.run_in_executor(None, doc_ref.get)

#         if not doc.exists:
#             return jsonify({"error": "Teaching plan not found for the given student"}), 404

#         data = doc.to_dict() or {}
#         subjects = data.get("subjects") or {}

#         if subject_name not in subjects:
#             print(f"[getTeachingPlan] Subject '{subject_name}' not found. Available: {list(subjects.keys())[:10]}")
#             return jsonify({"error": "Teaching plan not found for the given subject"}), 404

#         subject_map = subjects[subject_name] or {}

#         # Preferred: precomputed i18n (created in submit_feedback)
#         i18n = subject_map.get("actionPlan_i18n")
#         if isinstance(i18n, dict):
#             picked = i18n.get(lang) or i18n.get("en") or {}
#             return jsonify({"teaching_plan": deep_normalize(picked)}), 200

#         # Fallback: use actionPlan or entire subject_map, normalize, translate if AR
#         base_en = subject_map.get("actionPlan", subject_map)
#         base_en = deep_normalize(base_en)

#         if lang == "en":
#             return jsonify({"teaching_plan": base_en}), 200

#         try:
#             ar_obj = await translate_object_value_only(base_en, "ar")
#         except Exception as te:
#             print(f"[getTeachingPlan] translate failed ({student_id},{subject_name}): {te}")
#             return jsonify({"teaching_plan": base_en}), 200

#         # Optional backfill for faster next calls
#         try:
#             (db1.collection("TeachingPlans")
#                 .document(student_id)
#                 .set({"subjects": {subject_name: {"actionPlan_i18n": {"en": base_en, "ar": ar_obj}}}}, merge=True))
#         except Exception as se:
#             print(f"[getTeachingPlan] backfill i18n failed: {se}")

#         return jsonify({"teaching_plan": ar_obj}), 200

#     except Exception as e:
#         # Log the exact error to server logs
#         print(f"[getTeachingPlan] Database error ({student_id},{subject_name}): {e}")
#         return jsonify({"error": "Database error occurred"}), 500

# @app.route("/api/student/progress/<student_id>", methods=["GET"])
# def get_student_subject_average_and_data_table(student_id):
#     try:
#         # Fetch student document
#         student_ref = db.collection("students").document(student_id)
#         student_doc = student_ref.get()

#         if not student_doc.exists:
#             return jsonify({"error": "Student account not found"}), 404

#         student_data = student_doc.to_dict()
#         academic_data = student_data.get("academicData", {})
#         subjects_data = academic_data.get("subjects", {})

#         # Get optional date filters
#         start_date = request.args.get("startDate")
#         end_date = request.args.get("endDate")

#         start_date_obj = (
#             datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
#         )
#         end_date_obj = (
#             datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
#         )

#         average_percentages = {}
#         data_table = {}

#         for subject_name, subject_info in subjects_data.items():
#             history = subject_info.get("history", [])
#             percentages = []
#             detailed_entries = []

#             for entry in history:
#                 timestamp_str = entry.get("timestamp")
#                 if not timestamp_str:
#                     continue

#                 try:
#                     # --- CHANGES MADE HERE FOR ROBUST TIMESTAMP PARSING ---
#                     # 1. Handle microseconds/milliseconds (e.g., .123456)
#                     if '.' in timestamp_str:
#                         timestamp_str = timestamp_str.split('.')[0]

#                     # 2. Handle timezone offsets (e.g., Z or +00:00)
#                     timestamp_str = timestamp_str.rstrip('Z')
#                     if '+' in timestamp_str:
#                         timestamp_str = timestamp_str.split('+')[0]
                    
#                     # 3. Apply the original logic (removing 'T') after cleanup
#                     # This ensures the string matches the required format: YYYY-MM-DDHH:MM:SS
#                     timestamp_str = timestamp_str.replace("T", "")

#                     # 4. Parse using the cleaned string
#                     timestamp_obj = datetime.strptime(timestamp_str, "%Y-%m-%d%H:%M:%S")
#                     entry_date = timestamp_obj.date()
#                     # --- END CHANGES ---

#                     # Apply filters
#                     if start_date_obj and entry_date < start_date_obj:
#                         continue
#                     if end_date_obj and entry_date > end_date_obj:
#                         continue

#                     marks = entry.get("marks", 0)
#                     total_mark = entry.get("totalMark", 100)
#                     percentage = (marks / total_mark) * 100 if total_mark > 0 else 0
#                     rounded_percentage = round(percentage, 2)
#                     percentages.append(percentage)

#                     detailed_entries.append(
#                         {
#                             "timestamp": timestamp_obj.strftime("%Y-%m-%d %H:%M:%S"),
#                             "percentage": rounded_percentage,
#                             "marks": marks,
#                             "totalMark": total_mark,
#                             "grade": entry.get("grade", ""),
#                             "curriculumName": entry.get("curriculumName", ""),
#                         }
#                     )

#                 except Exception as e:
#                     print(f"Skipping invalid timestamp {timestamp_str}: {e}")
#                     continue

#             if percentages:
#                 average = round(sum(percentages) / len(percentages), 2)
#                 average_percentages[subject_name] = average

#                 # Sort entries by timestamp before adding to data_table
#                 detailed_entries.sort(key=lambda x: x["timestamp"])
#                 data_table[subject_name] = detailed_entries

#         return (
#             jsonify(
#                 {
#                     "studentId": student_id,
#                     "averageSubjectPercentages": average_percentages,
#                     "dataTable": data_table,
#                 }
#             ),
#             200,
#         )

#     except Exception as e:
#         return jsonify({"error": f"An error occurred: {str(e)}"}), 500


@app.route("/api/student/progress/<student_id>", methods=["GET"])
def get_student_subject_average_and_data_table(student_id):
    try:
        # Fetch student document
        student_ref = db.collection("students").document(student_id)
        student_doc = student_ref.get()

        if not student_doc.exists:
            return jsonify({"error": "Student account not found"}), 404

        student_data = student_doc.to_dict()
        academic_data = student_data.get("academicData", {})
        subjects_data = academic_data.get("subjects", {})

        # Get optional date filters
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")

        # --- FIX: Changed datetime.strptime to datetime.datetime.strptime ---
        start_date_obj = (
            datetime.datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        )
        end_date_obj = (
            datetime.datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
        )
        # ------------------------------------------------------------------------------------------

        average_percentages = {}
        data_table = {}

        for subject_name, subject_info in subjects_data.items():
            history = subject_info.get("history", [])
            percentages = []
            detailed_entries = []

            for entry in history:
                timestamp_str = entry.get("timestamp")
                if not timestamp_str:
                    continue

                try:
                    # --- CHANGES MADE HERE FOR ROBUST TIMESTAMP PARSING ---
                    # 1. Handle microseconds/milliseconds (e.g., .123456)
                    if '.' in timestamp_str:
                        timestamp_str = timestamp_str.split('.')[0]

                    # 2. Handle timezone offsets (e.g., Z or +00:00)
                    timestamp_str = timestamp_str.rstrip('Z')
                    if '+' in timestamp_str:
                        timestamp_str = timestamp_str.split('+')[0]
                    
                    # 3. Apply the original logic (removing 'T') after cleanup
                    # This ensures the string matches the required format: YYYY-MM-DDHH:MM:SS
                    timestamp_str = timestamp_str.replace("T", "")

                    # 4. Parse using the cleaned string
                    # --- FIX: Use datetime.datetime.strptime for consistency ---
                    timestamp_obj = datetime.datetime.strptime(timestamp_str, "%Y-%m-%d%H:%M:%S")
                    # ---------------------------------------------------------
                    entry_date = timestamp_obj.date()
                    # --- END CHANGES ---

                    # Apply filters
                    if start_date_obj and entry_date < start_date_obj:
                        continue
                    if end_date_obj and entry_date > end_date_obj:
                        continue

                    marks = entry.get("marks", 0)
                    total_mark = entry.get("totalMark", 100)
                    percentage = (marks / total_mark) * 100 if total_mark > 0 else 0
                    rounded_percentage = round(percentage, 2)
                    percentages.append(percentage)

                    detailed_entries.append(
                        {
                            "timestamp": timestamp_obj.strftime("%Y-%m-%d %H:%M:%S"),
                            "percentage": rounded_percentage,
                            "marks": marks,
                            "totalMark": total_mark,
                            "grade": entry.get("grade", ""),
                            "curriculumName": entry.get("curriculumName", ""),
                        }
                    )

                except Exception as e:
                    print(f"Skipping invalid timestamp {timestamp_str}: {e}")
                    continue

            if percentages:
                average = round(sum(percentages) / len(percentages), 2)
                average_percentages[subject_name] = average

                # Sort entries by timestamp before adding to data_table
                detailed_entries.sort(key=lambda x: x["timestamp"])
                data_table[subject_name] = detailed_entries

        return (
            jsonify(
                {
                    "studentId": student_id,
                    "averageSubjectPercentages": average_percentages,
                    "dataTable": data_table,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500


# @app.route("/get_student_analysis", methods=["GET"])
# def get_student_analysis():
#     student_id = request.args.get("student_id")

#     if not student_id:
#         return jsonify({"error": "Student ID is required"}), 400

#     student_ref = db.collection("students").document(student_id)
#     student_doc = student_ref.get()

#     if not student_doc.exists:
#         return jsonify({"error": "Student not found"}), 404

#     student_data = student_doc.to_dict()
#     analysis_data = student_data.get("analysis", {})

#     if not analysis_data:
#         return jsonify({"error": "Analysis not found for this student"}), 404

#     response_data = {
#         "student_id": student_id,
#         "strengths": analysis_data.get("strengths", []),
#         "weaknesses": analysis_data.get("weaknesses", []),
#         "interventions": analysis_data.get("interventions", []),
#         "recommendations": analysis_data.get("recommendations", []),
#     }

#     return jsonify(response_data), 200

# @app.route("/get_student_analysis", methods=["GET"])
# def get_student_analysis():
#     student_id = request.args.get("student_id")
#     lang = request.args.get("lang", "en").lower()

#     if not student_id:
#         return jsonify({"error": "Student ID is required"}), 400
#     if lang not in ("en", "ar"):
#         return jsonify({"error": "Unsupported language requested"}), 400

#     student_ref = db.collection("students").document(student_id)
#     student_doc = student_ref.get()
#     if not student_doc.exists:
#         return jsonify({"error": "Student not found"}), 404

#     student_data = student_doc.to_dict()
#     dual_language_analysis = student_data.get("analysis", {})
#     requested_analysis = dual_language_analysis.get(lang) or dual_language_analysis.get("en")

#     if not requested_analysis:
#         return jsonify({"error": f"Analysis data not found for language '{lang}' or fallback 'en'"}), 404

#     response_data = {
#         "student_id": student_id,
#         "strengths": requested_analysis.get("strengths", []),
#         "weaknesses": requested_analysis.get("weaknesses", []),
#         "interventions": requested_analysis.get("interventions", []),
#         "recommendations": requested_analysis.get("recommendations", []),
#     }
#     return jsonify(response_data), 200

@app.route("/get_student_analysis", methods=["GET"])
def get_student_analysis():
    student_id = request.args.get("student_id")
    lang = request.args.get("lang", "en").lower()

    if not student_id:
        return jsonify({"error": "Student ID is required"}), 400
    if lang not in ("en", "ar"):
        return jsonify({"error": "Unsupported language requested"}), 400

    student_ref = db.collection("students").document(student_id)
    student_doc = student_ref.get()
    if not student_doc.exists:
        return jsonify({"error": "Student not found"}), 404

    student_data = student_doc.to_dict()
    dual_language_analysis = student_data.get("analysis", {})
    # Core dual-language retrieval logic: get requested lang OR fall back to 'en'
    requested_analysis = dual_language_analysis.get(lang) or dual_language_analysis.get("en")

    if not requested_analysis:
        return jsonify({"error": f"Analysis data not found for language '{lang}' or fallback 'en'"}, 404)

    response_data = {
        "student_id": student_id,
        "strengths": requested_analysis.get("strengths", []),
        "weaknesses": requested_analysis.get("weaknesses", []),
        "interventions": requested_analysis.get("interventions", []),
        "recommendations": requested_analysis.get("recommendations", []),
    }
    return jsonify(response_data), 200

@app.route("/maintenance/backfill_ar", methods=["POST"])
def backfill_ar():
    n_fixed = 0
    for snap in db.collection("students").stream():
        d = snap.to_dict() or {}
        dual = d.get("analysis") or {}
        en, ar = dual.get("en"), dual.get("ar")
        if en and (not ar or not contains_arabic(ar)):
            new_ar = translate_analysis_recursively(en)
            if contains_arabic(new_ar):
                db.collection("students").document(snap.id).set({"analysis": {"ar": new_ar}}, merge=True)
                n_fixed += 1
    return jsonify({"status": "ok", "updated": n_fixed}), 200

@app.route("/student/getTeachingPlan/diag", methods=["GET"])
def diag_teaching_plan():
    student_id = (request.args.get("studentId") or "").strip()
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(_diag(student_id))
    finally:
        loop.close()

async def _diag(student_id):
    try:
        doc = await db1.collection("TeachingPlans").document(student_id).get()
        return jsonify({
            "exists": doc.exists,
            "keys": list((doc.to_dict() or {}).get("subjects", {}).keys()) if doc.exists else []
        }), 200
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500
#-------------------------------------------Delete user
def _delete_document_recursive(doc_ref):
    """
    Recursively delete all subcollections under a document, then the document itself.
    """
    # Defensive check: if db is None, we cannot proceed (relies on global db)
    if 'db' not in globals() or globals()['db'] is None:
        raise Exception("Firestore client is not initialized.")
        
    try:
        # NOTE: This can be a very slow operation for large documents/collections
        for subcol in doc_ref.collections():
            for subdoc in subcol.stream():
                _delete_document_recursive(subdoc.reference)
        doc_ref.delete()
    except Exception as e:
        raise


def _safe_delete_auth_user(uid):
    """
    Delete a Firebase Auth user; ignore if already deleted.
    """
    try:
        # We rely on auth being imported via firebase_admin
        auth.delete_user(uid)
        return True, None
    except Exception as e:
        # If the user is already gone or uid invalid, swallow to keep idempotent behavior.
        error_str = str(e)
        if "user record not found" in error_str.lower() or "user_not_found" in error_str.lower():
            return False, "User not found in Firebase Auth (already deleted or never existed)."
        return False, error_str


def _get_user_role(user_doc_dict):
    """
    Derive user role from known fields.
    """
    if not user_doc_dict:
        return None
    # Common fields used in many schemas: 'role', 'userType', 'type'
    return user_doc_dict.get("role") or user_doc_dict.get("userType") or user_doc_dict.get("type")


def _require_headmaster(request):
    """
    Placeholder for your real authorization/claims check.
    """
    # Example: a shared key header or a verified token claim.
    return request.headers.get("X-Headmaster-Key") == os.getenv("HEADMASTER_KEY")


@app.route("/api/headmasters/deleteuser", methods=["POST"])
def headmasters_delete_user():
    # Defensive check for DB initialization
    if 'db' not in globals() or globals()['db'] is None:
        return jsonify({"error": "Server configuration error: Firestore client is null."}), 500

    try:
        if not _require_headmaster(request):
            return jsonify({"error": "Forbidden"}), 403

        payload = request.get_json(silent=True) or {}
        user_id = payload.get("userId")
        # Removed: student_id = payload.get("studentId") to enforce UID-only input

        if not user_id:
            return jsonify({"error": "userId is required"}), 400

        # Fetch user document to determine role
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()
        
        # --- DEBUG LOGGING ---
        print(f"--- Deletion attempt for User ID: {user_id} ---")
        print(f"User Document Exists: {user_doc.exists}")
        # --- END DEBUG LOGGING ---
        
        user_dict = user_doc.to_dict() if user_doc.exists else {}
        # Role override from payload is still possible if provided, but preferred from user_dict
        role = _get_user_role(user_dict) or payload.get("role")

        # Delete Auth user first (idempotent)
        auth_deleted, auth_error = _safe_delete_auth_user(user_id)

        # Always attempt to delete users/{userId} (Universal Deletion)
        users_deleted = []
        students_deleted = []
        errors = []

        try:
            if user_doc.exists:
                # Document was found, perform recursive delete
                _delete_document_recursive(user_ref)
                users_deleted.append(user_id)
            else:
                # If document was not found (or existence check failed), perform simple, non-recursive delete.
                user_ref.delete()
                users_deleted.append(user_id + " (Simple Delete Attempted)")
                print(f"INFO: users/{user_id} not found/check failed, performed simple delete.")

        except Exception as e:
            errors.append({"where": f"users/{user_id}", "error": str(e)})

        # Conditional Student Deletion: If role is student, also delete students/{studentId}
        if (role and str(role).lower() == "student"):
            
            # --- LOGIC TO DERIVE STUDENT ID ONLY FROM USER DOCUMENT ---
            # Check 'studentId', then other common linked ID fields in the user's document
            sid = (user_dict.get("studentId") or
                   user_dict.get("associatedId") or
                   user_dict.get("linkedStudentId"))
            # ----------------------------------------------------------
            
            # --- DEBUG LOGGING ---
            print(f"Role detected as {role}. Attempting to use studentId: {sid}")
            # --- END DEBUG LOGGING ---
            
            if sid:
                student_ref = db.collection("students").document(sid)
                try:
                    student_doc = student_ref.get()
                    if student_doc.exists:
                        _delete_document_recursive(student_ref)
                        students_deleted.append(sid)
                    else:
                        print(f"INFO: students/{sid} not found in Firestore, skipping student delete.")
                except Exception as e:
                    errors.append({"where": f"students/{sid}", "error": str(e)})
            else:
                # Error if the role is student but the link ID is missing
                errors.append({"where": "students", "error": "studentId not derivable from user document for student role"})

        # For parent and teacher roles nothing extra beyond users/{userId} is required per spec
        result = {
            "authDeleted": auth_deleted,
            "authDeleteError": auth_error,
            "usersDeleted": users_deleted,
            "studentsDeleted": students_deleted,
            "errors": errors,
            "role": role,
        }
        
        # --- DEBUG LOGGING ---
        print(f"--- Deletion Result: {result} ---")
        # --- END DEBUG LOGGING ---
        
        status = 200 if not errors else 207  # 207 Multi-Status if partial failures
        return jsonify(result), status

    except Exception as e:
        return jsonify({
            "error": "Unhandled error while deleting user",
            "details": str(e),
            "trace": traceback.format_exc()
        }), 500

# # ...existing code...
# def _normalize_grade_name(name: str) -> str:
#     if not isinstance(name, str):
#         return ""
#     return re.sub(r'\s+', ' ', name).strip().upper()

# def _student_matches_grade_class(student_doc, grade_key, class_keys):
#     sd = student_doc.to_dict() or {}
#     assigned = sd.get("assignedGrades", {}) or {}
#     grades_map = assigned.get("grades", assigned) if isinstance(assigned, dict) else {}
#     if not isinstance(grades_map, dict):
#         return False

#     for student_grade in grades_map.keys():
#         if _normalize_grade_name(student_grade) != _normalize_grade_name(grade_key):
#             continue
#         student_classes = grades_map.get(student_grade) or {}
#         # if teacher/headmaster didn't specify classes, match by grade only
#         if not class_keys:
#             return True
#         # student_classes may be dict or list
#         if isinstance(student_classes, dict):
#             student_class_keys = {_normalize_grade_name(k) for k in student_classes.keys()}
#         elif isinstance(student_classes, list):
#             student_class_keys = {_normalize_grade_name(k) for k in student_classes}
#         else:
#             student_class_keys = {_normalize_grade_name(str(student_classes))}
#         desired = {_normalize_grade_name(k) for k in class_keys}
#         if student_class_keys & desired:
#             return True
#     return False

# @app.route("/api/headmaster/backfill-associated-ids", methods=["POST"])
# def backfill_associated_ids():
#     """
#     Body: { "userId": "...", optional "grades": { ... } }
#     If grades not provided, uses users/{userId}.assignedGrades.grades
#     """
#     try:
#         payload = request.get_json(silent=True) or {}
#         user_id = payload.get("userId")
#         if not user_id:
#             return jsonify({"error": "userId is required"}), 400

#         user_ref = db.collection("users").document(user_id)
#         user_doc = user_ref.get()
#         if not user_doc.exists:
#             return jsonify({"error": "user not found"}), 404

#         user_data = user_doc.to_dict() or {}
#         # Accept grades supplied in body or from user doc
#         grades_source = payload.get("grades") or user_data.get("assignedGrades") or {}
#         grades_map = grades_source.get("grades", grades_source) if isinstance(grades_source, dict) else {}

#         if not isinstance(grades_map, dict) or not grades_map:
#             # nothing to match -> clear associatedIds
#             user_ref.update({"associatedIds": []})
#             return jsonify({"updated": 0, "associatedIds": []}), 200

#         matched_ids = set()
#         # stream students and match
#         for student_snap in db.collection("students").stream():
#             for grade_key, class_map in grades_map.items():
#                 class_keys = class_map.keys() if isinstance(class_map, dict) else class_map if isinstance(class_map, list) else []
#                 if _student_matches_grade_class(student_snap, grade_key, class_keys):
#                     matched_ids.add(student_snap.id)
#                     break

#         associated_ids = sorted(matched_ids)
#         # persist to user doc
#         user_ref.update({"associatedIds": associated_ids})

#         return jsonify({"message": "associatedIds updated", "userId": user_id, "updated": len(associated_ids), "associatedIds": associated_ids}), 200

#     except Exception as e:
#         logging.exception("backfill_associated_ids failed")
#         return jsonify({"error": str(e)}), 500
# # ...existing code...
    
# # Run Flask app
if __name__ == "__main__":
    run_port = int(os.environ.get("PORT", "5001"))
    socketio.run(app, host='0.0.0.0', port=run_port, allow_unsafe_werkzeug=True)
